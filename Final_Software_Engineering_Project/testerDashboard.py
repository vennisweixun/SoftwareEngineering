import sys
import sqlite3
import os
import datetime
from datetime import datetime, timedelta, date
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *

try:
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure

    MATPLOTLIB_AVAILABLE = True
except ImportError:
    print("Warning: matplotlib not available. Charts will be disabled.")
    MATPLOTLIB_AVAILABLE = False
import pandas as pd
from fpdf import FPDF
import platform
from openpyxl.utils import get_column_letter
import yagmail
from PIL import Image
from PyQt5.QtWidgets import QApplication

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
db_path = os.path.join(BASE_DIR, "testing_system.db")

print(f"Database full path: {db_path}")
print(f"✓ Using standardized database schema compatible with databaseSetup1.py")


# === Chat Database Functions ===
def create_chat_tables():
    """Create necessary chat tables in testing_system.db"""
    try:
        with sqlite3.connect(db_path, timeout=30.0) as conn:
            cursor = conn.cursor()

            # Create message table (matching databaseSetup1.py)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS message (
                    message_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    role_id INTEGER,
                    username TEXT,
                    role TEXT,
                    message TEXT,
                    message_type TEXT,
                    read_status INTEGER DEFAULT 0,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(user_id) REFERENCES users(user_id),
                    FOREIGN KEY(role_id) REFERENCES roles(role_id)
                )
            """)

            # Create message_notifications table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS message_notifications (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    message_id INTEGER,
                    tagged_user_id INTEGER,
                    tagged_username TEXT,
                    is_read INTEGER DEFAULT 0,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(message_id) REFERENCES message(message_id)
                )
            """)

            # Create private_messages table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS private_messages (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sender_user_id INTEGER,
                    sender_username TEXT,
                    receiver_username TEXT,
                    sender_role TEXT,
                    message_text TEXT,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """)

            conn.commit()
        print("✓ Chat tables created successfully in testing_system.db")
        return True
    except Exception as e:
        print(f"Error creating chat tables: {e}")
        return False


def get_all_messages():
    """Get all public messages from the message table"""
    try:
        with sqlite3.connect(db_path, timeout=30.0) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT username, role, message, timestamp, message_id
                FROM message
                WHERE role IN ('superadmin', 'admin', 'tester', 'owner') 
                AND message NOT LIKE '[ACTIVITY]%'
                AND (message_type = 'public' OR message_type IS NULL)
                ORDER BY timestamp ASC
            """)
            messages = cursor.fetchall()
        return messages
    except Exception as e:
        print(f"Error getting messages: {e}")
        return []


def insert_message(user_id, username, role, role_id, message_text, message_type='public'):
    """Insert a new message into the message table"""
    try:
        with sqlite3.connect(db_path, timeout=30.0) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO message (user_id, role_id, username, role, message, message_type, timestamp)
                VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            """, (user_id, role_id, username, role, message_text, message_type))
            message_id = cursor.lastrowid
            conn.commit()
        return message_id
    except Exception as e:
        print(f"Error inserting message: {e}")
        return None


def get_all_chat_users(current_username):
    """Get all users for private messaging - testers can chat with all roles"""
    try:
        with sqlite3.connect(db_path, timeout=30.0) as conn:
            cursor = conn.cursor()

            # Get users from testing_system.db
            cursor.execute("""
                SELECT u.username, r.role_name, COALESCE(u.fullname, u.username) as fullname 
                FROM users u 
                LEFT JOIN roles r ON u.role = r.role_id 
                WHERE u.username != ? 
                ORDER BY r.role_name, u.username
            """, (current_username,))
            users = cursor.fetchall()
        return users
    except Exception as e:
        print(f"Error getting chat users: {e}")
        return []


def insert_private_message(sender_user_id, sender_username, receiver_username, sender_role, message_text):
    """Insert a private message into message table and create notification"""
    try:
        with sqlite3.connect(db_path, timeout=30.0) as conn:
            cursor = conn.cursor()

            # Get sender's role_id
            cursor.execute("SELECT role FROM users WHERE user_id = ?", (sender_user_id,))
            role_result = cursor.fetchone()
            sender_role_id = role_result[0] if role_result else 4  # Default to tester role

            # Insert into message table with message_type as 'private'
            cursor.execute("""
                INSERT INTO message (user_id, role_id, username, role, message, message_type, timestamp)
                VALUES (?, ?, ?, ?, ?, 'private', CURRENT_TIMESTAMP)
            """, (
            sender_user_id, sender_role_id, sender_username, sender_role, f"@{receiver_username}: {message_text}"))

            message_id = cursor.lastrowid

            # Get receiver's user_id
            cursor.execute("SELECT user_id FROM users WHERE username = ?", (receiver_username,))
            receiver_result = cursor.fetchone()
            receiver_user_id = receiver_result[0] if receiver_result else None

            # Create notification for the receiver
            if receiver_user_id:
                cursor.execute("""
                    INSERT INTO message_notifications (message_id, tagged_user_id, tagged_username, created_at)
                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                """, (message_id, receiver_user_id, receiver_username))

            # Also keep the old private_messages table for backward compatibility
            cursor.execute("""
                INSERT INTO private_messages 
                (sender_user_id, sender_username, receiver_username, sender_role, message_text)
                VALUES (?, ?, ?, ?, ?)
            """, (sender_user_id, sender_username, receiver_username, sender_role, message_text))

            conn.commit()
        return True
    except Exception as e:
        print(f"Error inserting private message: {e}")
        return False


def get_private_messages(username1, username2):
    """Get private messages between two users from private_messages table directly"""
    try:
        with sqlite3.connect(db_path, timeout=30.0) as conn:
            cursor = conn.cursor()

            # Query the private_messages table directly as specified by user
            cursor.execute("""
                SELECT sender_username, receiver_username, message_text, timestamp, sender_role
                FROM private_messages 
                WHERE (sender_username = ? AND receiver_username = ?) 
                   OR (sender_username = ? AND receiver_username = ?)
                ORDER BY timestamp ASC
                LIMIT 50
            """, (username1, username2, username2, username1))

            messages = cursor.fetchall()

            # Return in the expected format: (sender, receiver, message_text, timestamp, role)
            return messages if messages else []

    except Exception as e:
        print(f"Error getting private messages from private_messages table: {e}")
        return []


def format_malaysia_time(timestamp_str=None):
    """Format timestamp to Malaysia time"""
    try:
        if timestamp_str:
            if 'T' in timestamp_str or '-' in timestamp_str:
                dt = datetime.strptime(timestamp_str.replace('Z', '+00:00'), "%Y-%m-%d %H:%M:%S%z")
            else:
                try:
                    dt = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
                except:
                    dt = datetime.now()
        else:
            dt = datetime.now()
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def mark_messages_as_read(user_id, other_username):
    """Mark private messages as read when user opens a chat - simplified for private_messages table"""
    try:
        # For simplified implementation with private_messages table
        # Since we're now using the private_messages table directly,
        # we can implement read status later if needed
        print(f"Chat opened between user {user_id} and {other_username}")
        return True
    except Exception as e:
        print(f"Error marking messages as read: {e}")
        return False


def get_unread_message_count(user_id):
    """Get count of unread messages for a user - simplified for private_messages table"""
    try:
        # For simplified private_messages table, return 0 for now
        # This can be enhanced later with read status tracking if needed
        return 0
    except Exception as e:
        print(f"Error getting unread message count: {e}")
        return 0


def get_unread_messages_by_user(user_id):
    """Get unread message count grouped by sender - simplified for private_messages table"""
    try:
        # For simplified private_messages table, return empty dict for now
        # This can be enhanced later with read status tracking if needed
        return {}
    except Exception as e:
        print(f"Error getting unread messages by user: {e}")
        return {}


def create_database_tables():
    """Create all database tables with proper schema - using standardized schema from databaseSetup1.py"""
    try:
        with sqlite3.connect(db_path, timeout=30.0) as conn:
            cursor = conn.cursor()

            # Enable foreign keys
            cursor.execute("PRAGMA foreign_keys = ON")

            # Note: Tables should already exist from databaseSetup1.py initialization
            # This function now just ensures they exist, using the standardized schema

            # Create roles table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS roles (
                    role_id INTEGER PRIMARY KEY,
                    role_name TEXT
                )
            """)

            # Create branches table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS branches (
                    branch_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    branch_name TEXT
                )
            """)

            # Create racklocations table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS racklocations (
                    rack_location_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    rack_location_name TEXT
                )
            """)

            # Create assignment_status table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS assignment_status (
                    status_id INTEGER PRIMARY KEY,
                    status_name TEXT
                )
            """)

            # Create progress_status table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS progress_status (
                    status_id INTEGER PRIMARY KEY,
                    status_name TEXT
                )
            """)

            # Create test_results table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS test_results (
                    result_id INTEGER PRIMARY KEY,
                    result_name TEXT
                )
            """)

            # Create standardized users table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    user_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT,
                    password INTEGER CHECK(LENGTH(password) >= 4),
                    role INTEGER,
                    fullname TEXT,
                    email TEXT,
                    phone_no TEXT,
                    branch_id INTEGER,
                    FOREIGN KEY(branch_id) REFERENCES branches(branch_id),
                    FOREIGN KEY(role) REFERENCES roles(role_id)
                )
            """)

            # Create standardized products table with owner_id and tester_id separation
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS products (
                    product_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    owner_id INTEGER,
                    tester_id INTEGER,
                    product_name TEXT,
                    product_desc TEXT,
                    product_image TEXT,
                    arrival_date DATETIME,
                    branch_id INTEGER,
                    batch TEXT,
                    rack_location_id INTEGER,
                    sku INTEGER CHECK(sku > 0),
                    manufacture_date DATETIME,
                    expired_date DATETIME,
                    barcode TEXT,
                    barcode_image TEXT,
                    excel_name TEXT,
                    status TEXT DEFAULT 'pending',
                    location TEXT,
                    FOREIGN KEY(owner_id) REFERENCES users(user_id),
                    FOREIGN KEY(tester_id) REFERENCES users(user_id),
                    FOREIGN KEY(branch_id) REFERENCES branches(branch_id),
                    FOREIGN KEY(rack_location_id) REFERENCES racklocations(rack_location_id)
                )
            """)

            # Create standardized testing table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS testing (
                    test_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    product_id INTEGER,
                    assignment_status_id INTEGER,
                    test_start DATETIME,
                    test_end DATETIME,
                    progress_status_id INTEGER,
                    test_result_id INTEGER,
                    test_image TEXT,
                    FOREIGN KEY(user_id) REFERENCES users(user_id),
                    FOREIGN KEY(product_id) REFERENCES products(product_id),
                    FOREIGN KEY(assignment_status_id) REFERENCES assignment_status(status_id),
                    FOREIGN KEY(progress_status_id) REFERENCES progress_status(status_id),
                    FOREIGN KEY(test_result_id) REFERENCES test_results(result_id)
                )
            """)

            # Create standardized product_tester_assignments table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS product_tester_assignments (
                    assignment_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id INTEGER,
                    tester_id INTEGER,
                    assigned_by INTEGER,
                    assigned_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(product_id) REFERENCES products(product_id),
                    FOREIGN KEY(tester_id) REFERENCES users(user_id),
                    FOREIGN KEY(assigned_by) REFERENCES users(user_id)
                )
            """)

            # Create notifications table (renamed from the old one)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS notifications (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    message TEXT,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                    read INTEGER DEFAULT 0,
                    FOREIGN KEY(user_id) REFERENCES users(user_id)
                )
            """)

            conn.commit()

            # Ensure lookup data exists
            insert_lookup_data(cursor)
            conn.commit()

            print("Database tables verified/created successfully with standardized schema.")

            # Create chat tables
            create_chat_tables()
    except sqlite3.Error as e:
        print(f"Database error occurred: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")


def insert_lookup_data(cursor):
    """Insert essential lookup data for status tables and roles"""

    # Insert roles (as specified by standardized schema)
    roles = [
        (1, 'superadmin'),
        (2, 'admin'),
        (3, 'owner'),
        (4, 'tester')
    ]
    cursor.executemany("INSERT OR IGNORE INTO roles (role_id, role_name) VALUES (?, ?)", roles)

    # Insert assignment status
    assignment_statuses = [
        (0, 'Pending'),
        (1, 'Assigned'),
        (2, 'Cancelled')
    ]
    cursor.executemany("INSERT OR IGNORE INTO assignment_status (status_id, status_name) VALUES (?, ?)",
                       assignment_statuses)

    # Insert progress status
    progress_statuses = [
        (0, 'On Hold'),
        (1, 'In Progress'),
        (2, 'Testing Complete'),
        (3, 'Sample Expired')
    ]
    cursor.executemany("INSERT OR IGNORE INTO progress_status (status_id, status_name) VALUES (?, ?)",
                       progress_statuses)

    # Insert test results
    test_results = [
        (0, 'Rejected'),
        (1, 'Under Review'),
        (2, 'Approved')
    ]
    cursor.executemany("INSERT OR IGNORE INTO test_results (result_id, result_name) VALUES (?, ?)", test_results)


class DateEdit(QDateEdit):
    """Custom DateEdit widget to match original functionality"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setDate(QDate.currentDate())
        self.setCalendarPopup(True)
        self.setDisplayFormat("yyyy-MM-dd")


# --- Tester Chat Widget ---
class TesterChatWidget(QWidget):
    def __init__(self, user_info=None, parent=None):
        super().__init__(parent)
        self.user_info = user_info or {}
        self.active_pm_user = None
        self.active_pm_role = None
        self.selected_user_item = None
        self.unread_counts = {}
        self.last_message_count = {}  # Track message counts to detect new messages
        self.auto_scroll_enabled = True  # Auto-scroll to bottom for new messages

        # Configurable refresh rates (in milliseconds)
        self.GLOBAL_CHAT_REFRESH_RATE = 3000  # 3 seconds for global chat
        self.PRIVATE_CHAT_REFRESH_RATE = 2000  # 2 seconds for private chat (faster)
        self.USER_LIST_REFRESH_RATE = 5000  # 5 seconds for user list

        # Define role colors (purple theme for tester)
        self.role_colors = {
            'superadmin': '#FF9800',  # Orange
            'admin': '#4CAF50',  # Green
            'owner': '#2196F3',  # Blue
            'tester': '#9C27B0'  # Purple
        }

        # Initialize chat database tables
        create_chat_tables()

        self.setupUI()
        self.load_messages()
        self.load_available_users()
        self.load_unread_counts()

        # Set up multiple timers for different refresh rates
        self.setup_refresh_timers()

    def setup_refresh_timers(self):
        """Set up multiple timers with different refresh rates for optimal performance"""
        # Global chat refresh timer
        self.global_chat_timer = QTimer()
        self.global_chat_timer.timeout.connect(self.refresh_global_chat)
        self.global_chat_timer.start(self.GLOBAL_CHAT_REFRESH_RATE)

        # Private chat refresh timer (faster for active conversations)
        self.private_chat_timer = QTimer()
        self.private_chat_timer.timeout.connect(self.refresh_private_chat)
        self.private_chat_timer.start(self.PRIVATE_CHAT_REFRESH_RATE)

        # User list refresh timer (slower, as it changes less frequently)
        self.user_list_timer = QTimer()
        self.user_list_timer.timeout.connect(self.refresh_user_list)
        self.user_list_timer.start(self.USER_LIST_REFRESH_RATE)

        print(f"✓ Chat refresh timers initialized:")
        print(f"  - Global chat: {self.GLOBAL_CHAT_REFRESH_RATE}ms")
        print(f"  - Private chat: {self.PRIVATE_CHAT_REFRESH_RATE}ms")
        print(f"  - User list: {self.USER_LIST_REFRESH_RATE}ms")

    def refresh_global_chat(self):
        """Refresh only global chat messages"""
        try:
            self.update_refresh_status("🔄 Refreshing...")
            self.load_messages()
            self.update_refresh_status("🟢 Connected")
        except Exception as e:
            print(f"Error refreshing global chat: {e}")
            self.update_refresh_status("🔴 Error")

    def refresh_private_chat(self):
        """Refresh private messages if there's an active conversation"""
        try:
            if self.active_pm_user:
                # Store current scroll position
                scrollbar = self.pm_display.verticalScrollBar()
                was_at_bottom = scrollbar.value() >= (scrollbar.maximum() - 10)  # Allow small margin

                # Check if there are new messages before refreshing
                current_username = self.user_info.get('username', '')
                messages = get_private_messages(current_username, self.active_pm_user)
                current_message_count = len(messages) if messages else 0
                previous_count = self.last_message_count.get(self.active_pm_user, 0)

                # Only refresh if there are new messages or it's the first load
                if current_message_count != previous_count:
                    self.last_message_count[self.active_pm_user] = current_message_count
                    self.load_private_messages_inline(preserve_scroll=not was_at_bottom)

                    # Show notification if new messages arrived while user was scrolled up
                    if current_message_count > previous_count and not was_at_bottom:
                        self.show_new_message_indicator()

                    # Brief status update for new messages
                    if current_message_count > previous_count:
                        self.update_refresh_status("💬 New message")
                        QTimer.singleShot(2000, lambda: self.update_refresh_status("🟢 Connected"))
        except Exception as e:
            print(f"Error refreshing private chat: {e}")
            self.update_refresh_status("🔴 Error")

    def refresh_user_list(self):
        """Refresh user list and unread counts"""
        try:
            self.load_unread_counts()
            # Only refresh user list display if unread counts changed
            if hasattr(self, 'all_users'):
                self.populate_users_list(self.all_users)
        except Exception as e:
            print(f"Error refreshing user list: {e}")

    def show_new_message_indicator(self):
        """Show a subtle indicator when new messages arrive while user is scrolled up"""
        if hasattr(self, 'new_message_indicator'):
            return  # Already showing

        self.new_message_indicator = QPushButton("💬 New messages below - Click to scroll down")
        self.new_message_indicator.setStyleSheet("""
            QPushButton {
                background-color: #9C27B0;
                color: white;
                border: none;
                border-radius: 15px;
                padding: 8px 16px;
                font-size: 12px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #7B1FA2;
            }
        """)
        self.new_message_indicator.clicked.connect(self.scroll_to_bottom_and_hide_indicator)

        # Add to layout temporarily
        pm_layout = self.pm_display.parent().layout()
        pm_layout.insertWidget(pm_layout.indexOf(self.pm_display) + 1, self.new_message_indicator)

    def scroll_to_bottom_and_hide_indicator(self):
        """Scroll to bottom and hide the new message indicator"""
        scrollbar = self.pm_display.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

        if hasattr(self, 'new_message_indicator'):
            self.new_message_indicator.deleteLater()
            delattr(self, 'new_message_indicator')

    def setupUI(self):
        # Main layout with light purple theme
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(15)

        # Header with refresh status indicator
        header_layout = QHBoxLayout()
        title = QLabel("Chat / Announcement")
        title.setFont(QFont("Arial", 24, QFont.Bold))
        title.setStyleSheet("color: #5D3A00; background-color: transparent;")
        title.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(title)

        # Add refresh status indicator
        self.refresh_status = QLabel("🟢 Connected")
        self.refresh_status.setStyleSheet("""
            QLabel {
                color: #28a745;
                font-size: 12px;
                font-weight: 600;
                padding: 4px 8px;
                background-color: rgba(40, 167, 69, 0.1);
                border: 1px solid #28a745;
                border-radius: 12px;
            }
        """)
        header_layout.addWidget(self.refresh_status)

        main_layout.addLayout(header_layout)

        # Content layout - horizontal split
        content_layout = QHBoxLayout()
        content_layout.setSpacing(15)

        # Left side - Chat area (70% width)
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)

        # Global chat display
        chat_group = QGroupBox("Global Chat Room")
        chat_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #9C27B0;
                border-radius: 12px;
                margin: 8px 0px;
                padding-top: 20px;
                background: rgba(255, 255, 255, 0.9);
                font-size: 14px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 16px;
                padding: 8px 16px;
                color: white;
                background: #9C27B0;
                border-radius: 6px;
            }
        """)

        chat_layout = QVBoxLayout()
        self.chat_display = QTextEdit()
        self.chat_display.setReadOnly(True)
        self.chat_display.setMinimumHeight(300)
        self.chat_display.setStyleSheet("""
            QTextEdit {
                border: 1px solid #ddd;
                border-radius: 8px;
                background-color: rgba(255, 255, 255, 0.95);
                font-family: 'Segoe UI', Arial, sans-serif;
                font-size: 14px;
                padding: 12px;
            }
        """)
        chat_layout.addWidget(self.chat_display)

        # Add read-only notice for testers
        readonly_notice = QLabel(
            "📖 Global chat is view-only for testers. Use private messages to communicate with team members.")
        readonly_notice.setStyleSheet("""
        QLabel {
        color: #7B1FA2;
        font-size: 12px;
        font-style: italic;
        padding: 8px;
        background-color: #F3E5F5;
        border: 1px solid #CE93D8;
        border-radius: 6px;
        margin: 5px 0px;
        }
        """)
        readonly_notice.setWordWrap(True)
        chat_layout.addWidget(readonly_notice)

        chat_group.setLayout(chat_layout)
        left_layout.addWidget(chat_group)

        content_layout.addWidget(left_widget, 7)  # 70% width

        # Right side - Users and Private Chat (30% width)
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)

        # Users list
        self.users_group = QGroupBox("All Users")
        self.users_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #9C27B0;
                border-radius: 12px;
                margin: 8px 0px;
                padding-top: 20px;
                background: rgba(255, 255, 255, 0.9);
                font-size: 14px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 16px;
                padding: 8px 16px;
                color: white;
                background: #9C27B0;
                border-radius: 6px;
            }
        """)

        users_layout = QVBoxLayout()

        # Search bar
        search_layout = QHBoxLayout()
        search_label = QLabel("🔍")
        self.user_search = QLineEdit()
        self.user_search.setPlaceholderText("Search users...")
        self.user_search.textChanged.connect(self.search_users)
        self.user_search.setStyleSheet("""
            QLineEdit {
                font-size: 12px;
                padding: 8px;
                border: 1px solid #ccc;
                border-radius: 15px;
                background-color: white;
            }
        """)
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.user_search)
        users_layout.addLayout(search_layout)

        # Users list
        self.users_list = QListWidget()
        self.users_list.itemClicked.connect(self.start_private_chat)
        self.users_list.setMaximumHeight(200)
        self.users_list.setStyleSheet("""
            QListWidget {
                border: 1px solid #ddd;
                border-radius: 8px;
                background-color: white;
                font-size: 12px;
                selection-background-color: #9C27B0;
                outline: none;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid #f0f0f0;
                border-radius: 4px;
                margin: 1px;
            }
            QListWidget::item:selected {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                           stop:0 #9C27B0, stop:1 #7B1FA2);
                color: white;
                border: 2px solid #4A148C;
                font-weight: bold;
            }
            QListWidget::item:hover {
                background: #E1BEE7;
                border: 1px solid #BA68C8;
                border-radius: 4px;
            }
            QListWidget::item:selected:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                           stop:0 #7B1FA2, stop:1 #6A1B9A);
                color: white;
                border: 2px solid #4A148C;
            }
        """)
        users_layout.addWidget(self.users_list)

        pm_info = QLabel("💬 Chat with Team Members")
        pm_info.setStyleSheet("color: #666; font-size: 11px; font-style: italic;")
        users_layout.addWidget(pm_info)

        self.users_group.setLayout(users_layout)
        right_layout.addWidget(self.users_group)

        # Private message area
        pm_group = QGroupBox("Private Messages")
        pm_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #9C27B0;
                border-radius: 12px;
                margin: 8px 0px;
                padding-top: 20px;
                background: rgba(255, 255, 255, 0.9);
                font-size: 14px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 16px;
                padding: 8px 16px;
                color: white;
                background: #9C27B0;
                border-radius: 6px;
            }
        """)

        pm_layout = QVBoxLayout()

        # Active conversation indicator
        self.active_pm_label = QLabel("Select user to start messaging")
        self.active_pm_label.setStyleSheet("""
            QLabel {
                color: #888; 
                font-size: 10px; 
                padding: 4px; 
                border-bottom: 1px solid #eee;
                background-color: #f8f9fa;
                border-radius: 4px;
            }
        """)
        pm_layout.addWidget(self.active_pm_label)

        # Private message display
        self.pm_display = QTextEdit()
        self.pm_display.setReadOnly(True)
        self.pm_display.setMinimumHeight(180)
        self.pm_display.setMaximumHeight(220)
        self.pm_display.setPlaceholderText("Private messages will appear here...")
        self.pm_display.setStyleSheet("""
            QTextEdit {
                border: 1px solid #ddd;
                border-radius: 8px;
                background-color: white;
                font-size: 12px;
                padding: 8px;
            }
        """)
        pm_layout.addWidget(self.pm_display)

        # Private message input
        self.pm_input = QTextEdit()
        self.pm_input.setPlaceholderText("Type private message...")
        self.pm_input.setEnabled(False)
        self.pm_input.setMinimumHeight(50)
        self.pm_input.setMaximumHeight(70)
        self.pm_input.setStyleSheet("""
            QTextEdit {
                border: 1px solid #ccc;
                border-radius: 8px;
                padding: 8px;
                font-size: 12px;
                background-color: #f8f9fa;
            }
            QTextEdit:focus {
                border-color: #9C27B0;
                background-color: white;
            }
        """)
        pm_layout.addWidget(self.pm_input)

        # Send button
        self.pm_send_btn = QPushButton("Send Private Message")
        self.pm_send_btn.clicked.connect(self.send_private_message)
        self.pm_send_btn.setEnabled(False)
        self.pm_send_btn.setStyleSheet("""
            QPushButton {
                background: #9C27B0;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 8px 16px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background: #7B1FA2;
            }
            QPushButton:disabled {
                background: #ccc;
                color: #666;
            }
        """)
        pm_layout.addWidget(self.pm_send_btn)

        pm_group.setLayout(pm_layout)
        right_layout.addWidget(pm_group)

        content_layout.addWidget(right_widget, 3)  # 30% width
        main_layout.addLayout(content_layout)

        # Add keyboard shortcuts for better user experience
        self.setup_keyboard_shortcuts()

    def setup_keyboard_shortcuts(self):
        """Set up keyboard shortcuts for chat functionality"""
        try:
            # Enter key to send message in private chat
            self.pm_input.keyPressEvent = self.pm_input_key_press_event

            # Ctrl+R to refresh manually
            refresh_shortcut = QShortcut(QKeySequence("Ctrl+R"), self)
            refresh_shortcut.activated.connect(self.manual_refresh_all)

            # Ctrl+D to clear chat input
            clear_shortcut = QShortcut(QKeySequence("Ctrl+D"), self)
            clear_shortcut.activated.connect(lambda: self.pm_input.clear())

            print("✓ Keyboard shortcuts initialized")
        except Exception as e:
            print(f"Error setting up keyboard shortcuts: {e}")

    def pm_input_key_press_event(self, event):
        """Handle key press events in private message input"""
        if event.key() == Qt.Key_Return and event.modifiers() == Qt.NoModifier:
            # Enter key sends message
            self.send_private_message()
        elif event.key() == Qt.Key_Return and event.modifiers() == Qt.ShiftModifier:
            # Shift+Enter adds new line
            QTextEdit.keyPressEvent(self.pm_input, event)
        else:
            # Default behavior for other keys
            QTextEdit.keyPressEvent(self.pm_input, event)

    def manual_refresh_all(self):
        """Manually refresh all chat components"""
        try:
            print("🔄 Manual refresh triggered")
            self.refresh_global_chat()
            self.refresh_private_chat()
            self.refresh_user_list()

            # Show brief confirmation
            if hasattr(self, 'active_pm_label'):
                original_text = self.active_pm_label.text()
                self.active_pm_label.setText("🔄 Refreshed!")
                QTimer.singleShot(1500, lambda: self.active_pm_label.setText(original_text))
        except Exception as e:
            print(f"Error during manual refresh: {e}")

    def configure_refresh_rates(self, global_rate=None, private_rate=None, user_list_rate=None):
        """Allow configuration of refresh rates (in milliseconds)"""
        try:
            if global_rate is not None:
                self.GLOBAL_CHAT_REFRESH_RATE = max(1000, global_rate)  # Minimum 1 second
                if hasattr(self, 'global_chat_timer'):
                    self.global_chat_timer.setInterval(self.GLOBAL_CHAT_REFRESH_RATE)

            if private_rate is not None:
                self.PRIVATE_CHAT_REFRESH_RATE = max(500, private_rate)  # Minimum 0.5 seconds
                if hasattr(self, 'private_chat_timer'):
                    self.private_chat_timer.setInterval(self.PRIVATE_CHAT_REFRESH_RATE)

            if user_list_rate is not None:
                self.USER_LIST_REFRESH_RATE = max(2000, user_list_rate)  # Minimum 2 seconds
                if hasattr(self, 'user_list_timer'):
                    self.user_list_timer.setInterval(self.USER_LIST_REFRESH_RATE)

            print(f"✓ Refresh rates updated:")
            print(f"  - Global: {self.GLOBAL_CHAT_REFRESH_RATE}ms")
            print(f"  - Private: {self.PRIVATE_CHAT_REFRESH_RATE}ms")
            print(f"  - User list: {self.USER_LIST_REFRESH_RATE}ms")
        except Exception as e:
            print(f"Error configuring refresh rates: {e}")

    def update_refresh_status(self, status_text):
        """Update the refresh status indicator"""
        try:
            if hasattr(self, 'refresh_status'):
                self.refresh_status.setText(status_text)

                # Update colors based on status
                if "🟢" in status_text or "Connected" in status_text:
                    color = "#28a745"  # Green
                    bg_color = "rgba(40, 167, 69, 0.1)"
                elif "🔄" in status_text or "Refreshing" in status_text:
                    color = "#ffc107"  # Yellow
                    bg_color = "rgba(255, 193, 7, 0.1)"
                elif "💬" in status_text or "New message" in status_text:
                    color = "#9C27B0"  # Purple
                    bg_color = "rgba(156, 39, 176, 0.1)"
                elif "🔴" in status_text or "Error" in status_text:
                    color = "#dc3545"  # Red
                    bg_color = "rgba(220, 53, 69, 0.1)"
                else:
                    color = "#6c757d"  # Gray
                    bg_color = "rgba(108, 117, 125, 0.1)"

                self.refresh_status.setStyleSheet(f"""
                    QLabel {{
                        color: {color};
                        font-size: 12px;
                        font-weight: 600;
                        padding: 4px 8px;
                        background-color: {bg_color};
                        border: 1px solid {color};
                        border-radius: 12px;
                    }}
                """)
        except Exception as e:
            print(f"Error updating refresh status: {e}")

    def load_messages(self):
        """Load and display global chat messages"""
        messages = get_all_messages()

        # Save scroll position
        scrollbar = self.chat_display.verticalScrollBar()
        was_at_bottom = scrollbar.value() == scrollbar.maximum()

        self.chat_display.clear()

        for message in messages:
            username, role, text, timestamp, message_id = message

            # Format timestamp
            time_str = format_malaysia_time(timestamp)

            # Get role color and icon
            role_color = self.role_colors.get(role.lower(), '#666666')
            role_icons = {
                'superadmin': '👑',
                'admin': '📢',
                'owner': '🏢',
                'tester': '🧪'
            }
            role_icon = role_icons.get(role.lower(), '👤')

            # Format message bubble with purple theme
            bubble_style = (
                f"background: linear-gradient(135deg, {role_color}15 0%, {role_color}08 100%); "
                f"border: 2px solid {role_color}30; border-radius: 12px; "
                f"margin: 8px 0; padding: 12px;"
            )

            formatted_message = (
                f"<div style='{bubble_style}'>"
                f"<div style='color: {role_color}; font-weight: bold; font-size: 13px; margin-bottom: 4px;'>"
                f"{role_icon} {username} ({role.title()})"
                f"</div>"
                f"<div style='color: #666; font-size: 11px; margin-bottom: 8px;'>{time_str}</div>"
                f"<div style='font-size: 14px; color: #333;'>{text}</div>"
                f"</div>"
            )

            self.chat_display.append(formatted_message)

        # Restore scroll position
        if was_at_bottom:
            scrollbar.setValue(scrollbar.maximum())

    def send_global_message(self):
        """Send a global message to the chat room"""
        message_text = self.message_input.toPlainText().strip()
        if not message_text:
            return

        user_id = self.user_info.get('user_id', 0)
        username = self.user_info.get('username', '')
        role = 'tester'
        role_id = 4  # Tester role ID

        success = insert_message(user_id, username, role, role_id, message_text)

        if success:
            self.message_input.clear()
            self.load_messages()
        else:
            QMessageBox.warning(self, "Error", "Failed to send message")

    def load_available_users(self):
        """Load users for private messaging"""
        current_username = self.user_info.get('username', '')
        users = get_all_chat_users(current_username)

        self.all_users = users
        self.populate_users_list(users)

    def load_unread_counts(self):
        """Load unread message counts for each user"""
        user_id = self.user_info.get('user_id', 0)
        if user_id:
            self.unread_counts = get_unread_messages_by_user(user_id)

    def populate_users_list(self, users):
        """Populate the users list with unread message indicators"""
        self.users_list.clear()
        selected_username = None

        # Remember the currently selected user
        if self.selected_user_item and self.active_pm_user:
            selected_username = self.active_pm_user

        for username, role_name, fullname in users:
            display_name = fullname if fullname else username

            # Use role-specific icons
            role_icons = {
                'superadmin': '👑',
                'admin': '📢',
                'tester': '🧪',
                'owner': '🏢'
            }
            role_icon = role_icons.get(role_name.lower(), '👤')

            # Check for unread messages
            unread_count = self.unread_counts.get(username, 0)
            unread_indicator = ""
            if unread_count > 0:
                unread_indicator = f" 🔴({unread_count})"

            item_text = f"{role_icon} {username} ({role_name.title()}){unread_indicator}"
            item = QListWidgetItem(item_text)
            item.setData(Qt.UserRole, {'username': username, 'role_name': role_name, 'fullname': fullname})

            # Make text bold if there are unread messages
            if unread_count > 0:
                font = item.font()
                font.setBold(True)
                item.setFont(font)
                # Set text color to indicate unread messages
                item.setForeground(QColor('#d32f2f'))  # Red color for unread

            self.users_list.addItem(item)

            # Restore selection if this was the previously selected user
            if selected_username and username == selected_username:
                item.setSelected(True)
                self.selected_user_item = item
                self.users_list.setCurrentItem(item)

    def search_users(self):
        """Filter users based on search text"""
        search_text = self.user_search.text().strip().lower()

        if not search_text:
            if hasattr(self, 'all_users'):
                self.populate_users_list(self.all_users)
        else:
            if hasattr(self, 'all_users'):
                filtered_users = []
                for username, role_name, fullname in self.all_users:
                    if (search_text in username.lower() or
                            (fullname and search_text in fullname.lower())):
                        filtered_users.append((username, role_name, fullname))
                self.populate_users_list(filtered_users)

    def start_private_chat(self, item):
        """Start private chat with selected user"""
        # Track the selected item for visual feedback
        self.selected_user_item = item

        user_data = item.data(Qt.UserRole)
        target_username = user_data['username']
        target_role = user_data.get('role_name', 'user')

        self.activate_private_chat(target_username, target_role)

    def activate_private_chat(self, target_username, target_role):
        """Activate private messaging with enhanced tracking"""
        self.active_pm_user = target_username
        self.active_pm_role = target_role

        # Mark messages as read when opening a chat
        user_id = self.user_info.get('user_id', 0)
        if user_id:
            mark_messages_as_read(user_id, target_username)
            # Refresh unread counts after marking as read
            self.load_unread_counts()
            self.populate_users_list(self.all_users)

        # Initialize message count tracking for this conversation
        current_username = self.user_info.get('username', '')
        messages = get_private_messages(current_username, target_username)
        self.last_message_count[target_username] = len(messages) if messages else 0

        # Hide any existing new message indicator
        if hasattr(self, 'new_message_indicator'):
            self.new_message_indicator.deleteLater()
            delattr(self, 'new_message_indicator')

        # Update UI
        self.active_pm_label.setText(f"💬 Chat with {target_username}")
        self.active_pm_label.setStyleSheet("""
            QLabel {
                color: #9C27B0; 
                font-size: 12px; 
                font-weight: 600;
                padding: 4px; 
                border-bottom: 1px solid #CE93D8;
                background-color: #F3E5F5;
                border-radius: 4px;
            }
        """)
        self.pm_input.setEnabled(True)
        self.pm_input.setPlaceholderText(f"Type message to {target_username}...")
        self.pm_send_btn.setEnabled(True)

        # Load messages (force scroll to bottom for new conversation)
        self.load_private_messages_inline(preserve_scroll=False)

        print(f"✓ Activated private chat with {target_username} ({target_role})")

    def load_private_messages_inline(self, preserve_scroll=False):
        """Load private messages for active conversation with improved scroll handling"""
        if not self.active_pm_user:
            return

        try:
            # Store scroll position if needed
            scrollbar = self.pm_display.verticalScrollBar()
            original_scroll_pos = scrollbar.value()
            was_at_bottom = scrollbar.value() >= (scrollbar.maximum() - 10)

            current_username = self.user_info.get('username', '')
            messages = get_private_messages(current_username, self.active_pm_user)

            # Clear display
            self.pm_display.clear()

            if not messages:
                self.pm_display.setPlaceholderText(
                    f"No messages with {self.active_pm_user} yet. Start the conversation!")
                return

            # Track the number of messages for change detection
            message_count = len(messages)
            previous_count = self.last_message_count.get(self.active_pm_user, 0)

            # Load messages with styling
            for i, message in enumerate(messages):
                sender, receiver, text, timestamp, sender_role = message

                time_str = format_malaysia_time(timestamp)
                role_color = self.role_colors.get(sender_role.lower(), '#9C27B0')

                # Role icons
                role_icons = {
                    'superadmin': '👑',
                    'admin': '📢',
                    'owner': '🏢',
                    'tester': '🧪'
                }
                role_icon = role_icons.get(sender_role.lower(), '👤')

                # Check if this is a new message (arrived after previous count)
                is_new_message = i >= previous_count and message_count > previous_count
                new_indicator = " ✨" if is_new_message else ""

                bubble_style = (
                    f"background: linear-gradient(135deg, {role_color}15 0%, {role_color}08 100%); "
                    f"border: 2px solid {role_color}30; border-radius: 12px; "
                    f"margin: 6px 0; padding: 10px;"
                )

                # Add subtle animation class for new messages
                animation_style = ""
                if is_new_message:
                    animation_style = "animation: fadeIn 0.5s ease-in;"

                formatted_message = (
                    f"<div style='{bubble_style} {animation_style}'>"
                    f"<div style='color: {role_color}; font-weight: bold; font-size: 12px; margin-bottom: 2px;'>"
                    f"{role_icon} {sender}{new_indicator}"
                    f"</div>"
                    f"<div style='color: #666; font-size: 10px; margin-bottom: 6px;'>{time_str}</div>"
                    f"<div style='font-size: 13px; color: #333;'>{text}</div>"
                    f"</div>"
                )

                self.pm_display.append(formatted_message)

            # Handle scrolling based on preferences and new message status
            if preserve_scroll and not was_at_bottom:
                # Preserve scroll position for background updates
                QTimer.singleShot(50, lambda: scrollbar.setValue(original_scroll_pos))
            elif was_at_bottom or not preserve_scroll:
                # Auto-scroll to bottom for new messages or manual refresh
                QTimer.singleShot(50, lambda: scrollbar.setValue(scrollbar.maximum()))

            # Update message count tracking
            self.last_message_count[self.active_pm_user] = message_count

        except Exception as e:
            print(f"Error loading private messages: {e}")
            self.pm_display.clear()
            self.pm_display.setPlaceholderText(f"Error loading messages with {self.active_pm_user}. Please try again.")

    def send_private_message(self):
        """Send a private message with improved feedback"""
        if not self.active_pm_user:
            QMessageBox.warning(self, "No Chat Selected", "Please select a user to chat with first.")
            return

        message_text = self.pm_input.toPlainText().strip()
        if not message_text:
            QMessageBox.information(self, "Empty Message", "Please enter a message before sending.")
            return

        # Disable send button temporarily to prevent double-sending
        self.pm_send_btn.setEnabled(False)
        self.pm_send_btn.setText("Sending...")

        try:
            user_id = self.user_info.get('user_id', 0)
            username = self.user_info.get('username', '')
            role = 'tester'

            success = insert_private_message(user_id, username, self.active_pm_user, role, message_text)

            if success:
                self.pm_input.clear()
                # Force immediate refresh to show sent message
                self.load_private_messages_inline(preserve_scroll=False)

                # Update last message count for this conversation
                current_username = self.user_info.get('username', '')
                messages = get_private_messages(current_username, self.active_pm_user)
                self.last_message_count[self.active_pm_user] = len(messages) if messages else 0

                print(f"✓ Message sent to {self.active_pm_user}: {message_text[:50]}...")
            else:
                QMessageBox.warning(self, "Send Failed", "Failed to send private message. Please try again.")

        except Exception as e:
            print(f"Error sending message: {e}")
            QMessageBox.critical(self, "Error", f"An error occurred while sending the message:\n{str(e)}")
        finally:
            # Re-enable send button
            self.pm_send_btn.setEnabled(True)
            self.pm_send_btn.setText("Send Private Message")

    def closeEvent(self, event):
        """Stop all timers when widget is closed"""
        try:
            # Stop all refresh timers
            if hasattr(self, 'global_chat_timer'):
                self.global_chat_timer.stop()
            if hasattr(self, 'private_chat_timer'):
                self.private_chat_timer.stop()
            if hasattr(self, 'user_list_timer'):
                self.user_list_timer.stop()

            # Clean up any remaining UI elements
            if hasattr(self, 'new_message_indicator'):
                self.new_message_indicator.deleteLater()
                delattr(self, 'new_message_indicator')

            print("✓ Chat widget closed and timers stopped")
        except Exception as e:
            print(f"Error during chat widget cleanup: {e}")
        finally:
            event.accept()


class TesterHomePage(QMainWindow):
    def __init__(self, user_info=None):
        super().__init__()

        # Store user information from login
        self.user_info = user_info or {}

        # Initialize database tables
        create_database_tables()

        # Use user_id from login info if available, otherwise get from database
        if user_info and 'user_id' in user_info:
            self.user_id = user_info['user_id']
            print(f"✓ Tester Dashboard initialized for user: {user_info.get('username', 'Unknown')}")
            print(f"  - User ID: {self.user_id}")
            print(f"  - Full Name: {user_info.get('fullname', 'N/A')}")
            print(f"  - Role: {user_info.get('role_name', 'N/A')}")
        else:
            self.user_id = self.get_current_user_id()
            print(f"⚠️ No user info provided, using fallback user ID: {self.user_id}")

        # Auto-update testing table for assigned products
        self.auto_update_testing_records()

        self.attached_file_path = None
        self.init_ui()

    def auto_update_testing_records(self):
        """
        Auto-update testing table when tester enters dashboard.
        Creates testing records for all assigned products and sets progress_status_id to 1 (In Progress).
        """
        try:
            with sqlite3.connect(db_path, timeout=30.0) as conn:
                cursor = conn.cursor()

                print(f"🔄 Auto-updating testing records for tester ID: {self.user_id}")

                # Get all products assigned to this tester that don't have testing records yet
                cursor.execute("""
                    SELECT DISTINCT 
                        pta.product_id,
                        p.product_name,
                        p.batch
                    FROM product_tester_assignments pta
                    INNER JOIN products p ON pta.product_id = p.product_id
                    LEFT JOIN testing t ON t.product_id = pta.product_id AND t.user_id = pta.tester_id
                    WHERE pta.tester_id = ? AND t.test_id IS NULL
                """, (self.user_id,))

                untracked_products = cursor.fetchall()

                if untracked_products:
                    print(f"📝 Found {len(untracked_products)} assigned products without testing records")

                    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    created_count = 0

                    for product_id, product_name, batch in untracked_products:
                        try:
                            # Create testing record with:
                            # - user_id = current tester
                            # - product_id = assigned product
                            # - assignment_status_id = 1 (Assigned)
                            # - test_start = current timestamp
                            # - progress_status_id = 1 (In Progress)
                            # - test_result_id = 1 (Under Review - default status)
                            cursor.execute("""
                                INSERT INTO testing (
                                    user_id, 
                                    product_id, 
                                    assignment_status_id, 
                                    test_start, 
                                    progress_status_id, 
                                    test_result_id
                                ) VALUES (?, ?, ?, ?, ?, ?)
                            """, (self.user_id, product_id, 1, current_time, 1, 1))

                            created_count += 1
                            print(f"  ✓ Created testing record for Product {product_id}: {product_name}")

                        except sqlite3.Error as e:
                            print(f"  ✗ Failed to create testing record for Product {product_id}: {e}")
                            continue

                    # Commit all changes
                    conn.commit()
                    print(f"✅ Successfully created {created_count} testing records")

                    if created_count > 0:
                        # Send notification to tester about auto-started tests
                        try:
                            notification_message = f"Welcome! {created_count} assigned product(s) have been automatically set to 'In Progress' status. You can now begin testing."
                            cursor.execute("""
                                INSERT INTO notifications (user_id, message, timestamp, read)
                                VALUES (?, ?, ?, 0)
                            """, (self.user_id, notification_message, current_time))
                            conn.commit()
                            print(f"📬 Created welcome notification for tester")
                        except Exception as e:
                            print(f"⚠️ Could not create notification: {e}")

                else:
                    print(f"ℹ️ All assigned products already have testing records")

                # Also update any existing testing records that might be in 'Not Started' status to 'In Progress'
                cursor.execute("""
                    UPDATE testing 
                    SET progress_status_id = 1, test_start = COALESCE(test_start, ?)
                    WHERE user_id = ? 
                    AND product_id IN (
                        SELECT product_id FROM product_tester_assignments WHERE tester_id = ?
                    )
                    AND progress_status_id = 0
                """, (current_time, self.user_id, self.user_id))

                updated_count = cursor.rowcount
                if updated_count > 0:
                    conn.commit()
                    print(f"🔄 Updated {updated_count} existing records from 'On Hold' to 'In Progress'")

        except sqlite3.Error as e:
            print(f"✗ Database error during auto-update: {e}")
        except Exception as e:
            print(f"✗ Unexpected error during auto-update: {e}")

    def get_current_user_id(self):
        """Get the current tester user ID from database or session"""
        # If self.user_id is already set (e.g., supplied via login), use it directly
        if hasattr(self, 'user_id') and self.user_id:
            return self.user_id

        # Otherwise, try to fetch a tester ID from the database as a fallback
        try:
            with sqlite3.connect(db_path, timeout=30.0) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT user_id FROM users WHERE role = 4 LIMIT 1")  # role 4 = tester
                result = cursor.fetchone()
                return result[0] if result else 1
        except Exception as e:
            print(f"Error getting current user ID: {e}")
            return 1  # Default fallback

    def create_standard_title(self, title_text, subtitle_text=None):
        """Create a standardized title widget for all pages"""
        title_container = QWidget()
        title_layout = QVBoxLayout(title_container)
        title_layout.setContentsMargins(0, 0, 0, 20)
        title_layout.setSpacing(5)

        # Main title
        title_label = QLabel(title_text)
        title_label.setStyleSheet("""
            font-size: 32px; 
            font-weight: bold; 
            color: #2c3e50;
            margin: 0px;
            padding: 0px;
        """)
        title_label.setAlignment(Qt.AlignLeft)
        title_layout.addWidget(title_label)

        # Optional subtitle
        if subtitle_text:
            subtitle_label = QLabel(subtitle_text)
            subtitle_label.setStyleSheet("""
                font-size: 16px; 
                color: #6c757d;
                margin: 0px;
                padding: 0px;
                font-weight: 400;
            """)
            subtitle_label.setAlignment(Qt.AlignLeft)
            title_layout.addWidget(subtitle_label)

        # Add separator line
        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setFrameShadow(QFrame.Sunken)
        separator.setStyleSheet("""
            QFrame {
                color: #dee2e6;
                background-color: #dee2e6;
                height: 1px;
                border: none;
                margin: 10px 0px;
            }
        """)
        title_layout.addWidget(separator)

        return title_container

    def create_standard_page_layout(self):
        """Create a standardized page layout with consistent padding"""
        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(40, 30, 40, 30)
        content_layout.setSpacing(20)
        return content_widget, content_layout

    def init_ui(self):
        self.setWindowTitle("🧪 Medical Testing Dashboard")
        # Set window to be windowed (not fullscreen) with reasonable size
        self.setGeometry(150, 100, 1400, 900)
        self.setMinimumSize(1200, 700)  # Set minimum size to prevent too small windows

        # Set application icon and window styling
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f8f9fa;
            }
        """)

        # Create central widget and main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Create sidebar
        self.create_sidebar(main_layout)

        # Create main area
        self.create_main_area(main_layout)

        # Initialize with Product Assignments
        self.switch_frame("Product Assignments")

    def create_sidebar(self, main_layout):
        sidebar = QWidget()
        sidebar.setFixedWidth(320)
        sidebar.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                    stop:0 #667eea, stop:1 #764ba2);
            }
        """)

        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(20, 30, 20, 20)
        sidebar_layout.setSpacing(15)

        # Header section with user info
        header_widget = QWidget()
        header_layout = QVBoxLayout(header_widget)
        header_layout.setContentsMargins(0, 0, 0, 0)

        # Title
        title = QLabel("🧪 Testing Center")
        title.setStyleSheet("""
            color: white; 
            font-size: 28px; 
            font-weight: bold;
            margin-bottom: 5px;
        """)
        title.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(title)

        # User welcome message
        if self.user_info and self.user_info.get('fullname'):
            user_welcome = QLabel(f"Welcome, {self.user_info['fullname']}")
            user_welcome.setStyleSheet("""
                color: rgba(255, 255, 255, 0.9); 
                font-size: 14px;
                font-weight: 600;
                margin-bottom: 3px;
            """)
            user_welcome.setAlignment(Qt.AlignCenter)
            header_layout.addWidget(user_welcome)

        # Subtitle
        subtitle = QLabel("Medical Product Testing")
        subtitle.setStyleSheet("""
            color: rgba(255, 255, 255, 0.8); 
            font-size: 12px;
            margin-bottom: 20px;
        """)
        subtitle.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(subtitle)

        sidebar_layout.addWidget(header_widget)

        # Menu options with icons
        menu_options = [
            ("📋", "Product Assignments"),
            ("🔬", "Test Result Update"),
            ("📖", "Test History"),
            ("📈", "Test Analysis"),
            ("🔔", "Notifications"),
            ("💬", "Chat/Announcement"),
        ]

        self.buttons = {}
        for icon, opt in menu_options:
            btn = QPushButton(f"{icon}  {opt}")
            btn.setStyleSheet("""
                QPushButton {
                    background-color: rgba(255, 255, 255, 0.1);
                    color: white;
                    border: none;
                    border-radius: 12px;
                    padding: 18px 20px;
                    font-size: 16px;
                    font-weight: 500;
                    text-align: left;
                    margin: 3px 0px;
                }
                QPushButton:hover {
                    background-color: rgba(255, 255, 255, 0.2);
                    transform: translateX(5px);
                }
                QPushButton:pressed {
                    background-color: rgba(255, 255, 255, 0.3);
                }
                QPushButton:checked {
                    background-color: rgba(255, 255, 255, 0.25);
                    border-left: 4px solid #ffffff;
                }
            """)
            btn.setCheckable(True)
            btn.clicked.connect(lambda checked, o=opt: self.switch_frame(o))
            sidebar_layout.addWidget(btn)
            self.buttons[opt] = btn

        sidebar_layout.addStretch()

        # Logout button
        logout_btn = QPushButton("🚪 Logout")
        logout_btn.setStyleSheet("""
            QPushButton {
                background-color: rgba(220, 53, 69, 0.8);
                color: white;
                border: none;
                border-radius: 8px;
                padding: 12px 20px;
                font-size: 14px;
                font-weight: 600;
                margin: 10px 0px;
            }
            QPushButton:hover {
                background-color: rgba(220, 53, 69, 1.0);
            }
            QPushButton:pressed {
                background-color: rgba(176, 42, 55, 1.0);
            }
        """)
        logout_btn.clicked.connect(self.logout)
        sidebar_layout.addWidget(logout_btn)

        main_layout.addWidget(sidebar)

    def create_main_area(self, main_layout):
        self.main_frame = QWidget()
        self.main_frame.setStyleSheet("""
            QWidget {
                background-color: #f8f9fa;
                border-radius: 0px;
            }
        """)
        main_layout.addWidget(self.main_frame, 1)

    def clear_main_frame(self):
        layout = self.main_frame.layout()
        if layout is not None:
            while layout.count():
                child = layout.takeAt(0)
                if child.widget():
                    child.widget().deleteLater()
        else:
            layout = QVBoxLayout()
            self.main_frame.setLayout(layout)

    def switch_frame(self, section):
        # Update button states
        for btn_name, btn in self.buttons.items():
            btn.setChecked(btn_name == section)

        self.clear_main_frame()
        layout = self.main_frame.layout()

        if section == "Product Assignments":
            self.show_product_assignments(layout)
        elif section == "Test Result Update":
            self.show_test_result_update(layout)
        elif section == "Test History":
            self.show_test_history(layout)
        elif section == "Test Analysis":
            self.show_test_analysis(layout)
        elif section == "Notifications":
            self.show_notifications(layout)
        elif section == "Chat/Announcement":
            self.show_chat(layout)

    def logout(self):
        """Handle logout functionality"""
        reply = QMessageBox.question(self, 'Logout',
                                     'Are you sure you want to logout?',
                                     QMessageBox.Yes | QMessageBox.No,
                                     QMessageBox.No)

        if reply == QMessageBox.Yes:
            username = self.user_info.get('username', 'Unknown User')
            print(f"✓ User '{username}' logged out from Tester Dashboard")

            # Close the tester dashboard
            self.close()

            # Show the login window again
            self.show_login_window()

    def show_login_window(self):
        """Show the login window again after logout"""
        try:
            # Import here to avoid circular imports
            from user1 import SignInSignUpWindow

            self.login_window = SignInSignUpWindow()
            self.login_window.show()

            # Ensure the login window is brought to front and has focus
            self.login_window.raise_()
            self.login_window.activateWindow()

            print(f"✓ Login window opened successfully after logout")
        except Exception as e:
            print(f"✗ Error opening login window: {e}")
            # If we can't open login window, just exit the application
            QApplication.quit()

    def show_product_assignments(self, layout):
        # Create standardized page layout
        content_widget, content_layout = self.create_standard_page_layout()

        # Add standardized title
        title_widget = self.create_standard_title(
            "📋 My Assigned Products",
            "View and manage your assigned product testing tasks"
        )
        content_layout.addWidget(title_widget)

        # Action buttons section
        action_widget = QWidget()
        action_layout = QHBoxLayout(action_widget)
        action_layout.setContentsMargins(0, 0, 0, 0)
        action_layout.addStretch()

        # Refresh button
        refresh_btn = QPushButton("🔄 Refresh")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 12px 20px;
                font-size: 14px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
        """)
        # Use switch_frame to refresh so previous widgets are cleared first
        refresh_btn.clicked.connect(lambda: self.switch_frame("Product Assignments"))
        action_layout.addWidget(refresh_btn)

        content_layout.addWidget(action_widget)

        # Create table with modern styling
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Product ID", "Product Name", "Status", "Location", "Batch"])

        # Modern table styling
        table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #e9ecef;
                border-radius: 12px;
                gridline-color: #f1f3f4;
                font-size: 14px;
            }
            QTableWidget::item {
                padding: 12px;
                border-bottom: 1px solid #f1f3f4;
            }
            QTableWidget::item:selected {
                background-color: #e3f2fd;
                color: #1976d2;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                color: #495057;
                font-weight: 600;
                font-size: 14px;
                padding: 15px;
                border: none;
                border-bottom: 2px solid #dee2e6;
            }
            QHeaderView::section:first {
                border-top-left-radius: 12px;
            }
            QHeaderView::section:last {
                border-top-right-radius: 12px;
            }
        """)

        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.verticalHeader().setVisible(False)

        # Set column widths
        header = table.horizontalHeader()
        header.setStretchLastSection(True)
        header.resizeSection(0, 150)
        header.resizeSection(1, 500)
        header.resizeSection(2, 200)
        header.resizeSection(3, 250)
        header.resizeSection(4, 150)

        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute("""
                    SELECT 
                        P.product_id,
                        P.product_name, 
                        COALESCE(ps.status_name, 'Not Started') as test_status,
                        COALESCE(rl.rack_location_name, 'No Location') as location,
                        P.batch
                    FROM products P
                    INNER JOIN product_tester_assignments A ON P.product_id = A.product_id
                    LEFT JOIN testing T ON T.product_id = P.product_id AND T.user_id = A.tester_id
                    LEFT JOIN progress_status ps ON T.progress_status_id = ps.status_id
                    LEFT JOIN racklocations rl ON P.rack_location_id = rl.rack_location_id
                    WHERE A.tester_id = ?
                    ORDER BY P.product_name
                """, (self.user_id,))

            rows = cursor.fetchall()

            if rows:
                table.setRowCount(len(rows))
                for i, row in enumerate(rows):
                    product_id = row[0]

                    for j, value in enumerate(row):
                        item = QTableWidgetItem(str(value) if value else "")

                        # Add status-based styling
                        if j == 2:  # Status column
                            status = str(value) if value else "Not Started"
                            if status == "Testing Complete":
                                item.setBackground(QColor("#d4edda"))
                                item.setForeground(QColor("#155724"))
                            elif status == "In Progress":
                                item.setBackground(QColor("#fff3cd"))
                                item.setForeground(QColor("#856404"))
                            elif status == "On Hold":
                                item.setBackground(QColor("#f8d7da"))
                                item.setForeground(QColor("#721c24"))
                            else:
                                item.setBackground(QColor("#e2e3e5"))
                                item.setForeground(QColor("#383d41"))

                        table.setItem(i, j, item)

                    # Add action buttons in the last column
                    action_widget = QWidget()
                    action_layout = QHBoxLayout(action_widget)
                    action_layout.setContentsMargins(5, 5, 5, 5)
                    action_layout.setSpacing(5)

                    status = row[2] if row[2] else "Not Started"

                    if status == "Not Started":
                        # Start Testing button
                        start_btn = QPushButton("▶️ Start")
                        start_btn.setStyleSheet("""
                            QPushButton {
                                background-color: #28a745;
                                color: white;
                                border: none;
                                border-radius: 6px;
                                padding: 8px 15px;
                                font-size: 12px;
                                font-weight: 600;
                            }
                            QPushButton:hover {
                                background-color: #218838;
                            }
                            QPushButton:pressed {
                                background-color: #1e7e34;
                            }
                        """)
                        start_btn.clicked.connect(lambda checked, pid=product_id: self.quick_start_testing(pid))
                        action_layout.addWidget(start_btn)

                    elif status == "In Progress":
                        # Update Results button
                        update_btn = QPushButton("📝 Update")
                        update_btn.setStyleSheet("""
                            QPushButton {
                                background-color: #007bff;
                                color: white;
                                border: none;
                                border-radius: 6px;
                                padding: 8px 15px;
                                font-size: 12px;
                                font-weight: 600;
                            }
                            QPushButton:hover {
                                background-color: #0056b3;
                            }
                            QPushButton:pressed {
                                background-color: #004085;
                            }
                        """)
                        update_btn.clicked.connect(lambda checked: self.switch_frame("Test Result Update"))
                        action_layout.addWidget(update_btn)

                    elif status in ["Testing Complete", "Sample Expired"]:
                        # View Details button
                        view_btn = QPushButton("👁️ View")
                        view_btn.setStyleSheet("""
                            QPushButton {
                                background-color: #6c757d;
                                color: white;
                                border: none;
                                border-radius: 6px;
                                padding: 8px 15px;
                                font-size: 12px;
                                font-weight: 600;
                            }
                            QPushButton:hover {
                                background-color: #5a6268;
                            }
                            QPushButton:pressed {
                                background-color: #545b62;
                            }
                        """)
                        view_btn.clicked.connect(lambda checked: self.switch_frame("Test History"))
                        action_layout.addWidget(view_btn)

                    elif status == "On Hold":
                        # Resume button
                        resume_btn = QPushButton("🔄 Resume")
                        resume_btn.setStyleSheet("""
                            QPushButton {
                                background-color: #ffc107;
                                color: white;
                                border: none;
                                border-radius: 6px;
                                padding: 8px 15px;
                                font-size: 12px;
                                font-weight: 600;
                            }
                            QPushButton:hover {
                                background-color: #e0a800;
                            }
                            QPushButton:pressed {
                                background-color: #d39e00;
                            }
                        """)
                        resume_btn.clicked.connect(lambda checked, pid=product_id: self.resume_testing(pid))
                        action_layout.addWidget(resume_btn)

                    action_layout.addStretch()
                    table.setCellWidget(i, 5, action_widget)

                # Add summary info
                summary_widget = QWidget()
                summary_layout = QHBoxLayout(summary_widget)
                summary_layout.setContentsMargins(0, 20, 0, 0)

                total_label = QLabel(f"📊 Total Products: {len(rows)}")
                total_label.setStyleSheet("""
                    color: #6c757d;
                    font-size: 16px;
                    font-weight: 500;
                """)
                summary_layout.addWidget(total_label)
                summary_layout.addStretch()

                content_layout.addWidget(table)
                content_layout.addWidget(summary_widget)
            else:
                # No data state
                no_data_widget = QWidget()
                no_data_layout = QVBoxLayout(no_data_widget)
                no_data_layout.setAlignment(Qt.AlignCenter)

                no_data_icon = QLabel("📝")
                no_data_icon.setStyleSheet("font-size: 72px;")
                no_data_icon.setAlignment(Qt.AlignCenter)
                no_data_layout.addWidget(no_data_icon)

                no_data_label = QLabel("No products assigned yet")
                no_data_label.setAlignment(Qt.AlignCenter)
                no_data_label.setStyleSheet("""
                    font-size: 24px; 
                    color: #6c757d;
                    font-weight: 500;
                    margin: 20px;
                """)
                no_data_layout.addWidget(no_data_label)

                no_data_desc = QLabel("Contact your supervisor to get product assignments")
                no_data_desc.setAlignment(Qt.AlignCenter)
                no_data_desc.setStyleSheet("""
                    font-size: 16px; 
                    color: #adb5bd;
                """)
                no_data_layout.addWidget(no_data_desc)

                content_layout.addWidget(no_data_widget)

        except sqlite3.Error as e:
            # Error state
            error_widget = QWidget()
            error_layout = QVBoxLayout(error_widget)
            error_layout.setAlignment(Qt.AlignCenter)

            error_icon = QLabel("⚠️")
            error_icon.setStyleSheet("font-size: 72px;")
            error_icon.setAlignment(Qt.AlignCenter)
            error_layout.addWidget(error_icon)

            error_label = QLabel(f"Error loading assignments")
            error_label.setStyleSheet("""
                color: #dc3545; 
                font-size: 24px; 
                font-weight: 600;
                margin: 20px;
            """)
            error_label.setAlignment(Qt.AlignCenter)
            error_layout.addWidget(error_label)

            error_desc = QLabel(f"Details: {str(e)}")
            error_desc.setStyleSheet("""
                color: #6c757d; 
                font-size: 14px;
            """)
            error_desc.setAlignment(Qt.AlignCenter)
            error_layout.addWidget(error_desc)

            content_layout.addWidget(error_widget)
        finally:
            if conn:
                conn.close()

        layout.addWidget(content_widget)

    def quick_start_testing(self, product_id):
        """
        Quick start testing - sets status to In Progress with actual start timestamp.
        Note: Products are now automatically set to In Progress when assigned,
        so this method mainly serves to update the start time if needed.
        """
        try:
            with sqlite3.connect(db_path, timeout=10.0) as conn:
                cursor = conn.cursor()
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # Check if test record already exists
                cursor.execute("SELECT test_id, progress_status_id FROM testing WHERE user_id = ? AND product_id = ?",
                               (self.user_id, product_id))
                existing_test = cursor.fetchone()

                if existing_test:
                    test_id, current_status = existing_test

                    if current_status == 1:  # Already In Progress
                        # Get product name for message
                        cursor.execute("SELECT product_name FROM products WHERE product_id = ?", (product_id,))
                        product_result = cursor.fetchone()
                        product_name = product_result[0] if product_result else f"Product {product_id}"

                        QMessageBox.information(self, "Already In Progress",
                                                f"✅ Testing is already in progress for:\n{product_name}\n\n"
                                                f"Go to 'Test Result Update' to complete the testing.")
                        return

                    # Update existing record to In Progress (for products that were on hold, etc.)
                    cursor.execute("""
                        UPDATE testing 
                        SET progress_status_id = 1, test_start = ?, test_result_id = 1
                        WHERE test_id = ?
                    """, (now, test_id))
                    print(f"  ↻ Updated existing test record to In Progress")
                else:
                    # Create new test record with "In Progress" status (shouldn't happen with auto-assignment)
                    cursor.execute("""
                        INSERT INTO testing (user_id, product_id, test_start, progress_status_id, test_result_id)
                        VALUES (?, ?, ?, 1, 1)
                    """, (self.user_id, product_id, now))
                    print(f"  ✚ Created new test record with In Progress status")

                conn.commit()

                # Get product name for confirmation
                cursor.execute("SELECT product_name FROM products WHERE product_id = ?", (product_id,))
                product_result = cursor.fetchone()
                product_name = product_result[0] if product_result else f"Product {product_id}"

                QMessageBox.information(self, "Testing Started",
                                        f"✅ Testing started for:\n{product_name}\n\nStatus: In Progress\nStart Time: {now}")

                # Refresh the current view
                self.show_product_assignments(self.main_frame.layout())

        except sqlite3.Error as e:
            print(f"✗ Database error starting test: {e}")
            QMessageBox.critical(self, "Database Error", f"Failed to start testing:\n{str(e)}")
        except Exception as e:
            print(f"✗ Unexpected error starting test: {e}")
            QMessageBox.critical(self, "Error", f"An unexpected error occurred:\n{str(e)}")

    def resume_testing(self, product_id):
        """Resume testing from On Hold status back to In Progress"""
        try:
            with sqlite3.connect(db_path, timeout=10.0) as conn:
                cursor = conn.cursor()

                # Update status from On Hold (0) to In Progress (1)
                cursor.execute("""
                    UPDATE testing 
                    SET progress_status_id = 1, test_result_id = 1
                    WHERE user_id = ? AND product_id = ? AND progress_status_id = 0
                """, (self.user_id, product_id))

                if cursor.rowcount > 0:
                    conn.commit()

                    # Get product name for confirmation
                    cursor.execute("SELECT product_name FROM products WHERE product_id = ?", (product_id,))
                    product_result = cursor.fetchone()
                    product_name = product_result[0] if product_result else f"Product {product_id}"

                    QMessageBox.information(self, "Testing Resumed",
                                            f"✅ Testing resumed for:\n{product_name}\n\nStatus: In Progress")

                    # Refresh the current view
                    self.show_product_assignments(self.main_frame.layout())
                else:
                    QMessageBox.warning(self, "Resume Failed",
                                        "Could not resume testing. Product may not be on hold or may not exist.")

        except sqlite3.Error as e:
            print(f"✗ Database error resuming test: {e}")
            QMessageBox.critical(self, "Database Error", f"Failed to resume testing:\n{str(e)}")
        except Exception as e:
            print(f"✗ Unexpected error resuming test: {e}")
            QMessageBox.critical(self, "Error", f"An unexpected error occurred:\n{str(e)}")

    def show_test_result_update(self, layout):
        # Create standardized page layout
        content_widget, content_layout = self.create_standard_page_layout()

        # Add standardized title
        title_widget = self.create_standard_title(
            "🔬 Update Test Results",
            "Complete testing for products currently in progress - add final results and evidence"
        )
        content_layout.addWidget(title_widget)

        # Form container with card styling
        form_card = QWidget()
        form_card.setStyleSheet("""
            QWidget {
                background-color: white;
                border-radius: 15px;
                border: 1px solid #e9ecef;
            }
        """)
        form_card_layout = QVBoxLayout(form_card)
        form_card_layout.setContentsMargins(30, 30, 30, 30)
        form_card_layout.setSpacing(25)

        # Form title
        form_title = QLabel("📝 Test Information")
        form_title.setStyleSheet("""
            font-size: 20px;
            font-weight: 600;
            color: #495057;
            margin-bottom: 15px;
        """)
        form_card_layout.addWidget(form_title)

        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignLeft)
        form_layout.setFormAlignment(Qt.AlignLeft)
        form_layout.setSpacing(20)

        # Product selection
        self.product_combo = QComboBox()
        self.product_combo.setMinimumWidth(400)
        self.product_combo.setStyleSheet("""
            QComboBox {
                background-color: #f8f9fa;
                border: 2px solid #e9ecef;
                border-radius: 8px;
                padding: 12px 15px;
                font-size: 14px;
                color: #495057;
            }
            QComboBox:focus {
                border-color: #80bdff;
                background-color: white;
            }
            QComboBox::drop-down {
                border: none;
                width: 30px;
            }
            QComboBox::down-arrow {
                image: none;
                border: none;
                width: 12px;
                height: 12px;
                background: #6c757d;
            }
        """)

        # Load products assigned to this tester
        self.load_assigned_products()

        # Connect product selection change to update start time
        self.product_combo.currentIndexChanged.connect(self.update_start_time_display)

        product_label = QLabel("🧪 Select Product:")
        product_label.setStyleSheet("""
            font-size: 16px;
            font-weight: 600;
            color: #495057;
        """)

        # Create a container for product selection with refresh button
        product_container = QWidget()
        product_layout = QHBoxLayout(product_container)
        product_layout.setContentsMargins(0, 0, 0, 0)
        product_layout.setSpacing(10)

        product_layout.addWidget(self.product_combo)

        # Add refresh button
        refresh_products_btn = QPushButton("🔄")
        refresh_products_btn.setFixedSize(40, 40)
        refresh_products_btn.setToolTip("Refresh product list")
        refresh_products_btn.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:pressed {
                background-color: #0f6674;
            }
        """)
        refresh_products_btn.clicked.connect(self.load_assigned_products)
        product_layout.addWidget(refresh_products_btn)

        form_layout.addRow(product_label, product_container)

        # Test start time display (read-only)
        self.start_time_label = QLabel("Not selected")
        self.start_time_label.setStyleSheet("""
            QLabel {
                background-color: #e9ecef;
                border: 2px solid #dee2e6;
                border-radius: 8px;
                padding: 12px 15px;
                font-size: 14px;
                color: #495057;
                font-weight: 500;
            }
        """)

        start_time_label_text = QLabel("🕐 Test Started:")
        start_time_label_text.setStyleSheet("""
            font-size: 16px;
            font-weight: 600;
            color: #495057;
        """)
        form_layout.addRow(start_time_label_text, self.start_time_label)

        # Status selection
        self.status_combo = QComboBox()
        self.status_combo.setStyleSheet("""
            QComboBox {
                background-color: #f8f9fa;
                border: 2px solid #e9ecef;
                border-radius: 8px;
                padding: 12px 15px;
                font-size: 14px;
                color: #495057;
            }
            QComboBox:focus {
                border-color: #80bdff;
                background-color: white;
            }
        """)

        # Load status options from database
        try:
            with sqlite3.connect(db_path, timeout=30.0) as conn:
                cursor = conn.cursor()

                # Only load valid transitions from "In Progress" (status_id = 1)
                # Valid transitions: On Hold (0), Testing Complete (2), Sample Expired (3)
                valid_transitions = [
                    (0, "On Hold"),  # Temporary pause
                    (2, "Testing Complete"),  # Finished successfully
                    (3, "Sample Expired")  # Sample went bad during testing
                ]

                for status_id, status_name in valid_transitions:
                    self.status_combo.addItem(status_name, status_id)

                print("✓ Loaded valid status transitions from In Progress")

        except sqlite3.Error as e:
            print(f"Error loading statuses: {e}")
            # Fallback to default options if database fails
            self.status_combo.addItem("On Hold", 0)
            self.status_combo.addItem("Testing Complete", 2)
            self.status_combo.addItem("Sample Expired", 3)

        status_label = QLabel("📊 Test Status:")
        status_label.setStyleSheet("""
            font-size: 16px;
            font-weight: 600;
            color: #495057;
        """)
        form_layout.addRow(status_label, self.status_combo)

        # Test result
        self.result_text = QTextEdit()
        self.result_text.setMaximumHeight(120)
        self.result_text.setPlaceholderText("Enter detailed test results, observations, and notes...")
        self.result_text.setStyleSheet("""
            QTextEdit {
                background-color: #f8f9fa;
                border: 2px solid #e9ecef;
                border-radius: 8px;
                padding: 12px;
                font-size: 14px;
                color: #495057;
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QTextEdit:focus {
                border-color: #80bdff;
                background-color: white;
            }
        """)

        result_label = QLabel("📝 Test Notes:")
        result_label.setStyleSheet("""
            font-size: 16px;
            font-weight: 600;
            color: #495057;
        """)
        form_layout.addRow(result_label, self.result_text)

        # Complete the form layout
        form_card_layout.addLayout(form_layout)
        content_layout.addWidget(form_card)

        # Image upload card
        image_card = QWidget()
        image_card.setStyleSheet("""
            QWidget {
                background-color: white;
                border-radius: 15px;
                border: 1px solid #e9ecef;
            }
        """)
        image_card_layout = QVBoxLayout(image_card)
        image_card_layout.setContentsMargins(30, 30, 30, 30)
        image_card_layout.setSpacing(20)

        # Image upload title
        image_title = QLabel("📷 Test Evidence")
        image_title.setStyleSheet("""
            font-size: 20px;
            font-weight: 600;
            color: #495057;
            margin-bottom: 15px;
        """)
        image_card_layout.addWidget(image_title)

        # Image preview and upload
        image_container = QHBoxLayout()

        # Image preview
        self.image_label = QLabel()
        self.image_label.setFixedSize(300, 200)
        self.image_label.setStyleSheet("""
            QLabel {
                border: 2px dashed #dee2e6;
                border-radius: 12px;
                background-color: #f8f9fa;
                color: #6c757d;
                font-size: 14px;
            }
        """)
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setText("📷\nNo image selected")

        # Upload controls
        upload_controls = QVBoxLayout()
        upload_controls.setSpacing(15)

        self.image_path_label = QLabel("No file selected")
        self.image_path_label.setStyleSheet("""
            color: #6c757d;
            font-size: 14px;
            font-style: italic;
        """)

        attach_btn = QPushButton("📁 Choose File")
        attach_btn.clicked.connect(self.attach_file)
        attach_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 12px 20px;
                font-size: 14px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
            QPushButton:pressed {
                background-color: #495057;
            }
        """)

        upload_controls.addWidget(self.image_path_label)
        upload_controls.addWidget(attach_btn)
        upload_controls.addStretch()

        image_container.addWidget(self.image_label)
        image_container.addLayout(upload_controls)
        image_card_layout.addLayout(image_container)

        content_layout.addWidget(image_card)

        # Action buttons
        button_container = QWidget()
        button_layout = QHBoxLayout(button_container)
        button_layout.setContentsMargins(0, 20, 0, 0)
        button_layout.setSpacing(15)

        # Cancel button
        cancel_btn = QPushButton("↩️ Cancel")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 15px 30px;
                font-size: 16px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        cancel_btn.clicked.connect(self.clear_form)

        # Update button
        update_btn = QPushButton("✅ Update Test Status")
        update_btn.clicked.connect(self.update_test_status)
        update_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 15px 30px;
                font-size: 16px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #218838;
            }
            QPushButton:pressed {
                background-color: #1e7e34;
            }
        """)

        button_layout.addStretch()
        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(update_btn)

        content_layout.addWidget(button_container)
        content_layout.addStretch()

        layout.addWidget(content_widget)

    def update_start_time_display(self):
        """Update the start time display when product selection changes"""
        try:
            current_data = self.product_combo.currentData()
            if current_data and isinstance(current_data, dict) and 'test_start' in current_data:
                test_start = current_data['test_start']
                if test_start:
                    # Format the start time nicely
                    try:
                        start_dt = datetime.strptime(test_start, "%Y-%m-%d %H:%M:%S")
                        formatted_start = start_dt.strftime("%Y-%m-%d at %H:%M:%S")
                        self.start_time_label.setText(f"✅ {formatted_start}")
                        self.start_time_label.setStyleSheet("""
                            QLabel {
                                background-color: #d4edda;
                                border: 2px solid #c3e6cb;
                                border-radius: 8px;
                                padding: 12px 15px;
                                font-size: 14px;
                                color: #155724;
                                font-weight: 600;
                            }
                        """)
                    except:
                        self.start_time_label.setText(f"✅ {test_start}")
                        self.start_time_label.setStyleSheet("""
                            QLabel {
                                background-color: #d4edda;
                                border: 2px solid #c3e6cb;
                                border-radius: 8px;
                                padding: 12px 15px;
                                font-size: 14px;
                                color: #155724;
                                font-weight: 600;
                            }
                        """)
                else:
                    self.start_time_label.setText("⚠️ No start time recorded")
                    self.start_time_label.setStyleSheet("""
                        QLabel {
                            background-color: #fff3cd;
                            border: 2px solid #ffeaa7;
                            border-radius: 8px;
                            padding: 12px 15px;
                            font-size: 14px;
                            color: #856404;
                            font-weight: 500;
                        }
                    """)
            else:
                self.start_time_label.setText("Not selected")
                self.start_time_label.setStyleSheet("""
                    QLabel {
                        background-color: #e9ecef;
                        border: 2px solid #dee2e6;
                        border-radius: 8px;
                        padding: 12px 15px;
                        font-size: 14px;
                        color: #495057;
                        font-weight: 500;
                    }
                """)
        except Exception as e:
            print(f"Error updating start time display: {e}")
            self.start_time_label.setText("Error loading start time")

    def load_assigned_products(self):
        """Load only products that are currently In Progress for testing updates"""
        try:
            # Clear existing items
            self.product_combo.clear()

            print(f"🔍 Loading In Progress products for tester ID: {self.user_id}")

            with sqlite3.connect(db_path, timeout=30.0) as conn:
                cursor = conn.cursor()

                # Only get products with "In Progress" status (progress_status_id = 1)
                cursor.execute("""
                    SELECT DISTINCT
                        P.product_id,
                        P.product_name,
                        P.batch,
                        T.test_start,
                        COALESCE(rl.rack_location_name, 'No Location') as location
                    FROM products P
                    INNER JOIN product_tester_assignments A ON P.product_id = A.product_id
                    INNER JOIN testing T ON T.product_id = P.product_id AND T.user_id = A.tester_id
                    LEFT JOIN racklocations rl ON P.rack_location_id = rl.rack_location_id
                    INNER JOIN progress_status ps ON T.progress_status_id = ps.status_id
                    WHERE A.tester_id = ? AND ps.status_name = 'In Progress'
                    ORDER BY T.test_start ASC
                """, (self.user_id,))

                products = cursor.fetchall()

                if products:
                    print(f"✓ Found {len(products)} products in progress")

                    # Add a default "Select Product" option
                    self.product_combo.addItem("-- Select In Progress Product --", None)

                    # Add each in-progress product with detailed information
                    for product in products:
                        product_id, product_name, batch, test_start, location = product

                        # Create a descriptive display text
                        display_text = f"[{product_id}] {product_name}"
                        if batch:
                            display_text += f" (Batch: {batch})"
                        if test_start:
                            # Format start date nicely
                            try:
                                start_dt = datetime.strptime(test_start, "%Y-%m-%d %H:%M:%S")
                                start_date = start_dt.strftime("%Y-%m-%d")
                                display_text += f" - Started: {start_date}"
                            except:
                                display_text += f" - Started: {test_start}"

                        # Store the product_id and start time as user data for easy retrieval
                        self.product_combo.addItem(display_text, {'product_id': product_id, 'test_start': test_start})

                        print(f"  - Product: {product_name} (ID: {product_id}, Started: {test_start})")
                else:
                    print("⚠️ No products currently in progress")
                    self.product_combo.addItem("No products currently in progress", None)

                    # Show a helpful message
                    QMessageBox.information(self, "No Active Tests",
                                            "You currently have no products in progress.\n\n"
                                            "Go to 'Product Assignments' and click '▶️ Start' on a product to begin testing first.")

        except sqlite3.Error as e:
            print(f"✗ Database error loading in-progress products: {e}")
            self.product_combo.clear()
            self.product_combo.addItem("Error loading products", None)
            QMessageBox.critical(self, "Database Error",
                                 f"Failed to load in-progress products:\n{str(e)}\n\n"
                                 f"Please check your database connection and try again.")
        except Exception as e:
            print(f"✗ Unexpected error loading in-progress products: {e}")
            self.product_combo.clear()
            self.product_combo.addItem("Error loading products", None)
            QMessageBox.critical(self, "Error",
                                 f"An unexpected error occurred:\n{str(e)}")

    def clear_form(self):
        """Clear all form fields"""
        if hasattr(self, 'product_combo'):
            self.product_combo.setCurrentIndex(0)
        if hasattr(self, 'status_combo'):
            self.status_combo.setCurrentIndex(0)
        if hasattr(self, 'result_text'):
            self.result_text.clear()
        if hasattr(self, 'start_time_label'):
            self.start_time_label.setText("Not selected")
            self.start_time_label.setStyleSheet("""
                QLabel {
                    background-color: #e9ecef;
                    border: 2px solid #dee2e6;
                    border-radius: 8px;
                    padding: 12px 15px;
                    font-size: 14px;
                    color: #495057;
                    font-weight: 500;
                }
            """)
        if hasattr(self, 'image_label'):
            self.image_label.clear()
            self.image_label.setText("📷\nNo image selected")
        if hasattr(self, 'image_path_label'):
            self.image_path_label.setText("No file selected")
        self.attached_file_path = None

    def show_progress_tracker(self, layout):
        # Create standardized page layout
        content_widget, content_layout = self.create_standard_page_layout()

        # Add standardized title
        title_widget = self.create_standard_title(
            "📊 Progress Tracker",
            "Monitor your testing progress and completion statistics"
        )
        content_layout.addWidget(title_widget)

        if not MATPLOTLIB_AVAILABLE:
            error_label = QLabel("⚠️ Charts unavailable - matplotlib not installed")
            error_label.setStyleSheet("""
                font-size: 18px;
                color: #dc3545;
                padding: 20px;
                background-color: #f8d7da;
                border: 1px solid #f5c6cb;
                border-radius: 8px;
            """)
            content_layout.addWidget(error_label)
            layout.addWidget(content_widget)
            return

        # Create matplotlib figure
        fig = Figure(figsize=(10, 6))
        ax = fig.add_subplot(111)

        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()
            cursor.execute("""
                SELECT ps.status_name, COUNT(*) 
                FROM testing t
                LEFT JOIN progress_status ps ON t.progress_status_id = ps.status_id
                WHERE t.user_id = ? 
                GROUP BY ps.status_name
            """, (self.get_current_user_id(),))
            status_data = cursor.fetchall()

            if status_data and len(status_data) > 0:
                try:
                    statuses, counts = zip(*status_data)
                    # Filter out None values and ensure we have valid data
                    valid_data = [(s, c) for s, c in zip(statuses, counts) if s is not None and c is not None and c > 0]

                    if valid_data:
                        statuses, counts = zip(*valid_data)
                        color_map = {
                            "Testing Complete": "green",
                            "In Progress": "blue",
                            "On Hold": "orange",
                            "Sample Expired": "red",
                            "Not Started": "gray"
                        }
                        bar_colors = [color_map.get(status, 'black') for status in statuses]
                        ax.bar(statuses, counts, color=bar_colors)

                        total = sum(counts)
                        completed = sum(c for s, c in zip(statuses, counts) if s == "Testing Complete")
                        progress_percent = int((completed / total) * 100) if total > 0 else 0

                        progress_label = QLabel(f"Total Progress: {progress_percent}%")
                        progress_label.setStyleSheet("""
                            font-size: 20px; 
                            padding: 5px;
                            color: #2c3e50;
                            font-weight: 600;
                        """)
                        content_layout.addWidget(progress_label)
                    else:
                        ax.bar(['No Valid Data'], [1], color=['lightgray'])
                        ax.set_title("No Valid Test Data")
                        no_data_label = QLabel("No valid test data available.")
                        no_data_label.setStyleSheet("""
                            font-size: 16px; 
                            padding: 5px;
                            color: #6c757d;
                        """)
                        content_layout.addWidget(no_data_label)
                except Exception as e:
                    print(f"Error processing chart data: {e}")
                    ax.bar(['Error'], [1], color=['red'])
                    ax.set_title("Error Processing Data")
                    error_label = QLabel("Error processing chart data.")
                    error_label.setStyleSheet("""
                        font-size: 16px; 
                        padding: 5px;
                        color: #dc3545;
                    """)
                    content_layout.addWidget(error_label)
            else:
                ax.bar(['No Data'], [1], color=['lightgray'])
                ax.set_title("No Test Data Available")
                no_data_label = QLabel("No test data available yet.")
                no_data_label.setStyleSheet("""
                    font-size: 16px; 
                    padding: 5px;
                    color: #6c757d;
                """)
                content_layout.addWidget(no_data_label)

        except sqlite3.Error as e:
            ax.bar(['Error'], [0], color=['red'])
            error_label = QLabel(f"Database error: {e}")
            error_label.setStyleSheet("color: red; font-size: 12px; padding: 5px;")
            content_layout.addWidget(error_label)
        finally:
            if conn:
                conn.close()

        ax.set_title("Test Result Summary")
        ax.set_ylabel("Number of Tests")

        canvas = FigureCanvas(fig)
        content_layout.addWidget(canvas)

        # Add to main layout
        layout.addWidget(content_widget)

    def show_test_history(self, layout):
        # Create standardized page layout
        content_widget, content_layout = self.create_standard_page_layout()

        # Add standardized title
        title_widget = self.create_standard_title(
            "📖 Test History",
            "View your complete testing history and export reports"
        )
        content_layout.addWidget(title_widget)

        # Export buttons layout
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)

        export_excel_btn = QPushButton("Export to Excel")
        export_excel_btn.setFixedSize(120, 30)
        export_excel_btn.setStyleSheet("""
            QPushButton {
                background-color: #9C27B0;  /* Purple */
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 15px;
            }
            QPushButton:hover {
                background-color: #7B1FA2;
            }
            QPushButton:pressed {
                background-color: #4A148C;
            }
        """)
        export_excel_btn.clicked.connect(lambda: self.export_history("excel"))
        button_layout.addWidget(export_excel_btn)

        export_pdf_btn = QPushButton("Export to PDF")
        export_pdf_btn.setFixedSize(120, 30)
        export_pdf_btn.setStyleSheet(export_excel_btn.styleSheet())
        export_pdf_btn.clicked.connect(lambda: self.export_history("pdf"))
        button_layout.addWidget(export_pdf_btn)

        button_layout.addStretch()

        content_layout.addLayout(button_layout)

        # Test history table
        self.history_table = QTableWidget()
        self.history_table.setColumnCount(5)
        self.history_table.setHorizontalHeaderLabels(["Product", "Status", "Start Date", "End Date", "Result"])
        self.history_table.setAlternatingRowColors(True)
        self.history_table.setStyleSheet("alternate-background-color: #f2f2f2; background-color: white;")
        self.history_table.horizontalHeader().setStretchLastSection(True)
        self.history_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.history_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.history_table.setEditTriggers(QTableWidget.NoEditTriggers)

        try:
            with sqlite3.connect(db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT P.product_name, COALESCE(ps.status_name, 'Not Started') as test_status, 
                           T.test_start, T.test_end, COALESCE(tr.result_name, 'No Result') as test_result
                    FROM testing T
                    INNER JOIN products P ON T.product_id = P.product_id
                    LEFT JOIN progress_status ps ON T.progress_status_id = ps.status_id
                    LEFT JOIN test_results tr ON T.test_result_id = tr.result_id
                    WHERE T.user_id = ?
                    ORDER BY T.test_start DESC
                """, (self.get_current_user_id(),))

                history_rows = cursor.fetchall()

                if history_rows:
                    self.history_table.setRowCount(len(history_rows))
                    for i, row in enumerate(history_rows):
                        for j, value in enumerate(row):
                            item = QTableWidgetItem(str(value) if value else "N/A")
                            item.setTextAlignment(Qt.AlignCenter)
                            self.history_table.setItem(i, j, item)
                else:
                    QMessageBox.information(self, "No Data", "You have no test history yet.")

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"Failed to load test history: {e}")

        content_layout.addWidget(self.history_table)

        # Add to main layout
        layout.addWidget(content_widget)

    def show_test_analysis(self, layout):
        """Combined page for Progress Tracker and Test Analysis.
        Heavy widgets (charts, stats) are hidden by default and can be toggled
        with a button to improve initial loading performance and UI cleanliness."""
        # Base layout
        content_widget, content_layout = self.create_standard_page_layout()

        # Page title
        title_widget = self.create_standard_title(
            "📈 Test Analysis",
            "Review overall progress and detailed analytics in one place"
        )
        content_layout.addWidget(title_widget)

        # Analytics widgets are shown immediately (toggle removed)
        analysis_container = QWidget()
        analysis_container.setVisible(True)
        analysis_layout = QVBoxLayout(analysis_container)
        analysis_layout.setSpacing(25)

        # 1️⃣  (Progress tracker bar chart removed as per latest requirements)

        # 2️⃣ Detailed Summary & Charts (existing analysis)
        # -- Summary group
        summary_group = QGroupBox("📊 Summary Statistics")
        summary_group.setStyleSheet("""
            QGroupBox {
                font-size: 18px;
                font-weight: 600;
                color: #495057;
                background-color: white;
                border: 2px solid #e9ecef;
                border-radius: 12px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        summary_layout = QVBoxLayout(summary_group)

        chart_group = QGroupBox("📈 Visual Analytics")
        chart_group.setStyleSheet(summary_group.styleSheet())
        chart_layout = QVBoxLayout(chart_group)

        # Calculate stats
        status_counts = {"On Hold": 0, "In Progress": 0, "Testing Complete": 0, "Sample Expired": 0, "Not Started": 0}
        total_duration = 0
        duration_count = 0
        product_count = {}
        try:
            with sqlite3.connect(db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT COALESCE(ps.status_name, 'Not Started'), T.test_start, T.test_end, P.product_name
                    FROM testing T
                    INNER JOIN products P ON T.product_id = P.product_id
                    LEFT JOIN progress_status ps ON T.progress_status_id = ps.status_id
                    WHERE T.user_id = ?
                """, (self.get_current_user_id(),))
                rows = cursor.fetchall()
                for status, start, end, product in rows:
                    if status in status_counts:
                        status_counts[status] += 1
                    if start and end:
                        try:
                            start_dt = datetime.strptime(start, '%Y-%m-%d')
                            end_dt = datetime.strptime(end, '%Y-%m-%d')
                            dur = (end_dt - start_dt).days
                            if dur >= 0:
                                total_duration += dur
                                duration_count += 1
                        except:
                            pass
                    if product:
                        product_count[product] = product_count.get(product, 0) + 1
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"Failed to fetch analysis data: {e}")
            return

        total_tests = sum(status_counts.values())
        avg_duration = round(total_duration / duration_count, 2) if duration_count else 0

        summary_text = f"""🧪 Total Tests: {total_tests}
✅ Testing Complete: {status_counts['Testing Complete']}
⏳ In Progress: {status_counts['In Progress']}
⏸️ On Hold: {status_counts['On Hold']}
⏰ Sample Expired: {status_counts['Sample Expired']}
💤 Not Started: {status_counts['Not Started']}
⏱️ Avg Duration (for completed): {avg_duration} days"""
        summary_label = QLabel(summary_text)
        summary_label.setStyleSheet("""
            font-family: 'Courier New', monospace; 
            padding: 20px;
            font-size: 16px;
            line-height: 1.6;
            background-color: #f8f9fa;
            border-radius: 8px;
            color: #495057;
        """)
        summary_layout.addWidget(summary_label)

        # Charts
        if MATPLOTLIB_AVAILABLE:
            try:
                fig_detail = Figure(figsize=(10, 7))
                fig_detail.patch.set_facecolor('white')
                # Pie chart only
                ax1 = fig_detail.add_subplot(111)
                non_zero = {k: v for k, v in status_counts.items() if v > 0}
                if non_zero:
                    labels = list(non_zero.keys())
                    vals = list(non_zero.values())
                    colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7'][:len(labels)]
                    wedges, texts, autotexts = ax1.pie(vals, labels=labels, autopct='%1.1f%%', colors=colors,
                                                       startangle=90, textprops={'fontsize': 10})
                    for autotext in autotexts:
                        autotext.set_color('white')
                        autotext.set_fontweight('bold')
                else:
                    ax1.pie([1], labels=['No Test Data'], colors=['#E8E8E8'], autopct='')
                ax1.set_title('Test Status Distribution', fontsize=14, fontweight='bold', pad=20)

                fig_detail.tight_layout(pad=3.0)
                canvas_detail = FigureCanvas(fig_detail)
                chart_layout.addWidget(canvas_detail)
            except Exception as e:
                err_lbl = QLabel(f"⚠️ Error generating charts: {e}")
                err_lbl.setStyleSheet("color:#dc3545;")
                chart_layout.addWidget(err_lbl)
        else:
            nc_lbl = QLabel("⚠️ Charts unavailable - matplotlib not installed")
            nc_lbl.setStyleSheet("color:#dc3545;")
            chart_layout.addWidget(nc_lbl)

        # Assemble groups into container
        analysis_layout.addWidget(summary_group)
        analysis_layout.addWidget(chart_group)

        # Finalize container
        content_layout.addWidget(analysis_container)

        # Add to main frame layout
        layout.addWidget(content_widget)

    def show_notifications(self, layout):
        # Create standardized page layout
        content_widget, content_layout = self.create_standard_page_layout()

        # Add standardized title
        title_widget = self.create_standard_title(
            "🔔 Notifications",
            "View system notifications and important reminders"
        )
        content_layout.addWidget(title_widget)

        current_user_id = self.get_current_user_id()

        # Insert expiry reminders
        self.insert_expiry_reminders(current_user_id)

        # Create scrollable list for notifications (card style)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        list_container = QWidget()
        list_layout = QVBoxLayout(list_container)
        list_layout.setSpacing(12)
        scroll.setWidget(list_container)

        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()
            cursor.execute("""
                SELECT message, timestamp, read FROM notifications
                WHERE user_id = ?
                ORDER BY timestamp DESC
            """, (current_user_id,))
            notifications = cursor.fetchall()

            if notifications:
                for message, timestamp, read in notifications:
                    card = QFrame()
                    card.setFrameShape(QFrame.StyledPanel)
                    card.setStyleSheet(f"""
                        QFrame {{
                            background-color: {'#f1f3f4' if read else '#e8eaf6'};
                            border: 1px solid #d6d8db;
                            border-radius: 8px;
                            padding: 12px;
                        }}
                    """)
                    h = QHBoxLayout(card)
                    msg_lbl = QLabel(message)
                    msg_lbl.setWordWrap(True)
                    msg_lbl.setStyleSheet("font-size:14px;")
                    ts_lbl = QLabel(timestamp)
                    ts_lbl.setStyleSheet("color:#6c757d;font-size:12px;")
                    h.addWidget(msg_lbl, 1)
                    h.addWidget(ts_lbl)
                    list_layout.addWidget(card)

            # Mark all unread notifications as read
            cursor.execute("""
                UPDATE notifications SET read = 1
                WHERE user_id = ? AND read = 0
            """, (current_user_id,))
            conn.commit()

        except sqlite3.Error as e:
            error_label = QLabel(f"Error loading notifications: {e}")
            error_label.setStyleSheet("color: red; padding: 10px;")
            content_layout.addWidget(error_label)
            layout.addWidget(content_widget)
            return
        finally:
            if conn:
                conn.close()

        list_layout.addStretch()
        content_layout.addWidget(scroll)

        # Add to main layout
        layout.addWidget(content_widget)

    def show_chat(self, layout):
        """Show the chat/announcement interface with light purple theme"""
        # Create standardized page layout
        content_widget, content_layout = self.create_standard_page_layout()

        # Add standardized title
        title_widget = self.create_standard_title(
            "💬 Chat / Announcement",
            "Communicate with team members and view system announcements"
        )
        content_layout.addWidget(title_widget)

        # Create and add the chat widget
        self.chat_widget = TesterChatWidget(user_info=self.user_info, parent=self)
        content_layout.addWidget(self.chat_widget)

        # Add to main layout
        layout.addWidget(content_widget)

    def attach_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "📷 Select Test Evidence Image",
            "",
            "Image files (*.png *.jpg *.jpeg *.bmp *.gif);;All files (*.*)"
        )

        if file_path:
            self.attached_file_path = file_path
            # Extract filename for display
            filename = file_path.split('/')[-1] if '/' in file_path else file_path.split('\\')[-1]
            self.image_path_label.setText(f"✅ {filename}")
            self.image_path_label.setStyleSheet("""
                color: #28a745;
                font-size: 14px;
                font-weight: 600;
            """)

            # Load and display image
            try:
                pixmap = QPixmap(file_path)
                scaled_pixmap = pixmap.scaled(300, 200, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.image_label.setPixmap(scaled_pixmap)
                self.image_label.setStyleSheet("""
                    QLabel {
                        border: 2px solid #28a745;
                        border-radius: 12px;
                        background-color: white;
                    }
                """)
            except Exception as e:
                self.image_label.setText("❌\nError loading image")
                self.image_label.setStyleSheet("""
                    QLabel {
                        border: 2px solid #dc3545;
                        border-radius: 12px;
                        background-color: #f8f9fa;
                        color: #dc3545;
                        font-size: 14px;
                    }
                """)
                print(f"Error loading image: {e}")
        else:
            self.attached_file_path = None
            self.image_label.clear()
            self.image_label.setText("📷\nNo image selected")
            self.image_label.setStyleSheet("""
                QLabel {
                    border: 2px dashed #dee2e6;
                    border-radius: 12px;
                    background-color: #f8f9fa;
                    color: #6c757d;
                    font-size: 14px;
                }
            """)
            self.image_path_label.setText("No file selected")
            self.image_path_label.setStyleSheet("""
                color: #6c757d;
                font-size: 14px;
                font-style: italic;
            """)

    def update_test_status(self):
        # Get the selected product data from the dropdown's user data
        current_data = self.product_combo.currentData()
        selected_text = self.product_combo.currentText()

        # Extract product_id from the data structure
        if isinstance(current_data, dict) and 'product_id' in current_data:
            product_id = current_data['product_id']
            test_start = current_data.get('test_start')
        else:
            product_id = current_data  # Fallback for old data structure
            test_start = None

        # Validate product selection
        if not product_id or selected_text in ["-- Select In Progress Product --", "No products currently in progress",
                                               "Error loading products"]:
            QMessageBox.warning(self, "Selection Required",
                                "Please select a product that is currently in progress.\n\n"
                                "If no products are available, go to 'Product Assignments' and start testing on a product first.")
            return

        # Validate status selection
        status_id = self.status_combo.currentData()
        if status_id is None:
            QMessageBox.warning(self, "Selection Required", "Please select a final test status.")
            return

        # Get other form data
        test_result_text = self.result_text.toPlainText().strip()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        image_path = self.attached_file_path if hasattr(self, 'attached_file_path') else None

        print(f"🔄 Updating test status for Product ID: {product_id}, Status ID: {status_id}")

        try:
            with sqlite3.connect('testing_system.db', timeout=10.0) as conn:
                cursor = conn.cursor()

                # Since we only show In Progress products, we know the test record exists
                cursor.execute("SELECT test_id FROM testing WHERE user_id = ? AND product_id = ?",
                               (self.user_id, product_id))
                existing_test = cursor.fetchone()

                if not existing_test:
                    QMessageBox.critical(self, "Error",
                                         "Test record not found. This product may not be in progress.\n\n"
                                         "Please refresh and try again.")
                    return

                # Determine test_result_id and test_end based on final status
                if status_id == 2:  # Testing Complete
                    test_result_id = 1  # Under Review
                    test_end = now  # Set end time for completed tests
                    print(f"  📅 Setting test completion time: {test_end}")
                elif status_id == 3:  # Sample Expired
                    test_result_id = 0  # Rejected
                    test_end = now  # Set end time for expired samples
                    print(f"  📅 Setting expiry time: {test_end}")
                else:  # On Hold
                    test_result_id = 1  # Under Review (keeping it active)
                    test_end = None  # Don't set end time for on-hold tests
                    print(f"  ⏸️ Test put on hold - no end time set")

                # Update the existing test record
                print(f"  ↻ Updating test record (ID: {existing_test[0]}) to final status")

                if test_end:
                    # Update with test_end timestamp for completed/expired tests
                    cursor.execute("""
                        UPDATE testing
                        SET progress_status_id = ?, test_result_id = ?, test_end = ?, test_image = ?
                        WHERE test_id = ?
                    """, (status_id, test_result_id, test_end, image_path, existing_test[0]))
                else:
                    cursor.execute("""
                        UPDATE testing
                        SET progress_status_id = ?, test_result_id = ?, test_image = ?
                        WHERE test_id = ?
                    """, (status_id, test_result_id, image_path, existing_test[0]))

                conn.commit()

                # Get updated status name for confirmation
                cursor.execute("SELECT status_name FROM progress_status WHERE status_id = ?", (status_id,))
                status_result = cursor.fetchone()
                status_name = status_result[0] if status_result else "Unknown"

                print(f"✓ Test status updated successfully")

                # Show success message with details
                success_msg = f"Test status updated successfully!\n\n"
                success_msg += f"Product: {selected_text.split(']')[1].split(' (')[0].strip() if ']' in selected_text else selected_text}\n"
                success_msg += f"New Status: {self.status_combo.currentText()}\n"

                if test_end:
                    if status_id == 2:  # Testing Complete
                        success_msg += f"✅ Test Completed: {test_end}\n"
                        # Calculate and show testing duration if we have start time
                        if test_start:
                            try:
                                start_dt = datetime.strptime(test_start, "%Y-%m-%d %H:%M:%S")
                                end_dt = datetime.strptime(test_end, "%Y-%m-%d %H:%M:%S")
                                duration = end_dt - start_dt

                                # Format duration nicely
                                days = duration.days
                                hours, remainder = divmod(duration.seconds, 3600)
                                minutes, _ = divmod(remainder, 60)

                                duration_parts = []
                                if days > 0:
                                    duration_parts.append(f"{days} day{'s' if days != 1 else ''}")
                                if hours > 0:
                                    duration_parts.append(f"{hours} hour{'s' if hours != 1 else ''}")
                                if minutes > 0:
                                    duration_parts.append(f"{minutes} minute{'s' if minutes != 1 else ''}")

                                if duration_parts:
                                    duration_str = ", ".join(duration_parts)
                                    success_msg += f"⏱️ Testing Duration: {duration_str}\n"
                                else:
                                    success_msg += f"⏱️ Testing Duration: Less than 1 minute\n"
                            except Exception as e:
                                print(f"Error calculating duration: {e}")
                    elif status_id == 3:  # Sample Expired
                        success_msg += f"⏰ Sample Expired: {test_end}\n"
                    else:
                        success_msg += f"End Time: {test_end}\n"

                if image_path:
                    filename = image_path.split('/')[-1] if '/' in image_path else image_path.split('\\')[-1]
                    success_msg += f"Evidence: {filename}\n"
                if test_result_text:
                    success_msg += f"Notes: {test_result_text[:50]}{'...' if len(test_result_text) > 50 else ''}"

                QMessageBox.information(self, "Success", success_msg)

                # Clear form and reload products to show updated status
                self.clear_form()
                self.load_assigned_products()

        except sqlite3.Error as e:
            print(f"✗ Database error: {e}")
            QMessageBox.critical(self, "Database Error", f"Failed to update test status:\n{str(e)}")
        except Exception as e:
            print(f"✗ Unexpected error: {e}")
            QMessageBox.critical(self, "Error", f"An unexpected error occurred:\n{str(e)}")

    def export_history(self, table, export_type):
        """Export test history to Excel or PDF"""
        try:
            data = []
            # Extract data from QTableWidget
            for row in range(table.rowCount()):
                row_data = []
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)

            if not data:
                QMessageBox.warning(self, "No Data", "No test history to export")
                return

            # Extract column headers
            columns = []
            for col in range(table.columnCount()):
                header_item = table.horizontalHeaderItem(col)
                columns.append(header_item.text() if header_item else f"Column {col + 1}")

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"test_history_{timestamp}"

            if export_type == "excel":
                filename += ".xlsx"
                self.export_to_excel(columns, data, filename)
            elif export_type == "pdf":
                filename += ".pdf"
                self.export_to_pdf(columns, data, filename)

            QMessageBox.information(self, "Export Successful", f"Exported to {filename}")

            # Open file after export
            if platform.system() == "Windows":
                os.startfile(filename)
            elif platform.system() == "Darwin":  # macOS
                os.system(f"open {filename}")
            else:  # Linux
                os.system(f"xdg-open {filename}")

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export: {str(e)}")

    def export_to_excel(self, columns, data, filename):
        """Export data to Excel file"""
        df = pd.DataFrame(data, columns=columns)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')

            worksheet = writer.sheets['Sheet1']
            for i, col in enumerate(df.columns):
                col_width = max(df[col].astype(str).map(len).max(), len(col))
                worksheet.column_dimensions[get_column_letter(i + 1)].width = col_width + 2

    def export_to_pdf(self, columns, data, filename):
        pdf = FPDF()
        pdf.add_page()

        pdf.set_font("Arial" if platform.system() == "Windows" else "Helvetica", size=10)

        pdf.cell(200, 10, txt="Test History Report", ln=1, align='C')
        pdf.cell(200, 10, txt=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=1, align='C')

        col_widths = [40, 30, 30, 30, 60]

        # Header
        for i, col in enumerate(columns):
            pdf.cell(col_widths[i], 10, col, border=1)
        pdf.ln()

        # Rows
        for row in data:
            for i, item in enumerate(row):
                pdf.cell(col_widths[i], 10, str(item), border=1)
            pdf.ln()

        pdf.output(filename)

    def export_history(self, format_type):
        # Get column headers
        columns = [self.history_table.horizontalHeaderItem(i).text() for i in range(self.history_table.columnCount())]

        # Get table data
        data = []
        for row in range(self.history_table.rowCount()):
            row_data = []
            for col in range(self.history_table.columnCount()):
                item = self.history_table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)

        options = QFileDialog.Options()
        if format_type == "excel":
            filename, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)",
                                                      options=options)
            if filename:
                self.export_to_excel(columns, data, filename)
                QMessageBox.information(self, "Export", f"Excel file saved to:\n{filename}")
                self.open_file(filename)  # 👈 open after saving

        elif format_type == "pdf":
            filename, _ = QFileDialog.getSaveFileName(self, "Save PDF File", "", "PDF Files (*.pdf)", options=options)
            if filename:
                self.export_to_pdf(columns, data, filename)
                QMessageBox.information(self, "Export", f"PDF file saved to:\n{filename}")
                self.open_file(filename)  # 👈 open after saving

    def open_file(self, filepath):
        import subprocess
        import platform
        if platform.system() == 'Windows':
            os.startfile(filepath)
        elif platform.system() == 'Darwin':  # macOS
            subprocess.call(['open', filepath])
        else:  # Linux
            subprocess.call(['xdg-open', filepath])

    def insert_expiry_reminders(self, user_id):
        today = date.today()
        upcoming = today + timedelta(days=60)

        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Convert dates to strings to avoid deprecation warning
            today_str = today.strftime('%Y-%m-%d')
            upcoming_str = upcoming.strftime('%Y-%m-%d')

            cursor.execute("""
                SELECT p.product_id, p.expired_date, p.product_name
                FROM products p
                INNER JOIN product_tester_assignments a ON p.product_id = a.product_id
                WHERE a.tester_id = ? AND DATE(p.expired_date) BETWEEN ? AND ?
            """, (user_id, today_str, upcoming_str))

            rows = cursor.fetchall()

            for product_id, expiry_date, product_name in rows:
                # Prevent duplicate reminders
                cursor.execute("""
                    SELECT 1 FROM notifications
                    WHERE user_id = ? AND message LIKE ? AND DATE(timestamp) = ?
                """, (user_id, f"%{product_name}%expires%", today_str))
                exists = cursor.fetchone()

                if not exists:
                    message = f"Reminder: Product '{product_name}' (ID: {product_id}) expires on {expiry_date}. Please test it soon."
                    cursor.execute("""
                        INSERT INTO notifications (user_id, message, timestamp, read)
                        VALUES (?, ?, ?, 0)
                    """, (user_id, message, datetime.now()))

            conn.commit()

        except sqlite3.Error as e:
            print("Reminder insertion failed:", e)
            # Optional: Show error message to user if this is called from UI context
            if hasattr(self, 'parent') and self.parent:
                QMessageBox.critical(self.parent, "Database Error", f"Failed to insert expiry reminders: {e}")

        finally:
            if conn:
                conn.close()

    def test_insert_expiry_reminder(self):
        """Test method for expiry reminder insertion"""
        # Setup test data
        user_id = 1
        product_id = 200
        product_name = "Test Kit B"
        expiry_date = (date.today() + timedelta(days=1)).isoformat()

        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()
            cursor.execute("INSERT OR IGNORE INTO products (product_id, product_name, product_desc) VALUES (?, ?, ?)",
                           (product_id, product_name, "Test kit"))
            cursor.execute(
                "INSERT OR REPLACE INTO product_tester_assignments (product_id, tester_id) VALUES (?, ?)",
                (product_id, user_id))
            conn.commit()
            conn.close()

            # Run the method
            self.insert_expiry_reminders(user_id)

            # Check the notification
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()
            cursor.execute("SELECT message FROM notifications WHERE user_id = ? AND message LIKE ?",
                           (user_id, f"%{product_name}%"))
            result = cursor.fetchone()
            conn.close()

            if result:
                print("✅ Reminder successfully inserted:", result[0])
                # Optional: Show success message in UI
                if hasattr(self, 'parent') and self.parent:
                    QMessageBox.information(self.parent, "Test Successful",
                                            f"Reminder successfully inserted: {result[0]}")
                return True
            else:
                print("❌ Reminder not inserted.")
                # Optional: Show failure message in UI
                if hasattr(self, 'parent') and self.parent:
                    QMessageBox.warning(self.parent, "Test Failed", "Reminder not inserted.")
                return False

        except Exception as e:
            print(f"❌ Test failed with error: {e}")
            # Optional: Show error message in UI
            if hasattr(self, 'parent') and self.parent:
                QMessageBox.critical(self.parent, "Test Error", f"Test failed with error: {e}")
            return False

    def auto_start_testing_on_assignment(self, product_id, tester_id):
        """
        Automatically set product to 'In Progress' when assigned to a tester.
        This function should be called when a product is assigned to a tester.
        """
        try:
            with sqlite3.connect('testing_system.db', timeout=10.0) as conn:
                cursor = conn.cursor()
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # Check if test record already exists for this product and tester
                cursor.execute("SELECT test_id FROM testing WHERE user_id = ? AND product_id = ?",
                               (tester_id, product_id))
                existing_test = cursor.fetchone()

                if not existing_test:
                    # Create new test record with "In Progress" status (status_id = 1)
                    cursor.execute("""
                        INSERT INTO testing (user_id, product_id, test_start, progress_status_id, test_result_id)
                        VALUES (?, ?, ?, 1, 1)
                    """, (tester_id, product_id, now))

                    conn.commit()
                    print(f"✓ Auto-started testing for Product {product_id} assigned to Tester {tester_id}")
                    return True
                else:
                    print(f"⚠️ Test record already exists for Product {product_id} and Tester {tester_id}")
                    return False

        except sqlite3.Error as e:
            print(f"✗ Database error in auto-start testing: {e}")
            return False
        except Exception as e:
            print(f"✗ Unexpected error in auto-start testing: {e}")
            return False


if __name__ == '__main__':
    import sys
    from PyQt5.QtWidgets import QApplication

    try:
        app = QApplication(sys.argv)

        # Set application properties
        app.setApplicationName("Medical Tester Dashboard")
        app.setApplicationVersion("1.0")
        app.setOrganizationName("Medical Testing System")

        # Create and show main window
        window = TesterHomePage()
        window.show()

        # Start the application
        sys.exit(app.exec_())

    except Exception as e:
        print(f"Application startup error: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)

