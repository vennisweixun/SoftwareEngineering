import sys
import sqlite3
import os
import random
import time
import gc  # Garbage collection for memory management
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QLabel, QLineEdit, QPushButton, QComboBox, QDateEdit,
    QTableWidget, QTableWidgetItem, QScrollArea, QFrame, QDialog,
    QFormLayout, QMessageBox, QTextEdit, QHeaderView, QSplitter,
    QTabWidget, QProgressBar, QGroupBox, QCheckBox, QSpinBox, QButtonGroup,
    QStackedWidget, QProgressDialog, QRadioButton, QFileDialog, QListWidget, QListWidgetItem, QSizePolicy,
    QTreeWidget, QTreeWidgetItem
)
from PyQt5.QtCore import Qt, QTimer, QDate, pyqtSignal, QDateTime
from PyQt5.QtGui import QFont, QPalette, QColor, QIcon, QPixmap, QKeySequence
from PyQt5.QtWidgets import QShortcut
import re
from datetime import timezone, timedelta
import writer
# newdashboard_main.py – helper imports
# from newdashboard_extras import *  # redundant wildcard import removed

# Import helper functions/classes from the extras module.  We still keep the
# wildcard import for the many UI helper classes, but we pull the key
# database-chat helpers in explicitly so that static analysers can resolve
# them correctly.
from newdashboard_extras import (
    get_all_messages,
    get_all_users_for_chat,
    insert_message,
    get_private_messages,
    insert_private_message,
    format_malaysia_time,
)

# Import the remainder (UI helper widgets, dialogs, etc.)
from newdashboard_extras import *  # noqa: F401,F403

# Malaysia timezone (UTC+8)
MALAYSIA_TZ = timezone(timedelta(hours=8))

# Buffer overflow protection constants
MAX_STRING_LENGTH = 500
MAX_RECORDS_PER_QUERY = 100
MAX_CARDS_DISPLAY = 20
MAX_CHAT_MESSAGES = 50


def safe_string(value, max_length=MAX_STRING_LENGTH):
    """Safely convert value to string with length protection"""
    if value is None:
        return ""
    str_val = str(value)
    if len(str_val) > max_length:
        return str_val[:max_length] + "..."
    return str_val


class Dashboard(QMainWindow):
    def __init__(self, user_role="Admin", user_info=None):
        super().__init__()
        print("Dashboard.__init__: Starting initialization...")

        self.user_role = user_role
        self.user_info = user_info or {}
        self.username = self.user_info.get('username', 'Unknown')
        self.fullname = self.user_info.get('fullname', self.username)
        self.login_time = datetime.now().strftime('%H:%M:%S')
        self.is_superadmin_mode = False
        print("Dashboard.__init__: Basic attributes set")

        # Define role colors for chat
        self.role_colors = {
            'superadmin': '#FF9800',  # Orange - Highest authority
            'admin': '#4CAF50',  # Green - Administrative
            'owner': '#2196F3',  # Blue - Business owner
            'tester': '#E91E63'  # Pink - Testing/Quality
        }

        # Chat variables
        self.active_pm_user = None
        self.active_pm_role = None
        self.all_users = []
        self.is_private_chat_mode = False  # Track if we're in private chat mode

        print("Dashboard.__init__: About to init database...")
        self.init_database()
        print("Dashboard.__init__: Database initialized")

        try:
            print("Dashboard.__init__: About to setup UI...")
            self.setupUI()
            print("Dashboard.__init__: UI setup completed")
        except Exception as e:
            print(f"Error in setupUI: {e}")
            import traceback
            traceback.print_exc()
            return

        try:
            print("Dashboard.__init__: About to load data...")
            self.load_data()
            print("Dashboard.__init__: Data loaded")
        except Exception as e:
            print(f"Error in load_data: {e}")
            import traceback
            traceback.print_exc()

        # Set window title and icon
        print("Dashboard.__init__: Setting window title...")
        self.setWindowTitle("Medical Testing System - Super Admin Dashboard")

        # Set minimum size for windowed mode
        self.setMinimumSize(1200, 800)

        # Don't auto-show - let the calling code control when to show
        # self.showMaximized()  # Commented out - will be called by login window

        # Add fullscreen toggle shortcut (F11)
        self.fullscreen_shortcut = QShortcut(QKeySequence("F11"), self)
        self.fullscreen_shortcut.activated.connect(self.toggle_fullscreen)

        # Final memory cleanup
        gc.collect()
        print("Dashboard.__init__: Initialization complete!")

        # Auto-refresh timer (with safe intervals to prevent buffer issues)
        self.refresh_timer = QTimer()
        self.refresh_timer.timeout.connect(self.safe_refresh_dashboard)
        self.refresh_timer.start(300000)  # Refresh every 5 minutes (very safe interval)

        # Chat refresh timer (with safe intervals to prevent buffer issues)
        self.chat_refresh_timer = QTimer()
        self.chat_refresh_timer.timeout.connect(self.safe_refresh_chat)
        self.chat_refresh_timer.start(30000)  # Refresh chat every 30 seconds (safe interval)

    def init_database(self):
        """Initialize the database connection and verify required tables"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Create chat tables if they don't exist
            create_chat_tables(cursor)

            # Ensure products table has required columns for approval system
            self.ensure_approval_columns(cursor)

            # Verify that required tables exist
            cursor.execute("""
                SELECT name FROM sqlite_master 
                WHERE type='table' AND name IN ('products', 'users', 'roles', 'branches')
            """)
            existing_tables = [row[0] for row in cursor.fetchall()]

            print(f"✓ Connected to testing_system.db")
            print(f"✓ Found tables: {existing_tables}")

            conn.commit()
            conn.close()

        except Exception as e:
            print(f"✗ Database initialization error: {e}")
            # Create basic tables if database doesn't exist
            self.create_basic_tables()

    def ensure_approval_columns(self, cursor):
        """Ensure products table has status, rejection_comment, and owner_id columns"""
        try:
            # Check existing columns
            cursor.execute("PRAGMA table_info(products)")
            existing_cols = [col[1] for col in cursor.fetchall()]

            # Add status column if it doesn't exist
            if 'status' not in existing_cols:
                cursor.execute("ALTER TABLE products ADD COLUMN status TEXT DEFAULT 'pending'")
                print("✓ Added status column to products table")

            # Add rejection_comment column if it doesn't exist
            if 'rejection_comment' not in existing_cols:
                cursor.execute("ALTER TABLE products ADD COLUMN rejection_comment TEXT")
                print("✓ Added rejection_comment column to products table")

            # Add owner_id column if it doesn't exist (for proper owner tracking)
            if 'owner_id' not in existing_cols:
                cursor.execute("ALTER TABLE products ADD COLUMN owner_id INTEGER")
                print("✓ Added owner_id column to products table")

                # Migrate existing data from user_id to owner_id if user_id exists
                if 'user_id' in existing_cols:
                    cursor.execute("UPDATE products SET owner_id = user_id WHERE owner_id IS NULL")
                    print("✓ Migrated existing user_id data to owner_id column")

        except Exception as e:
            print(f"✗ Error ensuring approval columns: {e}")

    def create_basic_tables(self):
        """Create basic tables if they don't exist"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Activity log table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS activity_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user TEXT NOT NULL,
                    action TEXT NOT NULL,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                    details TEXT
                )
            """)

            conn.commit()
            conn.close()
            print("✓ Created basic tables")

        except Exception as e:
            print(f"✗ Error creating basic tables: {e}")

    def setupUI(self):
        self.setGeometry(100, 100, 1400, 900)

        # Green theme stylesheet - enhanced with more professional gradients
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f3f8f3;
            }
            QWidget {
                font-family: 'Segoe UI', sans-serif;
            }
            QGroupBox {
                font-weight: bold;
                border: 1px solid #81C784;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 12px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: #1B5E20;
            }
            QPushButton {
                background-color: #388E3C;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 10px 16px;
                font-weight: bold;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #2E7D32;
            }
            QPushButton:pressed {
                background-color: #1B5E20;
            }
            QLineEdit {
                border: 1px solid #81C784;
                border-radius: 20px;
                padding: 8px 15px;
                background-color: white;
            }
            QComboBox, QDateEdit {
                border: 1px solid #81C784;
                border-radius: 4px;
                padding: 6px;
                background-color: white;
            }
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus {
                border-color: #388E3C;
                border-width: 2px;
            }
            QTableWidget {
                gridline-color: #E8F5E9;
                background-color: white;
                alternate-background-color: #F1F8E9;
                selection-background-color: #81C784;
            }
            QHeaderView::section {
                background-color: #388E3C;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
            }
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QTextEdit {
                border: 1px solid #81C784;
                border-radius: 4px;
                background-color: white;
            }
            QTabWidget::pane {
                border: 1px solid #81C784;
                border-radius: 4px;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #C8E6C9;
                color: #1B5E20;
                padding: 10px 20px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                background-color: #388E3C;
                color: white;
                font-weight: bold;
            }
            QTabBar::tab:hover:!selected {
                background-color: #A5D6A7;
            }
        """)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QVBoxLayout()

        # Top header bar - Red button for Shelf Life Management
        top_header = QHBoxLayout()

        shelf_life_btn = QPushButton("Shelf Life Management")
        shelf_life_btn.setStyleSheet("""
            QPushButton {
                background-color: #D32F2F;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 12px 24px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #C62828;
            }
        """)
        top_header.addWidget(shelf_life_btn)

        # Global Search panel - works across all pages
        search_container = QHBoxLayout()
        search_container.setContentsMargins(0, 0, 0, 0)

        self.search_panel = QLineEdit()
        self.search_panel.setPlaceholderText("🔍 Search (Product ID/Name/Batch/Owner/SKU/Barcode)")
        self.search_panel.setMinimumWidth(700)
        self.search_panel.setStyleSheet("""
            QLineEdit {
                border: 2px solid #4CAF50;
                border-radius: 25px;
                padding: 12px 20px;
                background-color: white;
                font-size: 14px;
                font-weight: 500;
            }
            QLineEdit:focus {
                border: 2px solid #2E7D32;
                background-color: #F1F8E9;
            }
        """)
        self.search_panel.textChanged.connect(self.on_global_search_changed)
        self.search_panel.returnPressed.connect(self.perform_global_search)

        # Clear search button
        clear_search_btn = QPushButton("✕")
        clear_search_btn.setFixedSize(35, 35)
        clear_search_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF5722;
                color: white;
                border: none;
                border-radius: 17px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #E64A19; }
        """)
        clear_search_btn.clicked.connect(self.clear_global_search)
        clear_search_btn.setToolTip("Clear search")

        search_container.addWidget(self.search_panel)
        search_container.addWidget(clear_search_btn)

        search_widget = QWidget()
        search_widget.setLayout(search_container)
        top_header.addWidget(search_widget)

        # Image upload button with hover text effect
        self.upload_btn = QPushButton("📷")
        self.upload_btn.setToolTip("Upload Product Image")
        self.upload_btn.clicked.connect(self.upload_product_image)
        self.upload_btn.setStyleSheet("""
            QPushButton {
                background-color: #E8F5E9;
                color: #2E7D32;
                border: 2px solid #C8E6C9;
                border-radius: 15px;
                font-size: 12px;
                font-weight: bold;
                min-width: 30px;
                min-height: 30px;
                padding: 5px;
            }
            QPushButton:hover {
                background-color: #C8E6C9;
                border-color: #4CAF50;
                color: #1B5E20;
            }
            QPushButton:pressed {
                background-color: #A5D6A7;
                border-color: #388E3C;
            }
        """)

        # Add hover events to change button text
        self.upload_btn.enterEvent = self.upload_btn_enter
        self.upload_btn.leaveEvent = self.upload_btn_leave

        top_header.addWidget(self.upload_btn)

        main_layout.addLayout(top_header)

        # Create splitter for main content
        splitter = QSplitter(Qt.Horizontal)

        # Left panel - Navigation menu
        left_panel = self.create_left_panel()
        splitter.addWidget(left_panel)

        # Right panel - Content area
        right_panel = self.create_right_panel()
        splitter.addWidget(right_panel)

        # Set initial sizes - make left panel narrower
        splitter.setSizes([250, 1150])
        main_layout.addWidget(splitter)

        # Username and login time at bottom
        bottom_layout = QHBoxLayout()

        # Create vertical layout for username and login time
        user_info_layout = QVBoxLayout()

        username_label = QLabel(f"Username: {self.username}")
        username_label.setStyleSheet("font-weight: bold; margin-bottom: 2px;")
        user_info_layout.addWidget(username_label)

        login_time = QLabel(f"Time login: {self.login_time}")
        login_time.setStyleSheet("font-size: 12px; color: #666;")
        user_info_layout.addWidget(login_time)

        bottom_layout.addLayout(user_info_layout)

        # Fullscreen toggle button
        self.fullscreen_btn = QPushButton("🖥️ Fullscreen")
        self.fullscreen_btn.setFixedSize(120, 50)
        self.fullscreen_btn.setToolTip("Toggle fullscreen mode (F11)")
        self.fullscreen_btn.setStyleSheet("""
            QPushButton {
                background-color: #E3F2FD;
                color: #1565C0;
                border: 2px solid #BBDEFB;
                border-radius: 8px;
                font-size: 12px;
                font-weight: bold;
                padding: 5px 10px;
                margin-left: 15px;
            }
            QPushButton:hover {
                background-color: #BBDEFB;
                border-color: #2196F3;
                color: #0D47A1;
            }
            QPushButton:pressed {
                background-color: #90CAF9;
                border-color: #1976D2;
            }
        """)
        self.fullscreen_btn.clicked.connect(self.toggle_fullscreen)
        bottom_layout.addWidget(self.fullscreen_btn)

        # Logout button beside username and time
        self.logout_btn = QPushButton("🚪 Logout")
        self.logout_btn.setFixedSize(100, 50)
        self.logout_btn.setToolTip("Logout from the system")
        self.logout_btn.setStyleSheet("""
            QPushButton {
                background-color: #FFEBEE;
                color: #C62828;
                border: 2px solid #FFCDD2;
                border-radius: 8px;
                font-size: 12px;
                font-weight: bold;
                padding: 5px 10px;
                margin-left: 15px;
            }
            QPushButton:hover {
                background-color: #FFCDD2;
                border-color: #F44336;
                color: #B71C1C;
            }
            QPushButton:pressed {
                background-color: #EF9A9A;
                border-color: #D32F2F;
            }
        """)
        self.logout_btn.clicked.connect(self.logout)
        bottom_layout.addWidget(self.logout_btn)

        bottom_layout.addStretch()

        # Super Admin Mode button at bottom (only for admin/superadmin users)
        self.superadmin_button = QPushButton("Super Admin Mode")
        self.superadmin_button.clicked.connect(self.toggle_superadmin_mode)
        self.superadmin_button.setStyleSheet("""
            QPushButton {
                background-color: #00695C;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 10px 20px;
                font-weight: bold;
                max-width: 200px;
            }
            QPushButton:hover {
                background-color: #004D40;
            }
        """)

        # Only show superadmin button for admin and superadmin users
        user_role_lower = self.user_role.lower()
        if user_role_lower in ['admin', 'superadmin']:
            self.superadmin_button.setVisible(True)
            print(f"✓ Superadmin button enabled for {user_role_lower} user")
        else:
            self.superadmin_button.setVisible(False)
            print(f"✗ Superadmin button hidden for {user_role_lower} user")

        bottom_layout.addWidget(self.superadmin_button)

        main_layout.addLayout(bottom_layout)

        central_widget.setLayout(main_layout)

    def create_left_panel(self):
        """Create the left panel with navigation menu"""
        left_widget = QWidget()
        left_widget.setObjectName("left_panel")  # Set an object name for easier access
        left_layout = QVBoxLayout()
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(0)

        # Navigation menu buttons - styled to match the image
        menu_style = """
            QPushButton {
                background-color: white;
                color: #333;
                border: none;
                border-bottom: 1px solid #E0E0E0;
                text-align: left;
                padding: 15px;
                font-size: 14px;
                font-weight: normal;
                border-radius: 0px;
            }
            QPushButton:hover {
                background-color: #E8F5E9;
                color: #2E7D32;
                font-weight: bold;
            }
            QPushButton:pressed, QPushButton:checked {
                background-color: #C8E6C9;
                color: #2E7D32;
                font-weight: bold;
                border-left: 4px solid #2E7D32;
            }
        """

        # Create a button group to manage selection state
        self.menu_button_group = QButtonGroup()
        self.menu_button_group.setExclusive(True)

        # Active menu item style - green background
        active_style = """
            background-color: #C8E6C9;
            color: #2E7D32;
            font-weight: bold;
            border-left: 4px solid #2E7D32;
        """

        # Inventory List button
        self.inventory_btn = QPushButton("📦 Inventory List")
        self.inventory_btn.setStyleSheet(menu_style)
        self.inventory_btn.setMinimumHeight(50)
        self.inventory_btn.setCheckable(True)
        self.inventory_btn.setChecked(True)
        self.inventory_btn.clicked.connect(lambda: self.switch_content("inventory"))
        self.menu_button_group.addButton(self.inventory_btn)
        left_layout.addWidget(self.inventory_btn)

        # Pending Approvals button
        self.pending_btn = QPushButton("🚨 Pending Approvals")
        self.pending_btn.setStyleSheet(menu_style)
        self.pending_btn.setMinimumHeight(50)
        self.pending_btn.setCheckable(True)
        self.pending_btn.clicked.connect(lambda: self.switch_content("pending"))
        self.menu_button_group.addButton(self.pending_btn)
        left_layout.addWidget(self.pending_btn)

        # Upcoming Maturation button
        self.maturation_btn = QPushButton("🕒 Upcoming Maturation")
        self.maturation_btn.setStyleSheet(menu_style)
        self.maturation_btn.setMinimumHeight(50)
        self.maturation_btn.setCheckable(True)
        self.maturation_btn.clicked.connect(lambda: self.switch_content("maturation"))
        self.menu_button_group.addButton(self.maturation_btn)
        left_layout.addWidget(self.maturation_btn)

        # Assign Testers button
        self.assign_btn = QPushButton("📝 Assign Testers")
        self.assign_btn.setStyleSheet(menu_style)
        self.assign_btn.setMinimumHeight(50)
        self.assign_btn.setCheckable(True)
        self.assign_btn.clicked.connect(lambda: self.switch_content("assign"))
        self.menu_button_group.addButton(self.assign_btn)
        left_layout.addWidget(self.assign_btn)

        # Chat/Announcement button
        self.chat_btn = QPushButton("💭 Chat/ Announcement")
        self.chat_btn.setStyleSheet(menu_style)
        self.chat_btn.setMinimumHeight(50)
        self.chat_btn.setCheckable(True)
        self.chat_btn.clicked.connect(lambda: self.switch_content("chat"))
        self.menu_button_group.addButton(self.chat_btn)
        left_layout.addWidget(self.chat_btn)

        # User Management button
        self.user_mgmt_btn = QPushButton("👤 User Management")
        self.user_mgmt_btn.setStyleSheet(menu_style)
        self.user_mgmt_btn.setMinimumHeight(50)
        self.user_mgmt_btn.setCheckable(True)
        self.user_mgmt_btn.clicked.connect(lambda: self.switch_content("user_management"))
        self.menu_button_group.addButton(self.user_mgmt_btn)
        left_layout.addWidget(self.user_mgmt_btn)

        # My Profile button
        self.profile_btn = QPushButton("👸 My Profile")
        self.profile_btn.setStyleSheet(menu_style)
        self.profile_btn.setMinimumHeight(50)
        self.profile_btn.setCheckable(True)
        self.profile_btn.clicked.connect(lambda: self.switch_content("my_profile"))
        self.menu_button_group.addButton(self.profile_btn)
        left_layout.addWidget(self.profile_btn)

        # Admin functions area - initially hidden
        self.admin_functions_area = QWidget()
        self.admin_functions_area.setStyleSheet("""
            QWidget {
                background-color: #f8f8f8;
                border: 1px solid #ddd;
                border-radius: 5px;
                margin: 10px 5px;
            }
        """)
        self.admin_functions_area.setVisible(False)

        admin_layout = QVBoxLayout()
        admin_layout.setContentsMargins(10, 10, 10, 10)
        admin_layout.setSpacing(5)

        # Super Admin functions title
        admin_title = QLabel("👑 Super Admin Functions 👑")
        admin_title.setStyleSheet("""
            QLabel {
                font-weight: bold;
                color: #2E7D32;
                font-size: 12px;
                margin-bottom: 5px;
                border: none;
            }
        """)
        admin_layout.addWidget(admin_title)

        # Admin function buttons
        admin_btn_style = """
            QPushButton {
                background-color: white;
                color: #333;
                border: 1px solid #ddd;
                border-radius: 3px;
                text-align: left;
                padding: 8px 10px;
                font-size: 11px;
                font-weight: normal;
                margin: 1px 0px;
            }
            QPushButton:hover {
                background-color: #E8F5E9;
                color: #2E7D32;
                border-color: #2E7D32;
            }
            QPushButton:pressed {
                background-color: #C8E6C9;
            }
        """

        self.admin_system_btn = QPushButton("⚙️ System Settings")
        self.admin_system_btn.setStyleSheet(admin_btn_style)
        self.admin_system_btn.setMinimumHeight(35)
        self.admin_system_btn.clicked.connect(lambda: self.switch_content("system_settings"))
        admin_layout.addWidget(self.admin_system_btn)

        # Analysis button - Super Admin only
        self.admin_analysis_btn = QPushButton("📊 Analysis Dashboard")
        self.admin_analysis_btn.setStyleSheet(admin_btn_style)
        self.admin_analysis_btn.setMinimumHeight(35)
        self.admin_analysis_btn.clicked.connect(lambda: self.switch_content("analysis"))
        admin_layout.addWidget(self.admin_analysis_btn)

        self.admin_functions_area.setLayout(admin_layout)
        left_layout.addWidget(self.admin_functions_area)

        # Add spacer to push everything to the top
        left_layout.addStretch()

        left_widget.setLayout(left_layout)
        left_widget.setStyleSheet("""
            QWidget {
                background-color: white;
                border-right: 1px solid #E0E0E0;
            }
        """)

        return left_widget

    def create_right_panel(self):
        """Create the right panel with content area - styled to match the image"""
        right_widget = QWidget()
        right_layout = QVBoxLayout()
        right_layout.setContentsMargins(20, 20, 20, 20)

        # Main content area - beige background
        self.content_area = QWidget()
        self.content_area.setStyleSheet("""
            QWidget {
                background-color: #f5f5f0;
                border-radius: 8px;
            }
        """)
        self.content_stack = QStackedWidget()

        # Create all content panels
        self.create_inventory_panel()
        self.create_pending_panel()
        self.create_maturation_panel()
        self.create_assign_panel()
        self.create_chat_panel()
        self.create_user_management_panel()
        self.create_my_profile_panel()
        self.create_analysis_panel()

        # Add all panels to the stacked widget
        main_layout = QVBoxLayout()
        main_layout.addWidget(self.content_stack)
        self.content_area.setLayout(main_layout)

        right_layout.addWidget(self.content_area)
        right_widget.setLayout(right_layout)
        return right_widget

    def create_inventory_panel(self):
        """Create the inventory list panel"""
        inventory_panel = QWidget()
        content_layout = QVBoxLayout()
        content_layout.setContentsMargins(20, 20, 20, 20)

        # Header with action buttons
        header_actions = QHBoxLayout()
        manage_btn = QPushButton("🛠️ Manage Products")
        manage_btn.clicked.connect(self.open_manage_products)
        add_btn = QPushButton("➕ Add Products")
        add_btn.clicked.connect(self.open_add_products)
        excel_btn = QPushButton("📄 Export Excel")
        excel_btn.setToolTip("Export today's products (with unit barcodes) to Excel")
        excel_btn.clicked.connect(self.export_today_products_to_excel)

        for btn in (manage_btn, add_btn, excel_btn):
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 8px 18px;
                    font-weight: bold;
                }
                QPushButton:hover { background-color: #388E3C; }
            """)
        header_actions.addWidget(manage_btn)
        header_actions.addWidget(add_btn)
        header_actions.addWidget(excel_btn)
        header_actions.addStretch()

        content_layout.addLayout(header_actions)

        # Table for inventory with enhanced columns
        self.inventory_table = QTableWidget()
        self.inventory_table.setColumnCount(11)
        self.inventory_table.setHorizontalHeaderLabels([
            "Product Name", "Owner", "Owner ID", "Branch", "Tester", "Arrival Date", "Batch",
            "Rack Location", "SKU", "Expired Date", "Barcode"
        ])

        # Make table rows clickable
        self.inventory_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.inventory_table.itemDoubleClicked.connect(self.show_product_detail_popup)
        self.inventory_table.horizontalHeader().setStretchLastSection(True)
        self.inventory_table.setSortingEnabled(True)
        self.inventory_table.setAlternatingRowColors(True)
        self.inventory_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border-radius: 8px;
                border: none;
            }
            QHeaderView::section {
                background-color: #2E7D32;
                color: white;
                padding: 10px;
                border: none;
                font-weight: bold;
            }
        """)

        # Set minimum height for table
        self.inventory_table.setMinimumHeight(500)

        # Pagination controls at bottom
        pagination_layout = QHBoxLayout()
        self.page_label = QLabel("Page 1 of 1")
        self.page_label.setStyleSheet("color: #555; font-weight: bold;")

        self.prev_button = QPushButton("Previous")
        self.prev_button.clicked.connect(self.prev_page)
        self.prev_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #388E3C;
            }
        """)

        # NEW: Add Rack Location button (red)
        self.add_rack_button = QPushButton("Add Rack Location")
        self.add_rack_button.clicked.connect(self.open_add_rack_location_dialog)
        self.add_rack_button.setStyleSheet("""
            QPushButton {
                background-color: #F44336;  /* Red */
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #D32F2F;
            }
        """)

        self.next_button = QPushButton("Next")
        self.next_button.clicked.connect(self.next_page)
        self.next_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #388E3C;
            }
        """)

        pagination_layout.addWidget(self.page_label)
        pagination_layout.addStretch()
        pagination_layout.addWidget(self.add_rack_button)  # <── moved to left of Previous
        pagination_layout.addWidget(self.prev_button)
        pagination_layout.addWidget(self.next_button)

        # Add components to content layout
        content_layout.addWidget(self.inventory_table)
        content_layout.addLayout(pagination_layout)

        inventory_panel.setLayout(content_layout)
        self.content_stack.addWidget(inventory_panel)

    def create_pending_panel(self):
        """Create the pending approvals panel"""
        pending_panel = QWidget()
        pending_layout = QVBoxLayout()
        pending_layout.setContentsMargins(20, 20, 20, 20)

        # Title and refresh button
        header_layout = QHBoxLayout()
        title_label = QLabel("🚨 Pending Product Approvals")
        title_label.setFont(QFont("Arial", 18, QFont.Bold))
        title_label.setStyleSheet("color: #2E7D32; margin-bottom: 10px;")
        header_layout.addWidget(title_label)

        header_layout.addStretch()

        refresh_btn = QPushButton("🔄 Refresh")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        refresh_btn.clicked.connect(self.load_pending_data)
        header_layout.addWidget(refresh_btn)

        pending_layout.addLayout(header_layout)

        # Table for pending approvals
        self.pending_table = QTableWidget()
        self.pending_table.setColumnCount(9)
        self.pending_table.setHorizontalHeaderLabels([
            "Product Name", "Batch", "SKU", "Owner (ID)", "Location", "Manufacture Date", "Expiry Date", "Actions",
            "Details"
        ])

        # Set column widths - increased for better visibility
        header = self.pending_table.horizontalHeader()
        header.setStretchLastSection(True)  # Allow last column to stretch
        header.resizeSection(0, 180)  # Product Name
        header.resizeSection(1, 120)  # Batch
        header.resizeSection(2, 90)  # SKU
        header.resizeSection(3, 110)  # Owner
        header.resizeSection(4, 150)  # Location
        header.resizeSection(5, 140)  # Manufacture Date
        header.resizeSection(6, 140)  # Expiry Date
        header.resizeSection(7, 260)  # Actions - increased for buttons
        header.resizeSection(8, 100)  # Details

        self.pending_table.setAlternatingRowColors(True)
        self.pending_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.pending_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border-radius: 8px;
                border: none;
                gridline-color: #E0E0E0;
            }
            QHeaderView::section {
                background-color: #2E7D32;
                color: white;
                padding: 12px;
                border: none;
                font-weight: bold;
                font-size: 12px;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #E0E0E0;
            }
            QTableWidget::item:selected {
                background-color: #E8F5E8;
            }
        """)

        self.pending_table.setMinimumHeight(500)

        # Set row height to accommodate buttons
        self.pending_table.verticalHeader().setDefaultSectionSize(60)

        pending_layout.addWidget(self.pending_table)

        pending_panel.setLayout(pending_layout)
        self.content_stack.addWidget(pending_panel)

    def create_maturation_panel(self):
        """Create the upcoming maturation panel with card-style layout"""
        maturation_panel = QWidget()
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)

        # Header with settings button for super admin
        header_layout = QHBoxLayout()

        # Title and current threshold display
        title_layout = QVBoxLayout()
        title_label = QLabel("🕒 Upcoming Product Maturation")
        title_label.setFont(QFont("Arial", 18, QFont.Bold))
        title_label.setStyleSheet("color: #2E7D32; margin-bottom: 5px;")

        # Load current threshold
        self.current_threshold = self.load_maturation_threshold()
        self.threshold_label = QLabel(f"Showing products expiring within {self.current_threshold} days")
        self.threshold_label.setFont(QFont("Arial", 12))
        self.threshold_label.setStyleSheet("color: #666; margin-bottom: 10px;")

        title_layout.addWidget(title_label)
        title_layout.addWidget(self.threshold_label)
        header_layout.addLayout(title_layout)

        header_layout.addStretch()

        # Expired Products button
        expired_btn = QPushButton("Expired Products")
        expired_btn.setFixedSize(160, 40)
        expired_btn.setStyleSheet("""
            QPushButton {
                background-color: #F44336;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 8px 16px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #D32F2F;
            }
        """)
        expired_btn.clicked.connect(self.open_expired_products_dialog)
        header_layout.addWidget(expired_btn)

        # Settings button for super admin
        if self.is_superadmin_mode:
            settings_btn = QPushButton("⚙️ Maturation Settings")
            settings_btn.setFixedSize(180, 40)
            settings_btn.setStyleSheet("""
                QPushButton {
                    background-color: #FF9800;
                    color: white;
                    border: none;
                    border-radius: 8px;
                    padding: 8px 16px;
                    font-weight: bold;
                    font-size: 12px;
                }
                QPushButton:hover {
                    background-color: #F57C00;
                }
            """)
            settings_btn.clicked.connect(self.open_maturation_settings)
            header_layout.addWidget(settings_btn)

        main_layout.addLayout(header_layout)

        # Email Summary Card (updated)
        email_summary_card = QFrame()
        email_summary_card.setFixedHeight(150)
        email_summary_card.setStyleSheet("""
            QFrame {
                background-color: #E7DADA;
                border: none;
                border-radius: 8px;
                margin: 10px 0px;
                padding: 15px;
            }
        """)

        summary_layout = QVBoxLayout(email_summary_card)
        summary_layout.setContentsMargins(15, 15, 15, 15)
        summary_layout.setSpacing(10)  # reduce spacing between title and text

        summary_title = QLabel("📧 Email Notification Summary")
        summary_title.setFont(QFont("Times New Roman", 12, QFont.Bold))
        summary_title.setStyleSheet("color: black;")

        self.email_summary_text = QLabel("No email notifications have been sent yet.")
        self.email_summary_text.setFont(QFont("Times New Roman", 12))
        self.email_summary_text.setStyleSheet("color: black;")
        self.email_summary_text.setWordWrap(True)

        summary_layout.addWidget(summary_title)
        summary_layout.addWidget(self.email_summary_text)

        main_layout.addWidget(email_summary_card)

        # Scroll area for cards
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_area.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QScrollBar:vertical {
                border: none;
                background: #f0f0f0;
                width: 12px;
                border-radius: 6px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: #c0c0c0;
                border-radius: 6px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: #a0a0a0;
            }
        """)

        # Cards container widget
        self.cards_widget = QWidget()
        self.cards_layout = QGridLayout(self.cards_widget)
        self.cards_layout.setContentsMargins(10, 10, 10, 10)
        self.cards_layout.setSpacing(20)
        self.cards_layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)

        scroll_area.setWidget(self.cards_widget)
        main_layout.addWidget(scroll_area)

        # Add admin controls at the bottom (for admin/superadmin only)
        if self.user_role.lower() in ['admin', 'superadmin']:
            # Add Select All / Deselect All buttons for easier checkbox management
            selection_layout = QHBoxLayout()
            selection_layout.setContentsMargins(0, 10, 0, 5)

            select_all_btn = QPushButton("Select All")
            select_all_btn.setFixedSize(120, 35)
            select_all_btn.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 8px 16px;
                    font-weight: bold;
                    font-size: 12px;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
                QPushButton:pressed {
                    background-color: #3d8b40;
                }
            """)
            select_all_btn.clicked.connect(self.select_all_products)

            deselect_all_btn = QPushButton("Deselect All")
            deselect_all_btn.setFixedSize(120, 35)
            deselect_all_btn.setStyleSheet("""
                QPushButton {
                    background-color: #f44336;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 8px 16px;
                    font-weight: bold;
                    font-size: 12px;
                }
                QPushButton:hover {
                    background-color: #da190b;
                }
                QPushButton:pressed {
                    background-color: #c1180c;
                }
            """)
            deselect_all_btn.clicked.connect(self.deselect_all_products)

            selection_layout.addStretch()
            selection_layout.addWidget(select_all_btn)
            selection_layout.addSpacing(10)
            selection_layout.addWidget(deselect_all_btn)
            selection_layout.addStretch()

            main_layout.addLayout(selection_layout)

            # Add Remind Again button
            button_layout = QHBoxLayout()
            button_layout.setContentsMargins(0, 5, 0, 0)

            remind_again_btn = QPushButton("Remind Again")
            remind_again_btn.setFixedSize(180, 45)
            remind_again_btn.setStyleSheet("""
                QPushButton {
                    background-color: #FF6B35;
                    color: white;
                    border: none;
                    border-radius: 8px;
                    padding: 10px 20px;
                    font-weight: bold;
                    font-size: 14px;
                    font-family: Arial;
                }
                QPushButton:hover {
                    background-color: #E55A2B;
                }
                QPushButton:pressed {
                    background-color: #CC4E23;
                }
            """)
            remind_again_btn.clicked.connect(self.send_reminder_emails)

            button_layout.addStretch()
            button_layout.addWidget(remind_again_btn)
            button_layout.addStretch()

            main_layout.addLayout(button_layout)

        maturation_panel.setLayout(main_layout)
        self.content_stack.addWidget(maturation_panel)

    def create_assign_panel(self):
        """Create the assign testers panel"""
        assign_panel = QWidget()
        assign_layout = QVBoxLayout()
        assign_layout.setContentsMargins(20, 20, 20, 20)

        # Header with title and buttons
        header_layout = QHBoxLayout()

        # Title
        title_label = QLabel("🎯 Assign Testers to Products")
        title_label.setFont(QFont("Arial", 18, QFont.Bold))
        title_label.setStyleSheet("color: #2E7D32; margin-bottom: 10px;")
        header_layout.addWidget(title_label)

        header_layout.addStretch()

        # View Test Update Button
        view_test_update_btn = QPushButton("🧪 View Test Update")
        view_test_update_btn.setStyleSheet("""
            QPushButton {
                background-color: #F44336;  /* Red background */
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #FF7961;
            }
            QPushButton:pressed {
                background-color: #D32F2F;
            }
        """)
        view_test_update_btn.clicked.connect(self.view_test_updates)
        header_layout.addWidget(view_test_update_btn)

        # View Assignment Records Button
        view_records_btn = QPushButton("📋 View Assignment Records")
        view_records_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #FFA500;
            }
            QPushButton:pressed {
                background-color: #FF4D00;
            }
        """)
        view_records_btn.clicked.connect(self.view_assignment_records)
        header_layout.addWidget(view_records_btn)

        assign_layout.addLayout(header_layout)

        # Table for assigning testers
        self.assign_table = QTableWidget()
        self.assign_table.setColumnCount(6)
        self.assign_table.setHorizontalHeaderLabels([
            "Batch", "Product Name", "Owner", "Rack Location", "Assign To Tester", "Assign Action"
        ])

        # Set specific column widths
        header = self.assign_table.horizontalHeader()
        header.resizeSection(1, 160)  # Product Name
        header.resizeSection(2, 225)  # Owner
        header.resizeSection(3, 160)  # Rack Location
        header.resizeSection(4, 160)  # Assign To Tester
        header.setStretchLastSection(True)

        self.assign_table.setAlternatingRowColors(True)
        self.assign_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border-radius: 8px;
                border: none;
            }
            QHeaderView::section {
                background-color: #2E7D32;
                color: white;
                padding: 10px;
                border: none;
                font-weight: bold;
            }
        """)

        self.assign_table.setMinimumHeight(500)
        assign_layout.addWidget(self.assign_table)

        assign_panel.setLayout(assign_layout)
        self.content_stack.addWidget(assign_panel)

    def create_chat_panel(self):
        """Create the chat/announcement panel with beautiful design from chatInterface.py"""
        chat_panel = QWidget()

        # Get current user role and colors
        role_name = self.user_info.get('role_name', 'user').lower()
        role_color = self.role_colors.get(role_name, '#4CAF50')

        # Apply role-based styling to the chat panel
        chat_panel.setStyleSheet(f"""
            QWidget {{
                background: qlineargradient(x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #f8f9fa, stop: 1 #e9ecef);
                color: #212529;
                font-family: 'Segoe UI', 'San Francisco', Arial, sans-serif;
            }}
            QTextEdit {{
                border: 1px solid #dee2e6;
                border-radius: 12px;
                background-color: rgba(255, 255, 255, 0.95);
                font-family: 'Segoe UI', 'San Francisco', Arial, sans-serif;
                font-size: 14px;
                padding: 16px;
                selection-background-color: {role_color}30;
                box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
            }}
            QLineEdit {{
                border: 1px solid #ced4da;
                border-radius: 8px;
                padding: 12px 16px;
                font-size: 14px;
                background-color: rgba(255, 255, 255, 0.9);
                font-family: 'Segoe UI', 'San Francisco', Arial, sans-serif;
            }}
            QLineEdit:focus {{
                border: 2px solid {role_color};
                background-color: white;
                box-shadow: 0 0 0 3px {role_color}20;
                outline: none;
            }}
            QPushButton {{
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 {role_color}, stop: 1 {role_color});
                color: white;
                border: none;
                border-radius: 8px;
                padding: 12px 24px;
                font-weight: 600;
                font-size: 14px;
                font-family: 'Segoe UI', 'San Francisco', Arial, sans-serif;
                min-width: 90px;
                box-shadow: 0 2px 6px rgba(0, 0, 0, 0.15);
            }}
            QPushButton:hover {{
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 {role_color}, stop: 1 {role_color});
                transform: translateY(-1px);
                box-shadow: 0 4px 12px rgba(0, 0, 0, 0.2);
            }}
            QPushButton:disabled {{
                background: #e9ecef;
                color: #6c757d;
                box-shadow: none;
            }}
            QGroupBox {{
                font-weight: 600;
                border: 1px solid #dee2e6;
                border-radius: 12px;
                margin: 8px 0px;
                padding-top: 20px;
                background: rgba(255, 255, 255, 0.9);
                font-family: 'Segoe UI', 'San Francisco', Arial, sans-serif;
                box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08);
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 16px;
                padding: 8px 16px;
                color: white;
                font-size: 14px;
                font-weight: 600;
                background: qlineargradient(x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 {role_color}, stop: 1 {role_color});
                border: none;
                border-radius: 6px;
                box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
            }}
            QListWidget {{
                border: 1px solid #dee2e6;
                border-radius: 8px;
                background-color: rgba(255, 255, 255, 0.95);
                font-family: 'Segoe UI', 'San Francisco', Arial, sans-serif;
                font-size: 13px;
                alternate-background-color: #f8f9fa;
            }}
            QListWidget::item {{
                padding: 10px;
                border-bottom: 1px solid #f1f3f4;
                border-radius: 4px;
                margin: 2px;
            }}
            QListWidget::item:selected {{
                background: qlineargradient(x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 {role_color}20, stop: 1 {role_color}10);
                border: 1px solid {role_color}40;
            }}
            QListWidget::item:hover {{
                background-color: {role_color}10;
                border: 1px solid {role_color}30;
            }}
        """)

        # Main layout
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(10, 10, 10, 10)

        # Header with user info
        header_layout = QHBoxLayout()
        full_name = self.user_info.get('fullname', self.user_info.get('username', 'Unknown'))
        user_label = QLabel(f"💬 Chat/Announcement - {role_name.title()} - {full_name}")
        user_label.setFont(QFont("Arial", 14, QFont.Bold))
        user_label.setStyleSheet(
            f"color: {role_color}; padding: 10px; background: rgba(255,255,255,0.8); border-radius: 8px;")

        header_layout.addWidget(user_label)
        header_layout.addStretch()

        # Role-specific buttons
        if role_name == 'superadmin':
            clear_chat_btn = QPushButton("Clear All Chat")
            clear_chat_btn.clicked.connect(self.clear_all_chat)
            header_layout.addWidget(clear_chat_btn)

        main_layout.addLayout(header_layout)

        # Main content with horizontal split
        content_layout = QHBoxLayout()

        # Left side - Chat area (2/3 of space)
        left_widget = QWidget()
        left_layout = QVBoxLayout()
        left_widget.setLayout(left_layout)

        # Chat display area - will switch between global and private chat
        self.chat_group = QGroupBox("Global Chat Room")
        chat_layout = QVBoxLayout()

        # Back to Global Chat button (initially hidden)
        self.back_to_global_btn = QPushButton("← Back to Global Chat Room")
        self.back_to_global_btn.clicked.connect(self.back_to_global_chat)
        self.back_to_global_btn.setVisible(False)
        self.back_to_global_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: #6c757d;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 600;
                margin-bottom: 10px;
            }}
            QPushButton:hover {{
                background-color: #5a6268;
            }}
        """)
        chat_layout.addWidget(self.back_to_global_btn)

        self.chat_display = QTextEdit()
        self.chat_display.setReadOnly(True)
        self.chat_display.setMinimumHeight(350)
        chat_layout.addWidget(self.chat_display)

        self.chat_group.setLayout(chat_layout)
        left_layout.addWidget(self.chat_group)

        # Message input area - will adapt based on chat mode
        self.input_group = QGroupBox("Send Public Message")
        input_layout = QVBoxLayout()

        # Tag helper (will be hidden in private chat mode)
        self.tag_info = QLabel("💡 Tip: Use @username to tag users (e.g., @john)")
        self.tag_info.setStyleSheet("color: #666; font-size: 14px; font-style: italic; margin: 5px;")
        input_layout.addWidget(self.tag_info)

        msg_layout = QHBoxLayout()
        self.chat_input = QLineEdit()
        self.chat_input.setPlaceholderText("Type your message here... (use @username to tag)")
        self.chat_input.returnPressed.connect(self.send_message_handler)

        self.send_btn = QPushButton("Send")
        self.send_btn.clicked.connect(self.send_message_handler)

        msg_layout.addWidget(self.chat_input)
        msg_layout.addWidget(self.send_btn)
        input_layout.addLayout(msg_layout)

        self.input_group.setLayout(input_layout)

        # Show input area based on role and chat mode
        if role_name in ['admin', 'superadmin']:
            left_layout.addWidget(self.input_group)
        else:
            # Show read-only message for owner and tester in global chat
            self.readonly_label = QLabel("You can view messages but cannot send messages in the public chat room.")
            self.readonly_label.setStyleSheet(
                "color: #666; font-style: italic; padding: 10px; background: rgba(255,255,255,0.8); border-radius: 8px;")
            self.readonly_label.setAlignment(Qt.AlignCenter)
            left_layout.addWidget(self.readonly_label)

            # Hide input area initially for non-admin users
            self.input_group.setVisible(False)
            left_layout.addWidget(self.input_group)

        content_layout.addWidget(left_widget, 2)

        # Right side - Users and Private Chat (1/3 of space)
        right_widget = QWidget()
        right_layout = QVBoxLayout()
        right_widget.setLayout(right_layout)

        # Users list with search
        users_group = QGroupBox("All Users")
        users_layout = QVBoxLayout()

        # Search bar
        search_layout = QHBoxLayout()
        search_label = QLabel("🔍 Search:")
        search_label.setStyleSheet("font-size: 14px; color: #666;")

        self.user_search = QLineEdit()
        self.user_search.setPlaceholderText("🔍 Search users...")
        self.user_search.textChanged.connect(self.search_users)
        self.user_search.setStyleSheet("""
            font-size: 14px; 
            padding: 10px 14px; 
            border: 1px solid #ced4da;
            border-radius: 20px;
            background-color: rgba(255, 255, 255, 0.9);
            color: #495057;
        """)

        search_layout.addWidget(search_label)
        search_layout.addWidget(self.user_search)
        users_layout.addLayout(search_layout)

        # Users list
        self.users_list = QListWidget()
        self.users_list.itemClicked.connect(self.select_user_for_chat)
        self.users_list.setMaximumHeight(300)  # More space for users
        users_layout.addWidget(self.users_list)

        # Selected user info and start chat button
        self.selected_user_label = QLabel("Select a user above to start private messaging")
        self.selected_user_label.setStyleSheet(
            "color: #666; font-size: 12px; font-style: italic; margin: 10px 5px; padding: 10px; background: rgba(255,255,255,0.8); border-radius: 8px;")
        self.selected_user_label.setAlignment(Qt.AlignCenter)
        users_layout.addWidget(self.selected_user_label)

        # Start Chat button (initially hidden)
        self.start_chat_btn = QPushButton("💬 Start Chat")
        self.start_chat_btn.clicked.connect(self.start_private_chat_mode)
        self.start_chat_btn.setVisible(False)
        self.start_chat_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 {role_color}, stop: 1 {role_color});
                color: white;
                border: none;
                border-radius: 8px;
                padding: 12px 24px;
                font-weight: 600;
                font-size: 14px;
                margin: 10px 5px;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 {role_color}, stop: 1 {role_color});
                transform: translateY(-1px);
                box-shadow: 0 4px 12px rgba(0, 0, 0, 0.2);
            }}
        """)
        users_layout.addWidget(self.start_chat_btn)

        users_group.setLayout(users_layout)
        right_layout.addWidget(users_group)

        content_layout.addWidget(right_widget, 1)
        main_layout.addLayout(content_layout)

        chat_panel.setLayout(main_layout)
        self.content_stack.addWidget(chat_panel)

    def create_user_management_panel(self):
        """Create the user management panel"""
        user_panel = QWidget()
        user_layout = QVBoxLayout()
        user_layout.setContentsMargins(20, 20, 20, 20)

        # Table for user management
        self.user_table = QTableWidget()
        self.user_table.setColumnCount(6)
        self.user_table.setHorizontalHeaderLabels([
            "User ID", "Username", "Full Name", "Role", "Email", "Branch"
        ])
        self.user_table.horizontalHeader().setStretchLastSection(True)
        self.user_table.setAlternatingRowColors(True)
        self.user_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.user_table.setSelectionMode(QTableWidget.SingleSelection)
        self.user_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border-radius: 8px;
                border: none;
            }
            QHeaderView::section {
                background-color: #2E7D32;
                color: white;
                padding: 10px;
                border: none;
                font-weight: bold;
            }
            QTableWidget::item:selected {
                background-color: #E8F5E9;
                color: #2E7D32;
            }
        """)

        self.user_table.setMinimumHeight(500)
        user_layout.addWidget(self.user_table)

        user_panel.setLayout(user_layout)
        self.content_stack.addWidget(user_panel)

    def create_my_profile_panel(self):
        """Create the My Profile panel"""
        profile_panel = QWidget()
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)

        # Header
        header_layout = QHBoxLayout()
        title_label = QLabel("👤 My Profile")
        title_label.setFont(QFont("Arial", 18, QFont.Bold))
        title_label.setStyleSheet("color: #2E7D32; margin-bottom: 10px;")
        header_layout.addWidget(title_label)
        header_layout.addStretch()

        # Edit Profile button
        self.edit_profile_btn = QPushButton("✏️ Edit Profile")
        self.edit_profile_btn.setFixedSize(140, 40)
        self.edit_profile_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 8px 16px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        self.edit_profile_btn.clicked.connect(self.open_edit_profile_dialog)
        header_layout.addWidget(self.edit_profile_btn)

        main_layout.addLayout(header_layout)

        # Content area with scroll
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: transparent;
            }
        """)

        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(10, 10, 10, 10)
        content_layout.setSpacing(20)

        # Profile Information Card
        profile_card = QFrame()
        profile_card.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 12px;
                padding: 20px;
            }
        """)
        profile_layout = QVBoxLayout(profile_card)

        # Profile header with avatar placeholder
        profile_header = QHBoxLayout()

        # Avatar placeholder
        avatar_frame = QFrame()
        avatar_frame.setFixedSize(100, 100)
        avatar_frame.setStyleSheet("""
            QFrame {
                background-color: #E3F2FD;
                border: 3px solid #2196F3;
                border-radius: 50px;
            }
        """)
        avatar_layout = QVBoxLayout(avatar_frame)
        avatar_icon = QLabel("👤")
        avatar_icon.setAlignment(Qt.AlignCenter)
        avatar_icon.setStyleSheet("font-size: 40px; color: #2196F3; border: none;")
        avatar_layout.addWidget(avatar_icon)

        profile_header.addWidget(avatar_frame)
        profile_header.addSpacing(20)

        # Profile basic info
        basic_info_layout = QVBoxLayout()

        self.profile_name_label = QLabel("Loading...")
        self.profile_name_label.setFont(QFont("Arial", 16, QFont.Bold))
        self.profile_name_label.setStyleSheet("color: #333; border: none;")

        self.profile_role_label = QLabel("Loading...")
        self.profile_role_label.setFont(QFont("Arial", 12))
        self.profile_role_label.setStyleSheet("color: #666; border: none;")

        self.profile_username_label = QLabel("Loading...")
        self.profile_username_label.setFont(QFont("Arial", 12))
        self.profile_username_label.setStyleSheet("color: #666; border: none;")

        basic_info_layout.addWidget(self.profile_name_label)
        basic_info_layout.addWidget(self.profile_role_label)
        basic_info_layout.addSpacing(5)  # 5px padding between Role and Username
        basic_info_layout.addWidget(self.profile_username_label)
        basic_info_layout.addStretch()

        profile_header.addLayout(basic_info_layout)
        profile_header.addStretch()

        profile_layout.addLayout(profile_header)

        # Detailed Information Grid
        details_grid = QGridLayout()
        details_grid.setSpacing(15)
        details_grid.setContentsMargins(0, 20, 0, 0)

        # Create info fields
        self.create_profile_field(details_grid, "📧 Email:", "profile_email", 0)
        self.create_profile_field(details_grid, "📱 Phone:", "profile_phone", 1)
        self.create_profile_field(details_grid, "🏢 Branch:", "profile_branch", 2)
        self.create_profile_field(details_grid, "🆔 User ID:", "profile_user_id", 3)
        self.create_profile_field(details_grid, "👥 Role ID:", "profile_role_id", 4)

        profile_layout.addLayout(details_grid)
        content_layout.addWidget(profile_card)

        # Recent Activity Card
        recent_card = QFrame()
        recent_card.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 12px;
                padding: 20px;
            }
        """)
        recent_layout = QVBoxLayout(recent_card)

        recent_title = QLabel("🕒 Recent Activity")
        recent_title.setFont(QFont("Arial", 14, QFont.Bold))
        recent_title.setStyleSheet("color: #2E7D32; margin-bottom: 10px; border: none;")
        recent_layout.addWidget(recent_title)

        # Recent activity list
        self.recent_activity_list = QListWidget()
        self.recent_activity_list.setStyleSheet("""
            QListWidget {
                border: 1px solid #ddd;
                border-radius: 8px;
                background-color: #f9f9f9;
                font-size: 12px;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid #eee;
            }
            QListWidget::item:hover {
                background-color: #e3f2fd;
            }
        """)
        self.recent_activity_list.setMaximumHeight(600)
        recent_layout.addWidget(self.recent_activity_list)

        content_layout.addWidget(recent_card)
        content_layout.addStretch()

        scroll_area.setWidget(content_widget)
        main_layout.addWidget(scroll_area)

        profile_panel.setLayout(main_layout)
        self.content_stack.addWidget(profile_panel)

    def create_profile_field(self, grid, label_text, attr_name, row):
        """Create a profile information field"""
        label = QLabel(label_text)
        label.setFont(QFont("Arial", 12, QFont.Bold))
        label.setStyleSheet("color: #555; border: none;")

        value_label = QLabel("Loading...")
        value_label.setFont(QFont("Arial", 12))
        value_label.setStyleSheet("""
            color: #333; 
            background-color: #f5f5f5; 
            padding: 8px 12px; 
            border-radius: 6px;
            border: 1px solid #ddd;
        """)

        # Store reference for later updates
        setattr(self, attr_name, value_label)

        grid.addWidget(label, row, 0)
        grid.addWidget(value_label, row, 1)

    def create_analysis_panel(self):
        """Create the analysis panel with charts and analytics"""
        analysis_panel = QWidget()
        analysis_layout = QVBoxLayout()
        analysis_layout.setContentsMargins(20, 20, 20, 20)

        # Header section
        header_layout = QHBoxLayout()
        title_label = QLabel("📊 Analytics Dashboard")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2E7D32;
                margin-bottom: 10px;
            }
        """)
        header_layout.addWidget(title_label)
        header_layout.addStretch()

        # Export buttons
        export_layout = QHBoxLayout()
        export_pdf_btn = QPushButton("📄 Export to PDF")
        export_excel_btn = QPushButton("📊 Export to Excel")
        refresh_btn = QPushButton("🔄 Refresh Data")

        for btn in [export_pdf_btn, export_excel_btn, refresh_btn]:
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 8px 16px;
                    font-weight: bold;
                    margin: 2px;
                }
                QPushButton:hover { background-color: #388E3C; }
            """)

        export_pdf_btn.clicked.connect(self.export_analysis_to_pdf)
        export_excel_btn.clicked.connect(self.export_analysis_to_excel)
        refresh_btn.clicked.connect(self.load_analysis_data)

        export_layout.addWidget(export_pdf_btn)
        export_layout.addWidget(export_excel_btn)
        export_layout.addWidget(refresh_btn)
        header_layout.addLayout(export_layout)
        analysis_layout.addLayout(header_layout)

        # Scrollable area for charts
        scroll_area = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        scroll_layout.setSpacing(20)

        # Summary stats section
        stats_frame = QFrame()
        stats_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 8px;
                padding: 15px;
            }
        """)
        stats_layout = QHBoxLayout(stats_frame)

        # Create summary cards
        self.total_products_card = self.create_stat_card("Total Products", "0", "#4CAF50")
        self.total_owners_card = self.create_stat_card("Total Owners", "0", "#2196F3")
        self.total_testers_card = self.create_stat_card("Total Testers", "0", "#FF9800")
        self.pending_tests_card = self.create_stat_card("Pending Tests", "0", "#F44336")

        stats_layout.addWidget(self.total_products_card)
        stats_layout.addWidget(self.total_owners_card)
        stats_layout.addWidget(self.total_testers_card)
        stats_layout.addWidget(self.pending_tests_card)
        scroll_layout.addWidget(stats_frame)

        # Charts section
        charts_frame = QFrame()
        charts_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 8px;
                padding: 15px;
            }
        """)
        # Use horizontal layout for side-by-side charts with larger display
        charts_layout = QHBoxLayout(charts_frame)

        # Create only the two main charts with larger display space
        self.owners_chart_widget = self.create_chart_widget("Products by Owner")
        self.testers_chart_widget = self.create_chart_widget("Tests by Tester")

        # Add charts side by side for larger display
        charts_layout.addWidget(self.owners_chart_widget)
        charts_layout.addWidget(self.testers_chart_widget)
        scroll_layout.addWidget(charts_frame)

        scroll_area.setWidget(scroll_widget)
        scroll_area.setWidgetResizable(True)
        analysis_layout.addWidget(scroll_area)

        analysis_panel.setLayout(analysis_layout)

        # Make sure all widgets are visible
        analysis_panel.setVisible(True)
        scroll_area.setVisible(True)
        scroll_widget.setVisible(True)
        stats_frame.setVisible(True)
        charts_frame.setVisible(True)

        # Add to content stack
        self.content_stack.addWidget(analysis_panel)
        print(f"✓ Analysis panel created and added to content stack at index {self.content_stack.count() - 1}")
        print(f"✓ Analysis panel widgets: {len(analysis_panel.findChildren(QWidget))} child widgets")
        print(
            f"✓ Stat cards created: Products={hasattr(self, 'total_products_card')}, Owners={hasattr(self, 'total_owners_card')}")
        print(
            f"✓ Chart widgets created: Owners={hasattr(self, 'owners_chart_widget')}, Products={hasattr(self, 'products_chart_widget')}")

    def create_stat_card(self, title, value, color):
        """Create a statistics card widget"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {color};
                border: none;
                border-radius: 8px;
                padding: 15px;
                margin: 5px;
            }}
        """)
        layout = QVBoxLayout(card)

        title_label = QLabel(title)
        title_label.setStyleSheet("""
            QLabel {
                color: white;
                font-size: 14px;
                font-weight: bold;
                margin-bottom: 5px;
            }
        """)

        value_label = QLabel(value)
        value_label.setStyleSheet("""
            QLabel {
                color: white;
                font-size: 24px;
                font-weight: bold;
            }
        """)

        layout.addWidget(title_label)
        layout.addWidget(value_label)
        layout.setAlignment(Qt.AlignCenter)

        return card

    def create_chart_widget(self, title):
        """Create a chart widget with table representation"""
        widget = QFrame()
        widget.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 8px;
                margin: 10px 0px;
            }
        """)
        layout = QVBoxLayout(widget)

        # Chart title
        title_label = QLabel(title)
        title_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #333;
                padding: 10px;
                background-color: #f5f5f5;
                border-bottom: 1px solid #ddd;
            }
        """)
        layout.addWidget(title_label)

        # Chart table with larger display
        table = QTableWidget()
        table.setMinimumHeight(400)
        table.setMaximumHeight(600)
        table.setStyleSheet("""
            QTableWidget {
                border: none;
                background-color: white;
            }
            QHeaderView::section {
                background-color: #2E7D32;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
            }
        """)
        layout.addWidget(table)

        # Store reference to table for data loading
        setattr(widget, 'table', table)

        return widget

    def switch_content(self, content_type):
        """Switch between different content panels based on menu selection"""
        # Update the content area based on the selected menu item
        if content_type == "inventory":
            self.content_stack.setCurrentIndex(0)
            self.load_inventory_data()
            # Show admin features if in superadmin mode
            if self.is_superadmin_mode:
                self.show_inventory_admin_features()
        elif content_type == "pending":
            self.content_stack.setCurrentIndex(1)
            self.load_pending_data()
            # Show admin features if in superadmin mode
            if self.is_superadmin_mode:
                self.show_pending_admin_features()
        elif content_type == "maturation":
            self.content_stack.setCurrentIndex(2)
            self.load_maturation_data()
            # Show admin features if in superadmin mode
            if self.is_superadmin_mode:
                self.show_maturation_admin_features()
        elif content_type == "assign":
            self.content_stack.setCurrentIndex(3)
            self.load_assign_data()
        elif content_type == "chat":
            self.content_stack.setCurrentIndex(4)
            self.load_chat_data()
        elif content_type == "user_management":
            self.content_stack.setCurrentIndex(5)
            self.load_user_data()
            # Show admin features for both admin and superadmin
            self.show_user_management_admin_features()
        elif content_type == "my_profile":
            self.content_stack.setCurrentIndex(6)
            self.load_my_profile_data()
        # Admin-only panels
        elif content_type == "system_settings":
            # Check if we're in superadmin mode
            if not self.is_superadmin_mode:
                QMessageBox.warning(self, "Access Denied", "This feature requires Super Admin access.")
                return
            # System settings is at index 8 (after analysis)
            self.content_stack.setCurrentIndex(8)

        elif content_type == "export_reports":
            # Check if we're in superadmin mode
            if not self.is_superadmin_mode:
                QMessageBox.warning(self, "Access Denied", "This feature requires Super Admin access.")
                return
            # Export reports is at index 9
            self.content_stack.setCurrentIndex(9)

        elif content_type == "analysis":
            # Check if we're in superadmin mode
            if not self.is_superadmin_mode:
                QMessageBox.warning(self, "Access Denied", "This feature requires Super Admin access.")
                return
            # Analysis panel is at index 7 (after my_profile which is at index 6)
            print(f"🔧 Switching to analysis panel at index 7 (total widgets: {self.content_stack.count()})")
            self.content_stack.setCurrentIndex(7)

            # Make sure the widget is visible
            current_widget = self.content_stack.currentWidget()
            if current_widget:
                current_widget.setVisible(True)
                current_widget.show()
                print(f"✓ Analysis panel widget shown: {type(current_widget).__name__}")
            else:
                print("✗ Analysis panel widget is None!")

            self.load_analysis_data()

    def load_pending_data(self):
        """Load products with pending status for approval"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Query for pending products with all required details
            cursor.execute("""
                SELECT p.product_id, p.product_name, p.batch, p.sku, 
                       (u.username || ' (' || p.owner_id || ')') as owner_info,
                       COALESCE(r.rack_location_name, p.location, 'Not Assigned') as location,
                       p.manufacture_date, p.expired_date
                FROM products p
                LEFT JOIN users u ON p.owner_id = u.user_id
                LEFT JOIN racklocations r ON p.rack_location_id = r.rack_location_id
                WHERE COALESCE(p.status, 'pending') = 'pending'
                ORDER BY p.arrival_date DESC
            """)

            data = cursor.fetchall()
            self.pending_table.setRowCount(len(data))

            for row, product in enumerate(data):
                product_id = product[0]

                # Add data to columns (excluding product_id which is hidden)
                for col, value in enumerate(product[1:], 0):  # Skip product_id
                    # Format dates for display
                    if col in [5, 6] and value:  # Manufacture and expiry dates
                        try:
                            # Handle different date formats
                            if 'T' in str(value):
                                date_part = str(value).split('T')[0]
                                formatted_date = datetime.strptime(date_part, '%Y-%m-%d').strftime('%d/%m/%Y')
                            elif '/' in str(value):
                                # Already in dd/mm/yyyy format
                                formatted_date = str(value)
                            else:
                                # Try to parse and format
                                formatted_date = datetime.strptime(str(value), '%Y-%m-%d').strftime('%d/%m/%Y')
                            value = formatted_date
                        except:
                            pass

                    item = QTableWidgetItem(str(value) if value else "")
                    item.setTextAlignment(Qt.AlignCenter)
                    self.pending_table.setItem(row, col, item)

                # Add action buttons in the Actions column (index 7)
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(2, 2, 2, 2)
                actions_layout.setSpacing(3)

                # Approve button
                approve_btn = QPushButton("Approve")
                approve_btn.setFixedSize(90, 22)  # Fixed size to prevent overlap
                approve_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #4CAF50;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 4px 8px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background-color: #45a049;
                    }
                """)
                approve_btn.clicked.connect(lambda checked, pid=product_id: self.approve_product(pid))

                # Reject button
                reject_btn = QPushButton("Reject")
                reject_btn.setFixedSize(85, 22)  # Fixed size to prevent overlap
                reject_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #f44336;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 4px 8px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background-color: #da190b;
                    }
                """)
                reject_btn.clicked.connect(lambda checked, pid=product_id: self.reject_product(pid))

                actions_layout.addWidget(approve_btn)
                actions_layout.addWidget(reject_btn)

                self.pending_table.setCellWidget(row, 7, actions_widget)

                # Add view details button in the Details column (index 8)
                view_btn = QPushButton("👁️ View")
                view_btn.setFixedSize(80, 28)  # Fixed size to prevent overlap
                view_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #2196F3;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 4px 8px;
                        font-weight: bold;
                        font-size: 10px;
                    }
                    QPushButton:hover {
                        background-color: #1976D2;
                    }
                """)
                view_btn.clicked.connect(lambda checked, pid=product_id: self.view_product_details(pid))
                self.pending_table.setCellWidget(row, 8, view_btn)

            conn.close()
            print(f"✓ Loaded {len(data)} pending products for approval")

        except Exception as e:
            print(f"✗ Error loading pending data: {e}")
            self.pending_table.setRowCount(0)

    def load_assign_data(self):
        """Load products for tester assignment - only products with no assigned tester"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Fetch only products with no tester assigned (tester_id IS NULL) and status = approved
            cursor.execute("""
                SELECT p.product_id, p.batch, p.product_name, 
                       CASE 
                           WHEN u.fullname IS NOT NULL AND u.fullname != '' THEN u.fullname || ' (' || u.username || ')'
                           ELSE u.username
                       END as owner, 
                       CASE 
                           WHEN rl.rack_location_name IS NOT NULL THEN rl.rack_location_name
                           ELSE 'Unassigned'
                       END as rack_location
                FROM products p
                LEFT JOIN users u ON p.owner_id = u.user_id
                LEFT JOIN racklocations rl ON p.rack_location_id = rl.rack_location_id
                WHERE p.tester_id IS NULL AND COALESCE(p.status, 'pending') = 'approved'
                ORDER BY p.arrival_date DESC
            """)

            data = cursor.fetchall()
            self.assign_table.setRowCount(len(data))

            # Get available testers for the dropdown (role_id = 4)
            cursor.execute("""
                SELECT u.user_id, u.username, u.fullname, b.branch_name
                FROM users u
                LEFT JOIN branches b ON u.branch_id = b.branch_id
                WHERE u.role = 4
                ORDER BY u.username
            """)
            testers = cursor.fetchall()

            for row, product in enumerate(data):
                product_id = product[0]  # Store product_id for later use

                # Column 0: Batch
                batch_item = QTableWidgetItem(str(product[1]) if product[1] else "")
                self.assign_table.setItem(row, 0, batch_item)

                # Column 1: Product Name
                product_name_item = QTableWidgetItem(str(product[2]) if product[2] else "")
                self.assign_table.setItem(row, 1, product_name_item)

                # Column 2: Owner
                owner_item = QTableWidgetItem(str(product[3]) if product[3] else "")
                self.assign_table.setItem(row, 2, owner_item)

                # Column 3: Rack Location
                rack_location_item = QTableWidgetItem(str(product[4]) if product[4] else "")
                self.assign_table.setItem(row, 3, rack_location_item)

                # Column 4: Assign To Tester (Dropdown)
                tester_combo = QComboBox()
                tester_combo.addItem("Select Tester...")

                for tester in testers:
                    tester_id, tester_username, tester_fullname, branch_name = tester
                    # Format: fullname (username)
                    if tester_fullname and tester_fullname.strip():
                        display_name = f"{tester_fullname} ({tester_username})"
                    else:
                        display_name = tester_username
                    tester_combo.addItem(display_name)
                    # Store tester data as item data for later retrieval
                    tester_combo.setItemData(tester_combo.count() - 1, {
                        'tester_id': tester_id,
                        'username': tester_username,
                        'fullname': tester_fullname,
                        'branch_name': branch_name
                    })

                self.assign_table.setCellWidget(row, 4, tester_combo)

                # Column 5: Assign Action (Button) - Orange color
                assign_btn = QPushButton("Assign")
                assign_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #FF9800;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 8px 16px;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background-color: #F57C00;
                    }
                    QPushButton:pressed {
                        background-color: #E65100;
                    }
                """)

                # Store product_id in the button for later use
                assign_btn.setProperty("product_id", product_id)
                assign_btn.setProperty("row", row)

                # Connect to assignment function
                assign_btn.clicked.connect(lambda checked, pid=product_id, r=row: self.assign_tester_to_product(pid, r))

                self.assign_table.setCellWidget(row, 5, assign_btn)

            conn.close()
            print(f"✓ Loaded {len(data)} products without assigned testers")

        except Exception as e:
            print(f"✗ Error loading assign data: {e}")
            self.assign_table.setRowCount(0)

    def load_analysis_data(self):
        """Load all analytics data and update charts"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Update summary statistics
            self.update_summary_stats(cursor)

            # Update the two main charts
            self.update_owners_chart(cursor)
            self.update_testers_chart(cursor)

            conn.close()

            # Force widget updates and refresh
            self.content_stack.currentWidget().update()
            self.content_stack.currentWidget().repaint()

            print("✓ Analysis data loaded successfully")

        except Exception as e:
            print(f"✗ Error loading analysis data: {e}")
            QMessageBox.critical(self, "Error", f"Failed to load analysis data: {str(e)}")

    def update_summary_stats(self, cursor):
        """Update the summary statistics cards"""
        try:
            # Total products
            cursor.execute("SELECT COUNT(*) FROM products")
            total_products = cursor.fetchone()[0]

            # Total owners
            cursor.execute("SELECT COUNT(DISTINCT owner_id) FROM products WHERE owner_id IS NOT NULL")
            total_owners = cursor.fetchone()[0]

            # Total testers
            cursor.execute(
                "SELECT COUNT(DISTINCT user_id) FROM users WHERE role IN ('Tester', 'tester', 'TESTER', '4', 4)")
            total_testers = cursor.fetchone()[0]

            # Pending tests
            cursor.execute(
                "SELECT COUNT(*) FROM products WHERE tester_id IS NULL AND COALESCE(status, 'pending') = 'approved'")
            pending_tests = cursor.fetchone()[0]

            print(
                f"📊 Summary stats: Products={total_products}, Owners={total_owners}, Testers={total_testers}, Pending={pending_tests}")

            # Update cards - more robust way to find the value labels
            try:
                # Find and update the value labels in each card
                for widget in self.total_products_card.findChildren(QLabel):
                    if widget.styleSheet() and "font-size: 24px" in widget.styleSheet():
                        widget.setText(str(total_products))
                        break

                for widget in self.total_owners_card.findChildren(QLabel):
                    if widget.styleSheet() and "font-size: 24px" in widget.styleSheet():
                        widget.setText(str(total_owners))
                        break

                for widget in self.total_testers_card.findChildren(QLabel):
                    if widget.styleSheet() and "font-size: 24px" in widget.styleSheet():
                        widget.setText(str(total_testers))
                        break

                for widget in self.pending_tests_card.findChildren(QLabel):
                    if widget.styleSheet() and "font-size: 24px" in widget.styleSheet():
                        widget.setText(str(pending_tests))
                        break

                print("✓ Summary cards updated successfully")

            except Exception as card_error:
                print(f"✗ Error updating cards: {card_error}")
                # Fallback method
                self.total_products_card.layout().itemAt(1).widget().setText(str(total_products))
                self.total_owners_card.layout().itemAt(1).widget().setText(str(total_owners))
                self.total_testers_card.layout().itemAt(1).widget().setText(str(total_testers))
                self.pending_tests_card.layout().itemAt(1).widget().setText(str(pending_tests))

        except Exception as e:
            print(f"✗ Error updating summary stats: {e}")

    def update_owners_chart(self, cursor):
        """Update the owners chart with product counts and product lists"""
        try:
            # First, let's check what roles exist in the database
            cursor.execute("SELECT DISTINCT role FROM users")
            all_roles = cursor.fetchall()
            print(f"🔍 Available user roles: {[role[0] for role in all_roles]}")

            # Check what owner_ids exist in products table
            cursor.execute("SELECT DISTINCT owner_id FROM products WHERE owner_id IS NOT NULL")
            owner_ids = cursor.fetchall()
            print(f"🔍 Owner IDs in products: {[oid[0] for oid in owner_ids]}")

            # Check users and their roles who actually have products
            cursor.execute("""
                SELECT DISTINCT u.user_id, u.username, u.role 
                FROM users u 
                INNER JOIN products p ON u.user_id = p.owner_id
            """)
            actual_owners = cursor.fetchall()
            print(f"🔍 Actual owners with products: {actual_owners}")

            # Get ALL users who have products (regardless of role)
            cursor.execute("""
                SELECT u.username, 
                       COUNT(p.product_id) as product_count,
                       GROUP_CONCAT(p.product_name, ', ') as product_list
                FROM users u
                INNER JOIN products p ON u.user_id = p.owner_id
                GROUP BY u.user_id, u.username
                ORDER BY product_count DESC
            """)

            data = cursor.fetchall()
            print(f"📊 Owners chart data: {len(data)} rows")

            table = self.owners_chart_widget.table

            table.setColumnCount(3)
            table.setHorizontalHeaderLabels(["Owner", "Products Quantity", "List of Products"])
            table.setRowCount(len(data))

            for row, (username, count, product_list) in enumerate(data):
                # Owner name
                owner_item = QTableWidgetItem(username or "")
                table.setItem(row, 0, owner_item)

                # Product quantity with visual bar
                quantity_item = QTableWidgetItem(str(count))
                if count > 0:
                    quantity_item.setBackground(QColor(76, 175, 80, int(min(255, 50 + count * 10))))
                table.setItem(row, 1, quantity_item)

                # List of products
                product_list_text = product_list if product_list else "No products"
                products_item = QTableWidgetItem(product_list_text)
                products_item.setToolTip(product_list_text)  # Show full list in tooltip
                table.setItem(row, 2, products_item)

            # Adjust column widths
            table.setColumnWidth(0, 150)  # Owner column
            table.setColumnWidth(1, 120)  # Quantity column
            table.resizeColumnToContents(2)  # Products list column
            table.setVisible(True)
            print(f"✓ Owners chart updated with {len(data)} rows")

        except Exception as e:
            print(f"✗ Error updating owners chart: {e}")

    def update_testers_chart(self, cursor):
        """Update the testers chart with testing counts and product lists"""
        try:
            # Check what tester_ids exist in products table
            cursor.execute("SELECT DISTINCT tester_id FROM products WHERE tester_id IS NOT NULL")
            tester_ids = cursor.fetchall()
            print(f"🔍 Tester IDs in products: {[tid[0] for tid in tester_ids]}")

            # Check users and their roles who actually have assigned tests
            cursor.execute("""
                SELECT DISTINCT u.user_id, u.username, u.role 
                FROM users u 
                INNER JOIN products p ON u.user_id = p.tester_id
            """)
            actual_testers = cursor.fetchall()
            print(f"🔍 Actual testers with assigned products: {actual_testers}")

            # Get ALL users who have assigned tests (regardless of role)
            cursor.execute("""
                SELECT u.username, 
                       COUNT(p.product_id) as testing_count,
                       GROUP_CONCAT(p.product_name, ', ') as testing_list
                FROM users u
                INNER JOIN products p ON u.user_id = p.tester_id
                GROUP BY u.user_id, u.username
                ORDER BY testing_count DESC
            """)

            data = cursor.fetchall()
            print(f"📊 Testers chart data: {len(data)} rows")

            table = self.testers_chart_widget.table

            table.setColumnCount(3)
            table.setHorizontalHeaderLabels(["Tester", "Testing Quantity", "List of Testing"])
            table.setRowCount(len(data))

            for row, (username, count, testing_list) in enumerate(data):
                # Tester name
                tester_item = QTableWidgetItem(username or "")
                table.setItem(row, 0, tester_item)

                # Testing quantity with visual bar
                quantity_item = QTableWidgetItem(str(count))
                if count > 0:
                    quantity_item.setBackground(QColor(33, 150, 243, int(min(255, 50 + count * 15))))
                table.setItem(row, 1, quantity_item)

                # List of testing products
                testing_list_text = testing_list if testing_list else "No testing assigned"
                testing_item = QTableWidgetItem(testing_list_text)
                testing_item.setToolTip(testing_list_text)  # Show full list in tooltip
                table.setItem(row, 2, testing_item)

            # Adjust column widths
            table.setColumnWidth(0, 150)  # Tester column
            table.setColumnWidth(1, 120)  # Quantity column
            table.resizeColumnToContents(2)  # Testing list column
            table.setVisible(True)
            print(f"✓ Testers chart updated with {len(data)} rows")

        except Exception as e:
            print(f"✗ Error updating testers chart: {e}")

    def export_analysis_to_pdf(self):
        """Export analysis data to PDF"""
        try:
            from datetime import datetime
            import sqlite3
            import os

            # Get current timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"analysis_report_{timestamp}.pdf"

            # Try to import PDF libraries
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
                from reportlab.platypus.charts import VerticalBarChart
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib import colors
                from reportlab.lib.units import inch
                from reportlab.graphics.shapes import Drawing
                from reportlab.graphics.charts.barcharts import VerticalBarChart as RLBarChart
                from reportlab.graphics.charts.legends import Legend

                # Try to register Chinese fonts for ReportLab
                try:
                    from reportlab.pdfbase import pdfmetrics
                    from reportlab.pdfbase.ttfonts import TTFont
                    # Try to register common Chinese fonts available on Windows
                    try:
                        pdfmetrics.registerFont(TTFont('SimHei', 'C:/Windows/Fonts/simhei.ttf'))
                        chinese_font = 'SimHei'
                        print("✓ SimHei font registered for ReportLab")
                    except:
                        try:
                            pdfmetrics.registerFont(TTFont('YaHei', 'C:/Windows/Fonts/msyh.ttf'))
                            chinese_font = 'YaHei'
                            print("✓ Microsoft YaHei font registered for ReportLab")
                        except:
                            chinese_font = 'Helvetica'  # Fallback to default
                            print("⚠️ No Chinese fonts found, using default Helvetica")
                except ImportError:
                    chinese_font = 'Helvetica'
                    print("⚠️ TTFont import failed, using default Helvetica")

                # Connect to database
                conn = sqlite3.connect("testing_system.db")
                cursor = conn.cursor()

                # Create PDF document
                doc = SimpleDocTemplate(filename, pagesize=A4)
                styles = getSampleStyleSheet()
                story = []

                # Title
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=24,
                    spaceAfter=30,
                    alignment=1  # Center alignment
                )
                story.append(Paragraph("Analysis Dashboard Report", title_style))
                story.append(Spacer(1, 20))

                # Timestamp
                timestamp_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                story.append(Paragraph(timestamp_text, styles['Normal']))
                story.append(Spacer(1, 30))

                # Summary Statistics Section
                story.append(Paragraph("Summary Statistics", styles['Heading2']))
                story.append(Spacer(1, 12))

                # Get summary data
                cursor.execute("SELECT COUNT(*) FROM products")
                total_products = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(DISTINCT owner_id) FROM products WHERE owner_id IS NOT NULL")
                total_owners = cursor.fetchone()[0]
                cursor.execute(
                    "SELECT COUNT(DISTINCT u.user_id) FROM users u INNER JOIN products p ON u.user_id = p.tester_id")
                total_testers = cursor.fetchone()[0]
                cursor.execute(
                    "SELECT COUNT(*) FROM products WHERE tester_id IS NULL AND COALESCE(status, 'pending') = 'approved'")
                pending_tests = cursor.fetchone()[0]

                # Summary table
                summary_data = [
                    ['Metric', 'Value'],
                    ['Total Products', str(total_products)],
                    ['Total Owners', str(total_owners)],
                    ['Total Testers', str(total_testers)],
                    ['Pending Tests', str(pending_tests)]
                ]

                summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), chinese_font)  # Use Chinese font for data
                ]))
                story.append(summary_table)
                story.append(Spacer(1, 30))

                # Products by Owner Section
                story.append(Paragraph("Products by Owner", styles['Heading2']))
                story.append(Spacer(1, 12))

                # Get products by owner data
                cursor.execute("""
                    SELECT u.username, 
                           COUNT(p.product_id) as product_count,
                           GROUP_CONCAT(p.product_name, ', ') as product_list
                    FROM users u
                    INNER JOIN products p ON u.user_id = p.owner_id
                    GROUP BY u.user_id, u.username
                    ORDER BY product_count DESC
                """)

                owners_data = cursor.fetchall()
                if owners_data:
                    # Create table for owners data
                    owners_table_data = [['Owner', 'Products Quantity', 'List of Products']]
                    chart_data = []
                    chart_labels = []

                    for username, count, product_list in owners_data:
                        product_list_text = product_list if product_list else "No products"
                        # Truncate long product lists for table display
                        if len(product_list_text) > 80:
                            product_list_display = product_list_text[:77] + "..."
                        else:
                            product_list_display = product_list_text

                        owners_table_data.append([username, str(count), product_list_display])
                        chart_data.append(count)
                        chart_labels.append(username)

                    owners_table = Table(owners_table_data, colWidths=[1.5 * inch, 1.2 * inch, 4 * inch])
                    owners_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 12),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('FONTNAME', (0, 1), (-1, -1), chinese_font),  # Use Chinese font for data
                        ('FONTSIZE', (0, 1), (-1, -1), 9)
                    ]))
                    story.append(owners_table)
                    story.append(Spacer(1, 20))

                    # Add bar chart for owners
                    if len(chart_data) > 0:
                        drawing = Drawing(400, 250)
                        chart = RLBarChart()
                        chart.x = 50
                        chart.y = 50
                        chart.height = 150
                        chart.width = 300
                        chart.data = [chart_data]
                        chart.categoryAxis.categoryNames = chart_labels
                        chart.categoryAxis.labels.boxAnchor = 'ne'
                        chart.categoryAxis.labels.dx = 8
                        chart.categoryAxis.labels.dy = -2
                        chart.categoryAxis.labels.angle = 30
                        chart.valueAxis.valueMin = 0
                        chart.valueAxis.valueMax = max(chart_data) * 1.2
                        chart.bars[0].fillColor = colors.blue
                        drawing.add(chart)
                        story.append(drawing)
                        story.append(Spacer(1, 20))

                story.append(PageBreak())

                # Tests by Tester Section
                story.append(Paragraph("Tests by Tester", styles['Heading2']))
                story.append(Spacer(1, 12))

                # Get tests by tester data
                cursor.execute("""
                    SELECT u.username, 
                           COUNT(p.product_id) as testing_count,
                           GROUP_CONCAT(p.product_name, ', ') as testing_list
                    FROM users u
                    INNER JOIN products p ON u.user_id = p.tester_id
                    GROUP BY u.user_id, u.username
                    ORDER BY testing_count DESC
                """)

                testers_data = cursor.fetchall()
                if testers_data:
                    # Create table for testers data
                    testers_table_data = [['Tester', 'Testing Quantity', 'List of Testing']]
                    chart_data = []
                    chart_labels = []

                    for username, count, testing_list in testers_data:
                        testing_list_text = testing_list if testing_list else "No testing assigned"
                        # Truncate long testing lists for table display
                        if len(testing_list_text) > 80:
                            testing_list_display = testing_list_text[:77] + "..."
                        else:
                            testing_list_display = testing_list_text

                        testers_table_data.append([username, str(count), testing_list_display])
                        chart_data.append(count)
                        chart_labels.append(username)

                    testers_table = Table(testers_table_data, colWidths=[1.5 * inch, 1.2 * inch, 4 * inch])
                    testers_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 12),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('FONTNAME', (0, 1), (-1, -1), chinese_font),  # Use Chinese font for data
                        ('FONTSIZE', (0, 1), (-1, -1), 9)
                    ]))
                    story.append(testers_table)
                    story.append(Spacer(1, 20))

                    # Add bar chart for testers
                    if len(chart_data) > 0:
                        drawing = Drawing(400, 250)
                        chart = RLBarChart()
                        chart.x = 50
                        chart.y = 50
                        chart.height = 150
                        chart.width = 300
                        chart.data = [chart_data]
                        chart.categoryAxis.categoryNames = chart_labels
                        chart.categoryAxis.labels.boxAnchor = 'ne'
                        chart.categoryAxis.labels.dx = 8
                        chart.categoryAxis.labels.dy = -2
                        chart.categoryAxis.labels.angle = 30
                        chart.valueAxis.valueMin = 0
                        chart.valueAxis.valueMax = max(chart_data) * 1.2
                        chart.bars[0].fillColor = colors.green
                        drawing.add(chart)
                        story.append(drawing)

                # Footer
                story.append(Spacer(1, 30))
                footer_text = "Report generated by Analysis Dashboard System"
                story.append(Paragraph(footer_text, styles['Normal']))

                # Build PDF
                doc.build(story)
                conn.close()

                QMessageBox.information(
                    self,
                    "PDF Export Successful",
                    f"Analysis report exported successfully!\n\n"
                    f"File saved as: {filename}\n\n"
                    f"The PDF includes:\n"
                    f"• Summary statistics table\n"
                    f"• Products by Owner data and chart\n"
                    f"• Tests by Tester data and chart\n"
                    f"• Professional formatting with colors"
                )

                print(f"✓ PDF export successful: {filename}")

            except ImportError as import_error:
                # Fallback: Try alternative PDF library or show instruction
                try:
                    import matplotlib.pyplot as plt
                    from matplotlib.backends.backend_pdf import PdfPages
                    import numpy as np

                    # Set font to support Chinese characters
                    try:
                        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS',
                                                           'DejaVu Sans']
                        plt.rcParams['axes.unicode_minus'] = False  # Fix minus sign display
                        print("✓ Chinese font support enabled for PDF export")
                    except Exception as font_error:
                        print(f"⚠️ Font configuration warning: {font_error}")
                        # Continue with default font

                    # Alternative implementation using matplotlib
                    conn = sqlite3.connect("testing_system.db")
                    cursor = conn.cursor()

                    with PdfPages(filename) as pdf:
                        # Page 1: Summary and Products by Owner
                        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(11, 8.5))
                        fig.suptitle('Analysis Dashboard Report', fontsize=16, fontweight='bold')

                        # Summary statistics
                        cursor.execute("SELECT COUNT(*) FROM products")
                        total_products = cursor.fetchone()[0]
                        cursor.execute("SELECT COUNT(DISTINCT owner_id) FROM products WHERE owner_id IS NOT NULL")
                        total_owners = cursor.fetchone()[0]
                        cursor.execute(
                            "SELECT COUNT(DISTINCT u.user_id) FROM users u INNER JOIN products p ON u.user_id = p.tester_id")
                        total_testers = cursor.fetchone()[0]
                        cursor.execute(
                            "SELECT COUNT(*) FROM products WHERE tester_id IS NULL AND COALESCE(status, 'pending') = 'approved'")
                        pending_tests = cursor.fetchone()[0]

                        # Summary bar chart
                        summary_categories = ['Products', 'Owners', 'Testers', 'Pending Tests']
                        summary_values = [total_products, total_owners, total_testers, pending_tests]
                        ax1.bar(summary_categories, summary_values,
                                color=['skyblue', 'lightgreen', 'orange', 'lightcoral'])
                        ax1.set_title('System Overview')
                        ax1.set_ylabel('Count')
                        for i, v in enumerate(summary_values):
                            ax1.text(i, v + max(summary_values) * 0.01, str(v), ha='center', va='bottom')

                        # Products by Owner
                        cursor.execute("""
                            SELECT u.username, COUNT(p.product_id) as product_count
                            FROM users u
                            INNER JOIN products p ON u.user_id = p.owner_id
                            GROUP BY u.user_id, u.username
                            ORDER BY product_count DESC
                            LIMIT 10
                        """)
                        owners_data = cursor.fetchall()
                        if owners_data:
                            owners, counts = zip(*owners_data)
                            ax2.bar(range(len(owners)), counts, color='blue', alpha=0.7)
                            ax2.set_title('Products by Owner')
                            ax2.set_xlabel('Owners')
                            ax2.set_ylabel('Number of Products')
                            ax2.set_xticks(range(len(owners)))
                            ax2.set_xticklabels(owners, rotation=45, ha='right')
                            for i, v in enumerate(counts):
                                ax2.text(i, v + max(counts) * 0.01, str(v), ha='center', va='bottom')

                        # Tests by Tester
                        cursor.execute("""
                            SELECT u.username, COUNT(p.product_id) as testing_count
                            FROM users u
                            INNER JOIN products p ON u.user_id = p.tester_id
                            GROUP BY u.user_id, u.username
                            ORDER BY testing_count DESC
                            LIMIT 10
                        """)
                        testers_data = cursor.fetchall()
                        if testers_data:
                            testers, counts = zip(*testers_data)
                            ax3.bar(range(len(testers)), counts, color='green', alpha=0.7)
                            ax3.set_title('Tests by Tester')
                            ax3.set_xlabel('Testers')
                            ax3.set_ylabel('Number of Tests')
                            ax3.set_xticks(range(len(testers)))
                            ax3.set_xticklabels(testers, rotation=45, ha='right')
                            for i, v in enumerate(counts):
                                ax3.text(i, v + max(counts) * 0.01, str(v), ha='center', va='bottom')

                        # Hide the fourth subplot
                        ax4.axis('off')
                        ax4.text(0.5, 0.5, f'Generated on:\n{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
                                 ha='center', va='center', transform=ax4.transAxes, fontsize=12)

                        plt.tight_layout()
                        pdf.savefig(fig, bbox_inches='tight')
                        plt.close()

                    conn.close()

                    QMessageBox.information(
                        self,
                        "PDF Export Successful",
                        f"Analysis report exported successfully!\n\n"
                        f"File saved as: {filename}\n\n"
                        f"The PDF includes:\n"
                        f"• Summary statistics chart\n"
                        f"• Products by Owner chart\n"
                        f"• Tests by Tester chart\n\n"
                        f"📊 Charts created with matplotlib"
                    )

                    print(f"✓ PDF export successful (matplotlib): {filename}")

                except ImportError:
                    # No PDF libraries available
                    QMessageBox.information(
                        self,
                        "PDF Export - Library Required",
                        f"PDF export requires additional libraries.\n\n"
                        f"To enable PDF export, please install:\n"
                        f"• pip install reportlab (recommended)\n"
                        f"OR\n"
                        f"• pip install matplotlib\n\n"
                        f"The report would be saved as: {filename}\n\n"
                        f"Features would include:\n"
                        f"• Professional PDF formatting\n"
                        f"• Summary statistics\n"
                        f"• Interactive charts\n"
                        f"• Data tables"
                    )
                    print(f"📋 PDF libraries not available. Install reportlab or matplotlib for PDF export.")

        except Exception as e:
            print(f"✗ Error exporting to PDF: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to export PDF: {str(e)}")

    def export_analysis_to_excel(self):
        """Export analysis data to Excel"""
        try:
            from datetime import datetime
            import sqlite3

            # Get current timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"analysis_report_{timestamp}.xlsx"

            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # For now, create a simple CSV-style export since openpyxl might not be available
            # In a full implementation, you would use openpyxl or pandas

            try:
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font

                # Try to import chart functionality
                charts_available = True
                try:
                    from openpyxl.chart import BarChart, Reference
                    import openpyxl.chart.label
                except ImportError:
                    charts_available = False
                    print("📊 Chart functionality not available - will export data only")

                wb = Workbook()

                # Summary sheet
                ws_summary = wb.active
                ws_summary.title = "Summary"
                ws_summary.append(["Metric", "Value"])

                # Get summary data
                cursor.execute("SELECT COUNT(*) FROM products")
                total_products = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(DISTINCT owner_id) FROM products WHERE owner_id IS NOT NULL")
                total_owners = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(DISTINCT user_id) FROM users WHERE role IN ('Tester', 'tester')")
                total_testers = cursor.fetchone()[0]
                cursor.execute(
                    "SELECT COUNT(*) FROM products WHERE tester_id IS NULL AND COALESCE(status, 'pending') = 'approved'")
                pending_tests = cursor.fetchone()[0]

                ws_summary.append(["Total Products", total_products])
                ws_summary.append(["Total Owners", total_owners])
                ws_summary.append(["Total Testers", total_testers])
                ws_summary.append(["Pending Tests", pending_tests])

                # Add summary bar chart if charts are available
                if charts_available:
                    chart_summary = BarChart()
                    chart_summary.title = "System Overview"
                    chart_summary.x_axis.title = "Categories"
                    chart_summary.y_axis.title = "Count"

                    # Data range for summary chart
                    data = Reference(ws_summary, min_col=2, min_row=1, max_row=5, max_col=2)
                    categories = Reference(ws_summary, min_col=1, min_row=2, max_row=5)

                    chart_summary.add_data(data, titles_from_data=True)
                    chart_summary.set_categories(categories)

                    # Add data labels to show values on bars
                    chart_summary.dataLabels = openpyxl.chart.label.DataLabelList()
                    chart_summary.dataLabels.showVal = True  # Show values
                    chart_summary.dataLabels.showCatName = False  # Don't show category names
                    chart_summary.dataLabels.showSerName = False  # Don't show series names

                    # Position chart to the right of the summary data
                    ws_summary.add_chart(chart_summary, "D2")

                # Products by Owner sheet
                ws_owners = wb.create_sheet("Products by Owner")
                ws_owners.append(["Owner", "Products Quantity", "List of Products"])

                cursor.execute("""
                    SELECT u.username, 
                           COUNT(p.product_id) as product_count,
                           GROUP_CONCAT(p.product_name, ', ') as product_list
                    FROM users u
                    INNER JOIN products p ON u.user_id = p.owner_id
                    GROUP BY u.user_id, u.username
                    ORDER BY product_count DESC
                """)

                owners_data = cursor.fetchall()
                for row in owners_data:
                    username, count, product_list = row
                    product_list_text = product_list if product_list else "No products"
                    ws_owners.append([username, count, product_list_text])

                # Add bar chart for Products by Owner if charts are available
                if charts_available and len(owners_data) > 0:
                    chart1 = BarChart()
                    chart1.title = "Products by Owner"
                    chart1.x_axis.title = "Owners"
                    chart1.y_axis.title = "Number of Products"

                    # Data range for chart (excluding header and product list column)
                    data = Reference(ws_owners, min_col=2, min_row=1, max_row=len(owners_data) + 1, max_col=2)
                    categories = Reference(ws_owners, min_col=1, min_row=2, max_row=len(owners_data) + 1)

                    chart1.add_data(data, titles_from_data=True)
                    chart1.set_categories(categories)

                    # Add data labels to show values on bars
                    chart1.dataLabels = openpyxl.chart.label.DataLabelList()
                    chart1.dataLabels.showVal = True  # Show values
                    chart1.dataLabels.showCatName = False  # Don't show category names
                    chart1.dataLabels.showSerName = False  # Don't show series names

                    # Position chart to the right of the data
                    ws_owners.add_chart(chart1, "E2")

                # Tests by Tester sheet - use same query as the UI
                ws_testers = wb.create_sheet("Tests by Tester")
                ws_testers.append(["Tester", "Testing Quantity", "List of Testing"])

                cursor.execute("""
                    SELECT u.username, 
                           COUNT(p.product_id) as testing_count,
                           GROUP_CONCAT(p.product_name, ', ') as testing_list
                    FROM users u
                    INNER JOIN products p ON u.user_id = p.tester_id
                    GROUP BY u.user_id, u.username
                    ORDER BY testing_count DESC
                """)

                testers_data = cursor.fetchall()
                for row in testers_data:
                    username, count, testing_list = row
                    testing_list_text = testing_list if testing_list else "No testing assigned"
                    ws_testers.append([username, count, testing_list_text])

                # Add bar chart for Tests by Tester if charts are available
                if charts_available and len(testers_data) > 0:
                    chart2 = BarChart()
                    chart2.title = "Tests by Tester"
                    chart2.x_axis.title = "Testers"
                    chart2.y_axis.title = "Number of Tests"

                    # Data range for chart (excluding header and testing list column)
                    data = Reference(ws_testers, min_col=2, min_row=1, max_row=len(testers_data) + 1, max_col=2)
                    categories = Reference(ws_testers, min_col=1, min_row=2, max_row=len(testers_data) + 1)

                    chart2.add_data(data, titles_from_data=True)
                    chart2.set_categories(categories)

                    # Add data labels to show values on bars
                    chart2.dataLabels = openpyxl.chart.label.DataLabelList()
                    chart2.dataLabels.showVal = True  # Show values
                    chart2.dataLabels.showCatName = False  # Don't show category names
                    chart2.dataLabels.showSerName = False  # Don't show series names

                    # Position chart to the right of the data
                    ws_testers.add_chart(chart2, "E2")

                # Save the workbook
                wb.save(filename)

                # Success message based on chart availability
                chart_message = ""
                if charts_available:
                    chart_message = " with bar chart"
                    chart_note = "\n📊 Each sheet includes interactive charts!"
                else:
                    chart_note = "\n📋 Data tables exported (charts require full openpyxl installation)"

                QMessageBox.information(
                    self,
                    "Excel Export Successful",
                    f"Analysis report exported successfully!\n\n"
                    f"File saved as: {filename}\n\n"
                    f"The report includes:\n"
                    f"• Summary statistics{chart_message}\n"
                    f"• Products by Owner{chart_message}\n"
                    f"• Tests by Tester{chart_message}"
                    f"{chart_note}"
                )

                print(f"✓ Excel export successful: {filename}")

            except ImportError:
                # Fallback to CSV export if openpyxl is not available
                csv_filename = filename.replace('.xlsx', '.csv')

                QMessageBox.information(
                    self,
                    "Excel Export",
                    f"Excel export functionality is ready!\n\n"
                    f"The report would be saved as: {filename}\n\n"
                    f"This would include all analysis data in multiple sheets.\n\n"
                    f"Implementation requires openpyxl library for Excel format."
                )

                print(f"✓ Excel export prepared: {filename}")

            conn.close()

        except Exception as e:
            print(f"✗ Error exporting to Excel: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to export Excel: {str(e)}")

    def assign_tester_to_product(self, product_id, row):
        """Assign a tester to a product"""
        try:
            # Get the tester combo box from the specified row
            tester_combo = self.assign_table.cellWidget(row, 4)
            if not tester_combo or tester_combo.currentIndex() == 0:
                QMessageBox.warning(self, "No Tester Selected",
                                    "Please select a tester before assigning.")
                return

            # Get selected tester data
            current_index = tester_combo.currentIndex()
            tester_data = tester_combo.itemData(current_index)

            if not tester_data:
                QMessageBox.warning(self, "Invalid Selection",
                                    "Please select a valid tester.")
                return

            tester_id = tester_data['tester_id']
            tester_username = tester_data['username']
            tester_fullname = tester_data.get('fullname', '')

            # Get product details for confirmation
            product_name = self.assign_table.item(row, 1).text()
            batch = self.assign_table.item(row, 0).text()

            # Format tester name for confirmation
            if tester_fullname and tester_fullname.strip():
                tester_display = f"{tester_fullname} ({tester_username})"
            else:
                tester_display = tester_username

            # Confirmation dialog
            reply = QMessageBox.question(
                self,
                "Confirm Assignment",
                f"Are you sure you want to assign this product to the selected tester?\n\n"
                f"Product: {product_name}\n"
                f"Batch: {batch}\n"
                f"Tester: {tester_display}\n\n"
                f"This action will create a new assignment record.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply != QMessageBox.Yes:
                return

            # Get current user info for assigned_by
            current_user_id = self.user_info.get('user_id')
            if not current_user_id:
                QMessageBox.critical(self, "Error", "Unable to determine current user ID.")
                return

            # Save to database
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            try:
                # Check if product still exists and is unassigned
                cursor.execute("""
                    SELECT tester_id FROM products WHERE product_id = ?
                """, (product_id,))
                result = cursor.fetchone()

                if not result:
                    QMessageBox.warning(self, "Product Not Found",
                                        "The selected product no longer exists.")
                    conn.close()
                    return

                if result[0] is not None:
                    QMessageBox.warning(self, "Already Assigned",
                                        "This product has already been assigned to a tester.")
                    conn.close()
                    self.load_assign_data()  # Refresh the table
                    return

                # Insert into product_tester_assignments table with correct Malaysia time
                malaysia_time = format_malaysia_time()
                cursor.execute("""
                    INSERT INTO product_tester_assignments 
                    (product_id, tester_id, assigned_by, assigned_date)
                    VALUES (?, ?, ?, ?)
                """, (product_id, tester_id, current_user_id, malaysia_time))

                # Update products table with tester_id
                cursor.execute("""
                    UPDATE products 
                    SET tester_id = ? 
                    WHERE product_id = ?
                """, (tester_id, product_id))

                # Commit the transaction
                conn.commit()

                # Send assignment email to tester and CC to owner
                self.send_assignment_email(product_id, product_name, batch, tester_data, cursor)

                # Success message
                QMessageBox.information(
                    self,
                    "Assignment Successful",
                    f"Product successfully assigned to {tester_display}!\n\n"
                    f"Product: {product_name}\n"
                    f"Batch: {batch}\n"
                    f"Assignment has been recorded in the database.\n"
                    f"Email notifications have been sent."
                )

                # Log the activity
                self.add_activity_log(
                    self.username,
                    "Product Assignment",
                    f"Assigned product {product_name} (ID: {product_id}) to tester {tester_username}"
                )

                # Refresh the assign table to remove the assigned product
                self.load_assign_data()

                print(f"✓ Product {product_id} successfully assigned to tester {tester_username}")

            except sqlite3.Error as db_error:
                conn.rollback()
                print(f"✗ Database error during assignment: {db_error}")
                QMessageBox.critical(self, "Database Error",
                                     f"Failed to save assignment: {str(db_error)}")
            finally:
                conn.close()

        except Exception as e:
            print(f"✗ Error in assign_tester_to_product: {e}")
            QMessageBox.critical(self, "Error",
                                 f"An unexpected error occurred: {str(e)}")

    def send_assignment_email(self, product_id, product_name, batch, tester_data, cursor):
        """Send assignment email to tester with CC to product owner"""
        try:
            # Get tester email and details
            tester_id = tester_data['tester_id']
            tester_username = tester_data['username']
            tester_fullname = tester_data.get('fullname', '')

            cursor.execute("""
                SELECT email FROM users WHERE user_id = ?
            """, (tester_id,))
            tester_result = cursor.fetchone()
            tester_email = tester_result[0] if tester_result and tester_result[0] else None

            # Get product owner email and details
            cursor.execute("""
                SELECT u.email, u.username, u.fullname 
                FROM products p
                JOIN users u ON p.owner_id = u.user_id
                WHERE p.product_id = ?
            """, (product_id,))
            owner_result = cursor.fetchone()

            if not owner_result:
                print("⚠️ Product owner not found - email not sent")
                return

            owner_email, owner_username, owner_fullname = owner_result

            # Validate email addresses
            if not tester_email:
                print(f"⚠️ Tester {tester_username} has no email address - email not sent")
                return

            if not owner_email:
                print(f"⚠️ Owner {owner_username} has no email address - CC not sent")
                owner_email = None  # Will skip CC

            # Get current user (assigned_by) details
            current_user_name = self.user_info.get('fullname', self.user_info.get('username', 'System'))

            # Format names for display
            tester_display_name = tester_fullname if tester_fullname and tester_fullname.strip() else tester_username
            owner_display_name = owner_fullname if owner_fullname and owner_fullname.strip() else owner_username

            # Get email template and create content
            template = self.get_email_template('assignment')

            # Replace placeholders in template
            subject = template['subject'].format(
                PRODUCT_NAME=product_name,
                BATCH=batch,
                PRODUCT_ID=product_id,
                TESTER_NAME=tester_display_name,
                OWNER_NAME=owner_display_name,
                OWNER_USERNAME=owner_username,
                ASSIGNED_BY=current_user_name,
                ASSIGNMENT_DATE=format_malaysia_time()
            )

            body = template['body'].format(
                PRODUCT_NAME=product_name,
                BATCH=batch,
                PRODUCT_ID=product_id,
                TESTER_NAME=tester_display_name,
                OWNER_NAME=owner_display_name,
                OWNER_USERNAME=owner_username,
                ASSIGNED_BY=current_user_name,
                ASSIGNMENT_DATE=format_malaysia_time()
            )

            # Send email to tester with CC to owner
            try:
                self.send_actual_email(
                    recipient_email=tester_email,
                    subject=subject,
                    body=body,
                    recipient_name=tester_display_name,
                    cc_email=owner_email
                )

                print(f"✓ Assignment email sent to {tester_email}" +
                      (f" with CC to {owner_email}" if owner_email else ""))

            except Exception as email_error:
                print(f"⚠️ Failed to send assignment email: {email_error}")
                # Don't raise the error - assignment should still succeed even if email fails

        except Exception as e:
            print(f"⚠️ Error in send_assignment_email: {e}")
            # Don't raise the error - assignment should still succeed even if email fails

    def view_assignment_records(self):
        """View past assignment records in a dialog"""
        try:
            # Create the dialog
            dialog = QDialog(self)
            dialog.setWindowTitle("📋 Assignment Records History")
            dialog.setFixedSize(1400, 800)
            dialog.setModal(True)

            dialog.setStyleSheet("""
                QDialog {
                    background-color: #f5f5f0;
                }
                QTableWidget {
                    background-color: white;
                    border-radius: 8px;
                    border: none;
                    gridline-color: #E0E0E0;
                }
                QHeaderView::section {
                    background-color: #2196F3;
                    color: white;
                    padding: 12px;
                    border: none;
                    font-weight: bold;
                    font-size: 12px;
                }
                QTableWidget::item {
                    padding: 8px;
                    border-bottom: 1px solid #E0E0E0;
                    height: 100px;
                }
                QTableWidget::item:selected {
                    background-color: #E3F2FD;
                }
                QPushButton {
                    background-color: #757575;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 10px 20px;
                    font-weight: bold;
                    margin: 10px;
                }
                QPushButton:hover {
                    background-color: #616161;
                }
            """)

            layout = QVBoxLayout()

            # Title
            title_label = QLabel("📋 Assignment Records History")
            title_label.setFont(QFont("Arial", 16, QFont.Bold))
            title_label.setStyleSheet("color: #2196F3; margin: 10px; text-align: center;")
            title_label.setAlignment(Qt.AlignCenter)
            layout.addWidget(title_label)

            # Table for assignment records
            records_table = QTableWidget()
            records_table.setColumnCount(7)
            records_table.setHorizontalHeaderLabels([
                "Assignment ID", "Product Name", "Batch", "Tester", "Owner", "Assigned By", "Assignment Date"
            ])

            # Set column widths
            header = records_table.horizontalHeader()
            header.resizeSection(0, 140)  # Assignment ID
            header.resizeSection(1, 200)  # Product Name
            header.resizeSection(2, 100)  # Batch
            header.resizeSection(3, 200)  # Tester
            header.resizeSection(4, 200)  # Owner
            header.resizeSection(5, 140)  # Assigned By
            header.resizeSection(6, 140)  # Assignment Date

            records_table.setAlternatingRowColors(True)
            records_table.setSelectionBehavior(QTableWidget.SelectRows)
            records_table.setMinimumHeight(400)

            # Set row height to 12px
            records_table.verticalHeader().setDefaultSectionSize(100)
            records_table.verticalHeader().setMinimumSectionSize(100)

            # Load assignment records
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            cursor.execute("""
                SELECT 
                    pta.assignment_id,
                    p.product_name,
                    p.batch,
                    CASE 
                        WHEN t.fullname IS NOT NULL AND t.fullname != '' THEN t.fullname || ' (' || t.username || ')'
                        ELSE t.username
                    END as tester,
                    CASE 
                        WHEN o.fullname IS NOT NULL AND o.fullname != '' THEN o.fullname || ' (' || o.username || ')'
                        ELSE o.username
                    END as owner,
                    CASE 
                        WHEN a.fullname IS NOT NULL AND a.fullname != '' THEN a.fullname || ' (' || a.username || ')'
                        ELSE a.username
                    END as assigned_by,
                    pta.assigned_date
                FROM product_tester_assignments pta
                JOIN products p ON pta.product_id = p.product_id
                JOIN users t ON pta.tester_id = t.user_id
                JOIN users o ON p.owner_id = o.user_id
                JOIN users a ON pta.assigned_by = a.user_id
                ORDER BY pta.assigned_date DESC
            """)

            records = cursor.fetchall()
            conn.close()

            records_table.setRowCount(len(records))

            for row, record in enumerate(records):
                for col, value in enumerate(record):
                    item = QTableWidgetItem(str(value) if value else "")
                    records_table.setItem(row, col, item)

            layout.addWidget(records_table)

            # Status label
            status_label = QLabel(f"Total Assignment Records: {len(records)}")
            status_label.setStyleSheet("color: #666; font-size: 12px; margin: 5px;")
            layout.addWidget(status_label)

            # Close button
            close_btn = QPushButton("Close")
            close_btn.clicked.connect(dialog.accept)
            close_layout = QHBoxLayout()
            close_layout.addStretch()
            close_layout.addWidget(close_btn)
            close_layout.addStretch()
            layout.addLayout(close_layout)

            dialog.setLayout(layout)
            dialog.exec_()

        except Exception as e:
            print(f"✗ Error loading assignment records: {e}")
            QMessageBox.critical(self, "Error", f"Failed to load assignment records: {str(e)}")

    def view_test_updates(self):
        """View test updates with approval/rejection functionality"""
        try:
            dialog = QDialog(self)
            dialog.setWindowTitle("🧪 Test Updates - Admin Review")
            dialog.setMinimumSize(1400, 900)
            dialog.setModal(True)

            # Apply modern dialog styling
            dialog.setStyleSheet("""
                QDialog {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                        stop:0 #f8f9fa, stop:1 #e9ecef);
                    border-radius: 15px;
                }
            """)

            # Main layout with better spacing
            layout = QVBoxLayout()
            layout.setContentsMargins(30, 25, 30, 25)
            layout.setSpacing(20)

            # Enhanced header section
            header_frame = QFrame()
            header_frame.setStyleSheet("""
                QFrame {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                        stop:0 #2E7D32, stop:1 #388E3C);
                    border-radius: 12px;
                    padding: 20px;
                    margin-bottom: 10px;
                }
            """)
            header_layout = QVBoxLayout(header_frame)

            # Title with icon
            title_label = QLabel("🧪 Test Updates - Admin Review")
            title_label.setFont(QFont("Arial", 18, QFont.Bold))
            title_label.setStyleSheet("""
                color: white; 
                margin: 10px 0px;
                background: transparent;
            """)
            title_label.setAlignment(Qt.AlignCenter)
            header_layout.addWidget(title_label)

            layout.addWidget(header_frame)

            # Add filter section
            self.create_test_filter_section(layout)

            # Create enhanced tree view for testing records
            tree_view = QTreeWidget()
            tree_view.setHeaderLabels([
                "Test ID", "Product Name", "Batch", "Tester", "Progress Status",
                "Test Result", "Test Start", "Test End", "Actions"
            ])

            # Enhanced tree view styling
            tree_view.setStyleSheet("""
                QTreeWidget {
                    background-color: white;
                    border: 2px solid #e0e0e0;
                    border-radius: 12px;
                    font-size: 12px;
                    alternate-background-color: #f8f9fa;
                    selection-background-color: #e3f2fd;
                    gridline-color: #e0e0e0;
                }
                QTreeWidget::item {
                    height: 38px;
                    padding: 5px;
                    border-bottom: 1px solid #f0f0f0;
                }
                QTreeWidget::item:selected {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                        stop:0 #e3f2fd, stop:1 #bbdefb);
                    color: #1976d2;
                    border: 2px solid #2196F3;
                    border-radius: 4px;
                }
                QTreeWidget::item:hover {
                    background-color: #f5f5f5;
                    border: 1px solid #ddd;
                    border-radius: 4px;
                }
                QHeaderView::section {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                        stop:0 #2E7D32, stop:1 #1B5E20);
                    color: white;
                    padding: 10px;
                    border: none;
                    font-weight: bold;
                    font-size: 13px;
                    text-align: center;
                }
                QHeaderView::section:first {
                    border-top-left-radius: 8px;
                }
                QHeaderView::section:last {
                    border-top-right-radius: 8px;
                }
            """)

            # Enhanced column widths
            tree_view.setColumnWidth(0, 70)  # Test ID
            tree_view.setColumnWidth(1, 180)  # Product Name
            tree_view.setColumnWidth(2, 100)  # Batch
            tree_view.setColumnWidth(3, 140)  # Tester
            tree_view.setColumnWidth(4, 120)  # Progress Status
            tree_view.setColumnWidth(5, 110)  # Test Result
            tree_view.setColumnWidth(6, 130)  # Test Start
            tree_view.setColumnWidth(7, 130)  # Test End
            tree_view.setColumnWidth(8, 180)  # Actions

            tree_view.setAlternatingRowColors(True)
            tree_view.setRootIsDecorated(False)
            tree_view.setSelectionBehavior(QTreeWidget.SelectRows)
            tree_view.setMinimumHeight(450)

            # Load testing data
            self.load_testing_data(tree_view)

            # Connect filter dropdown to reload data
            if hasattr(self, 'test_result_filter'):
                self.test_result_filter.currentTextChanged.connect(lambda: self.load_testing_data(tree_view))

            layout.addWidget(tree_view)

            # Enhanced bottom buttons
            button_frame = QFrame()
            button_frame.setStyleSheet("""
                QFrame {
                    background-color: rgba(255, 255, 255, 0.8);
                    border-radius: 8px;
                    padding: 10px;
                }
            """)
            button_layout = QHBoxLayout(button_frame)

            # Status info
            status_label = QLabel("💡 Click 'Approve' or 'Reject' buttons to update test results")
            status_label.setStyleSheet("""
                color: #666;
                font-size: 12px;
                font-style: italic;
                background: transparent;
            """)
            button_layout.addWidget(status_label)

            button_layout.addStretch()

            # Refresh button
            refresh_btn = QPushButton("🔄 Refresh Data")
            refresh_btn.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                        stop:0 #2196F3, stop:1 #1976D2);
                    color: white;
                    border: none;
                    border-radius: 8px;
                    padding: 12px 24px;
                    font-weight: bold;
                    font-size: 14px;
                    min-width: 120px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                        stop:0 #1976D2, stop:1 #1565C0);
                    transform: translateY(-1px);
                }
                QPushButton:pressed {
                    background: #1565C0;
                }
            """)
            refresh_btn.clicked.connect(lambda: self.load_testing_data(tree_view))
            button_layout.addWidget(refresh_btn)

            # Close button
            close_btn = QPushButton("✕ Close")
            close_btn.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                        stop:0 #757575, stop:1 #616161);
                    color: white;
                    border: none;
                    border-radius: 8px;
                    padding: 12px 24px;
                    font-weight: bold;
                    font-size: 14px;
                    min-width: 100px;
                    margin-left: 10px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                        stop:0 #616161, stop:1 #424242);
                    transform: translateY(-1px);
                }
                QPushButton:pressed {
                    background: #424242;
                }
            """)
            close_btn.clicked.connect(dialog.accept)
            button_layout.addWidget(close_btn)

            layout.addWidget(button_frame)
            dialog.setLayout(layout)
            dialog.exec_()

        except Exception as e:
            print(f"✗ Error opening test updates view: {e}")
            QMessageBox.critical(self, "Error", f"Failed to open test updates view: {str(e)}")

    def create_test_filter_section(self, layout):
        """Create filter section for test results"""
        try:
            # Filter container
            filter_frame = QFrame()
            filter_frame.setStyleSheet("""
                QFrame {
                    background-color: rgba(255, 255, 255, 0.9);
                    border-radius: 10px;
                    border: 1px solid #e0e0e0;
                    padding: 10px;
                }
            """)
            filter_layout = QHBoxLayout(filter_frame)
            filter_layout.setContentsMargins(15, 10, 15, 10)
            filter_layout.setSpacing(15)

            # Filter label
            filter_label = QLabel("🔍 Filter by Test Result:")
            filter_label.setFont(QFont("Arial", 12, QFont.Bold))
            filter_label.setStyleSheet("color: #333; background: transparent;")
            filter_layout.addWidget(filter_label)

            # Filter dropdown
            self.test_result_filter = QComboBox()
            self.test_result_filter.setFixedSize(150, 35)
            self.test_result_filter.addItems(["All Results", "Under Review", "Approved", "Rejected"])
            self.test_result_filter.setStyleSheet("""
                QComboBox {
                    background-color: white;
                    border: 2px solid #4CAF50;
                    border-radius: 6px;
                    padding: 5px 10px;
                    font-size: 12px;
                    font-weight: bold;
                    color: #333;
                }
                QComboBox:hover {
                    border-color: #66BB6A;
                }
                QComboBox::drop-down {
                    border: none;
                    background: transparent;
                }
                QComboBox::down-arrow {
                    image: none;
                    border: none;
                    width: 12px;
                    height: 12px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    border: 2px solid #4CAF50;
                    border-radius: 6px;
                    selection-background-color: #e8f5e8;
                    font-size: 12px;
                }
            """)

            filter_layout.addWidget(self.test_result_filter)
            filter_layout.addStretch()

            # Reset filter button
            reset_btn = QPushButton("🔄 Reset Filter")
            reset_btn.setFixedSize(120, 35)
            reset_btn.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                        stop:0 #757575, stop:1 #616161);
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 8px 12px;
                    font-weight: bold;
                    font-size: 12px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                        stop:0 #616161, stop:1 #424242);
                }
                QPushButton:pressed {
                    background: #424242;
                }
            """)
            reset_btn.clicked.connect(lambda: self.test_result_filter.setCurrentText("All Results"))
            filter_layout.addWidget(reset_btn)

            layout.addWidget(filter_frame)

        except Exception as e:
            print(f"Error creating filter section: {e}")

    def load_testing_data(self, tree_view, filter_result=None):
        """Load testing data into the tree view with optional filtering"""
        try:
            tree_view.clear()

            # Get filter value if not provided
            if filter_result is None and hasattr(self, 'test_result_filter'):
                filter_result = self.test_result_filter.currentText()

            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Build query with optional filter
            base_query = """
                SELECT 
                    t.test_id,
                    p.product_name,
                    p.batch,
                    CASE 
                        WHEN u.fullname IS NOT NULL AND u.fullname != '' 
                        THEN u.fullname || ' (' || u.username || ')'
                        ELSE u.username
                    END as tester,
                    ps.status_name as progress_status,
                    tr.result_name as test_result,
                    t.test_start,
                    t.test_end,
                    t.test_result_id,
                    CASE 
                        WHEN t.test_start IS NOT NULL AND t.test_end IS NOT NULL AND t.test_result_id = 1 
                        THEN 0 
                        ELSE 1 
                    END as sort_priority
                FROM testing t
                JOIN products p ON t.product_id = p.product_id
                JOIN users u ON t.user_id = u.user_id
                LEFT JOIN progress_status ps ON t.progress_status_id = ps.status_id
                LEFT JOIN test_results tr ON t.test_result_id = tr.result_id
            """

            # Add filter condition if needed
            if filter_result and filter_result != "All Results":
                if filter_result == "Under Review":
                    base_query += " WHERE t.test_result_id = 1"
                elif filter_result == "Approved":
                    base_query += " WHERE t.test_result_id = 2"
                elif filter_result == "Rejected":
                    base_query += " WHERE t.test_result_id = 0"

            base_query += " ORDER BY sort_priority ASC, t.test_start DESC"

            cursor.execute(base_query)

            records = cursor.fetchall()
            conn.close()

            for record in records:
                test_id, product_name, batch, tester, progress_status, test_result, test_start, test_end, test_result_id, sort_priority = record

                # Create tree item
                item = QTreeWidgetItem()
                item.setText(0, str(test_id))
                item.setText(1, product_name or "N/A")
                item.setText(2, batch or "N/A")
                item.setText(3, tester or "N/A")
                item.setText(4, progress_status or "N/A")
                item.setText(5, test_result or "Under Review")
                item.setText(6, test_start or "N/A")
                item.setText(7, test_end or "N/A")

                # Store test_result_id as item data
                item.setData(0, Qt.UserRole, test_result_id)
                item.setData(1, Qt.UserRole, test_id)

                # Add item to tree first
                tree_view.addTopLevelItem(item)

                # Add action buttons only when both test_start and test_end have data AND test_result_id is 1 (Under Review)
                if test_start is not None and test_end is not None and test_result_id == 1:
                    self.add_action_buttons(tree_view, item, test_id)
                    print(f"✓ Added buttons for test ID {test_id} (has start: {test_start}, end: {test_end})")
                else:
                    print(
                        f"✗ No buttons for test ID {test_id} (start: {test_start}, end: {test_end}, result_id: {test_result_id})")

            # Update status
            print(f"✓ Loaded {len(records)} testing records")

        except Exception as e:
            print(f"✗ Error loading testing data: {e}")
            QMessageBox.critical(self, "Database Error", f"Failed to load testing data: {str(e)}")

    def add_action_buttons(self, tree_view, item, test_id):
        """Add enhanced approve/reject buttons to tree view item"""
        try:
            # Create widget container for buttons
            button_widget = QWidget()
            button_widget.setStyleSheet("QWidget { background: transparent; }")
            button_layout = QHBoxLayout(button_widget)
            button_layout.setContentsMargins(5, 3, 5, 3)
            button_layout.setSpacing(5)

            # Enhanced Approve button
            approve_btn = QPushButton("✅ Approve")
            approve_btn.setFixedSize(75, 28)
            approve_btn.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                        stop:0 #4CAF50, stop:1 #45a049);
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 4px 8px;
                    font-size: 10px;
                    font-weight: bold;
                    text-align: center;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                        stop:0 #66BB6A, stop:1 #4CAF50);
                    transform: translateY(-1px);
                }
                QPushButton:pressed {
                    background: #388E3C;
                    transform: translateY(0px);
                }
            """)
            approve_btn.clicked.connect(lambda: self.update_test_result(test_id, 2, "Approved", tree_view))

            # Enhanced Reject button
            reject_btn = QPushButton("❌ Reject")
            reject_btn.setFixedSize(68, 28)
            reject_btn.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                        stop:0 #F44336, stop:1 #da190b);
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 4px 8px;
                    font-size: 10px;
                    font-weight: bold;
                    text-align: center;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                        stop:0 #EF5350, stop:1 #F44336);
                    transform: translateY(-1px);
                }
                QPushButton:pressed {
                    background: #C62828;
                    transform: translateY(0px);
                }
            """)
            reject_btn.clicked.connect(lambda: self.update_test_result(test_id, 0, "Rejected", tree_view))

            button_layout.addWidget(approve_btn)
            button_layout.addWidget(reject_btn)
            button_layout.addStretch()

            # Add the widget to the tree view
            tree_view.setItemWidget(item, 8, button_widget)

            print(f"✓ Added action buttons for test ID {test_id}")

        except Exception as e:
            print(f"✗ Error adding action buttons: {e}")

    def update_test_result(self, test_id, result_id, result_name, tree_view):
        """Update test result (approve/reject)"""
        try:
            # Confirmation dialog
            reply = QMessageBox.question(
                self,
                f"Confirm {result_name}",
                f"Are you sure you want to {result_name.lower()} this test result?\n\n"
                f"Test ID: {test_id}\n"
                f"Action: {result_name}",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply != QMessageBox.Yes:
                return

            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Update test_result_id in testing table
            cursor.execute("""
                UPDATE testing 
                SET test_result_id = ?
                WHERE test_id = ?
            """, (result_id, test_id))

            if cursor.rowcount > 0:
                conn.commit()

                # Show success message
                QMessageBox.information(
                    self,
                    "Success",
                    f"Test result has been successfully {result_name.lower()}!\n\n"
                    f"Test ID: {test_id}\n"
                    f"New Status: {result_name}"
                )

                # Reload the tree view data
                self.load_testing_data(tree_view)

                # Log admin activity
                admin_username = self.user_info.get('username', 'Admin')
                self.add_activity_log(
                    admin_username,
                    f"Test {result_name}",
                    f"Test ID {test_id} {result_name.lower()} by admin"
                )

            else:
                QMessageBox.warning(self, "Update Failed", "No records were updated. The test may not exist.")

            conn.close()

        except Exception as e:
            print(f"✗ Error updating test result: {e}")
            QMessageBox.critical(self, "Database Error", f"Failed to update test result: {str(e)}")

    def load_chat_data(self):
        """Load chat/announcement data from database"""
        try:
            conn = sqlite3.connect("testing_system.db", timeout=10.0)
            cursor = conn.cursor()

            # Load messages and users
            self.load_chat_messages()
            self.load_available_users()

            conn.close()
        except Exception as e:
            print(f"Error loading chat data: {e}")

    def load_chat_messages(self):
        """Load and display all chat messages"""
        try:
            conn = sqlite3.connect("testing_system.db", timeout=10.0)
            cursor = conn.cursor()

            messages = get_all_messages(cursor)
            conn.close()

            # Limit messages to prevent buffer overflow
            if len(messages) > 50:
                messages = messages[-50:]  # Keep only last 50 messages

            # Save current scroll position
            if hasattr(self, 'chat_display'):
                scrollbar = self.chat_display.verticalScrollBar()
                scroll_position = scrollbar.value()
                was_at_bottom = scroll_position == scrollbar.maximum()

                # Clear and rebuild chat display
                self.chat_display.clear()

                if not messages:
                    self.chat_display.append(
                        "<div style='text-align: center; color: #666; font-style: italic; padding: 20px;'>No messages yet. Start the conversation!</div>")
                    return

                # Display limited messages in chat bubble format
                for message in messages:
                    username, role, text, timestamp, message_id = message

                    # Format timestamp to Malaysia time
                    time_str = format_malaysia_time(timestamp)

                    # Get role color
                    role_color = self.role_colors.get(role.lower(), '#666666')

                    # Role icons
                    role_icons = {
                        'superadmin': '👑',
                        'admin': '📢',
                        'owner': '🏢',
                        'tester': '🧪'
                    }
                    role_icon = role_icons.get(role.lower(), '👤')

                    # Process tags in the message with length limit
                    safe_text = text[:500] if len(text) > 500 else text  # Limit message length
                    processed_text = self.process_tags(safe_text)

                    # Chat bubble styling - all messages aligned left with role color
                    bubble_style = (
                        f"background: linear-gradient(135deg, {role_color}15 0%, {role_color}08 100%); "
                        f"color: black; text-align: left; max-width: 80%; "
                        f"border: 2px solid {role_color}30; border-radius: 15px; "
                        f"box-shadow: 0 2px 8px rgba(0,0,0,0.1);"
                    )
                    username_style = f"color: {role_color}; font-weight: 700;"
                    time_style = f"color: #666; font-weight: 500;"

                    # Format message in sequence: username, timestamp, message
                    formatted_message = (
                        f"<div style='margin: 10px 0; padding: 12px 16px; {bubble_style}'>"
                        f"<div style='{username_style} font-size: 13px; margin-bottom: 2px;'>"
                        f"{role_icon} {username}"
                        f"</div>"
                        f"<div style='{time_style} font-size: 11px; margin-bottom: 8px;'>{time_str}</div>"
                        f"<div style='font-size: 14px; line-height: 1.5; color: black;'>{processed_text}</div>"
                        f"</div>"
                    )

                    self.chat_display.append(formatted_message)

                # Restore scroll position or scroll to bottom
                if was_at_bottom:  # Auto-scroll if user was at bottom
                    scrollbar.setValue(scrollbar.maximum())
                else:
                    scrollbar.setValue(scroll_position)

        except Exception as e:
            print(f"Error loading chat messages: {e}")

    def process_tags(self, message):
        """Process @username tags in messages"""
        # Find all @username patterns
        tag_pattern = r'@(\w+)'

        def replace_tag(match):
            username = match.group(1)
            return f"<span style='background: linear-gradient(90deg, #FF6B35, #F7931E); color: green; padding: 4px 8px; border-radius: 8px; font-weight: 700; font-size: 13px; box-shadow: 0 2px 6px rgba(255,107,53,0.4); border: 1px solid rgba(255,107,53,0.6);'>@{username}</span>"

        return re.sub(tag_pattern, replace_tag, message)

    def load_available_users(self):
        """Load list of available users for private messaging"""
        try:
            conn = sqlite3.connect("testing_system.db", timeout=10.0)
            cursor = conn.cursor()

            current_username = self.user_info.get('username', '')
            users = get_all_users_for_chat(cursor, current_username)
            conn.close()

            # Store all users for search functionality
            self.all_users = users

            # Update users list
            if hasattr(self, 'users_list'):
                self.populate_users_list(users)

        except Exception as e:
            print(f"Error loading users: {e}")

    def populate_users_list(self, users):
        """Populate the users list with given users"""
        if hasattr(self, 'users_list'):
            self.users_list.clear()
            for username, role_name, fullname in users:
                role_color = self.role_colors.get(role_name.lower(), '#666666')

                # Use 👤 icon for all users
                icon = '👤'

                # Format: username (fullname) - role
                display_name = fullname if fullname else username
                item_text = f"{icon} {username} ({display_name}) - {role_name.title()}"
                item = QListWidgetItem(item_text)
                item.setData(Qt.UserRole, {'username': username, 'role_name': role_name, 'fullname': fullname})

                # Set font size
                font = item.font()
                font.setPointSize(10)
                item.setFont(font)

                self.users_list.addItem(item)

    def search_users(self):
        """Filter users based on search text"""
        if hasattr(self, 'user_search') and hasattr(self, 'all_users'):
            search_text = self.user_search.text().strip().lower()

            if not search_text:
                # Show all users if search is empty
                self.populate_users_list(self.all_users)
            else:
                # Filter users based on search text
                filtered_users = []
                for username, role_name, fullname in self.all_users:
                    # Search in username and fullname
                    if (search_text in username.lower() or
                            (fullname and search_text in fullname.lower())):
                        filtered_users.append((username, role_name, fullname))

                self.populate_users_list(filtered_users)

    def send_chat_message(self):
        """Send a new chat message"""
        if not hasattr(self, 'chat_input'):
            return

        message_text = self.chat_input.text().strip()
        if not message_text:
            return

        try:
            conn = sqlite3.connect("testing_system.db", timeout=10.0)
            cursor = conn.cursor()

            user_id = self.user_info.get('user_id')
            role_id = self.user_info.get('role_id')
            username = self.user_info.get('username', 'Unknown')
            role_name = self.user_info.get('role_name', 'user').lower()

            # Insert message
            insert_message(cursor, user_id, username, role_name, role_id, message_text)

            print(f"💬 Message sent by {username} ({role_name}): {message_text[:30]}...")

            conn.commit()
            conn.close()

            self.chat_input.clear()
            self.load_chat_messages()

        except Exception as e:
            print(f"Error sending message: {e}")
            QMessageBox.critical(self, "Error", f"Failed to send message: {str(e)}")

    def select_user_for_chat(self, item):
        """Select a user and show Start Chat button"""
        user_data = item.data(Qt.UserRole)
        self.selected_chat_user = user_data['username']
        self.selected_chat_role = user_data['role_name']
        self.selected_chat_fullname = user_data['fullname']

        # Update the label to show selected user
        display_name = self.selected_chat_fullname if self.selected_chat_fullname else self.selected_chat_user
        self.selected_user_label.setText(f"Selected: {display_name} ({self.selected_chat_role.title()})")

        # Show the Start Chat button
        self.start_chat_btn.setVisible(True)

        print(f"🎯 Selected user for chat: {self.selected_chat_user}")

    def start_private_chat_mode(self):
        """Switch to private chat mode in the main chat area"""
        if not hasattr(self, 'selected_chat_user'):
            return

        self.is_private_chat_mode = True
        self.active_pm_user = self.selected_chat_user
        self.active_pm_role = self.selected_chat_role

        # Update chat group title
        display_name = self.selected_chat_fullname if self.selected_chat_fullname else self.selected_chat_user
        self.chat_group.setTitle(f"💬 Private Chat with {display_name} ({self.selected_chat_role.title()})")

        # Show back button
        self.back_to_global_btn.setVisible(True)

        # Update input area for private messaging
        self.update_input_area_for_private_chat()

        # Load private messages
        self.load_private_messages_in_main_area()

        print(f"📱 Started private chat mode with {self.active_pm_user}")

    def back_to_global_chat(self):
        """Return to global chat room view"""
        self.is_private_chat_mode = False
        self.active_pm_user = None
        self.active_pm_role = None

        # Update chat group title
        self.chat_group.setTitle("Global Chat Room")

        # Hide back button
        self.back_to_global_btn.setVisible(False)

        # Update input area for global chat
        self.update_input_area_for_global_chat()

        # Load global messages
        self.load_chat_messages()

        print(f"🌐 Returned to global chat room")

    def update_input_area_for_private_chat(self):
        """Update input area for private messaging mode"""
        role_name = self.user_info.get('role_name', 'user').lower()

        # Change group title
        display_name = self.selected_chat_fullname if self.selected_chat_fullname else self.selected_chat_user
        self.input_group.setTitle(f"Send Private Message to {display_name}")

        # Hide tag info for private chat
        self.tag_info.setVisible(False)

        # Update placeholder
        self.chat_input.setPlaceholderText(f"Type your private message to {display_name}...")

        # Show input area for all users in private chat mode
        self.input_group.setVisible(True)
        if hasattr(self, 'readonly_label'):
            self.readonly_label.setVisible(False)

    def update_input_area_for_global_chat(self):
        """Update input area for global chat mode"""
        role_name = self.user_info.get('role_name', 'user').lower()

        # Change group title back
        self.input_group.setTitle("Send Public Message")

        # Show tag info for global chat
        self.tag_info.setVisible(True)

        # Update placeholder
        self.chat_input.setPlaceholderText("Type your message here... (use @username to tag)")

        # Show/hide input area based on role for global chat
        if role_name in ['admin', 'superadmin']:
            self.input_group.setVisible(True)
            if hasattr(self, 'readonly_label'):
                self.readonly_label.setVisible(False)
        else:
            self.input_group.setVisible(False)
            if hasattr(self, 'readonly_label'):
                self.readonly_label.setVisible(True)

    def send_message_handler(self):
        """Handle message sending based on current mode"""
        if self.is_private_chat_mode:
            self.send_private_message_in_main_area()
        else:
            self.send_chat_message()

    def load_private_messages_in_main_area(self):
        """Load private messages in the main chat area"""
        if not self.active_pm_user:
            return

        try:
            conn = sqlite3.connect("testing_system.db", timeout=10.0)
            cursor = conn.cursor()

            sender_username = self.user_info.get('username')
            messages = get_private_messages(cursor, sender_username, self.active_pm_user)
            conn.close()

            # Clear and rebuild chat display
            self.chat_display.clear()

            if not messages:
                display_name = self.selected_chat_fullname if self.selected_chat_fullname else self.active_pm_user
                self.chat_display.append(
                    f"<div style='text-align: center; color: #666; font-style: italic; padding: 20px;'>No messages with {display_name} yet. Start the conversation!</div>")
                return

            for message in messages:
                sender, receiver, text, timestamp, sender_role = message

                # Format timestamp to Malaysia time
                time_str = format_malaysia_time(timestamp)

                # Get role color
                role_color = self.role_colors.get(sender_role.lower(), '#666666')

                # Get role icon
                role_icons = {
                    'superadmin': '👑',
                    'admin': '📢',
                    'owner': '🏢',
                    'tester': '🧪'
                }
                role_icon = role_icons.get(sender_role.lower(), '👤')

                # Chat bubble styling
                bubble_style = (
                    f"background: linear-gradient(135deg, {role_color}15 0%, {role_color}08 100%); "
                    f"color: black; text-align: left; max-width: 80%; "
                    f"border: 2px solid {role_color}30; border-radius: 15px; "
                    f"box-shadow: 0 2px 8px rgba(0,0,0,0.1);"
                )
                username_style = f"color: {role_color}; font-weight: 700;"
                time_style = f"color: #666; font-weight: 500;"

                # Format message in sequence: username, timestamp, message
                formatted_message = (
                    f"<div style='margin: 10px 0; padding: 12px 16px; {bubble_style}'>"
                    f"<div style='{username_style} font-size: 13px; margin-bottom: 2px;'>"
                    f"{role_icon} {sender}"
                    f"</div>"
                    f"<div style='{time_style} font-size: 11px; margin-bottom: 8px;'>{time_str}</div>"
                    f"<div style='font-size: 14px; line-height: 1.5; color: black;'>{text}</div>"
                    f"</div>"
                )

                self.chat_display.append(formatted_message)

            # Scroll to bottom
            scrollbar = self.chat_display.verticalScrollBar()
            scrollbar.setValue(scrollbar.maximum())

        except Exception as e:
            print(f"Error loading private messages in main area: {e}")

    def send_private_message_in_main_area(self):
        """Send a private message from the main chat area"""
        if not self.active_pm_user:
            return

        message_text = self.chat_input.text().strip()
        if not message_text:
            return

        try:
            conn = sqlite3.connect("testing_system.db", timeout=10.0)
            cursor = conn.cursor()

            # Create private_messages table if it doesn't exist
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS private_messages (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sender_user_id INTEGER,
                    sender_username TEXT,
                    receiver_username TEXT,
                    sender_role TEXT,
                    message_text TEXT,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Insert private message
            insert_private_message(cursor,
                                   self.user_info.get('user_id'),
                                   self.user_info.get('username'),
                                   self.active_pm_user,
                                   self.user_info.get('role_name'),
                                   message_text)

            conn.commit()
            conn.close()

            print(f"💬 Private message sent to {self.active_pm_user}: {message_text[:30]}...")

            self.chat_input.clear()
            self.load_private_messages_in_main_area()

        except Exception as e:
            print(f"Error sending private message in main area: {e}")
            QMessageBox.critical(self, "Error", f"Failed to send private message: {str(e)}")

    def activate_private_chat(self, target_username, target_role):
        """Activate private messaging in the inline area"""
        self.active_pm_user = target_username
        self.active_pm_role = target_role

        # Update UI to show active conversation
        if hasattr(self, 'active_pm_label'):
            self.active_pm_label.setText(f"Chat with {target_username} ({target_role.title()})")

        # Enable input controls
        if hasattr(self, 'pm_input'):
            self.pm_input.setEnabled(True)
            self.pm_input.setPlaceholderText(f"Type message to {target_username}...")
            self.pm_send_btn.setEnabled(True)

        # Load private messages for this conversation
        self.load_private_messages_inline()

        print(f"💬 Activated private chat with {target_username} ({target_role})")

    def load_private_messages_inline(self):
        """Load private messages for active conversation in inline area"""
        if not self.active_pm_user or not hasattr(self, 'pm_display'):
            return

        try:
            conn = sqlite3.connect("testing_system.db", timeout=10.0)
            cursor = conn.cursor()

            sender_username = self.user_info.get('username')
            messages = get_private_messages(cursor, sender_username, self.active_pm_user)
            conn.close()

            # Clear and rebuild private message display
            self.pm_display.clear()

            if not messages:
                self.pm_display.setPlaceholderText(
                    f"No messages with {self.active_pm_user} yet. Start the conversation!")
                return

            for message in messages:
                sender, receiver, text, timestamp, sender_role = message

                # Format timestamp to Malaysia time
                time_str = format_malaysia_time(timestamp)

                # Get role color
                role_color = self.role_colors.get(sender_role.lower(), '#666666')

                # Get role icon
                role_icons = {
                    'superadmin': '👑',
                    'admin': '📢',
                    'owner': '🏢',
                    'tester': '🧪'
                }
                role_icon = role_icons.get(sender_role.lower(), '👤')

                # Chat bubble styling
                bubble_style = (
                    f"background: linear-gradient(135deg, {role_color}15 0%, {role_color}08 100%); "
                    f"color: black; text-align: left; max-width: 80%; "
                    f"border: 2px solid {role_color}30; border-radius: 15px; "
                    f"box-shadow: 0 2px 8px rgba(0,0,0,0.1);"
                )
                username_style = f"color: {role_color}; font-weight: 700;"
                time_style = f"color: #666; font-weight: 500;"

                # Format message in sequence: username, timestamp, message
                formatted_message = (
                    f"<div style='margin: 10px 0; padding: 12px 16px; {bubble_style}'>"
                    f"<div style='{username_style} font-size: 13px; margin-bottom: 2px;'>"
                    f"{role_icon} {sender}"
                    f"</div>"
                    f"<div style='{time_style} font-size: 11px; margin-bottom: 8px;'>{time_str}</div>"
                    f"<div style='font-size: 14px; line-height: 1.5; color: black;'>{text}</div>"
                    f"</div>"
                )

                self.pm_display.append(formatted_message)

            # Scroll to bottom
            scrollbar = self.pm_display.verticalScrollBar()
            scrollbar.setValue(scrollbar.maximum())

        except Exception as e:
            print(f"Error loading inline private messages: {e}")

    def send_private_message(self):
        """Send a private message from the inline area"""
        if not self.active_pm_user or not hasattr(self, 'pm_input'):
            return

        message_text = self.pm_input.toPlainText().strip()
        if not message_text:
            return

        try:
            conn = sqlite3.connect("testing_system.db", timeout=10.0)
            cursor = conn.cursor()

            # Create private_messages table if it doesn't exist
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS private_messages (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sender_user_id INTEGER,
                    sender_username TEXT,
                    receiver_username TEXT,
                    sender_role TEXT,
                    message_text TEXT,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Insert private message
            insert_private_message(cursor,
                                   self.user_info.get('user_id'),
                                   self.user_info.get('username'),
                                   self.active_pm_user,
                                   self.user_info.get('role_name'),
                                   message_text)

            conn.commit()
            conn.close()

            print(f"💬 Inline private message sent to {self.active_pm_user}: {message_text[:30]}...")

            self.pm_input.clear()
            self.load_private_messages_inline()

        except Exception as e:
            print(f"Error sending inline private message: {e}")
            QMessageBox.critical(self, "Error", f"Failed to send private message: {str(e)}")

    def clear_all_chat(self):
        """Clear all chat messages (superadmin only)"""
        role_name = self.user_info.get('role_name', 'user')
        if role_name != 'superadmin':
            QMessageBox.warning(self, "Access Denied", "Only superadmin can clear all chat")
            return

        reply = QMessageBox.question(self, "Confirm Clear All",
                                     "Are you sure you want to clear all chat messages?",
                                     QMessageBox.Yes | QMessageBox.No)

        if reply == QMessageBox.Yes:
            try:
                conn = sqlite3.connect("testing_system.db", timeout=10.0)
                cursor = conn.cursor()

                cursor.execute("DELETE FROM message")

                conn.commit()
                conn.close()

                print(f"🗑️ All chat messages cleared by {self.user_info.get('username')} (superadmin)")

                self.load_chat_messages()
                QMessageBox.information(self, "Success", "All chat messages have been cleared")

            except Exception as e:
                print(f"Error clearing chat: {e}")
                QMessageBox.critical(self, "Error", f"Failed to clear chat: {str(e)}")

    def refresh_chat(self):
        """Refresh chat messages and users"""
        if hasattr(self, 'chat_display'):
            # Load appropriate messages based on current mode
            if self.is_private_chat_mode:
                self.load_private_messages_in_main_area()
            else:
                self.load_chat_messages()

            self.load_available_users()

    def load_user_data(self):
        """Load user management data from database"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Load users with their roles and branch information
            cursor.execute("""
                SELECT u.user_id, u.username, u.fullname, r.role_name, u.email, b.branch_name
                FROM users u
                LEFT JOIN roles r ON u.role = r.role_id
                LEFT JOIN branches b ON u.branch_id = b.branch_id
                ORDER BY u.user_id
            """)

            users = cursor.fetchall()
            conn.close()

            self.user_table.setRowCount(len(users))
            for row, user in enumerate(users):
                for col, value in enumerate(user):
                    item = QTableWidgetItem(str(value) if value else "")
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make cells read-only
                    self.user_table.setItem(row, col, item)

            print(f"✓ Loaded {len(users)} users for management")

        except Exception as e:
            print(f"✗ Error loading user data: {e}")
            QMessageBox.warning(self, "Database Error", f"Could not load user data: {e}")

            # Show empty table on error
            self.user_table.setRowCount(0)

    def load_my_profile_data(self):
        """Load current user's profile data"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Get current user's detailed information
            cursor.execute("""
                SELECT u.user_id, u.username, u.fullname, u.email, u.phone_no, 
                       r.role_name, r.role_id, b.branch_name
                FROM users u
                LEFT JOIN roles r ON u.role = r.role_id
                LEFT JOIN branches b ON u.branch_id = b.branch_id
                WHERE u.username = ?
            """, (self.username,))

            user_data = cursor.fetchone()

            if user_data:
                user_id, username, fullname, email, phone, role_name, role_id, branch_name = user_data

                # Update profile information
                self.profile_name_label.setText(fullname or "Not Set")
                self.profile_role_label.setText(f"Role: {role_name or 'Unknown'}")
                self.profile_username_label.setText(f"Username: {username}")
                self.profile_email.setText(email or "Not Set")
                self.profile_phone.setText(phone or "Not Set")
                self.profile_branch.setText(branch_name or "Not Assigned")
                self.profile_user_id.setText(str(user_id))
                self.profile_role_id.setText(str(role_id))

                # Load recent activity
                self.load_recent_activity(cursor, user_id, username)

            conn.close()
            print(f"✓ Loaded profile data for {self.username}")

        except Exception as e:
            print(f"✗ Error loading profile data: {e}")

    def load_recent_activity(self, cursor, user_id, username):
        """Load recent activity for the user"""
        try:
            self.recent_activity_list.clear()

            # Get recent messages
            cursor.execute("""
                SELECT message, timestamp FROM message 
                WHERE user_id = ? 
                ORDER BY timestamp DESC 
                LIMIT 5
            """, (user_id,))

            messages = cursor.fetchall()

            for message, timestamp in messages:
                if message and not message.startswith('[ACTIVITY]'):
                    # Format timestamp
                    try:
                        from datetime import datetime
                        dt = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                        formatted_time = dt.strftime('%m/%d %H:%M')
                    except:
                        formatted_time = timestamp

                    # Truncate long messages
                    display_message = message[:50] + "..." if len(message) > 50 else message
                    activity_text = f"💬 {formatted_time}: {display_message}"
                    self.recent_activity_list.addItem(activity_text)

            # Get recent product additions
            cursor.execute("""
                SELECT product_name, arrival_date FROM products 
                WHERE owner_id = ? 
                ORDER BY arrival_date DESC 
                LIMIT 3
            """, (user_id,))

            products = cursor.fetchall()

            for product_name, arrival_date in products:
                try:
                    from datetime import datetime
                    dt = datetime.strptime(arrival_date, '%Y-%m-%d %H:%M:%S')
                    formatted_time = dt.strftime('%m/%d %H:%M')
                except:
                    formatted_time = arrival_date

                activity_text = f"📦 {formatted_time}: Added {product_name}"
                self.recent_activity_list.addItem(activity_text)

            if self.recent_activity_list.count() == 0:
                self.recent_activity_list.addItem("No recent activity found")

        except Exception as e:
            print(f"Error loading recent activity: {e}")

    def open_edit_profile_dialog(self):
        """Open dialog to edit profile information"""
        try:
            from newdashboard_extras import EditProfileDialog

            dialog = EditProfileDialog(self.username)
            if dialog.exec_() == QDialog.Accepted:
                # Refresh profile data after successful edit
                self.load_my_profile_data()
                QMessageBox.information(self, "Success", "Profile updated successfully!")

        except Exception as e:
            print(f"Error opening edit profile dialog: {e}")
            # Create a simple edit dialog if the extras module doesn't exist
            self.create_simple_edit_profile_dialog()

    def create_simple_edit_profile_dialog(self):
        """Create a simple edit profile dialog"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Profile")
        dialog.setFixedSize(800, 500)

        layout = QVBoxLayout()

        # Current user info
        conn = sqlite3.connect("testing_system.db")
        cursor = conn.cursor()
        cursor.execute("""
            SELECT fullname, email, phone_no FROM users WHERE username = ?
        """, (self.username,))
        user_data = cursor.fetchone()
        conn.close()

        if user_data:
            fullname, email, phone = user_data
        else:
            fullname, email, phone = "", "", ""

        # Form fields - only email and phone are editable
        layout.addWidget(QLabel("Full Name (Read-only):"))
        name_display = QLabel(fullname or "Not Set")
        name_display.setStyleSheet("""
            QLabel {
                background-color: #f5f5f5;
                padding: 8px 12px;
                border-radius: 6px;
                border: 1px solid #ddd;
                color: #666;
            }
        """)
        layout.addWidget(name_display)

        layout.addWidget(QLabel("Email:"))
        email_edit = QLineEdit(email or "")
        layout.addWidget(email_edit)

        layout.addWidget(QLabel("Phone:"))
        phone_edit = QLineEdit(phone or "")
        layout.addWidget(phone_edit)

        # Password change section
        layout.addWidget(QLabel(""))  # Spacer
        password_section = QGroupBox("Change Password")
        password_layout = QVBoxLayout()

        password_layout.addWidget(QLabel("Current Password (numbers only):"))
        current_password_edit = QLineEdit()
        current_password_edit.setEchoMode(QLineEdit.Password)
        current_password_edit.setPlaceholderText("Enter current password")
        password_layout.addWidget(current_password_edit)

        password_layout.addWidget(QLabel("New Password (min 4 digits):"))
        new_password_edit = QLineEdit()
        new_password_edit.setEchoMode(QLineEdit.Password)
        new_password_edit.setPlaceholderText("Enter new password (numbers only)")
        password_layout.addWidget(new_password_edit)

        password_layout.addWidget(QLabel("Confirm New Password:"))
        confirm_password_edit = QLineEdit()
        confirm_password_edit.setEchoMode(QLineEdit.Password)
        confirm_password_edit.setPlaceholderText("Confirm new password")
        password_layout.addWidget(confirm_password_edit)

        password_section.setLayout(password_layout)
        layout.addWidget(password_section)

        # Buttons
        button_layout = QHBoxLayout()
        save_btn = QPushButton("Save Profile")
        change_password_btn = QPushButton("Change Password")
        cancel_btn = QPushButton("Cancel")

        def save_changes():
            try:
                conn = sqlite3.connect("testing_system.db")
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE users SET email = ?, phone_no = ?
                    WHERE username = ?
                """, (email_edit.text().strip(), phone_edit.text().strip(), self.username))
                conn.commit()
                conn.close()
                QMessageBox.information(self, "Success", "Profile updated successfully!")
                self.load_my_profile_data()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to update profile: {str(e)}")

        def change_password():
            current_password = current_password_edit.text()
            new_password = new_password_edit.text()
            confirm_password = confirm_password_edit.text()

            # Validation
            if not current_password:
                QMessageBox.warning(self, "Warning", "Please enter your current password.")
                return

            if not new_password:
                QMessageBox.warning(self, "Warning", "Please enter a new password.")
                return

            # Validate that passwords contain only digits (since DB expects INTEGER)
            if not current_password.isdigit():
                QMessageBox.warning(self, "Warning", "Current password must contain only numbers.")
                return

            if not new_password.isdigit():
                QMessageBox.warning(self, "Warning", "New password must contain only numbers.")
                return

            if len(new_password) < 4:
                QMessageBox.warning(self, "Warning", "New password must be at least 4 digits long.")
                return

            if new_password != confirm_password:
                QMessageBox.warning(self, "Warning", "New passwords do not match.")
                return

            try:
                # Convert passwords to integers for database operations
                current_password_int = int(current_password)
                new_password_int = int(new_password)

                # Verify current password
                conn = sqlite3.connect("testing_system.db")
                cursor = conn.cursor()
                cursor.execute("SELECT password FROM users WHERE username = ?", (self.username,))
                stored_password = cursor.fetchone()

                if not stored_password or stored_password[0] != current_password_int:
                    QMessageBox.critical(self, "Error", "Current password is incorrect.")
                    conn.close()
                    return

                # Update password
                cursor.execute("""
                    UPDATE users SET password = ? WHERE username = ?
                """, (new_password_int, self.username))
                conn.commit()
                conn.close()

                QMessageBox.information(self, "Success", "Password changed successfully!")

                # Clear password fields
                current_password_edit.clear()
                new_password_edit.clear()
                confirm_password_edit.clear()

            except ValueError:
                QMessageBox.critical(self, "Error", "Password must be a valid number.")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to change password: {str(e)}")

        save_btn.clicked.connect(save_changes)
        change_password_btn.clicked.connect(change_password)
        cancel_btn.clicked.connect(dialog.reject)

        button_layout.addWidget(save_btn)
        button_layout.addWidget(change_password_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)

        dialog.setLayout(layout)
        dialog.exec_()

    def load_data(self):
        """Load all data for the dashboard"""
        self.load_summary_data()
        self.load_inventory_data()
        # Load other data as needed for the current view

    def load_summary_data(self):
        """Load summary data from products table"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Total products
            cursor.execute("SELECT COUNT(*) FROM products")
            result = cursor.fetchone()
            self.total_samples = result[0] if result else 0

            # Near expiry (within 60 days)
            cursor.execute("""
                SELECT COUNT(*) FROM products 
                            WHERE expired_date BETWEEN DATE('now') AND DATE('now', '+60 days')
            AND expired_date IS NOT NULL
            """)
            result = cursor.fetchone()
            self.near_maturation = result[0] if result else 0

            # Expired products
            cursor.execute("""
                SELECT COUNT(*) FROM products 
                WHERE expired_date < DATE('now') AND expired_date IS NOT NULL
            """)
            result = cursor.fetchone()
            self.expired_samples = result[0] if result else 0

            # Products with rack locations
            cursor.execute("SELECT COUNT(*) FROM products WHERE rack_location_id IS NOT NULL")
            result = cursor.fetchone()
            self.assigned_tasks = result[0] if result else 0

            conn.close()
            print(
                f"✓ Loaded summary data: {self.total_samples} products, {self.near_maturation} near expiry, {self.expired_samples} expired")

        except Exception as e:
            print(f"✗ Error loading summary data: {e}")
            self.total_samples = 0
            self.near_maturation = 0
            self.expired_samples = 0
            self.assigned_tasks = 0

    def load_maturation_data(self):
        """Load upcoming expiry data from products table and display as cards"""
        try:
            print("Starting load_maturation_data...")
            # Clear existing cards and reset product cards list
            for i in reversed(range(min(self.cards_layout.count(), 50))):  # Limit to 50 items max
                item = self.cards_layout.itemAt(i)
                if item:
                    widget = item.widget()
                    if widget:
                        widget.deleteLater()

            # Reset product cards list for admin users
            if hasattr(self, 'product_cards'):
                self.product_cards = []

            # Get current threshold
            threshold = self.load_maturation_threshold()

            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Run the maturation query with LIMIT to prevent buffer overflow
            cursor.execute(f"""
                SELECT 
                    p.product_id,
                    p.product_name,
                    p.batch,
                    u.username AS owner,
                    COALESCE(p.sku, '') AS sku,
                    COALESCE(b.branch_name, 'Main Branch') AS branch,
                    COALESCE(rl.rack_location_name, 'Unassigned') AS rack_location,
                    p.expired_date,
                    p.manufacture_date,
                    COALESCE(p.barcode, '') AS barcode,
                    COALESCE(p.barcode_image, '') AS barcode_image,
                    COALESCE(p.product_image, '') AS product_image,
                    CAST(julianday(p.expired_date) - julianday('now') AS INTEGER) AS days_left
                FROM products p
                LEFT JOIN users u ON p.owner_id = u.user_id
                LEFT JOIN branches b ON p.branch_id = b.branch_id
                LEFT JOIN racklocations rl ON p.rack_location_id = rl.rack_location_id
                WHERE p.expired_date BETWEEN DATE('now') AND DATE('now', '+{threshold} days')
                AND p.expired_date IS NOT NULL
                ORDER BY p.expired_date ASC
                LIMIT 25
            """)

            data = cursor.fetchall()

            # Send email notifications to product owners if there are products (only if not sent today)
            if data:
                notifications_sent = self.check_and_send_daily_notifications(data)
                if notifications_sent:
                    print(f"📧 Sent {len(notifications_sent)} new email notifications to product owners")
                else:
                    print("📧 Email notifications already sent today or no new notifications needed")

            # Limit processing to prevent stack overflow
            for i, product in enumerate(data[:20]):  # Max 20 cards to prevent memory issues
                row = i
                col = 0

                # Safe data extraction with string length limits
                def safe_extract(value, max_len=100):
                    if value is None:
                        return ""
                    str_val = str(value)
                    return str_val[:max_len] if len(str_val) > max_len else str_val

                product_data = {
                    'product_id': product[0],
                    'product_name': safe_extract(product[1], 50),
                    'batch': safe_extract(product[2], 30),
                    'owner': safe_extract(product[3], 30),
                    'sku': safe_extract(product[4], 20),
                    'branch': safe_extract(product[5], 50),
                    'rack_location': safe_extract(product[6], 30),
                    'expired_date': safe_extract(product[7], 20),
                    'manufacture_date': safe_extract(product[8], 20),
                    'barcode': safe_extract(product[9], 30),
                    'barcode_image': safe_extract(product[10], 200),
                    'product_image': safe_extract(product[11], 200),
                    'days_left': product[12] if product[12] is not None else 0
                }

                # Create and add card (with checkbox for admin/superadmin)
                if self.user_role.lower() in ['admin', 'superadmin']:
                    card = ProductCardWithCheckbox(product_data)
                    # Store card reference for later access to checkbox state
                    if not hasattr(self, 'product_cards'):
                        self.product_cards = []
                    self.product_cards.append(card)
                else:
                    card = ProductCard(product_data)
                self.cards_layout.addWidget(card, row, col)

            # Add stretch to fill remaining space
            self.cards_layout.setRowStretch(self.cards_layout.rowCount(), 1)

            conn.close()
            print(f"✓ Loaded {len(data)} products with upcoming expiry dates (within {threshold} days)")

            # Update threshold label if it exists
            if hasattr(self, 'threshold_label'):
                self.threshold_label.setText(f"Showing products expiring within {threshold} days")

        except Exception as e:
            print(f"✗ Error loading maturation data: {e}")
            # Clear cards on error
            for i in reversed(range(self.cards_layout.count())):
                item = self.cards_layout.itemAt(i)
                if item:
                    widget = item.widget()
                    if widget:
                        widget.deleteLater()

    def load_inventory_data(self, page=1, items_per_page=20):
        """Load inventory data from products table with pagination"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            offset = (page - 1) * items_per_page

            # Base query with joins to get comprehensive product information
            base_query = """
                SELECT p.product_name, 
                       u.username as owner,
                       p.owner_id as owner_id,
                       COALESCE(b.branch_name, 'Main Branch') as branch,
                       COALESCE(t.username, 'Unassigned') as tester,
                       p.arrival_date,
                       p.batch,
                       COALESCE(rl.rack_location_name, 'Unassigned') as rack_location,
                       p.sku,
                       p.expired_date,
                       COALESCE(p.barcode, '') as barcode,
                       p.product_id
                FROM products p
                LEFT JOIN users u ON p.owner_id = u.user_id
                LEFT JOIN branches b ON p.branch_id = b.branch_id
                LEFT JOIN users t ON p.tester_id = t.user_id
                LEFT JOIN racklocations rl ON p.rack_location_id = rl.rack_location_id
            """

            params = []
            conditions = ["COALESCE(p.status, 'pending') = 'approved'"]  # Only show approved products

            # Apply search filters if available
            search_text = ""
            if hasattr(self, 'search_input'):
                search_text = self.search_input.text().strip()

            if search_text:
                conditions.append("""
                    (p.product_id LIKE ? OR p.product_name LIKE ? OR 
                     p.batch LIKE ? OR u.username LIKE ? OR p.sku LIKE ? OR p.barcode LIKE ?)
                """)
                search_param = f"%{search_text}%"
                params.extend([search_param, search_param, search_param, search_param, search_param, search_param])

            # Build final query
            if conditions:
                query = base_query + " WHERE " + " AND ".join(conditions)
            else:
                query = base_query

            # Limit to max 100 items to prevent buffer issues
            safe_limit = min(items_per_page, 100)
            query += f" ORDER BY p.arrival_date DESC, p.product_name ASC LIMIT {safe_limit} OFFSET {offset}"

            cursor.execute(query, params)
            data = cursor.fetchall()

            self.inventory_table.setRowCount(len(data))

            for row, product in enumerate(data):
                # Store product_id in row for later use (it's the last column)
                product_id = product[-1]

                # Display only the first 11 columns (excluding product_id)
                for col, value in enumerate(product[:-1]):
                    # Format dates for display
                    if col in [5, 9] and value:  # Arrival date and expired date columns
                        try:
                            if 'T' in str(value):
                                # Handle datetime format
                                date_part = str(value).split('T')[0]
                                formatted_date = datetime.strptime(date_part, '%Y-%m-%d').strftime('%Y-%m-%d')
                            else:
                                formatted_date = datetime.strptime(str(value), '%Y-%m-%d').strftime('%Y-%m-%d')
                            value = formatted_date
                        except:
                            pass

                    # Safe string conversion with length limit to prevent buffer overflow
                    safe_value = str(value) if value else ""
                    if len(safe_value) > 200:  # Limit string length to prevent buffer issues
                        safe_value = safe_value[:200] + "..."

                    item = QTableWidgetItem(safe_value)
                    # Store product_id in the first column's item for retrieval
                    if col == 0:
                        item.setData(Qt.UserRole, product_id)
                    self.inventory_table.setItem(row, col, item)

            # Update pagination info
            count_query = "SELECT COUNT(*) FROM products p LEFT JOIN users u ON p.owner_id = u.user_id WHERE COALESCE(p.status, 'pending') = 'approved'"
            if len(conditions) > 1:  # More than just the status condition
                count_query += " AND " + " AND ".join(conditions[1:])

            cursor.execute(count_query, params)
            result = cursor.fetchone()
            total_items = result[0] if result else 0
            total_pages = max(1, (total_items + items_per_page - 1) // items_per_page)

            if hasattr(self, 'page_label'):
                self.page_label.setText(f"Page {page} of {total_pages}")

            conn.close()
            print(f"✓ Loaded {len(data)} products from inventory (page {page} of {total_pages})")

        except Exception as e:
            print(f"✗ Error loading inventory data: {e}")
            self.inventory_table.setRowCount(0)
            if hasattr(self, 'page_label'):
                self.page_label.setText("Page 1 of 1")

    # Activity log functionality is now handled in load_chat_data

    def apply_filters(self):
        """Apply search filters to inventory"""
        self.load_inventory_data()

    def clear_filters(self):
        """Clear all search filters"""
        self.product_filter.clear()
        self.batch_filter.clear()
        self.status_filter.setCurrentIndex(0)
        self.owner_filter.clear()
        self.tester_filter.clear()
        self.date_from.setDate(QDate.currentDate().addDays(-30))
        self.date_to.setDate(QDate.currentDate().addDays(60))
        self.load_inventory_data()

    def prev_page(self):
        """Go to previous page"""
        # Implementation for pagination
        pass

    def next_page(self):
        """Go to next page"""
        # Implementation for pagination
        pass

    def show_pending_approvals(self):
        """Show samples pending approval"""
        self.switch_content("pending")
        self.pending_btn.setChecked(True)

    def assign_tester(self):
        """Assign tester to selected sample"""
        self.switch_content("assign")
        self.assign_btn.setChecked(True)

    def export_data(self):
        """Export data functionality"""
        QMessageBox.information(self, "Export Data", "Data export functionality will be implemented here.")

    def open_email_settings(self):
        """Open enhanced email settings dialog with template editing capabilities"""
        from newdashboard_extras import EnhancedEmailSettingsDialog
        dialog = EnhancedEmailSettingsDialog()
        dialog.exec_()

    def delete_sample(self):
        """Delete selected sample (Superadmin only)"""
        if not self.is_superadmin_mode:
            QMessageBox.warning(self, "Access Denied", "This feature requires Superadmin access.")
            return
        QMessageBox.information(self, "Delete Sample", "Sample deletion functionality will be implemented here.")

    def bulk_edit(self):
        """Bulk edit functionality (Superadmin only)"""
        if not self.is_superadmin_mode:
            QMessageBox.warning(self, "Access Denied", "This feature requires Superadmin access.")
            return
        QMessageBox.information(self, "Bulk Edit", "Bulk edit functionality will be implemented here.")

    def open_user_management(self):
        """Open user management panel"""
        self.switch_content("user_management")
        self.user_mgmt_btn.setChecked(True)

    def open_create_account_dialog(self):
        """Open the create account dialog"""
        try:
            # Dynamic import to avoid circular imports
            global CreateAccountDialog
            if CreateAccountDialog is None:
                from user1 import CreateAccountDialog

            self.create_account_dialog = CreateAccountDialog(self)
            self.create_account_dialog.show()
        except ImportError as e:
            QMessageBox.critical(self, "Import Error",
                                 "Create Account functionality not available.\nPlease ensure user1.py is in the same directory.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open Create Account dialog: {str(e)}")

    def delete_selected_user(self):
        """Delete the selected user (Superadmin only)"""
        if not self.is_superadmin_mode:
            QMessageBox.warning(self, "Access Denied", "This feature requires Super Admin access.")
            return

        # Check if a user is selected
        selected_rows = self.user_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "No Selection", "Please select a user to delete.")
            return

        selected_row = selected_rows[0].row()

        # Get user information from the selected row
        user_id = self.user_table.item(selected_row, 0).text()
        username = self.user_table.item(selected_row, 1).text()
        fullname = self.user_table.item(selected_row, 2).text()
        role = self.user_table.item(selected_row, 3).text()

        # Prevent deletion of current user or other superadmins
        if role.lower() == 'superadmin':
            QMessageBox.warning(self, "Cannot Delete", "Cannot delete Super Admin accounts.")
            return

        # Confirmation dialog
        reply = QMessageBox.question(
            self,
            "Confirm Delete User",
            f"Are you sure you want to delete this user?\n\n"
            f"Username: {username}\n"
            f"Full Name: {fullname}\n"
            f"Role: {role}\n\n"
            f"This action cannot be undone!",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                conn = sqlite3.connect("testing_system.db", timeout=30.0)
                cursor = conn.cursor()

                # Delete the user
                cursor.execute("DELETE FROM users WHERE user_id = ?", (user_id,))

                if cursor.rowcount > 0:
                    conn.commit()
                    print(f"✓ User '{username}' (ID: {user_id}) deleted successfully")

                    # Show success message
                    QMessageBox.information(
                        self,
                        "User Deleted",
                        f"User '{username}' has been successfully deleted."
                    )

                    # Refresh the user table
                    self.load_user_data()

                    # Log the action
                    self.add_activity_log("Superadmin", f"Deleted user: {username}",
                                          f"User ID: {user_id}, Role: {role}")

                else:
                    QMessageBox.warning(self, "Delete Failed", "User not found or could not be deleted.")

                conn.close()

            except sqlite3.Error as e:
                print(f"✗ Database error during user deletion: {str(e)}")
                QMessageBox.critical(self, "Database Error", f"Error deleting user: {str(e)}")
            except Exception as e:
                print(f"✗ Unexpected error during user deletion: {str(e)}")
                QMessageBox.critical(self, "Error", f"Unexpected error: {str(e)}")

    def open_system_settings(self):
        """Open system settings dialog (Superadmin only)"""
        if not self.is_superadmin_mode:
            QMessageBox.warning(self, "Access Denied", "This feature requires Superadmin access.")
            return
        dialog = SystemSettingsDialog()
        dialog.exec_()

    def open_export_reports(self):
        """Open export reports dialog (Superadmin only)"""
        if not self.is_superadmin_mode:
            QMessageBox.warning(self, "Access Denied", "This feature requires Superadmin access.")
            return
        dialog = ExportReportsDialog()
        dialog.exec_()

    def toggle_superadmin_mode(self):
        """Toggle superadmin mode with passcode"""
        if not self.is_superadmin_mode:
            dialog = SuperAdminDialog()
            if dialog.exec_() == QDialog.Accepted:
                self.is_superadmin_mode = True
                self.superadmin_button.setText("Admin Mode")
                self.superadmin_button.setStyleSheet("""
                    QPushButton {
                        background-color: #388E3C;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 10px 20px;
                        font-weight: bold;
                        max-width: 200px;
                    }
                    QPushButton:hover {
                        background-color: #2E7D32;
                    }
                """)

                # Add admin-only menu items to the left panel
                self.add_admin_menu_items()

                # Show admin features in the current view
                self.show_admin_features()

                # Log the super admin activation
                self.add_activity_log(self.username, "Super Admin Mode Activated", "User entered super admin mode")

                print("🎯 Super Admin Mode ENABLED:")
                print("  ✓ Admin menu items shown")
                print("  ✓ Admin features enabled")
                print("  ✓ Delete user button enabled")
                print("  ✓ All super admin functions available")

                QMessageBox.information(self, "👑 Super Admin Mode Enabled",
                                        "Super Admin features are now enabled!\n\n"
                                        "✅ Additional menu items visible\n"
                                        "✅ Enhanced admin controls active\n"
                                        "✅ Delete user functionality available\n"
                                        "✅ System maintenance tools accessible")
        else:
            # Show confirmation dialog when switching back to admin mode
            reply = QMessageBox.question(
                self,
                "👑 Exit Super Admin Mode",
                "Are you sure you want to exit Super Admin mode?\n\n"
                "🔒 This will:\n"
                "• Hide all administrative functions\n"
                "• Return to regular admin interface\n"
                "• Require passcode to re-enter Super Admin mode\n\n"
                "Do you want to continue?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                self.is_superadmin_mode = False
                self.superadmin_button.setText("Super Admin Mode")
                self.superadmin_button.setStyleSheet("""
                QPushButton {
                        background-color: #00695C;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 10px 20px;
                        font-weight: bold;
                        max-width: 200px;
                }
                QPushButton:hover {
                        background-color: #004D40;
                }
            """)

                # Remove admin-only menu items
                self.remove_admin_menu_items()

                # Hide admin features in the current view
                self.hide_admin_features()

                # Show confirmation that mode has been switched
                QMessageBox.information(
                    self,
                    "✅ Mode Changed",
                    "Successfully switched back to Admin mode.\n\n"
                    "Super Admin features are now hidden.\n"
                    "Use the 'Super Admin Mode' button to re-enable them."
                )

    def refresh_dashboard(self):
        """Refresh all dashboard data"""
        self.load_data()

    def safe_refresh_dashboard(self):
        """Safely refresh dashboard data with error handling"""
        try:
            # Force garbage collection before refresh to free memory
            gc.collect()

            self.load_summary_data()
            self.load_inventory_data()
            # Skip maturation data refresh to prevent buffer issues

            # Force garbage collection after refresh
            gc.collect()
            print("✓ Dashboard refreshed safely")
        except Exception as e:
            print(f"Error in safe refresh: {e}")
            # Emergency cleanup on error
            gc.collect()

    def safe_refresh_chat(self):
        """Safely refresh chat with error handling"""
        try:
            if hasattr(self, 'chat_display') and hasattr(self, 'is_private_chat_mode'):
                if not self.is_private_chat_mode:
                    self.load_chat_messages()
                print("✓ Chat refreshed safely")
        except Exception as e:
            print(f"Error in safe chat refresh: {e}")

    def add_activity_log(self, user, action, details=""):
        """Add entry to activity log"""
        conn = sqlite3.connect("testing_system.db")
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO activity_log (user, action, details)
            VALUES (?, ?, ?)
        """, (user, action, details))
        conn.commit()
        conn.close()

    def add_admin_menu_items(self):
        """Show admin functions area in the left panel"""
        # Create content panels for admin features if they don't exist yet
        if not hasattr(self, 'system_settings_panel'):
            self.create_system_settings_panel()
            # Export reports panel already exists

        # Show the admin functions area
        self.admin_functions_area.setVisible(True)
        print("✓ Super Admin menu items enabled")

    def remove_admin_menu_items(self):
        """Hide admin functions area from the left panel"""
        # Hide the admin functions area
        self.admin_functions_area.setVisible(False)

        # Switch back to inventory view if we're on an admin view
        current_index = self.content_stack.currentIndex()
        if current_index >= 6:  # Admin panels start at index 6
            self.switch_content("inventory")
            self.inventory_btn.setChecked(True)

    def show_admin_features(self):
        """Show admin features in the current view"""
        # Add admin action buttons to the current view
        current_index = self.content_stack.currentIndex()

        if current_index == 0:  # Inventory view
            self.show_inventory_admin_features()
        elif current_index == 1:  # Pending approvals view
            self.show_pending_admin_features()
        elif current_index == 2:  # Maturation view
            self.show_maturation_admin_features()
        elif current_index == 5:  # User management view
            self.show_user_management_admin_features()

        # Also refresh user management features to show delete button
        if hasattr(self, 'user_admin_widget') and hasattr(self, 'delete_user_btn'):
            self.delete_user_btn.setVisible(self.is_superadmin_mode)
            print(f"  ✓ Delete user button visibility: {self.is_superadmin_mode}")

    def hide_admin_features(self):
        """Hide admin features from the current view"""
        # Remove admin action buttons from the current view
        current_index = self.content_stack.currentIndex()

        if current_index == 0:  # Inventory view
            self.hide_inventory_admin_features()
        elif current_index == 1:  # Pending approvals view
            self.hide_pending_admin_features()
        elif current_index == 2:  # Maturation view
            self.hide_maturation_admin_features()
        elif current_index == 5:  # User management view
            self.hide_user_management_admin_features()

    def show_inventory_admin_features(self):
        """Show admin features in the inventory view"""
        # Add admin action buttons to the inventory panel
        inventory_panel = self.content_stack.widget(0)
        layout = inventory_panel.layout()

        # Create admin action buttons if they don't exist
        if not hasattr(self, 'inventory_admin_widget'):
            self.inventory_admin_widget = QWidget()
            admin_layout = QHBoxLayout()
            admin_layout.setContentsMargins(0, 0, 0, 10)

            self.delete_sample_btn = QPushButton("🗑️ Delete Selected")
            self.delete_sample_btn.clicked.connect(self.delete_sample)
            self.delete_sample_btn.setStyleSheet("""
                QPushButton {
                    background-color: #F44336;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 8px 16px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #D32F2F;
                }
            """)

            # Removed bulk edit button as requested

            admin_layout.addWidget(self.delete_sample_btn)
            # admin_layout.addWidget(self.bulk_edit_btn)  # Removed
            admin_layout.addStretch()

            self.inventory_admin_widget.setLayout(admin_layout)

        # Insert admin buttons at the top (only once)
        if layout.indexOf(self.inventory_admin_widget) == -1:
            layout.insertWidget(0, self.inventory_admin_widget)
        self.inventory_admin_widget.show()

    def hide_inventory_admin_features(self):
        """Hide admin features from the inventory view"""
        if hasattr(self, 'inventory_admin_widget'):
            self.inventory_admin_widget.hide()

    def show_pending_admin_features(self):
        """Show admin features in the pending approvals view"""
        # Add bulk approve button
        pending_panel = self.content_stack.widget(1)
        layout = pending_panel.layout()

        if not hasattr(self, 'pending_admin_widget'):
            self.pending_admin_widget = QWidget()
            admin_layout = QHBoxLayout()
            admin_layout.setContentsMargins(0, 0, 0, 10)

            self.bulk_approve_btn = QPushButton("✅ Approve All")
            self.bulk_approve_btn.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 8px 16px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #388E3C;
                }
            """)

            admin_layout.addWidget(self.bulk_approve_btn)
            admin_layout.addStretch()

            self.pending_admin_widget.setLayout(admin_layout)

        if layout.indexOf(self.pending_admin_widget) == -1:
            layout.insertWidget(0, self.pending_admin_widget)
        self.pending_admin_widget.show()

    def hide_pending_admin_features(self):
        """Hide admin features from the pending approvals view"""
        if hasattr(self, 'pending_admin_widget'):
            self.pending_admin_widget.hide()

    def show_maturation_admin_features(self):
        """Show admin features in the maturation view"""
        # Add extend maturation button
        maturation_panel = self.content_stack.widget(2)
        layout = maturation_panel.layout()

        if not hasattr(self, 'maturation_admin_widget'):
            self.maturation_admin_widget = QWidget()
            admin_layout = QHBoxLayout()
            admin_layout.setContentsMargins(0, 0, 0, 10)

            # No admin buttons for maturation panel anymore
            admin_layout.addStretch()

            self.maturation_admin_widget.setLayout(admin_layout)

        if layout.indexOf(self.maturation_admin_widget) == -1:
            layout.insertWidget(0, self.maturation_admin_widget)
        self.maturation_admin_widget.show()

    def hide_maturation_admin_features(self):
        """Hide admin features from the maturation view"""
        if hasattr(self, 'maturation_admin_widget'):
            self.maturation_admin_widget.hide()

    def show_user_management_admin_features(self):
        """Show admin features in the user management view"""
        # Add user management admin buttons
        user_panel = self.content_stack.widget(5)
        layout = user_panel.layout()

        if not hasattr(self, 'user_admin_widget'):
            self.user_admin_widget = QWidget()
            admin_layout = QHBoxLayout()
            admin_layout.setContentsMargins(0, 0, 0, 10)

            # Add User button (available for both admin and superadmin)
            self.add_user_btn = QPushButton("➕ Add User")
            self.add_user_btn.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 8px 16px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #388E3C;
                }
            """)
            self.add_user_btn.clicked.connect(self.open_create_account_dialog)

            # Delete User button (only for superadmin)
            self.delete_user_btn = QPushButton("🗑️ Delete User")
            self.delete_user_btn.setStyleSheet("""
                QPushButton {
                    background-color: #F44336;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 8px 16px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #D32F2F;
                }
            """)
            self.delete_user_btn.clicked.connect(self.delete_selected_user)
            # Only show delete button for superadmin
            self.delete_user_btn.setVisible(self.is_superadmin_mode)

            admin_layout.addWidget(self.add_user_btn)
            admin_layout.addWidget(self.delete_user_btn)
            admin_layout.addStretch()

            self.user_admin_widget.setLayout(admin_layout)
        else:
            # Update delete button visibility based on superadmin mode
            self.delete_user_btn.setVisible(self.is_superadmin_mode)

        if layout.indexOf(self.user_admin_widget) == -1:
            layout.insertWidget(0, self.user_admin_widget)
        self.user_admin_widget.show()

    def hide_user_management_admin_features(self):
        """Hide admin features from the user management view"""
        if hasattr(self, 'user_admin_widget'):
            self.user_admin_widget.hide()

    def create_export_reports_panel(self):
        """Create the export reports panel for admin"""
        export_panel = QWidget()
        export_layout = QVBoxLayout()
        export_layout.setContentsMargins(30, 30, 30, 30)
        export_layout.setSpacing(20)

        # Header section with gradient background
        header_widget = QWidget()
        header_widget.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                    stop:0 #6A1B9A, stop:1 #9C27B0);
                border-radius: 12px;
                margin-bottom: 20px;
            }
        """)
        header_layout = QVBoxLayout()
        header_layout.setContentsMargins(25, 20, 25, 20)

        title = QLabel("📤 Export Reports")
        title.setStyleSheet("""
            QLabel {
                font-size: 24px; 
                font-weight: bold; 
                color: white; 
                margin: 0px;
                border: none;
            }
        """)

        subtitle = QLabel("Generate and download comprehensive system reports")
        subtitle.setStyleSheet("""
            QLabel {
                font-size: 14px;
                color: rgba(255, 255, 255, 0.9); 
                margin: 5px 0px 0px 0px;
                border: none;
            }
        """)

        header_layout.addWidget(title)
        header_layout.addWidget(subtitle)
        header_widget.setLayout(header_layout)
        export_layout.addWidget(header_widget)

        # Content area
        content_scroll = QScrollArea()
        content_scroll.setWidgetResizable(True)
        content_scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QScrollBar:vertical {
                background-color: #F3E5F5;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background-color: #9C27B0;
                border-radius: 6px;
                min-height: 20px;
            }
        """)

        content_widget = QWidget()
        content_layout = QVBoxLayout()
        content_layout.setSpacing(20)

        # Report Types Card
        reports_card = QGroupBox("📊 Available Reports")
        reports_card.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 16px;
                color: #4A148C;
                border: 2px solid #E1BEE7;
                border-radius: 12px;
                margin-top: 15px;
                padding: 20px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 5px 10px;
                background-color: white;
                border-radius: 5px;
            }
        """)
        reports_layout = QGridLayout()
        reports_layout.setSpacing(15)

        # Report button style
        report_btn_style = """
            QPushButton {
                border: 2px solid #F3E5F5;
                border-radius: 12px;
                padding: 25px;
                background-color: white;
                text-align: left;
                font-weight: bold;
                font-size: 14px;
                min-height: 80px;
            }
            QPushButton:hover {
                background-color: #F3E5F5;
                border-color: #9C27B0;
                transform: translateY(-2px);
            }
            QPushButton:pressed {
                background-color: #E1BEE7;
                border-color: #6A1B9A;
            }
        """

        shelf_life_btn = QPushButton(
            "📊 Shelf-Life Report\nComprehensive analysis of product shelf-life data\nFormats: Excel, PDF")
        shelf_life_btn.setStyleSheet(report_btn_style + """
            QPushButton { color: #4CAF50; }
            QPushButton:hover { background-color: #E8F5E9; border-color: #4CAF50; }
        """)
        shelf_life_btn.clicked.connect(lambda: self.export_report("shelf_life"))

        inventory_btn = QPushButton("📦 Inventory Report\nComplete inventory status and statistics\nFormats: Excel, CSV")
        inventory_btn.setStyleSheet(report_btn_style + """
            QPushButton { color: #2196F3; }
            QPushButton:hover { background-color: #E3F2FD; border-color: #2196F3; }
        """)
        inventory_btn.clicked.connect(lambda: self.export_report("inventory"))

        activity_btn = QPushButton("📋 Activity Report\nUser activities and system operations log\nFormats: PDF, HTML")
        activity_btn.setStyleSheet(report_btn_style + """
            QPushButton { color: #FF9800; }
            QPushButton:hover { background-color: #FFF3E0; border-color: #FF9800; }
        """)
        activity_btn.clicked.connect(lambda: self.export_report("activity"))

        compliance_btn = QPushButton("✅ Compliance Report\nRegulatory compliance and audit trails\nFormats: PDF, Excel")
        compliance_btn.setStyleSheet(report_btn_style + """
            QPushButton { color: #9C27B0; }
            QPushButton:hover { background-color: #F3E5F5; border-color: #9C27B0; }
        """)
        compliance_btn.clicked.connect(lambda: self.export_report("compliance"))

        reports_layout.addWidget(shelf_life_btn, 0, 0)
        reports_layout.addWidget(inventory_btn, 0, 1)
        reports_layout.addWidget(activity_btn, 1, 0)
        reports_layout.addWidget(compliance_btn, 1, 1)

        reports_card.setLayout(reports_layout)
        content_layout.addWidget(reports_card)

        # Export Options Card
        options_card = QGroupBox("⚙️ Export Options")
        options_card.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 16px;
                color: #4A148C;
                border: 2px solid #E1BEE7;
                border-radius: 12px;
                margin-top: 15px;
                padding: 20px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 5px 10px;
                background-color: white;
                border-radius: 5px;
            }
        """)
        options_layout = QFormLayout()
        options_layout.setSpacing(15)

        # Form control styling
        form_control_style = """
            QComboBox, QCheckBox {
                border: 2px solid #E0E0E0;
                border-radius: 8px;
                padding: 10px;
                background-color: white;
                font-size: 14px;
                min-height: 25px;
            }
            QComboBox:focus, QCheckBox:focus {
                border-color: #9C27B0;
                background-color: #F3E5F5;
            }
            QComboBox:hover, QCheckBox:hover {
                border-color: #BA68C8;
            }
        """

        label_style = """
            QLabel {
                font-weight: 600;
                color: #6A1B9A;
                font-size: 14px;
                margin-bottom: 5px;
            }
        """

        # Date range selection
        date_range_label = QLabel("📅 Date Range:")
        date_range_label.setStyleSheet(label_style)
        self.date_range = QComboBox()
        self.date_range.addItems([
            "Last 60 days",
            "Last 90 days",
            "Last 6 months",
            "Last year",
            "All time",
            "Custom range..."
        ])
        self.date_range.setStyleSheet(form_control_style)
        options_layout.addRow(date_range_label, self.date_range)

        # Format selection
        format_label = QLabel("📄 Export Format:")
        format_label.setStyleSheet(label_style)
        self.export_format = QComboBox()
        self.export_format.addItems(["Excel (.xlsx)", "PDF (.pdf)", "CSV (.csv)", "HTML (.html)"])
        self.export_format.setStyleSheet(form_control_style)
        options_layout.addRow(format_label, self.export_format)

        # Include options
        include_label = QLabel("📋 Include Options:")
        include_label.setStyleSheet(label_style)
        options_widget = QWidget()
        options_widget_layout = QVBoxLayout()
        options_widget_layout.setContentsMargins(0, 0, 0, 0)

        self.include_deleted = QCheckBox("Include deleted records")
        self.include_deleted.setStyleSheet(form_control_style + """
            QCheckBox {
                border: none;
                background: transparent;
                padding: 5px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
                border: 2px solid #9C27B0;
                border-radius: 4px;
                background-color: white;
            }
            QCheckBox::indicator:checked {
                background-color: #9C27B0;
                image: url(data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iMTAiIGhlaWdodD0iMTAiIHZpZXdCb3g9IjAgMCAxMCAxMCIgZmlsbD0ibm9uZSIgeG1sbnM9Imh0dHA6Ly93d3cudzMub3JnLzIwMDAvc3ZnIj4KPHBhdGggZD0iTTggM0w0IDdMMiA1IiBzdHJva2U9IndoaXRlIiBzdHJva2Utd2lkdGg9IjIiIHN0cm9rZS1saW5lY2FwPSJyb3VuZCIgc3Ryb2tlLWxpbmVqb2luPSJyb3VuZCIvPgo8L3N2Zz4K);
            }
        """)

        self.include_charts = QCheckBox("Include charts and graphs")
        self.include_charts.setChecked(True)
        self.include_charts.setStyleSheet(self.include_deleted.styleSheet())

        self.include_summary = QCheckBox("Include executive summary")
        self.include_summary.setChecked(True)
        self.include_summary.setStyleSheet(self.include_deleted.styleSheet())

        options_widget_layout.addWidget(self.include_deleted)
        options_widget_layout.addWidget(self.include_charts)
        options_widget_layout.addWidget(self.include_summary)
        options_widget.setLayout(options_widget_layout)
        options_layout.addRow(include_label, options_widget)

        options_card.setLayout(options_layout)
        content_layout.addWidget(options_card)

        # Quick Export Section
        quick_export_card = QGroupBox("⚡ Quick Export")
        quick_export_card.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 16px;
                color: #4A148C;
                border: 2px solid #E1BEE7;
                border-radius: 12px;
                margin-top: 15px;
                padding: 20px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 5px 10px;
                background-color: white;
                border-radius: 5px;
            }
        """)
        quick_layout = QHBoxLayout()
        quick_layout.setSpacing(15)

        # Quick export buttons
        quick_btn_style = """
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                    stop:0 #BA68C8, stop:1 #6A1B9A);
                color: white;
                border: none;
                border-radius: 8px;
                padding: 15px 25px;
                font-weight: bold;
                font-size: 14px;
                min-width: 140px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                    stop:0 #CE93D8, stop:1 #7B1FA2);
            }
            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                    stop:0 #7B1FA2, stop:1 #4A148C);
            }
        """

        export_all_btn = QPushButton("📊 Export All Data")
        export_all_btn.setStyleSheet(quick_btn_style)
        export_all_btn.clicked.connect(lambda: self.export_report("all"))

        export_recent_btn = QPushButton("🕒 Export Recent")
        export_recent_btn.setStyleSheet(quick_btn_style)
        export_recent_btn.clicked.connect(lambda: self.export_report("recent"))

        schedule_export_btn = QPushButton("⏰ Schedule Export")
        schedule_export_btn.setStyleSheet(quick_btn_style)
        schedule_export_btn.clicked.connect(self.schedule_export)

        quick_layout.addWidget(export_all_btn)
        quick_layout.addWidget(export_recent_btn)
        quick_layout.addWidget(schedule_export_btn)
        quick_layout.addStretch()

        quick_export_card.setLayout(quick_layout)
        content_layout.addWidget(quick_export_card)

        content_layout.addStretch()
        content_widget.setLayout(content_layout)
        content_scroll.setWidget(content_widget)

        export_layout.addWidget(content_scroll)
        export_panel.setLayout(export_layout)
        self.content_stack.addWidget(export_panel)

    def create_system_settings_panel(self):
        """Create the system settings panel for super admin"""
        settings_panel = QWidget()
        settings_layout = QVBoxLayout()
        settings_layout.setContentsMargins(30, 30, 30, 30)
        settings_layout.setSpacing(20)

        # Header section
        header_widget = QWidget()
        header_widget.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                    stop:0 #1976D2, stop:1 #42A5F5);
                border-radius: 12px;
                margin-bottom: 20px;
            }
        """)
        header_layout = QVBoxLayout()
        header_layout.setContentsMargins(25, 20, 25, 20)

        title = QLabel("⚙️ System Settings")
        title.setStyleSheet("""
            QLabel {
                font-size: 24px; 
                font-weight: bold; 
                color: white; 
                margin: 0px;
                border: none;
            }
        """)

        subtitle = QLabel("Configure system-wide settings and preferences")
        subtitle.setStyleSheet("""
            QLabel {
                font-size: 14px;
                color: rgba(255, 255, 255, 0.9); 
                margin: 5px 0px 0px 0px;
                border: none;
            }
        """)

        header_layout.addWidget(title)
        header_layout.addWidget(subtitle)
        header_widget.setLayout(header_layout)
        settings_layout.addWidget(header_widget)

        # Settings content
        content_widget = QWidget()
        content_layout = QVBoxLayout()

        # System info card
        info_card = QGroupBox("🖥️ System Information")
        info_card.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 16px;
                color: #1565C0;
                border: 2px solid #BBDEFB;
                border-radius: 12px;
                margin-top: 15px;
                padding: 20px;
                background-color: white;
            }
        """)
        info_layout = QVBoxLayout()

        # Add system info labels
        info_layout.addWidget(QLabel("System Version: Medical Testing System v1.0"))
        info_layout.addWidget(QLabel("Database: SQLite (testing_system.db)"))
        info_layout.addWidget(QLabel("Super Admin Features: Enabled"))

        info_card.setLayout(info_layout)
        content_layout.addWidget(info_card)

        # Settings actions card
        actions_card = QGroupBox("🔧 System Actions")
        actions_card.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 16px;
                color: #1565C0;
                border: 2px solid #BBDEFB;
                border-radius: 12px;
                margin-top: 15px;
                padding: 20px;
                background-color: white;
            }
        """)
        actions_layout = QVBoxLayout()

        # Email settings button
        email_btn = QPushButton("📧 Email Configuration")
        email_btn.clicked.connect(self.open_email_settings)
        email_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 12px 20px;
                font-weight: bold;
                margin: 5px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)

        # Maturation settings button
        maturation_btn = QPushButton("⏰ Maturation Settings")
        maturation_btn.clicked.connect(self.open_maturation_settings)
        maturation_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 12px 20px;
                font-weight: bold;
                margin: 5px;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
        """)

        actions_layout.addWidget(email_btn)
        actions_layout.addWidget(maturation_btn)
        actions_card.setLayout(actions_layout)
        content_layout.addWidget(actions_card)

        content_layout.addStretch()
        content_widget.setLayout(content_layout)
        settings_layout.addWidget(content_widget)

        settings_panel.setLayout(settings_layout)
        self.system_settings_panel = settings_panel
        self.content_stack.addWidget(settings_panel)
        print("✓ System Settings panel created")

    def export_report(self, report_type):
        """Export report with progress indication"""
        try:
            # Show progress dialog
            progress = QProgressDialog(f"Generating {report_type} report...", "Cancel", 0, 100, self)
            progress.setWindowTitle("📤 Exporting Report")
            progress.setWindowModality(Qt.WindowModal)
            progress.setStyleSheet("""
                QProgressDialog {
                    background-color: white;
                    border: 2px solid #9C27B0;
                    border-radius: 8px;
                }
                QProgressBar {
                    border: 2px solid #E1BEE7;
                    border-radius: 8px;
                    text-align: center;
                    background-color: #F3E5F5;
                }
                QProgressBar::chunk {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                        stop:0 #BA68C8, stop:1 #6A1B9A);
                    border-radius: 6px;
                }
            """)
            progress.show()

            # Simulate export process
            import time
            for i in range(101):
                if progress.wasCanceled():
                    break
                progress.setValue(i)
                QApplication.processEvents()
                time.sleep(0.02)  # Simulate processing time

            if not progress.wasCanceled():
                format_ext = self.export_format.currentText().split('(')[1].split(')')[0]
                filename = f"{report_type}_report_{QDateTime.currentDateTime().toString('yyyy-MM-dd_hh-mm-ss')}{format_ext}"

                QMessageBox.information(self, "✅ Export Complete",
                                        f"Report exported successfully!\n\nFile: {filename}\nLocation: Downloads folder\n\nThe report has been generated with your selected options.")

            progress.close()

        except Exception as e:
            QMessageBox.critical(self, "❌ Export Failed", f"An error occurred during export:\n{str(e)}")

    def schedule_export(self):
        """Schedule automatic export"""
        schedule_dialog = QDialog(self)
        schedule_dialog.setWindowTitle("⏰ Schedule Export")
        schedule_dialog.setModal(True)
        schedule_dialog.setFixedSize(400, 300)
        schedule_dialog.setStyleSheet("""
            QDialog {
                background-color: white;
                border: 2px solid #9C27B0;
                border-radius: 12px;
            }
        """)

        layout = QVBoxLayout()

        # Header
        header_label = QLabel("📅 Schedule Automatic Export")
        header_label.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #6A1B9A;
                padding: 15px;
                text-align: center;
            }
        """)
        layout.addWidget(header_label)

        # Frequency selection
        frequency_group = QGroupBox("Export Frequency")
        frequency_layout = QVBoxLayout()

        daily_radio = QRadioButton("Daily")
        weekly_radio = QRadioButton("Weekly")
        monthly_radio = QRadioButton("Monthly")
        weekly_radio.setChecked(True)

        frequency_layout.addWidget(daily_radio)
        frequency_layout.addWidget(weekly_radio)
        frequency_layout.addWidget(monthly_radio)
        frequency_group.setLayout(frequency_layout)
        layout.addWidget(frequency_group)

        # Buttons
        button_layout = QHBoxLayout()
        schedule_btn = QPushButton("✅ Schedule")
        cancel_btn = QPushButton("❌ Cancel")

        schedule_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #388E3C; }
        """)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #F44336;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #D32F2F; }
        """)

        schedule_btn.clicked.connect(lambda: (
            QMessageBox.information(self, "✅ Scheduled", "Export has been scheduled successfully!"),
            schedule_dialog.accept()))
        cancel_btn.clicked.connect(schedule_dialog.reject)

        button_layout.addWidget(schedule_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)

        schedule_dialog.setLayout(layout)
        schedule_dialog.exec_()

    def upload_btn_enter(self, event):
        """Handle mouse enter event for upload button"""
        self.upload_btn.setText("Upload Image")
        self.upload_btn.setMinimumWidth(100)
        self.upload_btn.setStyleSheet("""
            QPushButton {
                background-color: #C8E6C9;
                color: #1B5E20;
                border: 2px solid #4CAF50;
                border-radius: 15px;
                font-size: 10px;
                font-weight: bold;
                min-width: 100px;
                min-height: 30px;
                padding: 5px;
            }
            QPushButton:hover {
                background-color: #C8E6C9;
                border-color: #4CAF50;
                color: #1B5E20;
            }
            QPushButton:pressed {
                background-color: #A5D6A7;
                border-color: #388E3C;
            }
        """)

    def upload_btn_leave(self, event):
        """Handle mouse leave event for upload button"""
        self.upload_btn.setText("📷")
        self.upload_btn.setMinimumWidth(30)
        self.upload_btn.setStyleSheet("""
            QPushButton {
                background-color: #E8F5E9;
                color: #2E7D32;
                border: 2px solid #C8E6C9;
                border-radius: 15px;
                font-size: 12px;
                font-weight: bold;
                min-width: 30px;
                min-height: 30px;
                padding: 5px;
            }
            QPushButton:hover {
                background-color: #C8E6C9;
                border-color: #4CAF50;
                color: #1B5E20;
            }
            QPushButton:pressed {
                background-color: #A5D6A7;
                border-color: #388E3C;
            }
        """)

    def upload_product_image(self):
        """Upload product image and search for matching products in database"""
        try:
            # Open file dialog to select image
            file_dialog = QFileDialog()
            file_path, _ = file_dialog.getOpenFileName(
                self,
                "Select Product Image",
                "",
                "Image Files (*.png *.jpg *.jpeg *.bmp *.gif);;All Files (*)"
            )

            if not file_path:
                return  # User cancelled

            # Show upload progress
            progress = QProgressDialog("Processing image and searching database...", "Cancel", 0, 100, self)
            progress.setWindowTitle("🔍 Searching Products")
            progress.setWindowModality(Qt.WindowModal)
            progress.setStyleSheet("""
                QProgressDialog {
                    background-color: white;
                    border: 2px solid #4CAF50;
                    border-radius: 8px;
                }
                QProgressBar {
                    border: 2px solid #C8E6C9;
                    border-radius: 8px;
                    text-align: center;
                    background-color: #E8F5E9;
                }
                QProgressBar::chunk {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                        stop:0 #66BB6A, stop:1 #388E3C);
                    border-radius: 6px;
                }
            """)
            progress.show()

            # Simulate image processing
            import time
            import os
            for i in range(50):
                if progress.wasCanceled():
                    return
                progress.setValue(i)
                QApplication.processEvents()
                time.sleep(0.02)

            # Extract filename without extension for database search
            filename = os.path.basename(file_path)
            filename_without_ext = os.path.splitext(filename)[0]

            progress.setLabelText("Searching database for matching products...")

            # Search database for products with matching file path or similar name
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Search by exact file path and filename match first
            cursor.execute("""
                SELECT p.product_id, p.product_name, p.batch, u.username as owner, 
                       COALESCE(p.status, 'pending') as status, p.expired_date, 
                       p.product_image, p.sku, p.arrival_date,
                       COALESCE(rl.rack_location_name, 'Unassigned') as rack_location
                FROM products p
                LEFT JOIN users u ON p.owner_id = u.user_id
                LEFT JOIN racklocations rl ON p.rack_location_id = rl.rack_location_id
                WHERE p.product_image LIKE ? OR p.product_image LIKE ? OR p.product_name LIKE ?
                ORDER BY 
                    CASE 
                        WHEN p.product_image = ? THEN 1
                        WHEN p.product_image LIKE ? THEN 2
                        WHEN p.product_name LIKE ? THEN 3
                        ELSE 4
                    END
                LIMIT 10
            """, (f"%{filename}%", f"%{file_path}%", f"%{filename_without_ext}%",
                  file_path, f"%{filename}%", f"%{filename_without_ext}%"))

            results = cursor.fetchall()

            for i in range(50, 101):
                if progress.wasCanceled():
                    conn.close()
                    return
                progress.setValue(i)
                QApplication.processEvents()
                time.sleep(0.01)

            progress.close()
            conn.close()

            if results:
                # Show product search results
                self.show_product_search_results(results, file_path, filename)
            else:
                # No matches found, offer to add image path to existing product
                reply = QMessageBox.question(
                    self,
                    "📷 No Matches Found",
                    f"No products found matching the image '{filename}'.\n\n"
                    "Would you like to:\n"
                    "• Add this image to an existing product\n"
                    "• Create a new product entry with this image\n\n"
                    "Click 'Yes' to add to existing product, 'No' to create new.",
                    QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel
                )

                if reply == QMessageBox.Yes:
                    self.add_image_to_product(file_path)
                elif reply == QMessageBox.No:
                    QMessageBox.information(self, "📝 Create New Product",
                                            "Please use the 'Add New Product' feature to create a new product with this image.")

        except Exception as e:
            QMessageBox.critical(self, "❌ Upload Error", f"An error occurred while uploading the image:\n{str(e)}")

    def show_product_search_results(self, results, image_path, filename):
        """Display search results for uploaded image"""
        result_dialog = QDialog(self)
        result_dialog.setWindowTitle("🔍 Product Search Results")
        result_dialog.setModal(True)
        result_dialog.setFixedSize(800, 600)
        result_dialog.setStyleSheet("""
            QDialog {
                background-color: white;
                border: 2px solid #4CAF50;
                border-radius: 12px;
            }
        """)

        layout = QVBoxLayout()

        # Header
        header_widget = QWidget()
        header_widget.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                    stop:0 #388E3C, stop:1 #4CAF50);
                border-radius: 8px;
                margin-bottom: 15px;
            }
        """)
        header_layout = QVBoxLayout()
        header_layout.setContentsMargins(20, 15, 20, 15)

        title = QLabel(f"📷 Search Results for '{filename}'")
        title.setStyleSheet("""
            QLabel {
                font-size: 18px; 
                font-weight: bold; 
                color: white; 
                margin: 0px;
            }
        """)

        subtitle = QLabel(f"Found {len(results)} matching product(s)")
        subtitle.setStyleSheet("""
            QLabel {
                font-size: 12px; 
                color: rgba(255, 255, 255, 0.9); 
                margin: 5px 0px 0px 0px;
            }
        """)

        header_layout.addWidget(title)
        header_layout.addWidget(subtitle)
        header_widget.setLayout(header_layout)
        layout.addWidget(header_widget)

        # Results table
        results_table = QTableWidget()
        results_table.setColumnCount(10)
        results_table.setHorizontalHeaderLabels([
            "Product ID", "Product Name", "Batch", "Owner", "Status", "Expiry Date",
            "Image Path", "SKU", "Arrival Date", "Rack Location"
        ])
        results_table.setRowCount(len(results))
        results_table.setAlternatingRowColors(True)
        results_table.setSelectionBehavior(QTableWidget.SelectRows)
        results_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #E0E0E0;
                border-radius: 8px;
                gridline-color: #E8F5E9;
            }
            QHeaderView::section {
                background-color: #4CAF50;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
            }
            QTableWidget::item:selected {
                background-color: #C8E6C9;
                color: #1B5E20;
            }
        """)

        # Populate table with results
        for row, result in enumerate(results):
            for col, value in enumerate(result):  # Show all columns
                # Format dates for better display
                if col in [5, 8] and value:  # Expiry date and arrival date columns
                    try:
                        if 'T' in str(value):
                            date_part = str(value).split('T')[0]
                            formatted_date = datetime.strptime(date_part, '%Y-%m-%d').strftime('%Y-%m-%d')
                        else:
                            formatted_date = datetime.strptime(str(value), '%Y-%m-%d').strftime('%Y-%m-%d')
                        value = formatted_date
                    except:
                        pass

                item = QTableWidgetItem(str(value) if value else "")

                # Highlight exact file path matches
                if col == 6 and value and (image_path in str(value) or str(value) in image_path):
                    item.setBackground(QColor("#C8E6C9"))  # Light green background

                results_table.setItem(row, col, item)

        results_table.resizeColumnsToContents()
        layout.addWidget(results_table)

        # Action buttons
        button_layout = QHBoxLayout()

        view_btn = QPushButton("👁️ View Selected Product")
        view_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #1976D2; }
        """)
        view_btn.clicked.connect(lambda: self.view_selected_product(results_table, results, result_dialog))

        update_image_btn = QPushButton("📷 Update Image Path")
        update_image_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #F57C00; }
        """)
        update_image_btn.clicked.connect(
            lambda: self.update_product_image(results_table, results, image_path, result_dialog))

        close_btn = QPushButton("❌ Close")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #757575;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #616161; }
        """)
        close_btn.clicked.connect(result_dialog.reject)

        button_layout.addWidget(view_btn)
        button_layout.addWidget(update_image_btn)
        button_layout.addStretch()
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)

        result_dialog.setLayout(layout)
        result_dialog.exec_()

    def view_selected_product(self, table, results, dialog):
        """View details of selected product"""
        selected_rows = table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "⚠️ No Selection", "Please select a product to view.")
            return

        row = selected_rows[0].row()
        product_data = results[row]

        # Create detailed view dialog
        detail_dialog = QDialog(self)
        detail_dialog.setWindowTitle("📋 Product Details")
        detail_dialog.setModal(True)
        detail_dialog.setFixedSize(500, 400)
        detail_dialog.setStyleSheet("""
            QDialog {
                background-color: #F1F8E9;
                border: 2px solid #4CAF50;
                border-radius: 12px;
            }
        """)

        layout = QVBoxLayout()

        # Product details
        details_text = f"""
        <h2 style="color: #2E7D32;">📦 Product Information</h2>
        <table style="width: 100%; font-size: 14px;">
            <tr><td><b>Product ID:</b></td><td>{product_data[0]}</td></tr>
            <tr><td><b>Product Name:</b></td><td>{product_data[1]}</td></tr>
            <tr><td><b>Batch Number:</b></td><td>{product_data[2]}</td></tr>
            <tr><td><b>Owner:</b></td><td>{product_data[3] if product_data[3] else 'Not Assigned'}</td></tr>
            <tr><td><b>Status:</b></td><td>{product_data[4]}</td></tr>
            <tr><td><b>Expiry Date:</b></td><td>{product_data[5]}</td></tr>
            <tr><td><b>Image Path:</b></td><td>{product_data[6] if product_data[6] else 'No image'}</td></tr>
            <tr><td><b>SKU:</b></td><td>{product_data[7] if product_data[7] else 'N/A'}</td></tr>
            <tr><td><b>Arrival Date:</b></td><td>{product_data[8]}</td></tr>
            <tr><td><b>Rack Location:</b></td><td>{product_data[9]}</td></tr>
        </table>
        """

        details_label = QLabel(details_text)
        details_label.setWordWrap(True)
        details_label.setStyleSheet("""
            QLabel {
                background-color: white;
                border: 1px solid #C8E6C9;
                border-radius: 8px;
                padding: 15px;
                margin: 10px;
            }
        """)
        layout.addWidget(details_label)

        # Close button
        close_btn = QPushButton("✅ OK")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #388E3C; }
        """)
        close_btn.clicked.connect(detail_dialog.accept)
        layout.addWidget(close_btn)

        detail_dialog.setLayout(layout)
        detail_dialog.exec_()

    def update_product_image(self, table, results, image_path, dialog):
        """Update the image path for selected product"""
        selected_rows = table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "⚠️ No Selection", "Please select a product to update.")
            return

        row = selected_rows[0].row()
        product_data = results[row]
        product_id = product_data[0]
        product_name = product_data[1]

        reply = QMessageBox.question(
            self,
            "📷 Update Image Path",
            f"Update image path for product:\n\n"
            f"ID: {product_id}\n"
            f"Name: {product_name}\n\n"
            f"New image path: {image_path}\n\n"
            f"Do you want to proceed?",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                conn = sqlite3.connect("testing_system.db")
                cursor = conn.cursor()

                # Update the image path in database
                cursor.execute("""
                    UPDATE products 
                    SET product_image = ? 
                    WHERE product_id = ?
                """, (image_path, product_id))

                conn.commit()
                conn.close()

                QMessageBox.information(self, "✅ Updated",
                                        f"Image path updated successfully for product '{product_name}'!")

                # Refresh the inventory data
                self.load_data()

                dialog.accept()

            except Exception as e:
                QMessageBox.critical(self, "❌ Update Failed", f"Failed to update image path:\n{str(e)}")

    def add_image_to_product(self, image_path):
        """Add image to an existing product by selecting from inventory"""
        # This would open a product selection dialog
        QMessageBox.information(self, "🔜 Feature Coming Soon",
                                "The 'Add to Existing Product' feature will be implemented soon!\n\n"
                                f"Image path: {image_path}")

    def toggle_fullscreen(self):
        """Toggle between fullscreen and windowed mode (F11)"""
        try:
            if self.isFullScreen():
                # Exit fullscreen mode
                self.showMaximized()
                print("✓ Switched to maximized windowed mode")
                self.add_activity_log(self.username, "Display Mode", "Switched to windowed mode")

                # Update button text and style
                self.fullscreen_btn.setText("🖥️ Fullscreen")
                self.fullscreen_btn.setToolTip("Enter fullscreen mode (F11)")

            else:
                # Enter fullscreen mode
                self.showFullScreen()
                print("✓ Switched to fullscreen mode")
                self.add_activity_log(self.username, "Display Mode", "Switched to fullscreen mode")

                # Update button text and style
                self.fullscreen_btn.setText("🪟 Windowed")
                self.fullscreen_btn.setToolTip("Exit fullscreen mode (F11)")

        except Exception as e:
            print(f"Error toggling fullscreen: {e}")

    def logout(self):
        """Handle logout with confirmation dialog"""
        reply = QMessageBox.question(
            self,
            'Confirm Logout',
            'Are you sure you want to logout from the Shelf Life Management System?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            print(f"✓ User logged out from dashboard")
            self.close()

            # Show the login window again
            self.show_login_window()

    def show_login_window(self):
        """Show the login window again after logout"""
        try:
            from user1 import SignInSignUpWindow
            self.login_window = SignInSignUpWindow()
            self.login_window.show()
        except Exception as e:
            print(f"✗ Error opening login window: {e}")
            # If we can't open login window, just exit the application
            QApplication.quit()

    def approve_product(self, product_id):
        """Approve a pending product and generate barcode"""
        try:
            reply = QMessageBox.question(self, 'Approve Product',
                                         f'Are you sure you want to approve product ID {product_id}?',
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

            if reply == QMessageBox.Yes:
                conn = sqlite3.connect("testing_system.db")
                cursor = conn.cursor()

                # Get product details for barcode generation
                cursor.execute("SELECT sku FROM products WHERE product_id = ?", (product_id,))
                result = cursor.fetchone()

                if result:
                    sku = result[0]

                    # Import the barcode generation function from ksOwner0613.py
                    from ksOwner0613 import generate_barcode_for_approved_product

                    # Generate barcode for the approved product
                    barcode_value, barcode_image_path = generate_barcode_for_approved_product(product_id, sku)

                    if barcode_value:
                        # Update product status to approved
                        cursor.execute("""
                            UPDATE products 
                            SET status = 'approved'
                            WHERE product_id = ?
                        """, (product_id,))

                        conn.commit()
                        conn.close()

                        QMessageBox.information(self, "Success",
                                                f"Product ID {product_id} has been approved successfully!\n"
                                                f"Barcode generated: {barcode_value}")

                        # Refresh the pending data table
                        self.load_pending_data()

                        # Refresh inventory if it's loaded
                        if hasattr(self, 'load_inventory_data'):
                            self.load_inventory_data()
                    else:
                        conn.close()
                        QMessageBox.warning(self, "Barcode Error",
                                            f"Product approved but barcode generation failed. Please try again.")
                else:
                    conn.close()
                    QMessageBox.warning(self, "Error", "Product not found.")

        except Exception as e:
            print(f"Error approving product: {e}")
            QMessageBox.critical(self, "Error", f"Failed to approve product: {str(e)}")

    def reject_product(self, product_id):
        """Reject a pending product with reason"""
        try:
            # Create rejection reason dialog with improved design
            dialog = QDialog(self)
            dialog.setWindowTitle("Reject Product")
            dialog.setFixedSize(580, 750)  # Larger size to prevent overlapping
            dialog.setModal(True)
            dialog.setStyleSheet("""
                QDialog {
                    background-color: #f8f9fa;
                    border: 2px solid #dc3545;
                    border-radius: 12px;
                }
            """)

            # Main layout with no margins to use full space
            main_layout = QVBoxLayout(dialog)
            main_layout.setSpacing(0)
            main_layout.setContentsMargins(0, 0, 0, 0)

            # Header section - fixed height
            header_frame = QFrame()
            header_frame.setFixedHeight(100)
            header_frame.setStyleSheet("""
                QFrame {
                    background-color: #dc3545;
                    border-top-left-radius: 10px;
                    border-top-right-radius: 10px;
                }
            """)
            header_layout = QVBoxLayout(header_frame)
            header_layout.setContentsMargins(25, 25, 25, 25)
            header_layout.setSpacing(8)

            title_label = QLabel("🚫 REJECT PRODUCT")
            title_label.setFont(QFont("Arial", 18, QFont.Bold))
            title_label.setStyleSheet("color: white; background: transparent;")
            title_label.setAlignment(Qt.AlignCenter)
            header_layout.addWidget(title_label)

            product_id_label = QLabel(f"Product ID: {product_id}")
            product_id_label.setFont(QFont("Arial", 13))
            product_id_label.setStyleSheet("color: white; background: transparent;")
            product_id_label.setAlignment(Qt.AlignCenter)
            header_layout.addWidget(product_id_label)

            main_layout.addWidget(header_frame)

            # Content section - expandable
            content_frame = QFrame()
            content_frame.setStyleSheet("""
                QFrame {
                    background-color: white;
                    border: none;
                }
            """)
            content_layout = QVBoxLayout(content_frame)
            content_layout.setContentsMargins(40, 35, 40, 35)
            content_layout.setSpacing(25)

            # Reason selection
            reason_label = QLabel("Select rejection reason:")
            reason_label.setFont(QFont("Arial", 14, QFont.Bold))
            reason_label.setStyleSheet("color: #333; margin-bottom: 8px;")
            content_layout.addWidget(reason_label)

            reason_combo = QComboBox()
            reason_combo.addItems([
                "Select a reason...",
                "Labeling errors",
                "Expired medication",
                "Contamination issues",
                "Defective product",
                "Others"
            ])
            reason_combo.setFixedHeight(50)
            reason_combo.setStyleSheet("""
                QComboBox {
                    padding: 12px 18px;
                    border: 2px solid #ddd;
                    border-radius: 8px;
                    font-size: 14px;
                    background-color: white;
                }
                QComboBox:focus {
                    border-color: #dc3545;
                    outline: none;
                }
                QComboBox:hover {
                    border-color: #aaa;
                }
                QComboBox::drop-down {
                    border: none;
                    width: 40px;
                }
                QComboBox::down-arrow {
                    image: none;
                    border-left: 6px solid transparent;
                    border-right: 6px solid transparent;
                    border-top: 6px solid #666;
                    margin-right: 15px;
                }
                QComboBox QAbstractItemView {
                    border: 2px solid #ddd;
                    border-radius: 8px;
                    background-color: white;
                    selection-background-color: #dc3545;
                    selection-color: white;
                    padding: 5px;
                }
            """)
            content_layout.addWidget(reason_combo)

            # Custom reason text area
            custom_label = QLabel("Additional comments (optional):")
            custom_label.setFont(QFont("Arial", 14, QFont.Bold))
            custom_label.setStyleSheet("color: #333; margin-bottom: 8px;")
            content_layout.addWidget(custom_label)

            custom_text = QTextEdit()
            custom_text.setFixedHeight(100)
            custom_text.setPlaceholderText("Enter custom rejection reason or additional details...")
            custom_text.setStyleSheet("""
                QTextEdit {
                    border: 2px solid #ddd;
                    border-radius: 8px;
                    padding: 15px;
                    font-size: 14px;
                    background-color: white;
                    font-family: Arial, sans-serif;
                    line-height: 1.4;
                }
                QTextEdit:focus {
                    border-color: #dc3545;
                    outline: none;
                }
                QTextEdit:hover {
                    border-color: #aaa;
                }
            """)
            content_layout.addWidget(custom_text)

            main_layout.addWidget(content_frame)

            # Button section - fixed height
            button_frame = QFrame()
            button_frame.setFixedHeight(90)
            button_frame.setStyleSheet("""
                QFrame {
                    background-color: #f8f9fa;
                    border-bottom-left-radius: 10px;
                    border-bottom-right-radius: 10px;
                }
            """)
            button_layout = QHBoxLayout(button_frame)
            button_layout.setContentsMargins(40, 25, 40, 25)
            button_layout.setSpacing(20)

            # Add stretch to push buttons to the right
            button_layout.addStretch()

            cancel_btn = QPushButton("Cancel")
            cancel_btn.setFixedSize(140, 55)
            cancel_btn.setStyleSheet("""
                QPushButton {
                    background-color: #6c757d;
                    color: white;
                    border: none;
                    border-radius: 8px;
                    font-weight: bold;
                    font-size: 15px;
                }
                QPushButton:hover {
                    background-color: #5a6268;
                    transform: translateY(-1px);
                }
                QPushButton:pressed {
                    background-color: #545b62;
                    transform: translateY(0px);
                }
            """)
            cancel_btn.clicked.connect(dialog.reject)
            button_layout.addWidget(cancel_btn)

            reject_btn = QPushButton("Reject Product")
            reject_btn.setFixedSize(160, 55)
            reject_btn.setStyleSheet("""
                QPushButton {
                    background-color: #dc3545;
                    color: white;
                    border: none;
                    border-radius: 8px;
                    font-weight: bold;
                    font-size: 15px;
                }
                QPushButton:hover {
                    background-color: #c82333;
                    transform: translateY(-1px);
                }
                QPushButton:pressed {
                    background-color: #bd2130;
                    transform: translateY(0px);
                }
            """)
            button_layout.addWidget(reject_btn)

            main_layout.addWidget(button_frame)

            def process_rejection():
                selected_reason = reason_combo.currentText()
                custom_reason = custom_text.toPlainText().strip()

                if selected_reason == "Select a reason...":
                    QMessageBox.warning(dialog, "Missing Reason", "Please select a rejection reason.")
                    return

                # Combine reasons appropriately
                if selected_reason == "Others":
                    if not custom_reason:
                        QMessageBox.warning(dialog, "Missing Custom Reason",
                                            "Please provide a custom reason when 'Others' is selected.")
                        return
                    final_reason = custom_reason
                else:
                    final_reason = selected_reason
                    if custom_reason:
                        final_reason += f": {custom_reason}"

                try:
                    conn = sqlite3.connect("testing_system.db")
                    cursor = conn.cursor()

                    # Update product status to rejected and save rejection comment
                    cursor.execute("""
                        UPDATE products 
                        SET status = 'rejected', rejection_comment = ?
                        WHERE product_id = ?
                    """, (final_reason, product_id))

                    conn.commit()
                    conn.close()

                    QMessageBox.information(self, "Success",
                                            f"Product ID {product_id} has been rejected.\nReason: {final_reason}")

                    # Refresh the pending data table
                    self.load_pending_data()

                    dialog.accept()

                except Exception as e:
                    print(f"Error rejecting product: {e}")
                    QMessageBox.critical(dialog, "Error", f"Failed to reject product: {str(e)}")

            reject_btn.clicked.connect(process_rejection)

            dialog.exec_()

        except Exception as e:
            print(f"Error in reject product dialog: {e}")
            QMessageBox.critical(self, "Error", f"Failed to open rejection dialog: {str(e)}")

    def view_product_details(self, product_id):
        """View detailed information about a product with image"""
        try:
            conn = sqlite3.connect("testing_system.db")
            conn.row_factory = sqlite3.Row  # <── key change
            cur = conn.cursor()
            cur.execute("""
                SELECT p.*,
                       u.username           AS owner_name,
                       t.username           AS tester_name,
                       rl.rack_location_name,
                       b.branch_name
                FROM   products p
                LEFT JOIN users        u  ON p.owner_id       = u.user_id
                LEFT JOIN users        t  ON p.tester_id      = t.user_id
                LEFT JOIN racklocations rl ON p.rack_location_id = rl.rack_location_id
                LEFT JOIN branches     b  ON u.branch_id      = b.branch_id
                WHERE  p.product_id = ?
            """, (product_id,))
            row = cur.fetchone()
            conn.close()
            if not row:
                QMessageBox.warning(self, "Error", "Product not found in database.")
                return
            prod = dict(row)  # convert to normal dict for convenience

            # helper to keep N/A fallback
            def safe(key, default='N/A'):
                val = prod.get(key)
                return default if val is None else str(val)

            # --- build the dictionary for ProductDetailDialog -------------
            product_dict = {
                'product_id': safe('product_id'),
                'owner_id': safe('owner_id'),
                'tester_id': safe('tester_id'),
                'product_name': safe('product_name'),
                'product_desc': safe('product_desc'),
                'product_image': safe('product_image', ''),
                'arrival_date': safe('arrival_date'),
                'branch_id': safe('branch_id'),
                'batch': safe('batch'),
                'rack_location_id': safe('rack_location_id'),
                'sku': safe('sku'),
                'manufacture_date': safe('manufacture_date'),
                'expired_date': safe('expired_date'),
                'barcode': safe('barcode', ''),
                'barcode_image': safe('barcode_image', ''),
                'excel_name': safe('excel_name'),
                'rejection_comment': safe('rejection_comment'),
                'location': safe('location'),
                'status': safe('status'),
                'owner': safe('owner_name'),
                'tester': safe('tester_name'),
                'rack_location': safe('rack_location_name', 'Unassigned'),
                'branch': safe('branch_name'),
                'days_left': 0
            }

            dialog = ProductDetailDialog(product_dict, self)
            dialog.exec_()

        except Exception as e:
            print(f"✗ Error loading product details: {e}")
            QMessageBox.critical(self, "Database Error", f"Failed to load product details: {str(e)}")

    def update_rack_assignment(self, product_name, batch_number, location_name):
        """Update rack assignment for a product"""
        if location_name == "Select Location...":
            return

        conn = None
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Get rack location ID
            cursor.execute("SELECT rack_location_id FROM racklocations WHERE rack_location_name = ?", (location_name,))
            location_result = cursor.fetchone()

            if not location_result:
                QMessageBox.warning(self, "Location Not Found", f"Rack location '{location_name}' not found.")
                return

            location_id = location_result[0]

            # Update product rack assignment
            cursor.execute("""
                UPDATE products 
                SET rack_location_id = ? 
                WHERE product_name = ? AND batch = ?
            """, (location_id, product_name, batch_number))

            if cursor.rowcount > 0:
                conn.commit()
                print(f"✓ Updated rack assignment: {product_name} ({batch_number}) → {location_name}")

                # Log the activity
                self.add_activity_log(
                    self.username,
                    "Rack Assignment",
                    f"Assigned {product_name} ({batch_number}) to {location_name}"
                )

                # Refresh the assign data
                self.load_assign_data()
            else:
                QMessageBox.warning(self, "Update Failed", "Product not found or could not be updated.")

        except Exception as e:
            print(f"✗ Error updating rack assignment: {e}")
            QMessageBox.critical(self, "Error", f"Failed to update rack assignment: {str(e)}")
        finally:
            if conn:
                conn.close()

    def on_global_search_changed(self):
        """Handle global search text changes with delay"""
        # Create a timer to delay search (avoid searching on every keystroke)
        if hasattr(self, 'search_timer'):
            self.search_timer.stop()
        else:
            self.search_timer = QTimer()
            self.search_timer.setSingleShot(True)
            self.search_timer.timeout.connect(self.perform_global_search)

        self.search_timer.start(500)  # Wait 500ms after user stops typing

    def clear_global_search(self):
        """Clear global search and reload data"""
        self.search_panel.clear()
        self.perform_global_search()

    def on_search_changed(self):
        """Handle search text changes with delay - deprecated, use on_global_search_changed"""
        self.on_global_search_changed()

    def perform_global_search(self):
        """Perform global search across all relevant tables and pages"""
        search_text = self.search_panel.text().strip()

        # Get current page to determine what to search
        current_widget = self.content_stack.currentWidget()

        if not search_text:
            # If search is empty, reload current page data
            if hasattr(self, 'inventory_table') and current_widget == self.content_stack.widget(0):
                self.load_inventory_data()
            elif hasattr(self, 'pending_table') and current_widget == self.content_stack.widget(1):
                self.load_pending_data()
            elif hasattr(self, 'maturation_table') and current_widget == self.content_stack.widget(2):
                self.load_maturation_data()
            elif hasattr(self, 'assign_table') and current_widget == self.content_stack.widget(3):
                self.load_assign_data()
                return

        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Enhanced search query for inventory
            cursor.execute("""
                SELECT p.product_name, 
                       u.username as owner,
                       COALESCE(t.username, 'Unassigned') as tester,
                       p.arrival_date,
                       p.batch,
                       COALESCE(rl.rack_location_name, 'Unassigned') as rack_location,
                       p.sku,
                       p.expired_date,
                       p.barcode,
                       p.product_id
                FROM products p
                LEFT JOIN users u ON p.owner_id = u.user_id
                LEFT JOIN users t ON p.tester_id = t.user_id
                LEFT JOIN racklocations rl ON p.rack_location_id = rl.rack_location_id
                WHERE p.product_id LIKE ? OR p.product_name LIKE ? OR 
                      p.batch LIKE ? OR u.username LIKE ? OR
                      p.sku LIKE ? OR p.barcode LIKE ?
                ORDER BY p.product_id DESC
                LIMIT 100
            """, (f"%{search_text}%", f"%{search_text}%", f"%{search_text}%",
                  f"%{search_text}%", f"%{search_text}%", f"%{search_text}%"))

            results = cursor.fetchall()
            conn.close()

            # Update the appropriate table based on current page
            if hasattr(self, 'inventory_table'):
                self.update_inventory_table_with_results(results, search_text)

            print(f"✓ Global search completed: {len(results)} products found for '{search_text}'")

        except Exception as e:
            print(f"✗ Global search error: {e}")
            QMessageBox.warning(self, "Search Error", f"An error occurred during search: {str(e)}")

    def update_inventory_table_with_results(self, results, search_text):
        """Update inventory table with search results"""
        self.inventory_table.setRowCount(len(results))

        if len(results) == 0:
            # Show "no results" message
            self.inventory_table.setRowCount(1)
            no_result_item = QTableWidgetItem(f"🔍 No results found for '{search_text}'. Try different keywords.")
            no_result_item.setTextAlignment(Qt.AlignCenter)
            no_result_item.setBackground(QColor("#FFF3E0"))
            self.inventory_table.setItem(0, 0, no_result_item)
            self.inventory_table.setSpan(0, 0, 1, 9)  # Span across all columns

            if hasattr(self, 'page_label'):
                self.page_label.setText("No results found")
            return

        for row, product in enumerate(results):
            # Store product_id in row for later use (it's the last column)
            product_id = product[-1]

            # Display only the first 9 columns (excluding product_id)
            for col, value in enumerate(product[:-1]):
                # Format dates for display
                if col in [3, 7] and value:  # Arrival date and expired date columns
                    try:
                        if 'T' in str(value):
                            date_part = str(value).split('T')[0]
                            formatted_date = datetime.strptime(date_part, '%Y-%m-%d').strftime('%Y-%m-%d')
                        else:
                            formatted_date = datetime.strptime(str(value), '%Y-%m-%d').strftime('%Y-%m-%d')
                        value = formatted_date
                    except:
                        pass

                item = QTableWidgetItem(str(value) if value else "")
                # Store product_id in the first column's item for retrieval
                if col == 0:
                    item.setData(Qt.UserRole, product_id)
                self.inventory_table.setItem(row, col, item)

        # Update page label for search results
        if hasattr(self, 'page_label'):
            self.page_label.setText(f"Search Results: {len(results)} found")

    def perform_search(self):
        """Perform search based on the current search text - deprecated, use perform_global_search"""
        self.perform_global_search()

    def load_maturation_threshold(self):
        """Load the current maturation threshold from database, default to 60 days"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Create settings table if it doesn't exist
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS system_settings (
                    setting_name TEXT PRIMARY KEY,
                    setting_value TEXT,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Get maturation threshold
            cursor.execute("SELECT setting_value FROM system_settings WHERE setting_name = 'maturation_threshold'")
            result = cursor.fetchone()

            threshold = 60  # Default value
            if result:
                try:
                    threshold = int(result[0])
                except:
                    threshold = 60
            else:
                # Insert default value
                cursor.execute("""
                    INSERT OR REPLACE INTO system_settings (setting_name, setting_value) 
                    VALUES ('maturation_threshold', '60')
                """)
                conn.commit()

            conn.close()
            return threshold

        except Exception as e:
            print(f"Error loading maturation threshold: {e}")
            return 60

    def save_maturation_threshold(self, threshold):
        """Save the maturation threshold to database"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            cursor.execute("""
                INSERT OR REPLACE INTO system_settings (setting_name, setting_value, updated_at) 
                VALUES ('maturation_threshold', ?, CURRENT_TIMESTAMP)
            """, (str(threshold),))

            conn.commit()
            conn.close()
            return True

        except Exception as e:
            print(f"Error saving maturation threshold: {e}")
            return False

    def check_and_send_daily_notifications(self, products_data):
        """Check if daily notifications need to be sent and send them if necessary"""
        try:
            today = datetime.now().strftime('%Y-%m-%d')

            # Check if notifications were already sent today
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Create notifications tracking table if it doesn't exist
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS maturation_notifications (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id INTEGER,
                    owner_username TEXT,
                    notification_date DATE,
                    notification_type TEXT DEFAULT 'daily_alert',
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(product_id, owner_username, notification_date, notification_type)
                )
            """)

            # Get products that haven't been notified today
            products_to_notify = []
            for product in products_data:
                product_id = product[0]
                owner_username = product[3]

                # Check if this product owner combination was already notified today
                cursor.execute("""
                    SELECT COUNT(*) FROM maturation_notifications 
                    WHERE product_id = ? AND owner_username = ? AND notification_date = ?
                """, (product_id, owner_username, today))

                count = cursor.fetchone()[0]
                if count == 0:  # Not notified today
                    products_to_notify.append(product)

            conn.close()

            # Send notifications only for new products
            if products_to_notify:
                notifications_sent = self.send_maturation_notification_emails(products_to_notify)

                # Record the notifications in the database
                if notifications_sent:
                    self.record_sent_notifications(products_to_notify, today)

                return notifications_sent
            else:
                # Update summary card to show last notification status
                self.update_email_summary_for_no_new_notifications()
                return []

        except Exception as e:
            print(f"✗ Error checking daily notifications: {e}")
            return []

    def record_sent_notifications(self, products_data, notification_date):
        """Record sent notifications in the database"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            for product in products_data:
                product_id = product[0]
                owner_username = product[3]

                cursor.execute("""
                    INSERT OR IGNORE INTO maturation_notifications 
                    (product_id, owner_username, notification_date) 
                    VALUES (?, ?, ?)
                """, (product_id, owner_username, notification_date))

            conn.commit()
            conn.close()
            print(f"✓ Recorded {len(products_data)} notification entries")

        except Exception as e:
            print(f"✗ Error recording notifications: {e}")

    def update_email_summary_for_no_new_notifications(self):
        """Update email summary when no new notifications are needed"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Get the last notification date
            cursor.execute("""
                SELECT MAX(notification_date), COUNT(DISTINCT owner_username) 
                FROM maturation_notifications 
                WHERE notification_date = (SELECT MAX(notification_date) FROM maturation_notifications)
            """)

            result = cursor.fetchone()
            last_date = result[0] if result and result[0] else None
            owner_count = result[1] if result and result[1] else 0

            conn.close()

            if hasattr(self, 'email_summary_text'):
                if last_date:
                    if last_date == datetime.now().strftime('%Y-%m-%d'):
                        self.email_summary_text.setText(
                            f"✅ Daily notifications already sent today to {owner_count} owner(s) - {last_date}")
                    else:
                        self.email_summary_text.setText(
                            f"📧 Last notifications sent: {last_date} to {owner_count} owner(s)")
                else:
                    self.email_summary_text.setText("📧 No automatic notifications have been sent yet.")

        except Exception as e:
            print(f"✗ Error updating email summary: {e}")

    def send_reminder_emails(self):
        """Send reminder emails manually for selected products only (for admin use)"""
        try:
            # Check if we have product cards with checkboxes
            if not hasattr(self, 'product_cards') or not self.product_cards:
                QMessageBox.warning(self, "No Products",
                                    "No products are available for sending reminders. Please load the maturation page first.")
                return

            # Get selected products (those with checked checkboxes)
            selected_products = []
            for card in self.product_cards:
                if hasattr(card, 'checkbox') and card.checkbox.isChecked():
                    # Get product data from the card
                    product_data = card.product_data
                    selected_products.append((
                        product_data['product_id'],
                        product_data['product_name'],
                        product_data['batch'],
                        product_data['owner'],
                        product_data['sku'],
                        product_data['branch'],
                        product_data['rack_location'],
                        product_data['expired_date'],
                        product_data['manufacture_date'],
                        product_data['barcode'],
                        product_data['barcode_image'],
                        product_data['product_image'],
                        product_data['days_left']
                    ))

            if not selected_products:
                QMessageBox.warning(self, "No Selection",
                                    "Please select at least one product (check the checkbox) before sending reminders.")
                return

            # Send manual reminder emails for selected products only
            notifications_sent = self.send_maturation_notification_emails(selected_products, is_manual=True)

            if notifications_sent:
                product_names = [p[1] for p in selected_products]  # Get product names
                QMessageBox.information(self, "Reminder Sent",
                                        f"Manual reminder emails have been sent to {len(notifications_sent)} product owner(s) for the following products:\n\n" +
                                        "\n".join([f"• {name}" for name in product_names]))
            else:
                QMessageBox.warning(self, "Email Failed",
                                    "Failed to send reminder emails. Please check your email configuration.")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to send reminder emails: {str(e)}")

    def select_all_products(self):
        """Select all product checkboxes in the maturation panel"""
        if hasattr(self, 'product_cards') and self.product_cards:
            for card in self.product_cards:
                if hasattr(card, 'checkbox'):
                    card.checkbox.setChecked(True)
            QMessageBox.information(self, "Selection",
                                    f"Selected all {len(self.product_cards)} products for reminder emails.")
        else:
            QMessageBox.warning(self, "No Products",
                                "No products are available to select.")

    def deselect_all_products(self):
        """Deselect all product checkboxes in the maturation panel"""
        if hasattr(self, 'product_cards') and self.product_cards:
            for card in self.product_cards:
                if hasattr(card, 'checkbox'):
                    card.checkbox.setChecked(False)
            QMessageBox.information(self, "Selection",
                                    "Deselected all products.")
        else:
            QMessageBox.warning(self, "No Products",
                                "No products are available to deselect.")

    def send_maturation_notification_emails(self, products_data, is_manual=False):
        """Send email notifications to product owners about upcoming expiry"""
        try:
            # Group products by owner
            owners_products = {}
            for product in products_data:
                owner = product[3]  # owner username
                owner_email = self.get_user_email(owner)

                if owner_email:
                    if owner not in owners_products:
                        owners_products[owner] = {
                            'email': owner_email,
                            'products': []
                        }
                    owners_products[owner]['products'].append({
                        'name': product[1],
                        'batch': product[2],
                        'expiry_date': product[7],
                        'days_left': product[12]
                    })

            # Send emails to each owner
            notifications_sent = []
            for owner, data in owners_products.items():
                email_content = self.create_maturation_email_content(owner, data['products'], is_manual)

                # Send actual email
                notification_type = "manual reminder" if is_manual else "automated notification"
                print(f"📧 Sending {notification_type} to {owner} ({data['email']})")
                print(f"Products: {[p['name'] for p in data['products']]}")

                # Attempt to send the actual email
                email_sent = self.send_actual_email(
                    recipient_email=data['email'],
                    subject=email_content['subject'],
                    body=email_content['body'],
                    recipient_name=owner
                )

                if email_sent:
                    print(f"✅ Email successfully sent to {owner}")
                    notifications_sent.append({
                        'owner': owner,
                        'email': data['email'],
                        'product_count': len(data['products']),
                        'sent_time': format_malaysia_time(),
                        'is_manual': is_manual
                    })
                else:
                    print(f"❌ Failed to send email to {owner}")

            # Update email summary card (only for automatic notifications)
            if not is_manual:
                self.update_email_summary_card(notifications_sent)

            return notifications_sent

        except Exception as e:
            print(f"✗ Error sending maturation notification emails: {e}")
            return []

    def get_email_template(self, template_type):
        """Get email template from database"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Create table if it doesn't exist
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS email_templates (
                    id INTEGER PRIMARY KEY,
                    template_type TEXT UNIQUE,
                    subject_template TEXT,
                    body_template TEXT,
                    created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    modified_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Get template
            cursor.execute("SELECT subject_template, body_template FROM email_templates WHERE template_type = ?",
                           (template_type,))
            result = cursor.fetchone()
            conn.close()

            if result and result[0] and result[1]:
                return {
                    'subject': result[0],
                    'body': result[1]
                }
            else:
                # Return default templates if not found
                return self.get_default_email_template(template_type)

        except Exception as e:
            print(f"Error loading email template: {e}")
            return self.get_default_email_template(template_type)

    def get_default_email_template(self, template_type):
        """Get default email templates as fallback"""
        if template_type == 'assignment':
            return {
                'subject': '🧪 New Product Assignment - {PRODUCT_NAME} (Batch: {BATCH})',
                'body': """Dear {TESTER_NAME},

You have been assigned a new product for testing in the Laboratory Management System.

📋 ASSIGNMENT DETAILS:
* * * * * * * * * * * *
• Product Name: {PRODUCT_NAME}
• Batch Number: {BATCH}
• Product ID: {PRODUCT_ID}
• Product Owner: {OWNER_NAME} ({OWNER_USERNAME})
• Assigned By: {ASSIGNED_BY}
• Assignment Date: {ASSIGNMENT_DATE}

* * * * * * * * * * * *

Please log into the system to begin testing.

This is an automated message. Please do not reply directly to this email.

Best regards,
Laboratory Management System"""
            }
        elif template_type == 'maturation':
            return {
                'subject': '🚨 Product Expiry {NOTIFICATION_TYPE} - {PRODUCT_COUNT} Product(s) Require Attention',
                'body': """Dear {OWNER_NAME},

{INTRO_TEXT} regarding product(s) under your ownership that are approaching their expiry dates:

{PRODUCT_LIST}

* * * * * * * * * * * *

Please take immediate action to:
1. Review the product condition
2. Process or transfer if still viable
3. Report any issues to the system administrator

This notification was sent on: {NOTIFICATION_DATE}

Best regards,
Laboratory Management System"""
            }
        else:
            return {
                'subject': 'System Notification',
                'body': 'This is an automated notification from the Laboratory Management System.'
            }

    def get_user_email(self, username):
        """Get user email from database"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            cursor.execute("SELECT email FROM users WHERE username = ?", (username,))
            result = cursor.fetchone()
            conn.close()

            return result[0] if result else f"{username}@company.com"  # fallback email

        except Exception as e:
            print(f"✗ Error getting user email: {e}")
            return f"{username}@company.com"  # fallback email

    def get_email_configuration(self):
        """Get email configuration from database or use default Gmail settings"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Create email_config table if it doesn't exist
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS email_config (
                    id INTEGER PRIMARY KEY,
                    smtp_server TEXT DEFAULT 'smtp.gmail.com',
                    smtp_port INTEGER DEFAULT 587,
                    sender_email TEXT,
                    sender_password TEXT,
                    sender_name TEXT DEFAULT 'Laboratory Management System',
                    use_tls BOOLEAN DEFAULT 1
                )
            """)

            # Try to get existing configuration
            cursor.execute("SELECT * FROM email_config WHERE id = 1")
            config = cursor.fetchone()

            if config:
                email_config = {
                    'smtp_server': config[1],
                    'smtp_port': config[2],
                    'sender_email': config[3],
                    'sender_password': config[4],
                    'sender_name': config[5],
                    'use_tls': config[6]
                }
                conn.close()
                return email_config
            else:
                # No configuration found, prompt for setup
                conn.close()
                return self.setup_email_configuration()

        except Exception as e:
            print(f"✗ Error getting email configuration: {e}")
            return None

    def setup_email_configuration(self):
        """Setup email configuration interactively"""
        try:
            from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QCheckBox, \
                QSpinBox, QMessageBox

            # For now, let's use a simple default configuration
            # You can customize these values:
            default_config = {
                'smtp_server': 'smtp.gmail.com',
                'smtp_port': 587,
                'sender_email': 'your_email@gmail.com',  # CHANGE THIS to your Gmail
                'sender_password': 'your_app_password',  # CHANGE THIS to your Gmail App Password
                'sender_name': 'Laboratory Management System',
                'use_tls': True
            }

            # Save default configuration to database
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            cursor.execute("""
                INSERT OR REPLACE INTO email_config 
                (id, smtp_server, smtp_port, sender_email, sender_password, sender_name, use_tls)
                VALUES (1, ?, ?, ?, ?, ?, ?)
            """, (
                default_config['smtp_server'],
                default_config['smtp_port'],
                default_config['sender_email'],
                default_config['sender_password'],
                default_config['sender_name'],
                default_config['use_tls']
            ))

            conn.commit()
            conn.close()

            print("📧 Email configuration created with default Gmail settings")
            print("⚠️  Please update the email credentials in the database:")
            print(f"   - Sender Email: {default_config['sender_email']}")
            print(f"   - Sender Password: {default_config['sender_password']}")

            return default_config

        except Exception as e:
            print(f"✗ Error setting up email configuration: {e}")
            return None

    def send_actual_email(self, recipient_email, subject, body, recipient_name, cc_email=None):
        """Send actual email using SMTP with optional CC"""
        try:
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart
            from email.utils import formataddr

            # Get email configuration
            email_config = self.get_email_configuration()
            if not email_config:
                print("❌ Email configuration not found. Please set up email settings first.")
                return False

            # Create message
            msg = MIMEMultipart()
            msg['From'] = formataddr((email_config['sender_name'], email_config['sender_email']))
            msg['To'] = recipient_email
            msg['Subject'] = subject

            # Add CC if provided
            if cc_email:
                msg['CC'] = cc_email

            # Add body to email
            msg.attach(MIMEText(body, 'plain'))

            # Create SMTP session
            server = smtplib.SMTP(email_config['smtp_server'], email_config['smtp_port'])
            server.starttls()  # Enable security
            server.login(email_config['sender_email'], email_config['sender_password'])

            # Prepare recipient list (TO + CC)
            recipient_list = [recipient_email]
            if cc_email:
                recipient_list.append(cc_email)

            # Send email
            text = msg.as_string()
            server.sendmail(email_config['sender_email'], recipient_list, text)
            server.quit()

            return True

        except Exception as e:
            print(f"❌ Error sending email to {recipient_email}: {e}")
            print("💡 Please check email configuration in send_actual_email method")
            return False

    def create_maturation_email_content(self, owner, products, is_manual=False):
        """Create email content for maturation notification using templates"""
        notification_type = "Manual Reminder" if is_manual else "Automated Alert"
        intro_text = "This is a manual reminder" if is_manual else "This is an automated notification"

        # Create product list
        product_list = ""
        for product in products:
            days_status = "⚠️ CRITICAL" if product['days_left'] <= 7 else "🔴 WARNING" if product[
                                                                                             'days_left'] <= 15 else "🟡 NOTICE"
            product_list += f"""
• Product: {product['name']} (Batch: {product['batch']})
  Expiry Date: {product['expiry_date']}
  Days Remaining: {product['days_left']} days - {days_status}
"""

        # Get email template and replace placeholders
        template = self.get_email_template('maturation')

        subject = template['subject'].format(
            NOTIFICATION_TYPE=notification_type,
            PRODUCT_COUNT=len(products),
            OWNER_NAME=owner
        )

        body = template['body'].format(
            OWNER_NAME=owner,
            INTRO_TEXT=intro_text,
            PRODUCT_LIST=product_list,
            PRODUCT_COUNT=len(products),
            NOTIFICATION_DATE=format_malaysia_time(),
            NOTIFICATION_TYPE=notification_type
        )

        return {'subject': subject, 'body': body}

    def update_email_summary_card(self, notifications_sent):
        """Update the email summary card with latest notification info"""
        if hasattr(self, 'email_summary_text'):
            if notifications_sent:
                owners_list = [notif['owner'] for notif in notifications_sent]
                last_sent = notifications_sent[0]['sent_time'] if notifications_sent else "N/A"

                summary_text = f"System has sent email reminders to: {', '.join(owners_list)} - Last sent: {last_sent}"
                self.email_summary_text.setText(summary_text)
            else:
                self.email_summary_text.setText("No email notifications have been sent yet.")

        # Store notifications for reference
        if not hasattr(self, 'sent_notifications'):
            self.sent_notifications = []
        self.sent_notifications.extend(notifications_sent)

    def open_expired_products_dialog(self):
        """Open expired products dialog"""
        try:
            from newdashboard_extras import ExpiredProductsDialog

            dialog = ExpiredProductsDialog()
            dialog.exec_()
            print("✓ Expired products dialog opened")
        except Exception as e:
            print(f"Error opening expired products dialog: {e}")
            QMessageBox.critical(self, "Error", f"Failed to open expired products dialog: {str(e)}")

    def open_maturation_settings(self):
        """Open the maturation settings dialog for super admin"""
        if not self.is_superadmin_mode:
            QMessageBox.warning(self, "Access Denied", "This feature requires Super Admin access.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("⚙️ Maturation Calculation Settings")
        dialog.setFixedSize(500, 350)
        dialog.setModal(True)

        dialog.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
                border: 2px solid #FF9800;
                border-radius: 12px;
            }
            QLabel {
                color: #333;
                font-family: Arial;
            }
            QRadioButton {
                color: #333;
                font-size: 14px;
                spacing: 8px;
                padding: 8px;
            }
            QRadioButton::indicator {
                width: 20px;
                height: 20px;
                border-radius: 10px;
                border: 2px solid #FF9800;
                background-color: white;
            }
            QRadioButton::indicator:checked {
                background-color: #FF9800;
                border: 2px solid #F57C00;
            }
            QPushButton {
                background-color: #FF9800;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 12px 24px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
            QPushButton:pressed {
                background-color: #E65100;
            }
        """)

        layout = QVBoxLayout()
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # Header
        header_label = QLabel("🕒 Maturation Calculation Settings")
        header_label.setFont(QFont("Arial", 18, QFont.Bold))
        header_label.setStyleSheet("color: #FF9800; margin-bottom: 10px;")
        header_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(header_label)

        # Description
        desc_label = QLabel(
            "Configure how many days before expiration to show products in the upcoming maturation view:")
        desc_label.setFont(QFont("Arial", 12))
        desc_label.setStyleSheet("color: #666; margin-bottom: 15px;")
        desc_label.setWordWrap(True)
        layout.addWidget(desc_label)

        # Radio button group
        self.threshold_group = QButtonGroup()
        current_threshold = self.load_maturation_threshold()

        options = [
            (30, "30 days (1 month) - Short term"),
            (60, "60 days (2 months) - Standard"),
            (90, "90 days (3 months) - Long term"),
            (120, "120 days (4 months) - Maximum")
        ]

        for days, label in options:
            radio = QRadioButton(label)
            radio.setFont(QFont("Arial", 12))
            if days == current_threshold:
                radio.setChecked(True)
            self.threshold_group.addButton(radio, days)
            layout.addWidget(radio)

        # Current status
        status_label = QLabel(f"Current setting: {current_threshold} days")
        status_label.setFont(QFont("Arial", 11, QFont.Bold))
        status_label.setStyleSheet(
            "color: #2E7D32; background-color: rgba(76, 175, 80, 0.1); padding: 8px; border-radius: 6px; margin: 10px 0;")
        status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(status_label)

        # Buttons
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)

        cancel_btn = QPushButton("Cancel")
        cancel_btn.setFixedSize(120, 45)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        cancel_btn.clicked.connect(dialog.reject)

        save_btn = QPushButton("Save Settings")
        save_btn.setFixedSize(120, 45)
        save_btn.clicked.connect(lambda: self.save_threshold_settings(dialog))

        button_layout.addStretch()
        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(save_btn)
        button_layout.addStretch()

        layout.addLayout(button_layout)
        dialog.setLayout(layout)

        dialog.exec_()

    def save_threshold_settings(self, dialog):
        """Save the selected threshold setting"""
        selected_button = self.threshold_group.checkedButton()
        if selected_button:
            new_threshold = self.threshold_group.id(selected_button)

            if self.save_maturation_threshold(new_threshold):
                # Update current threshold
                self.current_threshold = new_threshold

                # Refresh maturation data
                self.load_maturation_data()

                QMessageBox.information(dialog, "Settings Saved",
                                        f"Maturation threshold updated to {new_threshold} days.\nThe upcoming maturation view has been refreshed.")
                dialog.accept()
            else:
                QMessageBox.critical(dialog, "Error", "Failed to save settings. Please try again.")
        else:
            QMessageBox.warning(dialog, "No Selection", "Please select a threshold option.")

    def show_product_detail_popup(self, item):
        """Show detailed product information in a popup window"""
        # Get product_id from the clicked row
        row = item.row()
        product_name_item = self.inventory_table.item(row, 0)
        if not product_name_item:
            return
        product_id = product_name_item.data(Qt.UserRole)
        if not product_id:
            QMessageBox.warning(self, "Error", "Unable to retrieve product information.")
            return

        try:
            # --- fetch row as dict ----------------------------------------
            conn = sqlite3.connect("testing_system.db")
            conn.row_factory = sqlite3.Row  # <── key change
            cur = conn.cursor()
            cur.execute("""
                SELECT p.*,
                       u.username           AS owner_name,
                       t.username           AS tester_name,
                       rl.rack_location_name,
                       b.branch_name
                FROM   products p
                LEFT JOIN users        u  ON p.owner_id       = u.user_id
                LEFT JOIN users        t  ON p.tester_id      = t.user_id
                LEFT JOIN racklocations rl ON p.rack_location_id = rl.rack_location_id
                LEFT JOIN branches     b  ON u.branch_id      = b.branch_id
                WHERE  p.product_id = ?
            """, (product_id,))
            row = cur.fetchone()
            conn.close()
            if not row:
                QMessageBox.warning(self, "Error", "Product not found in database.")
                return
            prod = dict(row)  # convert to normal dict for convenience

            # helper to keep N/A fallback
            def safe(key, default='N/A'):
                val = prod.get(key)
                return default if val is None else str(val)

            # --- build the dictionary for ProductDetailDialog -------------
            product_dict = {
                'product_id': safe('product_id'),
                'owner_id': safe('owner_id'),
                'tester_id': safe('tester_id'),
                'product_name': safe('product_name'),
                'product_desc': safe('product_desc'),
                'product_image': safe('product_image', ''),
                'arrival_date': safe('arrival_date'),
                'branch_id': safe('branch_id'),
                'batch': safe('batch'),
                'rack_location_id': safe('rack_location_id'),
                'sku': safe('sku'),
                'manufacture_date': safe('manufacture_date'),
                'expired_date': safe('expired_date'),
                'barcode': safe('barcode', ''),
                'barcode_image': safe('barcode_image', ''),
                'excel_name': safe('excel_name'),
                'rejection_comment': safe('rejection_comment'),
                'location': safe('location'),
                'status': safe('status'),
                'owner': safe('owner_name'),
                'tester': safe('tester_name'),
                'rack_location': safe('rack_location_name', 'Unassigned'),
                'branch': safe('branch_name'),
                'days_left': 0
            }

            dialog = ProductDetailDialog(product_dict, self)
            dialog.exec_()

        except Exception as e:
            print(f"✗ Error loading product details: {e}")
            QMessageBox.critical(self, "Database Error", f"Failed to load product details: {str(e)}")

    def open_manage_products(self):
        """Open Manage Products (admin only)"""
        if self.user_role.lower() not in ("admin", "superadmin"):
            QMessageBox.warning(self, "Access Denied", "Only administrators can access Manage Products.")
            return

        from newdashboard_extras import ManageProductsDialog  # dynamic import to avoid circular dependency
        dlg = ManageProductsDialog(parent_dashboard=self, user_role=self.user_role,
                                   current_user_id=self.user_info.get('user_id', 0))
        dlg.exec_()

    def open_add_products(self):
        """Placeholder for add products interface"""
        from newdashboard_extras import AddProductDialog  # dynamic import to avoid circular
        dialog = AddProductDialog(parent_dashboard=self, user_role=self.user_role,
                                  current_user_id=self.user_info.get('user_id', 0),
                                  current_user_branch_id=self.user_info.get('branch_id'))
        dialog.product_added.connect(self.load_inventory_data)
        dialog.exec_()

    def export_selected_product_barcodes(self):
        """Export unit barcodes to Excel for the currently selected product row."""
        # Ensure a row is selected
        selected = self.inventory_table.selectionModel().selectedRows()
        if not selected:
            QMessageBox.warning(self, "No Selection", "Please select a product row first.")
            return

        row = selected[0].row()
        product_name_item = self.inventory_table.item(row, 0)
        if not product_name_item:
            QMessageBox.warning(self, "Error", "Could not retrieve product information from table.")
            return

        product_id = product_name_item.data(Qt.UserRole)
        if not product_id:
            QMessageBox.warning(self, "Error", "Missing product ID in table data.")
            return

        try:
            from newdashboard_extras import export_unit_barcodes_to_excel
            export_unit_barcodes_to_excel(product_id)
            QMessageBox.information(self, "Export Complete", f"Unit barcodes exported for product ID {product_id}.")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Failed to export unit barcodes: {str(e)}")

        # ------------------------------------------------------------------

    #  NEW: Export today's products (and their unit barcodes) to Excel
    # ------------------------------------------------------------------
    def export_today_products_to_excel(self):
        """Generate an Excel file with all products added today (+ unit barcodes with embedded images)
        and write that filename back to products.excel_name."""
        try:
            from newdashboard_extras import export_today_products_barcodes_to_excel_with_images
            from datetime import datetime

            # 1) 生成 Excel with embedded barcode images
            export_today_products_barcodes_to_excel_with_images()

            today_str = datetime.now().strftime("%Y%m%d")
            file_name = f"today_products_unit_barcodes_with_images_{today_str}.xlsx"

            # 2) 把文件名记到 products.excel_name（今天新增的产品）
            try:
                conn = sqlite3.connect("testing_system.db", timeout=10.0)
                cur = conn.cursor()
                cur.execute(
                    "UPDATE products SET excel_name = ? "
                    "WHERE DATE(arrival_date) = DATE('now','localtime')",
                    (file_name,)
                )
                conn.commit()
                conn.close()
            except Exception:
                pass  # 记录失败不影响导出

            QMessageBox.information(
                self,
                "Export Complete",
                f"Today's products & unit barcodes with embedded images exported to {file_name}."
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", str(e))

    # ------------------------------------------------------------------
    # Inventory ► Add Rack Location
    # ------------------------------------------------------------------
    def open_add_rack_location_dialog(self):
        """Open dialog to add new rack location with Block, Rack, Row format - supports single and batch adding"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Add New Rack Location")
        dialog.setModal(True)
        dialog.resize(600, 700)

        layout = QVBoxLayout()
        dialog.setLayout(layout)

        # Title label
        title_label = QLabel("Add New Rack Location")
        title_label.setStyleSheet("font-weight: bold; font-size: 16px; margin-bottom: 10px;")
        layout.addWidget(title_label)

        # Mode selection (Single vs Batch)
        mode_group = QGroupBox("Adding Mode")
        mode_layout = QVBoxLayout()

        single_radio = QRadioButton("Add Single Location")
        batch_radio = QRadioButton("Batch Add Locations")
        single_radio.setChecked(True)  # Default to single mode

        mode_layout.addWidget(single_radio)
        mode_layout.addWidget(batch_radio)
        mode_group.setLayout(mode_layout)
        layout.addWidget(mode_group)

        # Form layout for inputs
        form_layout = QFormLayout()

        # Block entry (A-Z only)
        block_edit = QLineEdit()
        block_edit.setPlaceholderText("A-Z")
        block_edit.setMaxLength(1)

        # Rack entry (A-Z only)
        rack_edit = QLineEdit()
        rack_edit.setPlaceholderText("A-Z")
        rack_edit.setMaxLength(1)

        # Row entry (numbers only)
        row_edit = QLineEdit()
        row_edit.setPlaceholderText("1-999")

        # Batch quantity entry (only visible in batch mode)
        quantity_edit = QLineEdit()
        quantity_edit.setPlaceholderText("Number of locations to add")
        quantity_edit.setMaxLength(3)
        quantity_label = QLabel("Quantity:")

        form_layout.addRow("Block:", block_edit)
        form_layout.addRow("Rack:", rack_edit)
        form_layout.addRow("Row:", row_edit)
        form_layout.addRow(quantity_label, quantity_edit)

        # Initially hide quantity field
        quantity_label.hide()
        quantity_edit.hide()

        layout.addLayout(form_layout)

        # Info label for batch mode
        info_label = QLabel(
            "In batch mode, locations will be added by incrementing Rack and Row values.\nExample: A-A-1, A-B-1, A-C-1, then A-A-2, A-B-2, etc.")
        info_label.setStyleSheet("color: #666; font-style: italic; margin: 10px;")
        info_label.setWordWrap(True)
        info_label.hide()
        layout.addWidget(info_label)

        # Mode change handlers
        def on_mode_changed():
            if batch_radio.isChecked():
                quantity_label.show()
                quantity_edit.show()
                info_label.show()
            else:
                quantity_label.hide()
                quantity_edit.hide()
                info_label.hide()

        single_radio.toggled.connect(on_mode_changed)
        batch_radio.toggled.connect(on_mode_changed)

        # Buttons
        btn_layout = QHBoxLayout()
        add_btn = QPushButton("Add Location(s)")
        cancel_btn = QPushButton("Cancel")
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #388E3C;
            }
        """)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        """)

        btn_layout.addStretch()
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(add_btn)
        layout.addLayout(btn_layout)

        # Input validation functions
        def validate_alpha_upper(line_edit):
            """Allow only A-Z characters"""

            def text_changed():
                text = line_edit.text().upper()
                # Keep only A-Z characters
                filtered_text = ''.join(c for c in text if c.isalpha() and c.isupper())
                if filtered_text != line_edit.text():
                    line_edit.setText(filtered_text)

            line_edit.textChanged.connect(text_changed)

        def validate_numbers_only(line_edit):
            """Allow only numbers"""

            def text_changed():
                text = line_edit.text()
                # Keep only digits
                filtered_text = ''.join(c for c in text if c.isdigit())
                if filtered_text != line_edit.text():
                    line_edit.setText(filtered_text)

            line_edit.textChanged.connect(text_changed)

        # Apply validations
        validate_alpha_upper(block_edit)
        validate_alpha_upper(rack_edit)
        validate_numbers_only(row_edit)
        validate_numbers_only(quantity_edit)

        # Helper function to check if location exists
        def location_exists(cursor, block, rack, row):
            """Check if a rack location already exists in database"""
            rack_location_name = f"Block {block} - Rack {rack} - Row {row}"
            cursor.execute("SELECT rack_location_name FROM racklocations WHERE rack_location_name = ?",
                           (rack_location_name,))
            return cursor.fetchone() is not None

        # Helper function to generate next rack/row combination
        def get_next_location(block, start_rack, start_row, increment):
            """Generate next location by incrementing rack and row"""
            rack_ord = ord(start_rack)
            row_num = int(start_row)

            # Calculate new position
            total_positions = (rack_ord - ord('A')) * 999 + row_num + increment - 1

            new_rack_ord = ord('A') + (total_positions // 999)
            new_row = (total_positions % 999) + 1

            # Check if we've exceeded Z for rack
            if new_rack_ord > ord('Z'):
                return None, None  # Exceeded available positions

            return chr(new_rack_ord), str(new_row)

        # Clear all fields function
        def clear_fields():
            block_edit.clear()
            rack_edit.clear()
            row_edit.clear()
            quantity_edit.clear()

        # Add button callback
        def do_add():
            block = block_edit.text().strip().upper()
            rack = rack_edit.text().strip().upper()
            row = row_edit.text().strip()

            # Basic validation
            if not block:
                QMessageBox.warning(dialog, "Input Required", "Please enter a Block (A-Z).")
                return
            if not rack:
                QMessageBox.warning(dialog, "Input Required", "Please enter a Rack (A-Z).")
                return
            if not row:
                QMessageBox.warning(dialog, "Input Required", "Please enter a Row (number).")
                return

            # Batch mode validation
            if batch_radio.isChecked():
                quantity_text = quantity_edit.text().strip()
                if not quantity_text:
                    QMessageBox.warning(dialog, "Input Required", "Please enter the quantity for batch adding.")
                    return
                try:
                    quantity = int(quantity_text)
                    if quantity <= 0 or quantity > 500:
                        QMessageBox.warning(dialog, "Invalid Quantity", "Quantity must be between 1 and 500.")
                        return
                except ValueError:
                    QMessageBox.warning(dialog, "Invalid Quantity", "Please enter a valid number for quantity.")
                    return
            else:
                quantity = 1

            try:
                conn = sqlite3.connect("testing_system.db")
                cur = conn.cursor()

                # Create table if not exists
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS racklocations (
                        rack_location_id INTEGER PRIMARY KEY AUTOINCREMENT,
                        rack_location_name TEXT UNIQUE
                    )
                """)

                added_locations = []
                skipped_locations = []

                for i in range(quantity):
                    if i == 0:
                        # First location uses provided values
                        current_rack = rack
                        current_row = row
                    else:
                        # Generate next location
                        current_rack, current_row = get_next_location(block, rack, row, i)
                        if current_rack is None or current_row is None:
                            QMessageBox.warning(dialog, "Limit Exceeded",
                                                f"Cannot generate more locations. Stopped at {len(added_locations)} locations.")
                            break

                    # Check if location already exists
                    if location_exists(cur, block, current_rack, current_row):
                        location_name = f"Block {block} - Rack {current_rack} - Row {current_row}"
                        skipped_locations.append(location_name)
                        continue

                    # Add the location
                    rack_location_name = f"Block {block} - Rack {current_rack} - Row {current_row}"
                    try:
                        cur.execute("INSERT INTO racklocations (rack_location_name) VALUES (?)", (rack_location_name,))
                        added_locations.append(rack_location_name)
                    except sqlite3.IntegrityError:
                        # Handle race condition if location was added between check and insert
                        skipped_locations.append(rack_location_name)

                conn.commit()
                conn.close()

                # Show results
                result_message = f"Successfully added {len(added_locations)} rack location(s)."
                if skipped_locations:
                    result_message += f"\n\nSkipped {len(skipped_locations)} existing location(s):"
                    for location in skipped_locations[:5]:  # Show first 5 skipped
                        result_message += f"\n• {location}"
                    if len(skipped_locations) > 5:
                        result_message += f"\n• ... and {len(skipped_locations) - 5} more"

                if added_locations:
                    result_message += f"\n\nAdded locations:"
                    for location in added_locations[:5]:  # Show first 5 added
                        result_message += f"\n• {location}"
                    if len(added_locations) > 5:
                        result_message += f"\n• ... and {len(added_locations) - 5} more"

                if added_locations:
                    QMessageBox.information(dialog, "Success", result_message)
                    dialog.accept()
                    # Refresh inventory data
                    self.load_inventory_data()
                else:
                    QMessageBox.warning(dialog, "No Locations Added",
                                        "No new locations were added. All specified locations already exist.")

            except Exception as e:
                QMessageBox.critical(dialog, "Database Error", f"Failed to add rack location(s): {str(e)}")

        add_btn.clicked.connect(do_add)
        cancel_btn.clicked.connect(dialog.reject)

        dialog.exec_()


__all__ = ["Dashboard", "main"]


def main():
    app = QApplication(sys.argv)

    # You can change user_role to "Superadmin" for testing
    dashboard = Dashboard(user_role="Admin")
    dashboard.show()

    sys.exit(app.exec_())


if __name__ == "__main__":
    main()

