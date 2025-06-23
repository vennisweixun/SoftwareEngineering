import sys
import sqlite3
import os
import random
import time
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QLabel, QLineEdit, QPushButton, QComboBox, QDateEdit,
    QTableWidget, QTableWidgetItem, QScrollArea, QFrame, QDialog,
    QFormLayout, QMessageBox, QTextEdit, QHeaderView, QSplitter,
    QTabWidget, QProgressBar, QGroupBox, QCheckBox, QSpinBox, QButtonGroup,
    QStackedWidget, QProgressDialog, QRadioButton, QFileDialog, QListWidget, QListWidgetItem, QSizePolicy
)
from PyQt5.QtCore import Qt, QTimer, QDate, pyqtSignal, QDateTime
from PyQt5.QtGui import QFont, QPalette, QColor, QIcon, QPixmap
import re
from datetime import timezone, timedelta
# from newdashboard_extras import generate_barcode_images
import pandas as pd  # Local import to avoid heavy dependency at startup

# Malaysia timezone (UTC+8)
MALAYSIA_TZ = timezone(timedelta(hours=8))


# Chat database functions
def verify_user_login(cursor, username, password):
    """Verify user login credentials"""
    cursor.execute("""
        SELECT u.user_id, u.username, r.role_name, u.email, u.phone_no, 
               b.branch_name, u.branch_id, r.role_id, u.fullname
        FROM users u
        LEFT JOIN roles r ON u.role = r.role_id
        LEFT JOIN branches b ON u.branch_id = b.branch_id
        WHERE u.username = ? AND u.password = ?
    """, (username, password))

    result = cursor.fetchone()
    if result:
        return {
            'user_id': result[0],
            'username': result[1],
            'role_name': result[2],
            'email': result[3],
            'phone_no': result[4],
            'branch_name': result[5],
            'branch_id': result[6],
            'role_id': result[7],
            'fullname': result[8]
        }
    return None


def get_all_messages(cursor, include_deleted=False):
    """Get all messages from the message table"""
    cursor.execute("""
        SELECT username, role, message, timestamp, message_id
        FROM message
        WHERE role IN ('admin', 'superadmin') 
        AND message NOT LIKE '[ACTIVITY]%'
        ORDER BY timestamp ASC
    """)
    return cursor.fetchall()


def insert_message(cursor, user_id, username, role, role_id, message_text):
    """Insert a new message into the message table"""
    cursor.execute("""
        INSERT INTO message (user_id, role_id, username, role, message, timestamp)
        VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
    """, (user_id, role_id, username, role, message_text))


def get_all_users_for_chat(cursor, current_username):
    """Get all users for chat user list"""
    cursor.execute("""
        SELECT u.username, r.role_name, u.fullname 
        FROM users u 
        JOIN roles r ON u.role = r.role_id 
        WHERE u.username != ?
        ORDER BY r.role_name, u.username
    """, (current_username,))
    return cursor.fetchall()


def get_private_messages(cursor, sender_username, receiver_username):
    """Get private messages between two users"""
    cursor.execute("""
        SELECT sender_username, receiver_username, message_text, timestamp, sender_role
        FROM private_messages 
        WHERE (sender_username = ? AND receiver_username = ?) 
           OR (sender_username = ? AND receiver_username = ?)
        ORDER BY timestamp ASC
        LIMIT 20
    """, (sender_username, receiver_username, receiver_username, sender_username))
    return cursor.fetchall()


def insert_private_message(cursor, sender_user_id, sender_username, receiver_username, sender_role, message_text):
    """Insert a private message"""
    cursor.execute("""
        INSERT INTO private_messages 
        (sender_user_id, sender_username, receiver_username, sender_role, message_text)
        VALUES (?, ?, ?, ?, ?)
    """, (sender_user_id, sender_username, receiver_username, sender_role, message_text))


def format_malaysia_time(timestamp_str=None):
    """Format timestamp to Malaysia time"""
    try:
        if timestamp_str:
            if 'T' in timestamp_str or '-' in timestamp_str:
                dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
            else:
                try:
                    dt = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
                    dt = dt.replace(tzinfo=MALAYSIA_TZ)
                except:
                    dt = datetime.now(MALAYSIA_TZ)
        else:
            dt = datetime.now(MALAYSIA_TZ)

        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except:
        return datetime.now(MALAYSIA_TZ).strftime("%Y-%m-%d %H:%M:%S")


def create_chat_tables(cursor):
    """Create necessary chat tables"""
    # Message table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS message (
            message_id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            role_id INTEGER,
            username TEXT,
            role TEXT,
            message TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Private messages table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS private_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sender_user_id INTEGER,
            sender_username TEXT,
            receiver_username TEXT,
            sender_role TEXT,
            message_text TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)


# Note: CreateAccountDialog will be imported dynamically to avoid circular imports
CreateAccountDialog = None


class SuperAdminDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Super Admin Access")
        self.setFixedSize(600, 300)
        self.setStyleSheet("""
            QDialog {
                background-color: #fff3e0;
            }
            QLabel {
                font-size: 14px;
                font-weight: bold;
                color: #e65100;
            }
            QLineEdit {
                border: 2px solid #ff8400;
                border-radius: 20px;
                padding: 12px;
                font-size: 14px;
                background-color: white;
            }
            QPushButton {
                background-color: #ff8400;
                color: #7E5433;
                border: none;
                border-radius: 4px;
                padding: 12px 24px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #e67600;
            }
            QPushButton:pressed {
                background-color: #cc6600;
            }
        """)

        layout = QVBoxLayout()
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(15)

        title_label = QLabel("Super Admin Access")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("color: #e65100; margin-bottom: 10px;")

        self.passcode_input = QLineEdit()
        self.passcode_input.setEchoMode(QLineEdit.Password)
        self.passcode_input.setPlaceholderText("Enter Super Admin Passcode")
        self.passcode_input.setMinimumHeight(45)

        self.confirm_button = QPushButton("Confirm Access")
        self.confirm_button.clicked.connect(self.check_passcode)
        self.confirm_button.setMinimumHeight(45)

        layout.addWidget(title_label)
        layout.addWidget(QLabel("Please enter your passcode to access Super Admin features:"))
        layout.addWidget(self.passcode_input)
        layout.addWidget(self.confirm_button)

        self.setLayout(layout)

    def check_passcode(self):
        # Simple passcode check - in real app, this should be more secure
        if self.passcode_input.text() == "123456":
            self.accept()
            self.is_superadmin_mode = True
        else:
            QMessageBox.warning(self, "Access Denied", "Invalid passcode!")


class SummaryCard(QFrame):
    def __init__(self, title, value, subtitle="", color="#4CAF50"):
        super().__init__()
        self.setFrameStyle(QFrame.Box)
        self.setStyleSheet(f"""
            QFrame {{
                background-color: white;
                border: 2px solid {color};
                border-radius: 10px;
                padding: 10px;
            }}
            QLabel {{
                border: none;
            }}
        """)

        layout = QVBoxLayout()

        title_label = QLabel(title)
        title_label.setFont(QFont("Arial", 12, QFont.Bold))
        title_label.setStyleSheet(f"color: {color};")

        value_label = QLabel(str(value))
        value_label.setFont(QFont("Arial", 24, QFont.Bold))
        value_label.setStyleSheet(f"color: {color};")
        value_label.setAlignment(Qt.AlignCenter)

        if subtitle:
            subtitle_label = QLabel(subtitle)
            subtitle_label.setFont(QFont("Arial", 10))
            subtitle_label.setStyleSheet("color: #666;")
            subtitle_label.setAlignment(Qt.AlignCenter)
            layout.addWidget(subtitle_label)

        layout.addWidget(title_label)
        layout.addWidget(value_label)

        self.setLayout(layout)
        self.setFixedHeight(120)


class ProductCardWithCheckbox(QFrame):
    def __init__(self, product_data, parent=None):
        super().__init__(parent)
        self.product_data = product_data
        self.setMinimumSize(800, 350)  # Compact size - width, height
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
        self.setFrameStyle(QFrame.StyledPanel)

        # Calculate days until expiration
        days_left = product_data.get('days_left', 0)

        # Set color based on urgency
        if days_left <= 7:
            border_color = "#F44336"  # Red - Very urgent
            bg_color = "#FFEBEE"
        elif days_left <= 15:
            border_color = "#FF9800"  # Orange - Urgent
            bg_color = "#FFF3E0"
        elif days_left <= 30:
            border_color = "#FFC107"  # Yellow - Warning
            bg_color = "#FFFDE7"
        else:
            border_color = "#4CAF50"  # Green - Normal
            bg_color = "#F1F8E9"

        self.setStyleSheet(f"""
            QFrame {{
                background-color: {bg_color};
                border: 2px solid {border_color};
                border-radius: 12px;
                margin: 5px;
            }}
            QFrame:hover {{
                box-shadow: 0 4px 15px rgba(0, 0, 0, 0.2);
                transform: translateY(-2px);
            }}
        """)

        self.init_ui()

    def init_ui(self):
        # Main horizontal layout - checkbox, image on left, details on right
        main_layout = QHBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(25)

        # Left side - Checkbox
        self.checkbox = QCheckBox()
        self.checkbox.setFixedSize(20, 20)
        self.checkbox.setStyleSheet("""
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
                border: 2px solid #4CAF50;
                border-radius: 3px;
                background-color: white;
            }
            QCheckBox::indicator:checked {
                background-color: #4CAF50;
                image: url(data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iMTIiIGhlaWdodD0iOSIgdmlld0JveD0iMCAwIDEyIDkiIGZpbGw9Im5vbmUiIHhtbG5zPSJodHRwOi8vd3d3LnczLm9yZy8yMDAwL3N2ZyI+CjxwYXRoIGQ9Ik0xIDQuNUw0LjUgOEwxMSAxIiBzdHJva2U9IndoaXRlIiBzdHJva2Utd2lkdGg9IjIiIHN0cm9rZS1saW5lY2FwPSJyb3VuZCIgc3Ryb2tlLWxpbmVqb2luPSJyb3VuZCIvPgo8L3N2Zz4K);
            }
            QCheckBox::indicator:hover {
                border-color: #45a049;
            }
        """)
        main_layout.addWidget(self.checkbox, 0, Qt.AlignTop)

        # Product image container
        image_container = QFrame()
        image_container.setFixedSize(180, 200)
        image_container.setStyleSheet("""
            QFrame {
                background-color: transparent;
                border: none;
            }
        """)

        image_layout = QVBoxLayout(image_container)
        image_layout.setContentsMargins(15, 15, 15, 15)
        image_layout.setAlignment(Qt.AlignCenter)

        # Product image
        product_image = QLabel()
        product_image.setFixedSize(220, 200)
        product_image.setAlignment(Qt.AlignCenter)
        product_image.setStyleSheet("""
            QLabel {
                background-color: transparent;
                border: none;
                border-radius: 0px;
            }
        """)

        # Load product image if available
        image_path = self.product_data.get('product_image', '')
        if image_path and os.path.exists(image_path):
            try:
                pixmap = QPixmap(image_path)
                if not pixmap.isNull():
                    scaled_pixmap = pixmap.scaled(218, 198, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    product_image.setPixmap(scaled_pixmap)
                    product_image.setText("")
                else:
                    product_image.setText("No Image")
                    product_image.setStyleSheet(product_image.styleSheet() + "color: #999; font-size: 12px;")
            except Exception as e:
                print(f"Error loading product image {image_path}: {e}")
                product_image.setText("No Image")
                product_image.setStyleSheet(product_image.styleSheet() + "color: #999; font-size: 12px;")
        else:
            product_image.setText("No Image")
            product_image.setStyleSheet(product_image.styleSheet() + "color: #999; font-size: 12px;")

        image_layout.addWidget(product_image)

        # Barcode image (if available)
        barcode_image_path = self.product_data.get('barcode_image', '')
        if barcode_image_path and os.path.exists(barcode_image_path):
            barcode_image = QLabel()
            barcode_image.setFixedHeight(35)
            barcode_image.setAlignment(Qt.AlignCenter)
            try:
                barcode_pixmap = QPixmap(barcode_image_path)
                if not barcode_pixmap.isNull():
                    scaled_barcode = barcode_pixmap.scaled(200, 33, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    barcode_image.setPixmap(scaled_barcode)
                    barcode_image.setStyleSheet(
                        "background-color: white; border: 1px solid #DDD; border-radius: 3px; margin-top: 5px;")
                    image_layout.addWidget(barcode_image)
            except:
                pass

        main_layout.addWidget(image_container)

        # Right side - Product details (same as original ProductCard)
        details_container = QFrame()
        details_layout = QVBoxLayout(details_container)
        details_layout.setContentsMargins(0, 0, 0, 0)
        details_layout.setSpacing(15)

        # Product name (large, bold)
        name_label = QLabel(self.product_data.get('product_name', 'Unknown Product'))
        name_label.setFont(QFont("Arial", 18, QFont.Bold))
        name_label.setStyleSheet("""
            color: #002D04;
            border: none;
            padding: 0px;
            margin: 0px;
        """)
        name_label.setAlignment(Qt.AlignCenter)
        name_label.setWordWrap(True)
        details_layout.addWidget(name_label)

        # Product info grid
        info_grid = QGridLayout()
        info_grid.setHorizontalSpacing(20)
        info_grid.setVerticalSpacing(4)

        label_style = """
            color: #000000;
            border: none;
            padding: 0px;
            margin: 0px;
        """
        value_style = """
            color: #333333;
            border: none;
            padding: 0px;
            margin: 0px;
        """

        def make_label(text, bold=False, is_value=False):
            label = QLabel(text)
            font = QFont("Arial", 13)
            if bold:
                font.setBold(True)
            label.setFont(font)
            label.setStyleSheet(value_style if is_value else label_style)

            if is_value:
                label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            else:
                label.setAlignment(Qt.AlignCenter)
            return label

        # Row 1
        info_grid.addWidget(make_label("Product ID:"), 0, 0)
        info_grid.addWidget(make_label(str(self.product_data.get('product_id', 'N/A')), is_value=True), 0, 1)
        info_grid.addWidget(make_label("Manufacture Date:"), 0, 2)
        info_grid.addWidget(make_label(self.product_data.get('manufacture_date', 'N/A'), is_value=True), 0, 3)

        # Row 2
        info_grid.addWidget(make_label("Branch:"), 1, 0)
        info_grid.addWidget(make_label(self.product_data.get('branch', 'Main Branch- Bayan Lepas'), is_value=True), 1,
                            1)
        info_grid.addWidget(make_label("Expiry Date:", bold=True), 1, 2)
        info_grid.addWidget(make_label(self.product_data.get('expired_date', 'N/A'), is_value=True, bold=True), 1, 3)

        # Row 3
        info_grid.addWidget(make_label("SKU:"), 2, 0)
        info_grid.addWidget(
            make_label(str(self.product_data.get('sku', self.product_data.get('product_id', 'N/A'))), is_value=True), 2,
            1)
        info_grid.addWidget(make_label("Rack Location:"), 2, 2)
        info_grid.addWidget(make_label(self.product_data.get('rack_location', 'Unassigned'), is_value=True), 2, 3)

        # Row 4
        info_grid.addWidget(make_label("Batch:"), 3, 0)
        info_grid.addWidget(make_label(self.product_data.get('batch', 'N/A'), is_value=True), 3, 1)
        info_grid.addWidget(make_label("Barcode:"), 3, 2)

        barcode_text = self.product_data.get('barcode', '')
        if barcode_text:
            barcode_value = make_label(barcode_text, is_value=True)
            barcode_value.setFont(QFont("Courier", 13))
        else:
            barcode_value = make_label("No Barcode", is_value=True)
            barcode_value.setStyleSheet("color: #999999; font-style: italic; border: none; padding: 0px; margin: 0px;")

        info_grid.addWidget(barcode_value, 3, 3)

        details_layout.addLayout(info_grid)

        # Barcode image display directly under the barcode number
        barcode_img_path = self.product_data.get('barcode_image', '')
        barcode_text = str(self.product_data.get('barcode', ''))

        possible_paths = []
        if barcode_img_path:
            # exact path stored
            possible_paths.append(barcode_img_path)
            # same name inside barcodes folder
            possible_paths.append(os.path.join('barcodes', os.path.basename(barcode_img_path)))

        # derive from barcode number if available
        if barcode_text and barcode_text not in ('N/A', ''):
            possible_paths.extend([
                f'barcode_{barcode_text}.png',
                f'barcode_{barcode_text}.png.png',
                os.path.join('barcodes', f'{barcode_text}.png'),
                os.path.join('barcodes', f'{barcode_text}.png.png'),
                os.path.join('barcodes', f'barcode_{barcode_text}.png'),
                os.path.join('barcodes', f'barcode_{barcode_text}.png.png')
            ])

        # include SKU prefix variants
        sku_value = str(self.product_data.get('sku', '')).strip()
        if sku_value and sku_value not in ('N/A', '') and barcode_text and barcode_text not in ('N/A', ''):
            possible_paths.extend([
                f'{sku_value}_{barcode_text}.png',
                f'{sku_value}_{barcode_text}.png.png',
                os.path.join('barcodes', f'{sku_value}_{barcode_text}.png'),
                os.path.join('barcodes', f'{sku_value}_{barcode_text}.png.png')
            ])

        resolved_barcode_img = next((p for p in possible_paths if os.path.isfile(p)), '')

        if resolved_barcode_img:
            barcode_img_label = QLabel()
            barcode_pixmap = QPixmap(resolved_barcode_img)
            if not barcode_pixmap.isNull():
                barcode_pixmap = barcode_pixmap.scaled(180, 60, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                barcode_img_label.setPixmap(barcode_pixmap)
                barcode_img_label.setAlignment(Qt.AlignCenter)
                # Place the image spanning full grid width beneath barcode row
                info_grid.addWidget(barcode_img_label, 4, 0, 1, 4)
        else:
            print(f'[INFO] Barcode image not found for: {barcode_text} (searched {len(possible_paths)} paths)')

        # Days left (prominent display at bottom)
        days_left = self.product_data.get('days_left', 0)
        if days_left < 0:
            days_text = f"⚠️ EXPIRED ({abs(days_left)} days ago)"
            days_color = "#D32F2F"
            days_bg = "#FFCDD2"
        elif days_left == 0:
            days_text = "⚠️ EXPIRES TODAY"
            days_color = "#D32F2F"
            days_bg = "#FFCDD2"
        elif days_left <= 7:
            days_text = f"🔴 {days_left} days left"
            days_color = "#D32F2F"
            days_bg = "#FFEBEE"
        elif days_left <= 15:
            days_text = f"🟠 {days_left} days left"
            days_color = "#F57C00"
            days_bg = "#FFF3E0"
        elif days_left <= 30:
            days_text = f"🟡 {days_left} days left"
            days_color = "#FBC02D"
            days_bg = "#FFFDE7"
        else:
            days_text = f"🟢 {days_left} days left"
            days_color = "#388E3C"
            days_bg = "#F1F8E9"

        days_label = QLabel(days_text)
        days_label.setFont(QFont("Arial", 16, QFont.Bold))
        days_label.setStyleSheet(f"""
            color: {days_color}; 
            background-color: {days_bg}; 
            padding: 12px 20px; 
            border-radius: 8px; 
            border: 2px solid {days_color};
            margin-top: 10px;
        """)
        days_label.setAlignment(Qt.AlignCenter)
        details_layout.addWidget(days_label)

        details_layout.addStretch()
        main_layout.addWidget(details_container)

        self.setLayout(main_layout)


class ProductCard(QFrame):
    def __init__(self, product_data, parent=None):
        super().__init__(parent)
        self.product_data = product_data
        self.setMinimumSize(800, 300)  # Compact size - width, height
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
        self.setFrameStyle(QFrame.StyledPanel)

        # Calculate days until expiration
        days_left = product_data.get('days_left', 0)

        # Set color based on urgency
        if days_left <= 7:
            border_color = "#F44336"  # Red - Very urgent
            bg_color = "#FFEBEE"
        elif days_left <= 15:
            border_color = "#FF9800"  # Orange - Urgent
            bg_color = "#FFF3E0"
        elif days_left <= 30:
            border_color = "#FFC107"  # Yellow - Warning
            bg_color = "#FFFDE7"
        else:
            border_color = "#4CAF50"  # Green - Normal
            bg_color = "#F1F8E9"

        self.setStyleSheet(f"""
            QFrame {{
                background-color: {bg_color};
                border: 2px solid {border_color};
                border-radius: 12px;
                margin: 5px;
            }}
            QFrame:hover {{
                box-shadow: 0 4px 15px rgba(0, 0, 0, 0.2);
                transform: translateY(-2px);
            }}
        """)

        self.init_ui()

    def init_ui(self):
        # Main horizontal layout - image on left, details on right
        main_layout = QHBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(25)

        # Left side - Product image
        image_container = QFrame()
        image_container.setFixedSize(180, 200)
        image_container.setStyleSheet("""
            QFrame {
                background-color: transparent;
                border: none;
            }
        """)

        image_layout = QVBoxLayout(image_container)
        image_layout.setContentsMargins(15, 15, 15, 15)
        image_layout.setAlignment(Qt.AlignCenter)

        # Product image
        product_image = QLabel()
        product_image.setFixedSize(220, 200)
        product_image.setAlignment(Qt.AlignCenter)
        product_image.setStyleSheet("""
            QLabel {
                background-color: transparent;
                border: none;
                border-radius: 0px;
            }
        """)

        # Load product image if available
        image_path = self.product_data.get('product_image', '')
        if image_path and os.path.exists(image_path):
            try:
                pixmap = QPixmap(image_path)
                if not pixmap.isNull():
                    scaled_pixmap = pixmap.scaled(218, 198, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    product_image.setPixmap(scaled_pixmap)
                    product_image.setText("")
                else:
                    product_image.setText("No Image")
                    product_image.setStyleSheet(product_image.styleSheet() + "color: #999; font-size: 12px;")
            except:
                product_image.setText("No Image")
                product_image.setStyleSheet(product_image.styleSheet() + "color: #999; font-size: 12px;")
        else:
            product_image.setText("No Image")
            product_image.setStyleSheet(product_image.styleSheet() + "color: #999; font-size: 12px;")

        image_layout.addWidget(product_image)

        # Barcode image (if available)
        barcode_image_path = self.product_data.get('barcode_image', '')
        if barcode_image_path and os.path.exists(barcode_image_path):
            barcode_image = QLabel()
            barcode_image.setFixedHeight(35)
            barcode_image.setAlignment(Qt.AlignCenter)
            try:
                barcode_pixmap = QPixmap(barcode_image_path)
                if not barcode_pixmap.isNull():
                    scaled_barcode = barcode_pixmap.scaled(200, 33, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    barcode_image.setPixmap(scaled_barcode)
                    barcode_image.setStyleSheet(
                        "background-color: white; border: 1px solid #DDD; border-radius: 3px; margin-top: 5px;")
                    image_layout.addWidget(barcode_image)
            except:
                pass

        main_layout.addWidget(image_container)

        # Right side - Product details
        details_container = QFrame()
        details_layout = QVBoxLayout(details_container)
        details_layout.setContentsMargins(0, 0, 0, 0)
        details_layout.setSpacing(15)

        # Product name (large, bold)
        name_label = QLabel(self.product_data.get('product_name', 'Unknown Product'))
        name_label.setFont(QFont("Arial", 18, QFont.Bold))
        name_label.setStyleSheet("""
            color: #002D04;
            border: none;
            padding: 0px;
            margin: 0px;
        """)
        name_label.setAlignment(Qt.AlignCenter)

        name_label.setWordWrap(True)
        details_layout.addWidget(name_label)

        # Product info grid
        info_grid = QGridLayout()
        info_grid.setHorizontalSpacing(20)  # 控制列与列的水平间距
        info_grid.setVerticalSpacing(4)  # 缩小行与行之间的垂直间距

        # 公用样式：移除边框、边距、设置字体颜色
        label_style = """
            color: #000000;
            border: none;
            padding: 0px;
            margin: 0px;
        """
        value_style = """
            color: #333333;
            border: none;
            padding: 0px;
            margin: 0px;
        """

        def make_label(text, bold=False, is_value=False):
            label = QLabel(text)
            font = QFont("Arial", 13)
            if bold:
                font.setBold(True)
            label.setFont(font)
            label.setStyleSheet(value_style if is_value else label_style)

            if is_value:
                label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            else:
                label.setAlignment(Qt.AlignCenter)
            return label

        # Row 1
        info_grid.addWidget(make_label("Product ID:"), 0, 0)
        info_grid.addWidget(make_label(str(self.product_data.get('product_id', 'N/A')), is_value=True), 0, 1)
        info_grid.addWidget(make_label("Manufacture Date:"), 0, 2)
        info_grid.addWidget(make_label(self.product_data.get('manufacture_date', 'N/A'), is_value=True), 0, 3)

        # Row 2
        info_grid.addWidget(make_label("Branch:"), 1, 0)
        info_grid.addWidget(make_label(self.product_data.get('branch', 'Main Branch- Bayan Lepas'), is_value=True), 1,
                            1)
        info_grid.addWidget(make_label("Expiry Date:", bold=True), 1, 2)
        info_grid.addWidget(make_label(self.product_data.get('expired_date', 'N/A'), is_value=True, bold=True), 1, 3)

        # Row 3
        info_grid.addWidget(make_label("SKU:"), 2, 0)
        info_grid.addWidget(
            make_label(str(self.product_data.get('sku', self.product_data.get('product_id', 'N/A'))), is_value=True), 2,
            1)
        info_grid.addWidget(make_label("Rack Location:"), 2, 2)
        info_grid.addWidget(make_label(self.product_data.get('rack_location', 'Unassigned'), is_value=True), 2, 3)

        # Row 4
        info_grid.addWidget(make_label("Batch:"), 3, 0)
        info_grid.addWidget(make_label(self.product_data.get('batch', 'N/A'), is_value=True), 3, 1)
        info_grid.addWidget(make_label("Barcode:"), 3, 2)

        barcode_text = self.product_data.get('barcode', '')
        if barcode_text:
            barcode_value = make_label(barcode_text, is_value=True)
            barcode_value.setFont(QFont("Courier", 13))
        else:
            barcode_value = make_label("No Barcode", is_value=True)
            barcode_value.setStyleSheet("color: #999999; font-style: italic; border: none; padding: 0px; margin: 0px;")

        info_grid.addWidget(barcode_value, 3, 3)

        details_layout.addLayout(info_grid)

        # Calculate days left until expiry
        expired_str = self.product_data.get('expired_date', '')
        days_left = 0
        if expired_str and expired_str not in ('N/A', ''):
            try:
                # Extract only date component if datetime string
                date_part = expired_str.split(' ')[0]
                exp_dt = datetime.strptime(date_part, '%Y-%m-%d')
                days_left = (exp_dt.date() - datetime.now().date()).days
            except Exception:
                pass  # keep default 0 if parsing fails

        # Days left (prominent display)
        if days_left < 0:
            days_text = f"⚠️ EXPIRED ({abs(days_left)} days ago)"
            days_color = "#D32F2F"
            days_bg = "#FFCDD2"
        elif days_left == 0:
            days_text = "⚠️ EXPIRES TODAY"
            days_color = "#D32F2F"
            days_bg = "#FFCDD2"
        elif days_left <= 7:
            days_text = f"🔴 {days_left} days left"
            days_color = "#D32F2F"
            days_bg = "#FFEBEE"
        elif days_left <= 15:
            days_text = f"🟠 {days_left} days left"
            days_color = "#F57C00"
            days_bg = "#FFF3E0"
        elif days_left <= 30:
            days_text = f"🟡 {days_left} days left"
            days_color = "#FBC02D"
            days_bg = "#FFFDE7"
        else:
            days_text = f"🟢 {days_left} days left"
            days_color = "#388E3C"
            days_bg = "#F1F8E9"

        days_label = QLabel(days_text)
        days_label.setFont(QFont("Arial", 14, QFont.Bold))  # smaller font
        days_label.setStyleSheet(f"""
            color: {days_color}; 
            background-color: {days_bg}; 
            padding: 8px 16px; 
            border-radius: 6px; 
            border: 1px solid {days_color};
            margin-top: 8px;
        """)
        days_label.setAlignment(Qt.AlignCenter)
        details_layout.addWidget(days_label)

        details_layout.addStretch()
        main_layout.addWidget(details_container)

        self.setLayout(main_layout)


class ExpiredProductsDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Expired Products")
        self.setFixedSize(1500, 650)
        self.setStyleSheet("""
            QDialog {
                background-color: #fff5f5;
                font-size: 14px;
            }
            QTableWidget {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 8px;
                gridline-color: #e0e0e0;
                font-size: 14px;
            }
            QHeaderView::section {
                background-color: #F44336;
                color: white;
                padding: 10px;
                border: none;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton {
                background-color: #F44336;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #D32F2F;
            }
            QLabel {
                font-size: 14px;
            }
        """)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Header
        header_layout = QHBoxLayout()
        title = QLabel("⚠️ Expired Products Report")
        title.setFont(QFont("Arial", 18, QFont.Bold))
        title.setStyleSheet("color: #D32F2F; margin-bottom: 10px;")
        header_layout.addWidget(title)
        header_layout.addStretch()

        # Refresh button
        refresh_btn = QPushButton("Refresh")
        refresh_btn.setFixedSize(120, 35)
        refresh_btn.clicked.connect(self.load_expired_products)
        header_layout.addWidget(refresh_btn)

        layout.addLayout(header_layout)

        # Info label
        info_label = QLabel("Products that have passed their expiration date:")
        info_label.setStyleSheet("color: #666; font-size: 12px; margin-bottom: 10px;")
        layout.addWidget(info_label)

        # Table for expired products
        self.expired_table = QTableWidget()
        self.expired_table.setColumnCount(9)
        self.expired_table.setHorizontalHeaderLabels([
            "Product Name", "Owner", "Branch", "Batch", "Expired Date",
            "Days Overdue", "SKU", "Rack Location", "Status"
        ])

        # Set column widths
        header = self.expired_table.horizontalHeader()
        header.setStretchLastSection(True)
        self.expired_table.setColumnWidth(0, 200)  # Product Name
        self.expired_table.setColumnWidth(1, 120)  # Owner
        self.expired_table.setColumnWidth(2, 100)  # Branch
        self.expired_table.setColumnWidth(3, 100)  # Batch
        self.expired_table.setColumnWidth(4, 120)  # Expired Date
        self.expired_table.setColumnWidth(5, 130)  # Days Overdue
        self.expired_table.setColumnWidth(6, 80)  # SKU
        self.expired_table.setColumnWidth(7, 120)  # Rack Location

        self.expired_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.expired_table.setAlternatingRowColors(True)
        self.expired_table.setSortingEnabled(True)

        layout.addWidget(self.expired_table)

        # Bottom buttons
        button_layout = QHBoxLayout()

        export_btn = QPushButton("📤 Export to Excel")
        export_btn.clicked.connect(self.export_expired_products)

        close_btn = QPushButton("✖️ Close")
        close_btn.clicked.connect(self.accept)

        button_layout.addStretch()
        button_layout.addWidget(export_btn)
        button_layout.addWidget(close_btn)

        layout.addLayout(button_layout)
        self.setLayout(layout)

        # Load data on initialization
        self.load_expired_products()

    def load_expired_products(self):
        """Load expired products from database"""
        try:
            import sqlite3
            from datetime import datetime

            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Query for expired products
            cursor.execute("""
                SELECT 
                    p.product_name,
                    p.owner_id,
                    b.branch_name,
                    p.batch,
                    p.expired_date,
                    p.sku,
                    rl.rack_location_name,  -- Correct column name
                    p.status,
                    julianday('now') - julianday(p.expired_date) as days_overdue
                FROM products p
                LEFT JOIN branches b ON p.branch_id = b.branch_id
                LEFT JOIN racklocations rl ON p.rack_location_id = rl.rack_location_id  -- Correct table name
                WHERE p.expired_date < date('now')
                    AND p.status = 'approved'
                ORDER BY p.expired_date DESC
                LIMIT 100
            """)

            results = cursor.fetchall()
            conn.close()

            # Clear existing data
            self.expired_table.setRowCount(0)

            if not results:
                # Show message if no expired products
                self.expired_table.setRowCount(1)
                no_data_item = QTableWidgetItem("No expired products found")
                no_data_item.setTextAlignment(Qt.AlignCenter)
                self.expired_table.setItem(0, 0, no_data_item)
                self.expired_table.setSpan(0, 0, 1, 9)
                return

            # Populate table
            self.expired_table.setRowCount(len(results))

            for row, data in enumerate(results):
                product_name, owner, branch, batch, expired_date, sku, rack_location, status, days_overdue = data

                # Safe string conversion
                def safe_str(value):
                    return str(value) if value is not None else "N/A"

                self.expired_table.setItem(row, 0, QTableWidgetItem(safe_str(product_name)))
                self.expired_table.setItem(row, 1, QTableWidgetItem(safe_str(owner)))
                self.expired_table.setItem(row, 2, QTableWidgetItem(safe_str(branch)))
                self.expired_table.setItem(row, 3, QTableWidgetItem(safe_str(batch)))
                self.expired_table.setItem(row, 4, QTableWidgetItem(safe_str(expired_date)))

                # Days overdue with color coding
                days_item = QTableWidgetItem(f"{int(days_overdue)} days")
                if days_overdue > 30:
                    days_item.setBackground(QColor("#FFCDD2"))  # Light red
                elif days_overdue > 7:
                    days_item.setBackground(QColor("#FFE0B2"))  # Light orange
                days_item.setTextAlignment(Qt.AlignCenter)
                self.expired_table.setItem(row, 5, days_item)

                self.expired_table.setItem(row, 6, QTableWidgetItem(safe_str(sku)))
                self.expired_table.setItem(row, 7, QTableWidgetItem(safe_str(rack_location)))

                # Status with color
                status_item = QTableWidgetItem(safe_str(status).upper())
                status_item.setBackground(QColor("#FFCDD2"))  # Light red for expired
                status_item.setTextAlignment(Qt.AlignCenter)
                self.expired_table.setItem(row, 8, status_item)

        except Exception as e:
            print(f"Error loading expired products: {e}")
            QMessageBox.critical(self, "Error", f"Failed to load expired products: {str(e)}")

    def export_expired_products(self):
        """Export expired products to Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from datetime import datetime
            import os

            wb = Workbook()
            ws = wb.active
            ws.title = "Expired Products"

            # Headers
            headers = [
                "Product Name", "Owner", "Branch", "Batch", "Expired Date",
                "Days Overdue", "SKU", "Rack Location", "Status"
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

            # Data
            for row in range(self.expired_table.rowCount()):
                for col in range(self.expired_table.columnCount()):
                    item = self.expired_table.item(row, col)
                    if item:
                        ws.cell(row=row + 2, column=col + 1, value=item.text())

            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"expired_products_{timestamp}.xlsx"
            wb.save(filename)

            QMessageBox.information(self, "Export Complete",
                                    f"Expired products exported to: {filename}")

        except Exception as e:
            print(f"Error exporting expired products: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to export: {str(e)}")


class EmailSettingsDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("📬 Email Settings")
        self.setFixedSize(1200, 900)
        self.setStyleSheet("""
            QDialog {
                background-color: #f0f8f0;
                font-size: 14px;
            }
            QLineEdit, QTextEdit, QComboBox, QSpinBox {
                border: 2px solid #4CAF50;
                border-radius: 5px;
                padding: 8px;
                background-color: white;
                font-size: 14px;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #4CAF50;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
                font-size: 14px;
            }
            QLabel {
                font-size: 14px;
            }
            QCheckBox {
                font-size: 14px;
            }
        """)

        layout = QVBoxLayout()

        # SMTP Settings
        smtp_group = QGroupBox("SMTP Configuration")
        smtp_layout = QFormLayout()

        self.smtp_server = QLineEdit()
        self.smtp_server.setPlaceholderText("smtp.gmail.com")
        smtp_layout.addRow("SMTP Server:", self.smtp_server)

        self.smtp_port = QSpinBox()
        self.smtp_port.setRange(1, 65535)
        self.smtp_port.setValue(587)
        smtp_layout.addRow("Port:", self.smtp_port)

        self.sender_email = QLineEdit()
        self.sender_email.setPlaceholderText("venniscc04@gmail.com")
        smtp_layout.addRow("Sender Email:", self.sender_email)

        self.sender_password = QLineEdit()
        self.sender_password.setEchoMode(QLineEdit.Password)
        self.sender_password.setPlaceholderText("beoj bywk xffl hqoo")
        smtp_layout.addRow("Password:", self.sender_password)

        smtp_group.setLayout(smtp_layout)
        layout.addWidget(smtp_group)

        # Notification Settings
        notification_group = QGroupBox("Notification Schedule")
        notification_layout = QFormLayout()

        self.notify_maturation = QCheckBox("Samples near maturation")
        self.notify_expired = QCheckBox("Expired samples")
        self.notify_assignments = QCheckBox("New task assignments")

        notification_layout.addRow("Enable notifications for:", self.notify_maturation)
        notification_layout.addRow("", self.notify_expired)
        notification_layout.addRow("", self.notify_assignments)

        self.notification_frequency = QComboBox()
        self.notification_frequency.addItems(["Daily", "Weekly", "Monthly"])
        notification_layout.addRow("Frequency:", self.notification_frequency)

        notification_group.setLayout(notification_layout)
        layout.addWidget(notification_group)

        # Email Template
        template_group = QGroupBox("Email Template")
        template_layout = QVBoxLayout()

        self.email_template = QTextEdit()
        self.email_template.setPlainText("""Subject: Medical Sample Status Update

Dear Team,

This is an automated notification about medical sample status updates:

{CONTENT}

Best regards,
Medical Sample Management System""")
        template_layout.addWidget(self.email_template)

        template_group.setLayout(template_layout)
        layout.addWidget(template_group)

        # Buttons
        button_layout = QHBoxLayout()
        save_button = QPushButton("Save Settings")
        save_button.clicked.connect(self.save_settings)
        test_button = QPushButton("Send Test Email")
        test_button.clicked.connect(self.send_test_email)
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(self.reject)

        button_layout.addWidget(save_button)
        button_layout.addWidget(test_button)
        button_layout.addWidget(cancel_button)

        layout.addLayout(button_layout)
        self.setLayout(layout)

    def save_settings(self):
        QMessageBox.information(self, "Settings Saved", "Email settings have been saved successfully!")
        self.accept()

    def send_test_email(self):
        QMessageBox.information(self, "Test Email", "Test email sent successfully!")


class UserManagementDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("👥 User Management")
        self.setFixedSize(700, 500)
        self.setStyleSheet("""
            QDialog {
                background-color: #f0f8f0;
            }
            QTableWidget {
                gridline-color: #C8E6C9;
                background-color: white;
                alternate-background-color: #f5f5f5;
                selection-background-color: #81C784;
            }
            QHeaderView::section {
                background-color: #4CAF50;
                color: white;
                padding: 8px;
                border: 1px solid #45a049;
                font-weight: bold;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)

        layout = QVBoxLayout()

        # User table
        self.user_table = QTableWidget()
        self.user_table.setColumnCount(5)
        self.user_table.setHorizontalHeaderLabels(["ID", "Username", "Role", "Email", "Last Login"])
        self.user_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.user_table)

        # Buttons
        button_layout = QHBoxLayout()
        add_user_btn = QPushButton("Add User")
        add_user_btn.clicked.connect(self.add_user)
        edit_user_btn = QPushButton("Edit User")
        edit_user_btn.clicked.connect(self.edit_user)
        delete_user_btn = QPushButton("Delete User")
        delete_user_btn.clicked.connect(self.delete_user)
        reset_password_btn = QPushButton("Reset Password")
        reset_password_btn.clicked.connect(self.reset_password)

        button_layout.addWidget(add_user_btn)
        button_layout.addWidget(edit_user_btn)
        button_layout.addWidget(delete_user_btn)
        button_layout.addWidget(reset_password_btn)

        layout.addLayout(button_layout)
        self.setLayout(layout)
        self.load_users()

    def load_users(self):
        """Load users from the database"""
        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Load users with their roles
            cursor.execute("""
                SELECT u.user_id, u.username, r.role_name, u.email, u.phone_no
                FROM users u
                LEFT JOIN roles r ON u.role = r.role_id
                ORDER BY u.user_id
            """)

            users = cursor.fetchall()
            conn.close()

            self.user_table.setRowCount(len(users))
            for row, user in enumerate(users):
                for col, value in enumerate(user):
                    item = QTableWidgetItem(str(value) if value is not None else "")
                    self.user_table.setItem(row, col, item)

            print(f"✓ Loaded {len(users)} users from the database.")

        except sqlite3.Error as e:
            print(f"✗ Database error in UserManagementDialog: {e}")
            QMessageBox.critical(self, "Database Error", f"Could not load users: {e}")
            self.user_table.setRowCount(0)

    def add_user(self):
        QMessageBox.information(self, "Add User", "Add user functionality will be implemented here.")

    def edit_user(self):
        QMessageBox.information(self, "Edit User", "Edit user functionality will be implemented here.")

    def delete_user(self):
        QMessageBox.information(self, "Delete User", "Delete user functionality will be implemented here.")

    def reset_password(self):
        QMessageBox.information(self, "Reset Password", "Password reset functionality will be implemented here.")


class SystemSettingsDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("⚙️ System Settings")
        self.setFixedSize(500, 400)
        self.setStyleSheet("""
            QDialog {
                background-color: #f0f8f0;
            }
            QLineEdit, QSpinBox, QComboBox {
                border: 2px solid #4CAF50;
                border-radius: 5px;
                padding: 8px;
                background-color: white;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #4CAF50;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
        """)

        layout = QVBoxLayout()

        # Expiration Rules
        expiration_group = QGroupBox("Expiration Rules")
        expiration_layout = QFormLayout()

        self.default_shelf_life = QSpinBox()
        self.default_shelf_life.setRange(1, 3650)
        self.default_shelf_life.setValue(365)
        self.default_shelf_life.setSuffix(" days")
        expiration_layout.addRow("Default Shelf Life:", self.default_shelf_life)

        self.warning_period = QSpinBox()
        self.warning_period.setRange(1, 365)
        self.warning_period.setValue(60)
        self.warning_period.setSuffix(" days")
        expiration_layout.addRow("Warning Period:", self.warning_period)

        expiration_group.setLayout(expiration_layout)
        layout.addWidget(expiration_group)

        # System Parameters
        system_group = QGroupBox("System Parameters")
        system_layout = QFormLayout()

        self.max_login_attempts = QSpinBox()
        self.max_login_attempts.setRange(1, 10)
        self.max_login_attempts.setValue(3)
        system_layout.addRow("Max Login Attempts:", self.max_login_attempts)

        self.session_timeout = QSpinBox()
        self.session_timeout.setRange(5, 480)
        self.session_timeout.setValue(60)
        self.session_timeout.setSuffix(" minutes")
        system_layout.addRow("Session Timeout:", self.session_timeout)

        self.backup_frequency = QComboBox()
        self.backup_frequency.addItems(["Daily", "Weekly", "Monthly"])
        system_layout.addRow("Backup Frequency:", self.backup_frequency)

        system_group.setLayout(system_layout)
        layout.addWidget(system_group)

        # Buttons
        button_layout = QHBoxLayout()
        save_button = QPushButton("Save Settings")
        save_button.clicked.connect(self.save_settings)
        reset_button = QPushButton("Reset to Defaults")
        reset_button.clicked.connect(self.reset_defaults)
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(self.reject)

        button_layout.addWidget(save_button)
        button_layout.addWidget(reset_button)
        button_layout.addWidget(cancel_button)

        layout.addLayout(button_layout)
        self.setLayout(layout)

    def save_settings(self):
        QMessageBox.information(self, "Settings Saved", "System settings have been saved successfully!")
        self.accept()

    def reset_defaults(self):
        self.default_shelf_life.setValue(365)
        self.warning_period.setValue(60)
        self.max_login_attempts.setValue(3)
        self.session_timeout.setValue(60)
        self.backup_frequency.setCurrentIndex(0)


class ExportReportsDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("📤 Export Full Reports")
        self.setFixedSize(450, 350)
        self.setStyleSheet("""
            QDialog {
                background-color: #f0f8f0;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
                text-align: left;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #4CAF50;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
        """)

        layout = QVBoxLayout()

        # Report Types
        reports_group = QGroupBox("Available Reports")
        reports_layout = QVBoxLayout()

        shelf_life_btn = QPushButton("📊 Shelf-Life Report (Excel)")
        shelf_life_btn.clicked.connect(lambda: self.export_report("shelf_life", "excel"))
        reports_layout.addWidget(shelf_life_btn)

        inventory_excel_btn = QPushButton("📦 Inventory Data (Excel)")
        inventory_excel_btn.clicked.connect(lambda: self.export_report("inventory", "excel"))
        reports_layout.addWidget(inventory_excel_btn)

        inventory_pdf_btn = QPushButton("📦 Inventory Report (PDF)")
        inventory_pdf_btn.clicked.connect(lambda: self.export_report("inventory", "pdf"))
        reports_layout.addWidget(inventory_pdf_btn)

        activity_excel_btn = QPushButton("📋 User Activity (Excel)")
        activity_excel_btn.clicked.connect(lambda: self.export_report("activity", "excel"))
        reports_layout.addWidget(activity_excel_btn)

        activity_pdf_btn = QPushButton("📋 Activity Report (PDF)")
        activity_pdf_btn.clicked.connect(lambda: self.export_report("activity", "pdf"))
        reports_layout.addWidget(activity_pdf_btn)

        all_data_btn = QPushButton("🗂️ Complete System Export (Excel)")
        all_data_btn.clicked.connect(lambda: self.export_report("complete", "excel"))
        reports_layout.addWidget(all_data_btn)

        reports_group.setLayout(reports_layout)
        layout.addWidget(reports_group)

        # Close button
        close_button = QPushButton("Close")
        close_button.clicked.connect(self.accept)
        layout.addWidget(close_button)

        self.setLayout(layout)

    def export_report(self, report_type, format_type):
        QMessageBox.information(self, "Export Started",
                                f"Exporting {report_type} report as {format_type.upper()}...")


class ProductDetailDialog(QDialog):
    """Dialog for showing complete product details. If *allow_recall* is False,
    the "Recall Product" button will be hidden (used e.g. from Pending page)."""

    def __init__(self, product_data, parent=None, allow_recall: bool = True):
        super().__init__(parent)
        self.product_data = product_data
        self.parent_dashboard = parent
        self.allow_recall = allow_recall
        self.setWindowTitle("Product Details")
        self.setFixedSize(800, 700)
        self.setModal(True)

        # Set dialog style
        self.setStyleSheet("""
            QDialog {
                background-color: #F5F5F5;
                border: 2px solid #4CAF50;
                border-radius: 15px;
            }
        """)

        self.init_ui()

    def init_ui(self):
        """Initialize the user interface"""
        layout = QVBoxLayout()
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(20)

        # Title
        title_label = QLabel("📦 Product Information")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2E7D32;
                margin-bottom: 10px;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # Product image display under the "Product Information" label
        product_img_path = self.product_data.get('product_image', '')
        product_img_label = QLabel()
        product_img_label.setAlignment(Qt.AlignCenter)
        if product_img_path and os.path.isfile(product_img_path):
            pixmap = QPixmap(product_img_path)
            if not pixmap.isNull():
                pixmap = pixmap.scaled(250, 250, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                product_img_label.setPixmap(pixmap)
        else:
            product_img_label.setText("No Product Image")
            product_img_label.setStyleSheet("color: #757575; font-style: italic;")
        layout.addWidget(product_img_label)

        # Create scrollable area for product details
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("""
            QScrollArea {
                border: 1px solid #E0E0E0;
                border-radius: 8px;
                background-color: white;
            }
        """)

        # Product details content
        details_widget = QWidget()
        details_layout = QVBoxLayout()
        details_layout.setContentsMargins(20, 20, 20, 20)

        # Product information grid
        info_grid = QGridLayout()
        info_grid.setSpacing(15)
        info_grid.setHorizontalSpacing(20)
        info_grid.setVerticalSpacing(12)

        # Row 1: Product ID and Manufacture Date
        id_label = QLabel("Product ID:")
        id_label.setFont(QFont("Arial", 12, QFont.Bold))
        id_label.setStyleSheet("color: #000000; padding: 4px;")
        id_value = QLabel(str(self.product_data.get('product_id', 'N/A')))
        id_value.setFont(QFont("Arial", 12))
        id_value.setStyleSheet("color: #333333; padding: 4px;")

        mfg_label = QLabel("Manufacture Date:")
        mfg_label.setFont(QFont("Arial", 12, QFont.Bold))
        mfg_label.setStyleSheet("color: #000000; padding: 4px;")
        mfg_value = QLabel(str(self.product_data.get('manufacture_date', 'N/A')))
        mfg_value.setFont(QFont("Arial", 12))
        mfg_value.setStyleSheet("color: #333333; padding: 4px;")

        info_grid.addWidget(id_label, 0, 0)
        info_grid.addWidget(id_value, 0, 1)
        info_grid.addWidget(mfg_label, 0, 2)
        info_grid.addWidget(mfg_value, 0, 3)

        # Row 2: Branch and Expiry Date
        branch_label = QLabel("Branch:")
        branch_label.setFont(QFont("Arial", 12, QFont.Bold))
        branch_label.setStyleSheet("color: #000000; padding: 4px;")
        # Get branch info from owner or use default
        branch_value = QLabel(str(self.product_data.get('branch', 'Main Branch- Bayan Lepas')))
        branch_value.setFont(QFont("Arial", 12))
        branch_value.setStyleSheet("color: #333333; padding: 4px;")

        expiry_label = QLabel("Expiry Date:")
        expiry_label.setFont(QFont("Arial", 12, QFont.Bold))
        expiry_label.setStyleSheet("color: #000000; padding: 4px;")
        expiry_value = QLabel(str(self.product_data.get('expired_date', 'N/A')))
        expiry_value.setFont(QFont("Arial", 12))
        expiry_value.setStyleSheet("color: #333333; padding: 4px;")

        info_grid.addWidget(branch_label, 1, 0)
        info_grid.addWidget(branch_value, 1, 1)
        info_grid.addWidget(expiry_label, 1, 2)
        info_grid.addWidget(expiry_value, 1, 3)

        # Row 3: SKU, Rack Location, Batch
        sku_label = QLabel("SKU:")
        sku_label.setFont(QFont("Arial", 12, QFont.Bold))
        sku_label.setStyleSheet("color: #000000; padding: 4px;")
        sku_value = QLabel(str(self.product_data.get('sku', self.product_data.get('product_id', 'N/A'))))
        sku_value.setFont(QFont("Arial", 12))
        sku_value.setStyleSheet("color: #333333; padding: 4px;")

        rack_label = QLabel("Rack Location:")
        rack_label.setFont(QFont("Arial", 12, QFont.Bold))
        rack_label.setStyleSheet("color: #000000; padding: 4px;")
        rack_value = QLabel(str(self.product_data.get('rack_location', 'Unassigned')))
        rack_value.setFont(QFont("Arial", 12))
        rack_value.setStyleSheet("color: #333333; padding: 4px;")

        info_grid.addWidget(sku_label, 2, 0)
        info_grid.addWidget(sku_value, 2, 1)
        info_grid.addWidget(rack_label, 2, 2)
        info_grid.addWidget(rack_value, 2, 3)

        # Row 4: Batch and Barcode
        batch_label = QLabel("Batch:")
        batch_label.setFont(QFont("Arial", 12, QFont.Bold))
        batch_label.setStyleSheet("color: #000000; padding: 4px;")
        batch_value = QLabel(str(self.product_data.get('batch', 'N/A')))
        batch_value.setFont(QFont("Arial", 12))
        batch_value.setStyleSheet("color: #333333; padding: 4px;")

        barcode_label = QLabel("Barcode:")
        barcode_label.setFont(QFont("Arial", 12, QFont.Bold))
        barcode_label.setStyleSheet("color: #000000; padding: 4px;")
        barcode_text = str(self.product_data.get('barcode', ''))
        if barcode_text and barcode_text != 'N/A':
            barcode_value = QLabel(barcode_text)
            barcode_value.setFont(QFont("Courier", 11, QFont.Bold))
            barcode_value.setStyleSheet("color: #333333; padding: 4px;")
        else:
            barcode_value = QLabel("No Barcode")
            barcode_value.setFont(QFont("Arial", 11))
            barcode_value.setStyleSheet("color: #999999; padding: 4px; font-style: italic;")

        info_grid.addWidget(batch_label, 3, 0)
        info_grid.addWidget(batch_value, 3, 1)
        info_grid.addWidget(barcode_label, 3, 2)
        info_grid.addWidget(barcode_value, 3, 3)

        # Barcode image display directly under the barcode number
        barcode_img_path = self.product_data.get('barcode_image', '')
        barcode_text = str(self.product_data.get('barcode', ''))

        possible_paths = []
        if barcode_img_path:
            # exact path stored
            possible_paths.append(barcode_img_path)
            # same name inside barcodes folder
            possible_paths.append(os.path.join('barcodes', os.path.basename(barcode_img_path)))

        # derive from barcode number if available
        if barcode_text and barcode_text not in ('N/A', ''):
            possible_paths.extend([
                f'barcode_{barcode_text}.png',
                f'barcode_{barcode_text}.png.png',
                os.path.join('barcodes', f'{barcode_text}.png'),
                os.path.join('barcodes', f'{barcode_text}.png.png'),
                os.path.join('barcodes', f'barcode_{barcode_text}.png'),
                os.path.join('barcodes', f'barcode_{barcode_text}.png.png')
            ])

        # include SKU prefix variants
        sku_value = str(self.product_data.get('sku', '')).strip()
        if sku_value and sku_value not in ('N/A', '') and barcode_text and barcode_text not in ('N/A', ''):
            possible_paths.extend([
                f'{sku_value}_{barcode_text}.png',
                f'{sku_value}_{barcode_text}.png.png',
                os.path.join('barcodes', f'{sku_value}_{barcode_text}.png'),
                os.path.join('barcodes', f'{sku_value}_{barcode_text}.png.png')
            ])

        resolved_barcode_img = next((p for p in possible_paths if os.path.isfile(p)), '')

        if resolved_barcode_img:
            barcode_img_label = QLabel()
            barcode_pixmap = QPixmap(resolved_barcode_img)
            if not barcode_pixmap.isNull():
                barcode_pixmap = barcode_pixmap.scaled(180, 60, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                barcode_img_label.setPixmap(barcode_pixmap)
                barcode_img_label.setAlignment(Qt.AlignCenter)
                # Place the image spanning full grid width beneath barcode row
                info_grid.addWidget(barcode_img_label, 4, 0, 1, 4)
        else:
            print(f'[INFO] Barcode image not found for: {barcode_text} (searched {len(possible_paths)} paths)')

        details_layout.addLayout(info_grid)

        # Calculate days left until expiry
        expired_str = self.product_data.get('expired_date', '')
        days_left = 0
        if expired_str and expired_str not in ('N/A', ''):
            try:
                # Extract only date component if datetime string
                date_part = expired_str.split(' ')[0]
                exp_dt = datetime.strptime(date_part, '%Y-%m-%d')
                days_left = (exp_dt.date() - datetime.now().date()).days
            except Exception:
                pass  # keep default 0 if parsing fails

        # Days left (prominent display)
        if days_left < 0:
            days_text = f"⚠️ EXPIRED ({abs(days_left)} days ago)"
            days_color = "#D32F2F"
            days_bg = "#FFCDD2"
        elif days_left == 0:
            days_text = "⚠️ EXPIRES TODAY"
            days_color = "#D32F2F"
            days_bg = "#FFCDD2"
        elif days_left <= 7:
            days_text = f"🔴 {days_left} days left"
            days_color = "#D32F2F"
            days_bg = "#FFEBEE"
        elif days_left <= 15:
            days_text = f"🟠 {days_left} days left"
            days_color = "#F57C00"
            days_bg = "#FFF3E0"
        elif days_left <= 30:
            days_text = f"🟡 {days_left} days left"
            days_color = "#FBC02D"
            days_bg = "#FFFDE7"
        else:
            days_text = f"🟢 {days_left} days left"
            days_color = "#388E3C"
            days_bg = "#F1F8E9"

        days_label = QLabel(days_text)
        days_label.setFont(QFont("Arial", 14, QFont.Bold))  # smaller font
        days_label.setStyleSheet(f"""
            color: {days_color}; 
            background-color: {days_bg}; 
            padding: 8px 16px; 
            border-radius: 6px; 
            border: 1px solid {days_color};
            margin-top: 8px;
        """)
        days_label.setAlignment(Qt.AlignCenter)
        details_layout.addWidget(days_label)

        details_layout.addStretch()

        # Set the layout on the details widget and add to scroll area
        details_widget.setLayout(details_layout)
        scroll.setWidget(details_widget)
        layout.addWidget(scroll)

        # Buttons
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)

        # Close button
        close_btn = QPushButton("✕ Close")
        close_btn.setFixedSize(120, 40)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #757575;
                color: white;
                border: none;
                border-radius: 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #616161; }
        """)
        close_btn.clicked.connect(self.accept)

        # Recall product button (optional)
        if self.allow_recall:
            recall_btn = QPushButton("⚠️ Recall Product")
            recall_btn.setFixedSize(150, 40)
            recall_btn.setStyleSheet("""
                QPushButton {
                    background-color: #F44336;
                    color: white;
                    border: none;
                    border-radius: 20px;
                    font-size: 14px;
                    font-weight: bold;
                }
                QPushButton:hover { background-color: #D32F2F; }
            """)
            recall_btn.clicked.connect(self.show_recall_dialog)

        button_layout.addStretch()
        button_layout.addWidget(close_btn)
        if self.allow_recall:
            button_layout.addWidget(recall_btn)
        button_layout.addStretch()

        layout.addLayout(button_layout)
        self.setLayout(layout)

    def add_field_to_grid(self, grid, label_text, value_text, row, is_recall=False):
        """Add a field to the grid layout"""
        label = QLabel(label_text)
        label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                color: #424242;
                min-width: 150px;
            }
        """)

        value = QLabel(value_text)
        if is_recall:
            value.setStyleSheet("""
                QLabel {
                    color: #D32F2F;
                    font-weight: bold;
                    background-color: #FFEBEE;
                    padding: 5px;
                    border-radius: 4px;
                    border: 1px solid #FFCDD2;
                }
            """)
        else:
            value.setStyleSheet("""
                QLabel {
                    color: #212121;
                    background-color: #F8F9FA;
                    padding: 5px;
                    border-radius: 4px;
                    border: 1px solid #E9ECEF;
                }
            """)

        grid.addWidget(label, row, 0)
        grid.addWidget(value, row, 1)

    def show_recall_dialog(self):
        """Show recall reason selection dialog"""
        recall_dialog = QDialog(self)
        recall_dialog.setWindowTitle("Product Recall")
        recall_dialog.setFixedSize(400, 250)
        recall_dialog.setModal(True)

        recall_dialog.setStyleSheet("""
            QDialog {
                background-color: #FFF3E0;
                border: 2px solid #FF9800;
                border-radius: 10px;
            }
        """)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        # Warning message
        warning_label = QLabel("⚠️ Warning: Product Recall")
        warning_label.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #E65100;
                margin-bottom: 10px;
            }
        """)
        warning_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(warning_label)

        # Instruction
        instruction_label = QLabel("Please select the reason for recalling this product:")
        instruction_label.setStyleSheet("color: #BF360C; font-weight: bold; margin-bottom: 10px;")
        layout.addWidget(instruction_label)

        # Reason dropdown
        self.reason_combo = QComboBox()
        self.reason_combo.addItems([
            "Labeling errors",
            "Expired medication",
            "Contamination issues",
            "Defective product",
            "Others"
        ])
        self.reason_combo.setStyleSheet("""
            QComboBox {
                border: 2px solid #FF9800;
                border-radius: 5px;
                padding: 8px;
                font-size: 14px;
                background-color: white;
            }
            QComboBox:hover {
                border-color: #F57C00;
            }
        """)
        layout.addWidget(self.reason_combo)

        # Buttons
        button_layout = QHBoxLayout()

        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #757575;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #616161; }
        """)
        cancel_btn.clicked.connect(recall_dialog.reject)

        confirm_btn = QPushButton("Confirm Recall")
        confirm_btn.setStyleSheet("""
            QPushButton {
                background-color: #F44336;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #D32F2F; }
        """)
        confirm_btn.clicked.connect(lambda: self.process_recall(recall_dialog))

        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(confirm_btn)
        layout.addLayout(button_layout)

        recall_dialog.setLayout(layout)
        recall_dialog.exec_()

    def process_recall(self, dialog):
        """Process the product recall"""
        reason = self.reason_combo.currentText()
        product_id = self.product_data[0]

        try:
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Update product with recall reason
            cursor.execute("""
                UPDATE products 
                SET rejection_comment = ?, status = 'Recalled'
                WHERE product_id = ?
            """, (reason, product_id))

            conn.commit()
            conn.close()

            dialog.accept()
            QMessageBox.information(self, "Recall Successful",
                                    f"Product has been recalled.\nReason: {reason}")

            # Refresh the parent dashboard
            if self.parent_dashboard:
                self.parent_dashboard.load_inventory_data()

            self.accept()

        except Exception as e:
            print(f"✗ Error processing recall: {e}")
            QMessageBox.critical(self, "Recall Error", f"Failed to process recall: {str(e)}")


# === Add Product Dialog ===
class AddProductDialog(QDialog):
    """Dialog for adding a new product (owner or admin view)."""

    product_added = pyqtSignal()  # emitted after successful save

    def __init__(self, parent_dashboard=None, user_role='Admin', current_user_id=None, current_user_branch_id=None):
        super().__init__(parent=parent_dashboard)
        self.dashboard = parent_dashboard  # reference to Dashboard for helpers
        self.user_role = user_role.lower()
        self.current_user_id = current_user_id or 0
        self.current_user_branch_id = current_user_branch_id
        self.setWindowTitle("Add Product")
        self.setFixedSize(700, 820)
        self.setModal(True)

        self._batch_sequence = 0  # will load lazily
        self._last_batch_date = datetime.now().date()
        self.uploaded_image_path = None
        self.photo_uploaded = False

        self.init_ui()

    # --------------------------- UI helpers ---------------------------
    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)

        title = QLabel("➕ Add Product")
        title.setFont(QFont("Arial", 20, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color:#2E7D32;")
        main_layout.addWidget(title)

        form_layout = QGridLayout()
        form_layout.setHorizontalSpacing(15)
        form_layout.setVerticalSpacing(10)

        # Product name
        form_layout.addWidget(QLabel("Product Name:"), 0, 0)
        self.product_name_edit = QLineEdit()
        form_layout.addWidget(self.product_name_edit, 0, 1, 1, 2)

        # Description
        form_layout.addWidget(QLabel("Description:"), 1, 0)
        self.desc_edit = QTextEdit()
        self.desc_edit.setFixedHeight(80)
        form_layout.addWidget(self.desc_edit, 1, 1, 1, 2)

        # Image upload
        form_layout.addWidget(QLabel("Product Image:"), 2, 0)
        img_hbox = QHBoxLayout()
        self.img_label = QLabel("No Image")
        self.img_label.setFixedSize(140, 140)
        self.img_label.setStyleSheet("background:#eee;border:1px solid #ccc;border-radius:6px;color:#777;")
        self.img_label.setAlignment(Qt.AlignCenter)
        img_hbox.addWidget(self.img_label)
        upload_btn = QPushButton("Upload")
        upload_btn.clicked.connect(self.upload_image)
        img_hbox.addWidget(upload_btn)
        form_layout.addLayout(img_hbox, 2, 1, 1, 2)

        # SKU
        form_layout.addWidget(QLabel("SKU:"), 3, 0)
        self.sku_edit = QLineEdit()
        form_layout.addWidget(self.sku_edit, 3, 1)

        # Arrival / Manufacture / Expiry dates
        self.arrival_date_edit = self.create_date_edit()
        self.mfg_date_edit = self.create_date_edit()
        self.exp_date_edit = self.create_date_edit()

        form_layout.addWidget(QLabel("Arrival Date:"), 4, 0)
        form_layout.addWidget(self.arrival_date_edit, 4, 1)
        form_layout.addWidget(QLabel("Manufacture Date:"), 5, 0)
        form_layout.addWidget(self.mfg_date_edit, 5, 1)
        form_layout.addWidget(QLabel("Expiry Date:"), 6, 0)
        form_layout.addWidget(self.exp_date_edit, 6, 1)

        # Rack location
        form_layout.addWidget(QLabel("Rack Location:"), 7, 0)
        self.rack_combo = QComboBox()
        form_layout.addWidget(self.rack_combo, 7, 1)
        self.load_available_rack_locations()

        # Admin-only: owner & tester assignment
        if self.user_role == 'admin' or self.user_role == 'superadmin':
            form_layout.addWidget(QLabel("Assign Owner:"), 8, 0)
            self.owner_combo = QComboBox()
            form_layout.addWidget(self.owner_combo, 8, 1)
            form_layout.addWidget(QLabel("Assign Tester:"), 9, 0)
            self.tester_combo = QComboBox()
            form_layout.addWidget(self.tester_combo, 9, 1)
            self.load_owner_tester_lists()
        else:
            self.owner_combo = None
            self.tester_combo = None

        # Batch (read-only)
        form_layout.addWidget(QLabel("Batch Code:"), 10, 0)
        self.batch_edit = QLineEdit()
        self.batch_edit.setReadOnly(True)
        form_layout.addWidget(self.batch_edit, 10, 1)
        self.update_batch_code()

        main_layout.addLayout(form_layout)

        # Save/Cancel buttons
        btn_box = QHBoxLayout()
        btn_box.addStretch()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        save_btn = QPushButton("Save")
        save_btn.clicked.connect(self.save_product)
        for b in (cancel_btn, save_btn):
            b.setStyleSheet(
                """QPushButton{background:#4CAF50;color:white;padding:8px 20px;border:none;border-radius:6px;font-weight:bold;}QPushButton:hover{background:#388E3C;}""")
        btn_box.addWidget(cancel_btn)
        btn_box.addWidget(save_btn)
        btn_box.addStretch()
        main_layout.addLayout(btn_box)

    def create_date_edit(self):
        de = QDateEdit()
        de.setCalendarPopup(True)
        de.setDisplayFormat("yyyy-MM-dd")
        de.setDate(QDate.currentDate())
        return de

    # --------------------------- Data helpers ---------------------------
    def load_available_rack_locations(self):
        self.rack_combo.clear()
        self.rack_combo.addItem("Select", None)
        try:
            conn = sqlite3.connect("testing_system.db")
            cur = conn.cursor()
            cur.execute("""
                SELECT rack_location_id, rack_location_name FROM racklocations
                WHERE rack_location_id NOT IN (SELECT DISTINCT rack_location_id FROM products WHERE rack_location_id IS NOT NULL)
                ORDER BY rack_location_name
            """)
            for rid, name in cur.fetchall():
                self.rack_combo.addItem(name, rid)
            conn.close()
        except Exception as e:
            print(f"Error loading racks: {e}")

    def load_owner_tester_lists(self):
        # owners (role_id 3) testers (role_id 4)
        self.owner_combo.clear();
        self.tester_combo.clear()
        self.owner_combo.addItem("Select", None)
        self.tester_combo.addItem("Select", None)
        try:
            conn = sqlite3.connect("testing_system.db");
            cur = conn.cursor()
            cur.execute(
                "SELECT user_id, fullname, role_name FROM users u JOIN roles r ON u.role = r.role_id WHERE r.role_id=3 ORDER BY fullname")
            for uid, fname, role_name in cur.fetchall():
                display = f"{fname} ({role_name.title()})" if fname else f"User{uid} ({role_name.title()})"
                self.owner_combo.addItem(display, uid)
            cur.execute(
                "SELECT user_id, fullname, role_name FROM users u JOIN roles r ON u.role = r.role_id WHERE r.role_id=4 ORDER BY fullname")
            for uid, fname, role_name in cur.fetchall():
                display = f"{fname} ({role_name.title()})" if fname else f"User{uid} ({role_name.title()})"
                self.tester_combo.addItem(display, uid)
            conn.close()
        except Exception as e:
            print(f"Error loading users list: {e}")

    def load_batch_sequence(self):
        # similar to ksOwner logic
        try:
            conn = sqlite3.connect("testing_system.db")
            cur = conn.cursor()
            cur.execute(
                "CREATE TABLE IF NOT EXISTS BatchSequence (date TEXT PRIMARY KEY, last_sequence INTEGER DEFAULT 0)")
            today = datetime.now().strftime("%Y-%m-%d")
            cur.execute("SELECT last_sequence FROM BatchSequence WHERE date=?", (today,))
            res = cur.fetchone()
            conn.close()
            return res[0] if res else 0
        except Exception:
            return 0

    def save_batch_sequence(self):
        try:
            conn = sqlite3.connect("testing_system.db");
            cur = conn.cursor()
            today = datetime.now().strftime("%Y-%m-%d")
            cur.execute("INSERT OR REPLACE INTO BatchSequence (date,last_sequence) VALUES (?,?)",
                        (today, self._batch_sequence))
            conn.commit();
            conn.close()
        except Exception as e:
            print(f"Error saving batch seq: {e}")

    def update_batch_code(self):
        today = datetime.now().date()
        if today != self._last_batch_date:
            self._batch_sequence = 0
            self._last_batch_date = today
        else:
            self._batch_sequence = self.load_batch_sequence()
        batch_letter = chr(ord('A') + self._batch_sequence)
        code = f"B{today.strftime('%Y%m%d')}{batch_letter}"
        self.batch_edit.setText(code)

    # --------------------------- actions ---------------------------
    def upload_image(self):
        fname, _ = QFileDialog.getOpenFileName(self, "Select Image", "", "Images (*.png *.jpg *.jpeg *.bmp)")
        if fname:
            pix = QPixmap(fname)
            if not pix.isNull():
                self.img_label.setPixmap(pix.scaled(140, 140, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                self.img_label.setText("")
                self.uploaded_image_path = fname
                self.photo_uploaded = True

    def validate(self):
        if not self.product_name_edit.text().strip():
            QMessageBox.warning(self, "Validation", "Product name required");
            return False
        if not self.desc_edit.toPlainText().strip():
            QMessageBox.warning(self, "Validation", "Description required");
            return False
        if not self.photo_uploaded:
            QMessageBox.warning(self, "Validation", "Please upload product image");
            return False
        if not self.sku_edit.text().strip():
            QMessageBox.warning(self, "Validation", "SKU required");
            return False
        if self.rack_combo.currentData() is None:
            QMessageBox.warning(self, "Validation", "Select rack location");
            return False
        if self.user_role in ('admin', 'superadmin'):
            if self.owner_combo.currentData() is None or self.tester_combo.currentData() is None:
                QMessageBox.warning(self, "Validation", "Select owner and tester");
                return False
        return True

    def save_product(self):
        if not self.validate():
            return
        # generate batch letter increment
        self._batch_sequence += 1
        self.save_batch_sequence()
        batch_code = self.batch_edit.text()

        prod = {
            'owner_id': self.owner_combo.currentData() if self.owner_combo else self.current_user_id,
            'tester_id': self.tester_combo.currentData() if self.tester_combo else None,
            'product_name': self.product_name_edit.text().strip(),
            'product_desc': self.desc_edit.toPlainText().strip(),
            'product_image': self.uploaded_image_path,
            'arrival_date': self.arrival_date_edit.date().toString('yyyy-MM-dd'),
            'manufacture_date': self.mfg_date_edit.date().toString('yyyy-MM-dd'),
            'expired_date': self.exp_date_edit.date().toString('yyyy-MM-dd'),
            'rack_location_id': self.rack_combo.currentData(),
            'batch': batch_code,
            'sku': self.sku_edit.text().strip()
        }
        try:
            conn = sqlite3.connect("testing_system.db");
            cur = conn.cursor()
            status_value = 'approved' if self.user_role in ('admin', 'superadmin') else 'pending'

            # Get owner's branch_id
            owner_branch_id = None
            if prod['owner_id']:
                # If admin/superadmin assigns owner, get that owner's branch_id
                cur.execute("SELECT branch_id FROM users WHERE user_id = ?", (prod['owner_id'],))
                result = cur.fetchone()
                owner_branch_id = result[0] if result else None
            else:
                # For non-admin users (owner role), use current user's branch_id
                owner_branch_id = self.current_user_branch_id

            cur.execute("""
                INSERT INTO products (owner_id, tester_id, product_name, product_desc, product_image, arrival_date,
                    manufacture_date, expired_date, location, rack_location_id, batch, sku, status, branch_id)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                prod['owner_id'], prod['tester_id'], prod['product_name'], prod['product_desc'], prod['product_image'],
                prod['arrival_date'], prod['manufacture_date'], prod['expired_date'], self.rack_combo.currentText(),
                prod['rack_location_id'], prod['batch'], prod['sku'], status_value, owner_branch_id
            ))
            product_id = cur.lastrowid

            # -------------------------------------------------------------
            # Auto-generate unit barcodes based on SKU (numeric quantity)
            # -------------------------------------------------------------
            try:
                unit_qty = int(prod['sku']) if prod['sku'] else 0
            except ValueError:
                unit_qty = 0

            if unit_qty > 0:
                unit_barcodes = generate_unit_barcodes(product_id, prod['batch'], unit_qty)
                if unit_barcodes:
                    insert_unit_barcodes_to_db(unit_barcodes)
                    generate_barcode_images(unit_barcodes)

            # If admin, generate barcode immediately
            if self.user_role in ('admin', 'superadmin'):
                try:
                    import random, string
                    import barcode
                    from barcode.writer import ImageWriter
                    code_val = ''.join(random.choices(string.digits, k=12))
                    barcode_class = barcode.get_barcode_class('code128')
                    code128 = barcode_class(code_val, writer=ImageWriter())
                    os.makedirs('barcodes', exist_ok=True)
                    file_base = os.path.join('barcodes', f"{prod['sku']}_{code_val}")
                    code128.save(file_base)
                    img_path = f"{file_base}.png"
                    cur.execute("UPDATE products SET barcode=?, barcode_image=? WHERE product_id=?",
                                (code_val, img_path, product_id))
                    conn.commit()
                except Exception as be:
                    print(f"Barcode generation failed: {be}")

            # If admin assigns tester, create assignment record
            if self.user_role in ('admin', 'superadmin') and prod['tester_id']:
                cur.execute("""
                    INSERT INTO product_tester_assignments (product_id, tester_id, assigned_by)
                    VALUES (?,?,?)
                """, (product_id, prod['tester_id'], self.current_user_id))
            conn.commit()
            conn.close()

            # === EMAIL NOTIFY ===
            if self.user_role in ('admin', 'superadmin') and self.dashboard:
                subj = "New Product Assigned"
                body_tpl = "Dear {fname},\n\nProduct {pname} (Batch {batch}) has been assigned to you for testing.\n\nRegards,\nMedical Testing System"

                def fetch_email_fullname(uid):
                    conn_f = sqlite3.connect("testing_system.db")
                    res = conn_f.execute("SELECT email, fullname FROM users WHERE user_id=?", (uid,)).fetchone()
                    conn_f.close()
                    return res  # (email, fullname)

                # owner
                if prod['owner_id']:
                    res = fetch_email_fullname(prod['owner_id'])
                    if res and res[0]:
                        email, fname = res
                        body = body_tpl.format(fname=fname or 'User', pname=prod['product_name'], batch=prod['batch'])
                        self.dashboard.send_actual_email(email, subj, body, fname)

                # tester
                if prod['tester_id']:
                    res = fetch_email_fullname(prod['tester_id'])
                    if res and res[0]:
                        email, fname = res
                        body = body_tpl.format(fname=fname or 'User', pname=prod['product_name'], batch=prod['batch'])
                        self.dashboard.send_actual_email(email, subj, body, fname)

            # ensure connection closed safely
            try:
                conn.close()
            except Exception:
                pass
            QMessageBox.information(self, "Success", "Product saved successfully.")

            self.product_added.emit()
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save product: {e}")


# === Manage Products Dialog ===
class EditProductDialog(QDialog):
    """Dialog to edit limited product fields."""

    product_updated = pyqtSignal()  # emitted after successful save

    def __init__(self, product_id, parent_dashboard=None):
        super().__init__(parent_dashboard)
        self.dashboard = parent_dashboard
        self.product_id = product_id
        self.uploaded_image_path = None
        self.photo_uploaded = False
        self.setWindowTitle("Edit Product")
        self.setFixedSize(700, 770)
        self.setModal(True)
        self.init_ui()
        self.load_product_data()

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)

        title = QLabel("📝 Edit Product")
        title.setFont(QFont("Arial", 20, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color:#1B5E20;")
        main_layout.addWidget(title)

        form_layout = QGridLayout()
        form_layout.setHorizontalSpacing(15)
        form_layout.setVerticalSpacing(10)

        # Product name
        form_layout.addWidget(QLabel("Product Name:"), 0, 0)
        self.product_name_edit = QLineEdit()
        form_layout.addWidget(self.product_name_edit, 0, 1, 1, 2)

        # Description
        form_layout.addWidget(QLabel("Description:"), 1, 0)
        self.desc_edit = QTextEdit()
        self.desc_edit.setFixedHeight(80)
        form_layout.addWidget(self.desc_edit, 1, 1, 1, 2)

        # Image upload
        form_layout.addWidget(QLabel("Product Image:"), 2, 0)
        img_hbox = QHBoxLayout()
        self.img_label = QLabel("No Image")
        self.img_label.setFixedSize(140, 140)
        self.img_label.setStyleSheet("background:#eee;border:1px solid #ccc;border-radius:6px;color:#777;")
        self.img_label.setAlignment(Qt.AlignCenter)
        img_hbox.addWidget(self.img_label)
        upload_btn = QPushButton("Upload")
        upload_btn.clicked.connect(self.upload_image)
        img_hbox.addWidget(upload_btn)
        form_layout.addLayout(img_hbox, 2, 1, 1, 2)

        # SKU
        form_layout.addWidget(QLabel("SKU:"), 3, 0)
        self.sku_edit = QLineEdit()
        form_layout.addWidget(self.sku_edit, 3, 1)

        # Rack location
        form_layout.addWidget(QLabel("Rack Location:"), 4, 0)
        self.rack_combo = QComboBox()
        form_layout.addWidget(self.rack_combo, 4, 1)

        main_layout.addLayout(form_layout)

        # Buttons
        btn_box = QHBoxLayout()
        btn_box.addStretch()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        save_btn = QPushButton("Save")
        save_btn.clicked.connect(self.save_product)
        for b in (cancel_btn, save_btn):
            b.setStyleSheet(
                """QPushButton{background:#388E3C;color:white;padding:8px 20px;border:none;border-radius:6px;font-weight:bold;}QPushButton:hover{background:#2E7D32;}""")
        btn_box.addWidget(cancel_btn)
        btn_box.addWidget(save_btn)
        btn_box.addStretch()
        main_layout.addLayout(btn_box)

    # --------------------------- Data helpers ---------------------------
    def load_rack_locations(self, current_rack_id):
        self.rack_combo.clear()
        self.rack_combo.addItem("Unassigned", None)
        try:
            conn = sqlite3.connect("testing_system.db")
            cur = conn.cursor()
            cur.execute("SELECT rack_location_id, rack_location_name FROM racklocations ORDER BY rack_location_name")
            for rid, name in cur.fetchall():
                self.rack_combo.addItem(name, rid)
            conn.close()
        except Exception as e:
            print(f"Error loading racks: {e}")
        # set current
        idx = self.rack_combo.findData(current_rack_id)
        if idx >= 0:
            self.rack_combo.setCurrentIndex(idx)

    def load_product_data(self):
        try:
            conn = sqlite3.connect("testing_system.db")
            cur = conn.cursor()
            cur.execute("""
                SELECT product_name, product_desc, product_image, sku, rack_location_id
                FROM products WHERE product_id=?
            """, (self.product_id,))
            row = cur.fetchone()
            conn.close()
            if not row:
                QMessageBox.warning(self, "Not Found", "Product not found in database.")
                self.reject()
                return
            name, desc, img_path, sku, rack_id = row
            self.product_name_edit.setText(name or "")
            self.desc_edit.setPlainText(desc or "")
            if img_path and os.path.isfile(img_path):
                pix = QPixmap(img_path)
                if not pix.isNull():
                    self.img_label.setPixmap(pix.scaled(140, 140, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                    self.img_label.setText("")
                    self.uploaded_image_path = img_path
            self.sku_edit.setText(str(sku or ""))
            self.load_rack_locations(rack_id)
        except Exception as e:
            print(f"Error loading product: {e}")
            QMessageBox.critical(self, "Error", f"Failed to load product: {e}")
            self.reject()

    def upload_image(self):
        fname, _ = QFileDialog.getOpenFileName(self, "Select Image", "", "Images (*.png *.jpg *.jpeg *.bmp)")
        if fname:
            pix = QPixmap(fname)
            if not pix.isNull():
                self.img_label.setPixmap(pix.scaled(140, 140, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                self.img_label.setText("")
                self.uploaded_image_path = fname
                self.photo_uploaded = True

    def validate(self):
        if not self.product_name_edit.text().strip():
            QMessageBox.warning(self, "Validation", "Product name required");
            return False
        if not self.desc_edit.toPlainText().strip():
            QMessageBox.warning(self, "Validation", "Description required");
            return False
        if not self.sku_edit.text().strip():
            QMessageBox.warning(self, "Validation", "SKU required");
            return False
        return True

    def save_product(self):
        if not self.validate():
            return
        try:
            conn = sqlite3.connect("testing_system.db")
            cur = conn.cursor()
            rack_id = self.rack_combo.currentData()
            rack_name = self.rack_combo.currentText() if rack_id else "Unassigned"
            cur.execute("""
                UPDATE products SET product_name=?, product_desc=?, product_image=?, sku=?, rack_location_id=?, location=?
                WHERE product_id=?
            """, (
                self.product_name_edit.text().strip(),
                self.desc_edit.toPlainText().strip(),
                self.uploaded_image_path,
                self.sku_edit.text().strip(),
                rack_id,
                rack_name,
                self.product_id
            ))
            conn.commit();
            conn.close()
            QMessageBox.information(self, "Updated", "Product updated successfully.")
            self.product_updated.emit()
            if self.dashboard:
                try:
                    self.dashboard.load_inventory_data()
                except Exception:
                    pass
            self.accept()
        except Exception as e:
            print(f"Error updating product: {e}")
            QMessageBox.critical(self, "Error", f"Failed to update product: {e}")


class ManageProductsDialog(QDialog):
    """Dialog to list products and allow editing."""

    def __init__(self, parent_dashboard=None, user_role='Admin', current_user_id=None):
        super().__init__(parent_dashboard)
        self.dashboard = parent_dashboard
        self.user_role = user_role.lower()
        self.current_user_id = current_user_id or 0
        self.setWindowTitle("Manage Products")
        self.setFixedSize(1450, 1200)
        self.setModal(True)
        self.init_ui()
        self.load_products()

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(10)

        # Search panel
        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search product name ...")
        self.search_input.textChanged.connect(self.filter_products)
        search_btn = QPushButton("Search")
        search_btn.clicked.connect(self.filter_products)
        clear_btn = QPushButton("Clear")
        clear_btn.clicked.connect(lambda: self.search_input.clear())
        for w in (self.search_input, search_btn, clear_btn):
            w.setStyleSheet(
                "QLineEdit{padding:6px;border:1px solid #81C784;border-radius:4px;} QPushButton{background:#388E3C;color:white;padding:6px 14px;border:none;border-radius:4px;font-weight:bold;} QPushButton:hover{background:#2E7D32;}")
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(search_btn)
        search_layout.addWidget(clear_btn)
        main_layout.addLayout(search_layout)

        header = QLabel("📋 Product Management")
        header.setFont(QFont("Arial", 18, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("color:#1B5E20;")
        main_layout.addWidget(header)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["ID", "Preview", "Name", "SKU", "Rack Location", "Actions"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet(
            "QTableWidget{background:white;} QHeaderView::section{background:#388E3C;color:white;font-weight:bold;}")
        self.table.setColumnHidden(0, True)  # hide ID
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        # Set preferred column widths for better readability
        self.table.setColumnWidth(1, 100)  # Preview image column
        self.table.setColumnWidth(2, 280)  # Name
        self.table.setColumnWidth(3, 120)  # SKU
        self.table.setColumnWidth(4, 220)  # Rack location
        self.table.setColumnWidth(5, 120)  # Actions
        self.table.verticalHeader().setDefaultSectionSize(70)
        main_layout.addWidget(self.table)

        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        close_btn.setStyleSheet(
            "QPushButton{background:#6c757d;color:white;padding:6px 20px;border:none;border-radius:6px;font-weight:bold;} QPushButton:hover{background:#5a6268;}")
        btn_box = QHBoxLayout();
        btn_box.addStretch();
        btn_box.addWidget(close_btn)
        main_layout.addLayout(btn_box)

    def load_products(self):
        try:
            conn = sqlite3.connect("testing_system.db")
            cur = conn.cursor()
            if self.user_role in ("admin", "superadmin"):
                cur.execute("""
                    SELECT p.product_id, p.product_name, p.product_image, p.sku, COALESCE(rl.rack_location_name,'Unassigned')
                    FROM products p
                    LEFT JOIN racklocations rl ON p.rack_location_id = rl.rack_location_id
                    ORDER BY p.product_id DESC
                """)
            else:
                cur.execute("""
                    SELECT p.product_id, p.product_name, p.product_image, p.sku, COALESCE(rl.rack_location_name,'Unassigned')
                    FROM products p
                    LEFT JOIN racklocations rl ON p.rack_location_id = rl.rack_location_id
                    WHERE p.owner_id = ?
                    ORDER BY p.product_id DESC
                """, (self.current_user_id,))
            rows = cur.fetchall();
            conn.close()
            self.table.setRowCount(len(rows))
            for r, (pid, name, img_path, sku, rack) in enumerate(rows):
                id_item = QTableWidgetItem(str(pid));
                self.table.setItem(r, 0, id_item)

                # Preview image
                preview_label = QLabel()
                preview_label.setFixedSize(60, 60)
                preview_label.setAlignment(Qt.AlignCenter)
                if img_path and os.path.isfile(img_path):
                    pix = QPixmap(img_path)
                    if not pix.isNull():
                        preview_label.setPixmap(pix.scaled(58, 58, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                    else:
                        preview_label.setText("No Img")
                else:
                    preview_label.setText("No Img")
                self.table.setCellWidget(r, 1, preview_label)

                self.table.setItem(r, 2, QTableWidgetItem(name or ""))
                self.table.setItem(r, 3, QTableWidgetItem(str(sku or "")))
                self.table.setItem(r, 4, QTableWidgetItem(rack or "Unassigned"))

                # Actions: Edit button
                edit_btn = QPushButton("Edit")
                edit_btn.setProperty("pid", pid)
                edit_btn.clicked.connect(self.open_edit_dialog)
                edit_btn.setStyleSheet(
                    "QPushButton{background:#AE957F;color:white;padding:4px 12px;border:none;border-radius:6px;font-weight:bold;} QPushButton:hover{background:#FF8F00;}")
                self.table.setCellWidget(r, 5, edit_btn)

        except Exception as e:
            print(f"Error loading products: {e}")
            QMessageBox.critical(self, "Error", f"Failed to load products: {e}")

    def filter_products(self):
        text = self.search_input.text().strip().lower()
        for row in range(self.table.rowCount()):
            item_name = self.table.item(row, 2)
            if not item_name:
                self.table.setRowHidden(row, False)
                continue
            name_lower = item_name.text().lower()
            self.table.setRowHidden(row, text not in name_lower)

    def open_edit_dialog(self):
        sender = self.sender()
        pid = sender.property("pid")
        dlg = EditProductDialog(pid, parent_dashboard=self.dashboard)
        dlg.product_updated.connect(self.load_products)
        dlg.exec_()


# === Create Account Dialog ===
class CreateAccountDialog(QDialog):
    """Dialog to create a new user account (admin only)."""

    account_created = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Create Account")
        self.setFixedSize(500, 400)  # Reduced height since we removed phone field
        self.setModal(True)
        self.init_ui()

    # ---------------- UI ----------------
    def init_ui(self):
        self.setStyleSheet("""
            QDialog { background:#f3f8f3; }
            QLabel   { font-size:14px; }
            QLineEdit,QComboBox { border:1px solid #81C784; border-radius:4px; padding:6px; background:white; }
            QPushButton { background:#388E3C; color:white; border:none; border-radius:4px; padding:8px 18px; font-weight:bold; }
            QPushButton:hover { background:#2E7D32; }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        title = QLabel("➕ New User Account")
        title.setFont(QFont("Arial", 18, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color:#1B5E20;")
        layout.addWidget(title)

        form = QGridLayout()
        form.setHorizontalSpacing(12)
        form.setVerticalSpacing(10)

        # Full Name
        form.addWidget(QLabel("Full Name:"), 0, 0)
        self.fullname_edit = QLineEdit()
        form.addWidget(self.fullname_edit, 0, 1)

        # Email
        form.addWidget(QLabel("Email:"), 1, 0)
        self.email_edit = QLineEdit()
        form.addWidget(self.email_edit, 1, 1)

        # Role
        form.addWidget(QLabel("Role:"), 2, 0)
        self.role_combo = QComboBox()
        form.addWidget(self.role_combo, 2, 1)

        # Branch (only for owner)
        form.addWidget(QLabel("Branch:"), 3, 0)
        self.branch_combo = QComboBox()
        form.addWidget(self.branch_combo, 3, 1)

        layout.addLayout(form)

        # Buttons
        btn_box = QHBoxLayout();
        btn_box.addStretch()
        cancel_btn = QPushButton("Cancel");
        cancel_btn.clicked.connect(self.reject)
        save_btn = QPushButton("Create");
        save_btn.clicked.connect(self.save_account)
        btn_box.addWidget(cancel_btn);
        btn_box.addWidget(save_btn);
        btn_box.addStretch()
        layout.addLayout(btn_box)

        # Load roles / branches
        self.load_roles()
        self.load_branches()
        self.role_combo.currentIndexChanged.connect(self.update_branch_state)
        self.update_branch_state()

    # ------------ Data helpers -------------
    def load_roles(self):
        try:
            conn = sqlite3.connect("testing_system.db");
            cur = conn.cursor()
            cur.execute("SELECT role_id, role_name FROM roles ORDER BY role_id")
            self.role_combo.clear()
            for rid, name in cur.fetchall():
                self.role_combo.addItem(name.title(), rid)
            conn.close()
        except Exception as e:
            print(f"Error loading roles: {e}")

    def load_branches(self):
        try:
            conn = sqlite3.connect("testing_system.db");
            cur = conn.cursor()
            cur.execute("SELECT branch_id, branch_name FROM branches ORDER BY branch_name")
            self.branch_combo.clear()
            for bid, name in cur.fetchall():
                self.branch_combo.addItem(name, bid)
            conn.close()
        except Exception as e:
            print(f"Error loading branches: {e}")

    def update_branch_state(self):
        role_id = self.role_combo.currentData()
        # assume role_id 3 is owner per existing code
        self.branch_combo.setEnabled(role_id == 3)

    # ------------ Generation helpers -----------
    def generate_unique_username(self, fullname, cur):
        base = ''.join(fullname.lower().split())
        if not base:
            base = 'user'
        suffix = 1
        username = base
        while cur.execute("SELECT 1 FROM users WHERE username=?", (username,)).fetchone():
            suffix += 1
            username = f"{base}{suffix}"
        return username

    def generate_password(self):
        return random.randint(100000, 999999)  # 6-digit passcode

    # ------------- actions ----------------
    def save_account(self):
        fullname = self.fullname_edit.text().strip()
        email = self.email_edit.text().strip()
        role_id = self.role_combo.currentData()
        branch_id = self.branch_combo.currentData() if self.branch_combo.isEnabled() else None

        # Basic validation
        if not fullname or not email:
            QMessageBox.warning(self, "Validation", "Full name and email are required")
            return
        if role_id == 3 and branch_id is None:
            QMessageBox.warning(self, "Validation", "Please select branch for owner role")
            return

        try:
            conn = sqlite3.connect("testing_system.db");
            cur = conn.cursor()
            username = self.generate_unique_username(fullname, cur)
            password = self.generate_password()

            cur.execute("""
                INSERT INTO users (username,password,role,fullname,email,branch_id)
                VALUES (?,?,?,?,?,?)
            """, (username, password, role_id, fullname, email, branch_id))
            conn.commit();
            conn.close()

            # Send welcome email if parent dashboard has email helper
            parent_dash = self.parent()
            if parent_dash and hasattr(parent_dash, 'send_actual_email'):
                subject = "Welcome to Shelf Life Management System"
                body = (f"Dear {fullname},\n\n"
                        f"Your account has been created in the Shelf Life Management System. Please use the credentials below to log in for the first time and change your password immediately.\n\n"
                        f"Username: {username}\n"
                        f"Temporary Passcode: {password}\n\n"
                        f"Regards,\nShelf Life Management System Admin")
                try:
                    parent_dash.send_actual_email(email, subject, body, fullname)
                except Exception as se:
                    print(f"Email send failed: {se}")

            QMessageBox.information(self, "Account Created",
                                    f"User created successfully!\n\nUsername: {username}\n\nAn email has been sent with login instructions.")
            self.account_created.emit()
            self.accept()
        except Exception as e:
            print(f"Error creating account: {e}")
            QMessageBox.critical(self, "Error", f"Failed to create account: {e}")


# Expose globally for dynamic import in dashboard
CreateAccountDialog = CreateAccountDialog


def generate_unit_barcodes(product_id: int, batch: str, num_units: int):
    # Basic validation
    if num_units <= 0:
        return []

    # Produce purely numeric barcodes: product_id (6 digits) + unit sequence (3 digits)
    barcodes = []
    for seq in range(1, num_units + 1):
        barcode_str = f"{product_id:06d}{seq:03d}"
        barcodes.append({
            "product_id": product_id,
            "seq_no": seq,
            "barcode": barcode_str
        })

    return barcodes


def insert_unit_barcodes_to_db(unit_barcodes):
    if not unit_barcodes:
        return 0  # Nothing to insert

    max_retries = 4
    delay = 0.8  # seconds between retries

    for attempt in range(max_retries):
        conn = None
        try:
            conn = sqlite3.connect("testing_system.db", timeout=10.0)
            conn.execute("PRAGMA busy_timeout = 8000")  # Wait up to 8 sec if locked
            conn.execute("PRAGMA journal_mode = WAL")  # Better concurrent writes

            cur = conn.cursor()

            # Ensure table exists with the UNIQUE constraint.
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS product_barcodes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id INTEGER,
                    seq_no INTEGER,
                    barcode TEXT,
                    UNIQUE(product_id, seq_no)
                )
                """
            )

            # Prepare bulk insertion.
            records = [
                (item["product_id"], item["seq_no"], item["barcode"])
                for item in unit_barcodes
                if {"product_id", "seq_no", "barcode"}.issubset(item)
            ]

            cur.executemany(
                """
                INSERT OR IGNORE INTO product_barcodes (product_id, seq_no, barcode)
                VALUES (?, ?, ?)
                """,
                records,
            )

            conn.commit()
            # Successful – return number of rows attempted (SQLite can't give accurate count when OR IGNORE)
            return cur.rowcount if cur.rowcount != -1 else len(records)

        except sqlite3.OperationalError as e:
            if "database is locked" in str(e).lower() and attempt < max_retries - 1:
                print("⚠️  Database locked when inserting unit barcodes – retrying...")
                time.sleep(delay)
                continue
            else:
                print(f"Error inserting unit barcodes: {e}")
                break
        finally:
            if conn:
                conn.close()

    return 0  # Failed after retries


def generate_barcode_images(barcodes_or_units, output_dir="barcodes"):
    """Generate PNG barcode images and update *product_barcodes* table.

    This helper works with *either* a plain list of barcode strings **or** the
    list of dictionaries returned by :pyfunc:`generate_unit_barcodes`.

    Parameters
    ----------
    barcodes_or_units : Sequence[str] | Sequence[dict]
        • If sequence elements are *str*, each element is taken as the barcode
          value and the function inserts/updates DB rows with *NULL* for
          ``product_id``/``seq_no``.
        • If sequence elements are *dict* (with keys ``product_id``,
          ``seq_no``, ``barcode``) the full information is used so that the
          UNIQUE( product_id, seq_no ) constraint is respected.

    output_dir : str, default ``"barcodes"``
        Destination directory for the generated PNG files. It is created if it
        does not already exist.

    Returns
    -------
    list[str]
        Absolute file paths to successfully generated images.
    """
    # ------------------------------------------------------------------
    # Normalise input: convert to list[str] *barcodes* and optional metadata
    # ------------------------------------------------------------------
    if not barcodes_or_units:
        return []

    # If the first element is a dict we presume the *unit-barcodes* structure
    unit_dicts = None
    if isinstance(barcodes_or_units[0], dict):
        unit_dicts = barcodes_or_units  # type: ignore
        barcodes = [d.get("barcode") for d in unit_dicts if d.get("barcode")]
    else:
        # Assume it's already a list/tuple of strings
        barcodes = [str(x) for x in barcodes_or_units]

    if not barcodes:
        return []

    # ------------------------------------------------------------------
    # Ensure output directory exists
    # ------------------------------------------------------------------
    os.makedirs(output_dir, exist_ok=True)

    # ------------------------------------------------------------------
    # Determine which barcode generator is available (python-barcode > qrcode)
    # ------------------------------------------------------------------
    generator = None
    try:
        import barcode  # type: ignore
        from barcode.writer import ImageWriter  # type: ignore
        generator = ("python-barcode", barcode.get_barcode_class("code128"), ImageWriter)
    except ImportError:
        try:
            import qrcode  # type: ignore
            generator = ("qrcode", qrcode, None)
        except ImportError:
            print(
                "⚠️  Neither 'python-barcode' nor 'qrcode' libraries are available. ``generate_barcode_images`` aborted.")
            return []

    generated_paths = []

    # ------------------------------------------------------------------
    # Open database connection to update image paths
    # ------------------------------------------------------------------
    conn = None
    try:
        conn = sqlite3.connect("testing_system.db", timeout=10.0)
        cur = conn.cursor()

        # Ensure product_barcodes table exists with barcode_image column
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS product_barcodes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER,
                seq_no INTEGER,
                barcode TEXT,
                barcode_image TEXT,
                UNIQUE(product_id, seq_no)
            )
            """
        )
        # If the column did not exist previously, attempt an ALTER (will fail harmlessly if already present)
        try:
            cur.execute("ALTER TABLE product_barcodes ADD COLUMN barcode_image TEXT")
        except sqlite3.OperationalError:
            pass  # column already exists

        # Iterate through barcodes and generate images
        for bc in barcodes:
            # Decide file path
            filename = f"{bc}.png"
            filepath = os.path.abspath(os.path.join(output_dir, filename))

            # Skip generation if file already exists; ensure DB has path
            if not os.path.isfile(filepath):
                try:
                    if generator[0] == "python-barcode":
                        cls, ImageWriter = generator[1], generator[2]
                        cls_instance = cls(bc, writer=ImageWriter())
                        cls_instance.save(filepath[:-4])  # library appends .png automatically
                        # python-barcode adds .png, ensure consistent
                        if not filepath.endswith(".png"):
                            filepath = filepath + ".png"
                    else:  # qrcode fallback – not ideal for barcodes, but acceptable as QR
                        qr = generator[1].QRCode(box_size=10, border=4)
                        qr.add_data(bc)
                        qr.make(fit=True)
                        img = qr.make_image(fill_color="black", back_color="white")
                        img.save(filepath)
                except Exception as gen_err:
                    print(f"❌ Failed to generate image for {bc}: {gen_err}")
                    continue

            # Record path in DB – update matching barcode row; create row if missing
            if unit_dicts is not None:
                # Find matching dict to supply product_id/seq_no
                match = next((d for d in unit_dicts if d.get("barcode") == bc), None)
                if match is not None:
                    cur.execute(
                        """
                        INSERT INTO product_barcodes (product_id, seq_no, barcode, barcode_image)
                        VALUES (?, ?, ?, ?)
                        ON CONFLICT(product_id, seq_no) DO UPDATE SET barcode_image=excluded.barcode_image
                        """,
                        (match["product_id"], match["seq_no"], bc, filepath),
                    )
                else:
                    cur.execute(
                        "INSERT OR IGNORE INTO product_barcodes (product_id, seq_no, barcode, barcode_image) VALUES (NULL, NULL, ?, ?)",
                        (bc, filepath),
                    )
            else:
                cur.execute(
                    "INSERT OR IGNORE INTO product_barcodes (product_id, seq_no, barcode, barcode_image) VALUES (NULL, NULL, ?, ?)",
                    (bc, filepath),
                )

        conn.commit()
    except sqlite3.Error as db_err:
        print(f"Database error in generate_barcode_images: {db_err}")
    finally:
        if conn:
            conn.close()

    return generated_paths


def export_product_with_unit_barcodes_to_excel(product_id: int):
    """Export a product and all its unit barcodes to an Excel file.

    Parameters
    ----------
    product_id : int
        ID of the product to export.

    The function creates ``product_<product_id>_unit_barcodes.xlsx`` in the
    current working directory. If the product does not exist, a warning is
    printed and nothing is generated.
    """
    import pandas as pd  # Local import to avoid heavy dependency at startup

    conn = None
    try:
        conn = sqlite3.connect("testing_system.db", timeout=10.0)
        cur = conn.cursor()

        # ------------------------------------------------------------------
        # Fetch product-level information
        # ------------------------------------------------------------------
        cur.execute(
            """
            SELECT p.product_id, p.product_name, p.batch, p.sku,
                   p.arrival_date, p.manufacture_date, p.expired_date,
                   COALESCE(rl.rack_location_name, p.location, '') AS location,
                   p.branch_id, COALESCE(p.status, '') AS status,
                   COALESCE(p.product_desc, '') AS description,
                   COALESCE(p.rejection_comment, '') AS rejection_comment
            FROM products p
            LEFT JOIN racklocations rl ON p.rack_location_id = rl.rack_location_id
            WHERE p.product_id = ?
            """,
            (product_id,),
        )
        prod = cur.fetchone()
        if not prod:
            print(f"⚠️  Product ID {product_id} not found – export skipped.")
            return

        (prod_id, prod_name, batch, sku, arrival_date, mfg_date, exp_date,
         location, branch_id, status, description, rejection_comment) = prod

        # ------------------------------------------------------------------
        # Fetch unit barcodes for this product
        # ------------------------------------------------------------------
        cur.execute(
            "SELECT seq_no, barcode FROM product_barcodes WHERE product_id = ? ORDER BY seq_no",
            (product_id,),
        )
        barcode_rows = cur.fetchall()
        if not barcode_rows:
            print(f"⚠️  No unit barcodes found for product {product_id}. Excel not created.")
            return

        # ------------------------------------------------------------------
        # Build rows for DataFrame
        # ------------------------------------------------------------------
        rows = []
        for seq_no, barcode in barcode_rows:
            rows.append({
                "product_id": prod_id,
                "product_name": prod_name,
                "batch": batch,
                "sku": sku,
                "arrival_date": arrival_date,
                "manufacture_date": mfg_date,
                "expired_date": exp_date,
                "location": location,
                "branch_id": branch_id,
                "status": status,
                "barcode": barcode,
                "description": description,
                "rejection_comment": rejection_comment,
            })

        df = pd.DataFrame(rows)
        file_name = f"product_{product_id}_unit_barcodes.xlsx"
        df.to_excel(file_name, index=False)
        print(f"✅ Excel exported: {file_name} ({len(rows)} rows)")
    except Exception as e:
        print(f"❌ Error exporting product {product_id} barcodes: {e}")
    finally:
        if conn:
            conn.close()


def export_unit_barcodes_to_excel(product_id: int):
    """Export only the unit-level barcode table for *product_id* to Excel.

    Output file: ``product_<product_id>_unit_barcodes_only.xlsx``
    Columns: ``seq_no, barcode, created_at, barcode_image``
    """
    import pandas as pd

    conn = None
    try:
        conn = sqlite3.connect("testing_system.db", timeout=10.0)
        cur = conn.cursor()

        # Ensure table has a created_at column (timestamp)
        try:
            cur.execute("ALTER TABLE product_barcodes ADD COLUMN created_at DATETIME DEFAULT CURRENT_TIMESTAMP")
        except sqlite3.OperationalError:
            pass  # column already exists

        cur.execute(
            "SELECT seq_no, barcode, created_at, barcode_image FROM product_barcodes WHERE product_id = ? ORDER BY seq_no",
            (product_id,),
        )
        rows = cur.fetchall()
        if not rows:
            print(f"⚠️  No unit barcodes found for product {product_id}. Export skipped.")
            return

        df = pd.DataFrame(rows, columns=["seq_no", "barcode", "created_at", "barcode_image"])
        file_name = f"product_{product_id}_unit_barcodes_only.xlsx"
        df.to_excel(file_name, index=False)
        print(f"✅ Unit barcodes Excel exported: {file_name} ({len(rows)} rows)")
    except Exception as e:
        print(f"❌ Error exporting unit barcodes for product {product_id}: {e}")
    finally:
        if conn:
            conn.close()


def export_today_products_barcodes_to_excel(top_n: int = 10, near_expiry_days: int = 60):
    """Export all products added today together with their unit barcodes *and* three
    useful charts into one Excel workbook.

    Charts generated (each on its own worksheet):
        1) 今日各产品单位数量 – column chart (TodaySummary sheet)
        2) 保质期到期分布（月） – area chart (ExpiryDist sheet)
        3) TOP-N 临期产品（≤ near_expiry_days） – column chart (TopNearExpiry sheet)

    After saving, the file name ``today_products_unit_barcodes_<YYYYMMDD>.xlsx``
    is written back to ``products.excel_name`` for all products added today.
    """
    import pandas as pd
    import sqlite3
    from datetime import datetime, timedelta

    conn: sqlite3.Connection | None = None
    try:
        conn = sqlite3.connect("testing_system.db", timeout=15.0)
        cur = conn.cursor()

        # Ensure product_barcodes table has created_at column (safety for older DBs)
        try:
            cur.execute("ALTER TABLE product_barcodes ADD COLUMN created_at DATETIME DEFAULT CURRENT_TIMESTAMP")
        except sqlite3.OperationalError:
            pass  # column already exists

        # ------------------------------------------------------------------
        # Guarantee unit-level barcode completeness for today's products first
        # ------------------------------------------------------------------
        cur.execute("SELECT product_id, COALESCE(sku, 0), COALESCE(batch, '') FROM products "
                    "WHERE DATE(arrival_date) = DATE('now','localtime')")
        for pid, sku_qty, batch in cur.fetchall():
            try:
                ensure_unit_barcodes_for_product(pid, int(sku_qty), batch)
            except Exception as prep_err:
                print(f"[WARN] ensure barcodes for product {pid}: {prep_err}")

        # ------------------------------------------------------------------
        # Fetch joined detail rows (products added today  + unit barcodes)
        # ------------------------------------------------------------------
        df = pd.read_sql_query(
            """
            SELECT p.product_id, p.product_name, p.sku, p.batch,
                   p.arrival_date, p.manufacture_date, p.expired_date,
                   COALESCE(rl.rack_location_name, p.location, '')   AS location,
                   p.branch_id, COALESCE(p.status, '')              AS status,
                   pb.seq_no, pb.barcode, pb.barcode_image, pb.created_at
            FROM   products p
            LEFT JOIN racklocations rl ON p.rack_location_id = rl.rack_location_id
            JOIN   product_barcodes pb  ON pb.product_id = p.product_id
            WHERE  DATE(p.arrival_date) = DATE('now','localtime')
            ORDER  BY p.product_id, pb.seq_no
            """,
            conn
        )

        if df.empty:
            print("⚠️  No products found for today – export skipped.")
            return

        # --------------------------------------------------------------
        # 2) Standardise date columns (strip time) & rename columns EN
        # --------------------------------------------------------------
        date_cols = [c for c in ["arrival_date", "manufacture_date", "expired_date"] if c in df.columns]
        for col in date_cols:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.date

        # Rename column headers to English friendly titles
        df.rename(columns={
            "product_id": "Product ID",
            "product_name": "Product Name",
            "sku": "SKU",
            "batch": "Batch",
            "arrival_date": "Arrival Date",
            "manufacture_date": "Manufacture Date",
            "expired_date": "Expiry Date",
            "location": "Location",
            "branch_id": "Branch ID",
            "status": "Status",
            "seq_no": "Seq No",
            "barcode": "Barcode",
            "barcode_image": "Barcode Image",
            "created_at": "Barcode Created",
        }, inplace=True)

        today_str = datetime.now().strftime("%Y%m%d")
        file_name = f"today_products_unit_barcodes_{today_str}.xlsx"

        # ------------------------------------------------------------------
        # Write to Excel and add charts
        # ------------------------------------------------------------------
        with pd.ExcelWriter(file_name, engine="xlsxwriter") as writer:
            wb = writer.book

            # A) Raw detail sheet -------------------------------------------------
            df.to_excel(writer, sheet_name="Barcodes", index=False)

            # B) 今日各产品单位数量 ---------------------------------------------
            pivot_today = (
                df.groupby("Product Name")["Barcode"]
                .count()
                .reset_index(name="Unit Count")
            )
            pivot_today.rename(columns={"Product Name": "Product"}, inplace=True)
            pivot_today.to_excel(writer, sheet_name="TodaySummary", index=False)
            ws_today = writer.sheets["TodaySummary"]

            chart_today = wb.add_chart({"type": "column"})
            max_row_today = len(pivot_today)
            chart_today.add_series({
                "name": "Unit Count",
                "categories": ["TodaySummary", 1, 0, max_row_today, 0],  # A column
                "values": ["TodaySummary", 1, 1, max_row_today, 1],  # B column
            })
            chart_today.set_title({"name": "Unit Quantity of Today's Products"})
            chart_today.set_x_axis({"name": "Product"})
            chart_today.set_y_axis({"name": "Quantity"})
            ws_today.insert_chart("D2", chart_today)

            # C) 保质期到期分布（月） -------------------------------------------
            month_dist = (
                df.dropna(subset=["Expiry Date"])
                .assign(Month=lambda x: pd.to_datetime(x["Expiry Date"]).dt.to_period("M").astype(str))
                .groupby("Month")["Barcode"]
                .count()
                .reset_index(name="Unit Count")
                .sort_values("Month")
            )
            month_dist.to_excel(writer, sheet_name="ExpiryDist", index=False)
            ws_exp = writer.sheets["ExpiryDist"]

            chart_exp = wb.add_chart({"type": "area"})
            max_row_exp = len(month_dist)
            chart_exp.add_series({
                "name": "Unit Count",
                "categories": ["ExpiryDist", 1, 0, max_row_exp, 0],
                "values": ["ExpiryDist", 1, 1, max_row_exp, 1],
            })
            chart_exp.set_title({"name": "Expiry Distribution by Month"})
            chart_exp.set_x_axis({"name": "Expiry Month"})
            chart_exp.set_y_axis({"name": "Quantity"})
            ws_exp.insert_chart("D2", chart_exp)

            # D) TOP-N 临期产品 --------------------------------------------------
            today_date = datetime.now().date()
            near_expiry_df = (
                df[pd.to_datetime(df["Expiry Date"], errors="coerce").dt.date.between(
                    today_date,
                    today_date + timedelta(days=near_expiry_days),
                    inclusive="both")]
                .groupby("Product Name")["Barcode"]
                .count()
                .reset_index(name="Unit Count")
                .sort_values("Unit Count", ascending=False)
                .head(top_n)
            )
            near_expiry_df.rename(columns={"Product Name": "Product"}, inplace=True)
            near_expiry_df.to_excel(writer, sheet_name="TopNearExpiry", index=False)
            ws_near = writer.sheets["TopNearExpiry"]

            chart_near = wb.add_chart({"type": "column"})
            max_row_near = len(near_expiry_df)
            chart_near.add_series({
                "name": f"<= {near_expiry_days} Days",
                "categories": ["TopNearExpiry", 1, 0, max_row_near, 0],
                "values": ["TopNearExpiry", 1, 1, max_row_near, 1],
            })
            chart_near.set_title({"name": f"TOP-{top_n} Near-Expiry Products (<= {near_expiry_days} Days)"})
            chart_near.set_x_axis({"name": "Product"})
            chart_near.set_y_axis({"name": "Quantity"})
            ws_near.insert_chart("D2", chart_near)

            # Workbook saved automatically on exit

        # ------------------------------------------------------------------
        # Persist excel file name back to DB for today's products
        # ------------------------------------------------------------------
        cur.execute(
            "UPDATE products SET excel_name = ? "
            "WHERE DATE(arrival_date) = DATE('now','localtime')",
            (file_name,)
        )
        conn.commit()
        print(f"✅ Excel exported with charts: {file_name}")

    except Exception as e:
        print(f"❌ Error exporting today's products: {e}")
    finally:
        if conn:
            conn.close()


# ------------------------------------------------------------------
# Helper: ensure a product has exactly *sku* rows in product_barcodes
# ------------------------------------------------------------------

def ensure_unit_barcodes_for_product(product_id: int, sku: int, batch: str):
    """Make sure *product_id* has *sku* unit-barcode rows.

    If missing rows are detected, they will be generated (barcode string =
    product_id(6) + seq_no(3)), inserted into the DB and corresponding PNG
    images produced.
    """
    if sku <= 0:
        return

    try:
        conn = sqlite3.connect("testing_system.db", timeout=10.0)
        cur = conn.cursor()

        cur.execute("SELECT seq_no FROM product_barcodes WHERE product_id = ?", (product_id,))
        existing = {row[0] for row in cur.fetchall()}

        missing_seqs = [seq for seq in range(1, sku + 1) if seq not in existing]

        if not missing_seqs:
            return  # already complete

        # Build missing units list
        units = []
        for seq in missing_seqs:
            bc = f"{product_id:06d}{seq:03d}"
            units.append({"product_id": product_id, "seq_no": seq, "barcode": bc})

        # Insert & generate images using helpers (they open their own connections)
        insert_unit_barcodes_to_db(units)
        generate_barcode_images(units)

    except sqlite3.Error as e:
        print(f"Error ensuring unit barcodes for product {product_id}: {e}")
    finally:
        if 'conn' in locals() and conn:
            conn.close()


def export_today_products_barcodes_to_excel_with_images(top_n: int = 10, near_expiry_days: int = 60):
    """Export all products added today together with their unit barcodes *and* three
    useful charts into one Excel workbook. This enhanced version embeds actual barcode
    images into the Excel file instead of just showing file paths.

    Charts generated (each on its own worksheet):
        1) 今日各产品单位数量 – column chart (TodaySummary sheet)
        2) 保质期到期分布（月） – area chart (ExpiryDist sheet)
        3) TOP-N 临期产品（≤ near_expiry_days） – column chart (TopNearExpiry sheet)

    After saving, the file name ``today_products_unit_barcodes_with_images_<YYYYMMDD>.xlsx``
    is written back to ``products.excel_name`` for all products added today.
    """
    import pandas as pd
    import sqlite3
    import os
    from datetime import datetime, timedelta

    conn: sqlite3.Connection | None = None
    try:
        conn = sqlite3.connect("testing_system.db", timeout=15.0)
        cur = conn.cursor()

        # Ensure product_barcodes table has created_at column (safety for older DBs)
        try:
            cur.execute("ALTER TABLE product_barcodes ADD COLUMN created_at DATETIME DEFAULT CURRENT_TIMESTAMP")
        except sqlite3.OperationalError:
            pass  # column already exists

        # ------------------------------------------------------------------
        # Guarantee unit-level barcode completeness for today's products first
        # ------------------------------------------------------------------
        cur.execute("SELECT product_id, COALESCE(sku, 0), COALESCE(batch, '') FROM products "
                    "WHERE DATE(arrival_date) = DATE('now','localtime')")
        for pid, sku_qty, batch in cur.fetchall():
            try:
                ensure_unit_barcodes_for_product(pid, int(sku_qty), batch)
            except Exception as prep_err:
                print(f"[WARN] ensure barcodes for product {pid}: {prep_err}")

        # ------------------------------------------------------------------
        # Fetch joined detail rows (products added today  + unit barcodes)
        # ------------------------------------------------------------------
        df = pd.read_sql_query(
            """
            SELECT p.product_id, p.product_name, p.sku, p.batch,
                   p.arrival_date, p.manufacture_date, p.expired_date,
                   COALESCE(rl.rack_location_name, p.location, '')   AS location,
                   p.branch_id, COALESCE(p.status, '')              AS status,
                   pb.seq_no, pb.barcode, pb.barcode_image, pb.created_at
            FROM   products p
            LEFT JOIN racklocations rl ON p.rack_location_id = rl.rack_location_id
            JOIN   product_barcodes pb  ON pb.product_id = p.product_id
            WHERE  DATE(p.arrival_date) = DATE('now','localtime')
            ORDER  BY p.product_id, pb.seq_no
            """,
            conn
        )

        if df.empty:
            print("⚠️  No products found for today – export skipped.")
            return

        # --------------------------------------------------------------
        # 2) Standardise date columns (strip time) & rename columns EN
        # --------------------------------------------------------------
        date_cols = [c for c in ["arrival_date", "manufacture_date", "expired_date"] if c in df.columns]
        for col in date_cols:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.date

        # Rename column headers to English friendly titles
        df.rename(columns={
            "product_id": "Product ID",
            "product_name": "Product Name",
            "sku": "SKU",
            "batch": "Batch",
            "arrival_date": "Arrival Date",
            "manufacture_date": "Manufacture Date",
            "expired_date": "Expiry Date",
            "location": "Location",
            "branch_id": "Branch ID",
            "status": "Status",
            "seq_no": "Seq No",
            "barcode": "Barcode",
            "barcode_image": "Barcode File Path",  # Keep original path in this column
            "created_at": "Barcode Created",
        }, inplace=True)

        # Add a new column for barcode images (will be filled with actual images in Excel)
        df["Barcode Image"] = ""  # Empty column, images will be inserted directly in Excel

        today_str = datetime.now().strftime("%Y%m%d")
        file_name = f"today_products_unit_barcodes_with_images_{today_str}.xlsx"

        # ------------------------------------------------------------------
        # Write to Excel with embedded images
        # ------------------------------------------------------------------
        with pd.ExcelWriter(file_name, engine="xlsxwriter") as writer:
            wb = writer.book

            # A) Raw detail sheet with embedded barcode images ---------------
            df.to_excel(writer, sheet_name="Barcodes", index=False)
            ws_barcodes = writer.sheets["Barcodes"]

            # Set zoom level to 55% when opening the Excel file
            ws_barcodes.set_zoom(55)

            # Set row height excel units
            ws_barcodes.set_default_row(70)

            # Set font size to 14px for all cells
            font_format = wb.add_format({'font_size': 20})
            ws_barcodes.set_column('A:Z', None, font_format)

            # Find the column indices
            barcode_file_path_col = df.columns.get_loc("Barcode File Path")
            barcode_img_col = df.columns.get_loc("Barcode Image")

            # Store temporary files for cleanup after Excel is written
            temp_files_to_cleanup = []

            # Insert barcode images in the Barcode Image column while keeping file paths in Barcode File Path column
            for idx, row in df.iterrows():
                barcode_img_path = row["Barcode File Path"]
                excel_row = idx + 1  # +1 because Excel is 1-indexed and we have headers

                print(f"Processing barcode image: {barcode_img_path}")

                if barcode_img_path and os.path.exists(barcode_img_path):
                    try:
                        # Don't need to clear the cell as it's already empty from DataFrame

                        # Resize image from 300x300 to smaller size before inserting
                        from PIL import Image
                        import tempfile

                        print(f"Opening image: {barcode_img_path}")
                        # Create a temporary resized image
                        with Image.open(barcode_img_path) as img:
                            # Resize to 100x100 pixels (smaller for Excel)
                            resized_img = img.resize((90, 90), Image.Resampling.LANCZOS)

                            # Save to temporary file with proper extension
                            temp_fd, temp_img_path = tempfile.mkstemp(suffix='.png', dir=os.getcwd())
                            os.close(temp_fd)  # Close the file descriptor
                            print(f"Saving resized image to: {temp_img_path}")
                            resized_img.save(temp_img_path, 'PNG')

                            # Verify the temp file exists before proceeding
                            if not os.path.exists(temp_img_path):
                                raise FileNotFoundError(f"Temporary file was not created: {temp_img_path}")

                            print(f"Temp file size: {os.path.getsize(temp_img_path)} bytes")

                            # Add to cleanup list for later
                            temp_files_to_cleanup.append(temp_img_path)

                        # Insert the resized image directly into the barcode image path column
                        print(f"Inserting image into Excel at row {excel_row}, col {barcode_img_col}")
                        ws_barcodes.insert_image(
                            excel_row, barcode_img_col,  # row, col (same column as the path)
                            temp_img_path,
                            {
                                'x_scale': 1.0,  # No additional scaling since we already resized
                                'y_scale': 1.0,
                                'x_offset': 2,  # Small offset from cell border
                                'y_offset': 2
                            }
                        )
                        print(f"Successfully inserted image for row {excel_row}")

                    except Exception as img_err:
                        print(f"Failed to insert image {barcode_img_path}: {img_err}")
                        import traceback
                        traceback.print_exc()
                        # Write "Image Error" text as fallback
                        ws_barcodes.write(excel_row, barcode_img_col, "Image Error")
                else:
                    print(f"Image file not found or empty path: {barcode_img_path}")
                    # Write "No Image" if path doesn't exist
                    ws_barcodes.write(excel_row, barcode_img_col, "No Image")

            # Set all column widths to 300px (approximately 38.5 Excel units)
            # Excel column width units
            total_columns = len(df.columns)
            for col_idx in range(total_columns):
                ws_barcodes.set_column(col_idx, col_idx, 25, font_format)

            # B) 今日各产品单位数量 ---------------------------------------------
            pivot_today = (
                df.groupby("Product Name")["Barcode"]
                .count()
                .reset_index(name="Unit Count")
            )
            pivot_today.rename(columns={"Product Name": "Product"}, inplace=True)
            pivot_today.to_excel(writer, sheet_name="TodaySummary", index=False)
            ws_today = writer.sheets["TodaySummary"]

            chart_today = wb.add_chart({"type": "column"})
            max_row_today = len(pivot_today)
            chart_today.add_series({
                "name": "Unit Count",
                "categories": ["TodaySummary", 1, 0, max_row_today, 0],  # A column
                "values": ["TodaySummary", 1, 1, max_row_today, 1],  # B column
            })
            chart_today.set_title({"name": "Unit Quantity of Today's Products"})
            chart_today.set_x_axis({"name": "Product"})
            chart_today.set_y_axis({"name": "Quantity"})
            ws_today.insert_chart("D2", chart_today)

            # C) 保质期到期分布（月） -------------------------------------------
            month_dist = (
                df.dropna(subset=["Expiry Date"])
                .assign(Month=lambda x: pd.to_datetime(x["Expiry Date"]).dt.to_period("M").astype(str))
                .groupby("Month")["Barcode"]
                .count()
                .reset_index(name="Unit Count")
                .sort_values("Month")
            )
            month_dist.to_excel(writer, sheet_name="ExpiryDist", index=False)
            ws_exp = writer.sheets["ExpiryDist"]

            chart_exp = wb.add_chart({"type": "area"})
            max_row_exp = len(month_dist)
            chart_exp.add_series({
                "name": "Unit Count",
                "categories": ["ExpiryDist", 1, 0, max_row_exp, 0],
                "values": ["ExpiryDist", 1, 1, max_row_exp, 1],
            })
            chart_exp.set_title({"name": "Expiry Distribution by Month"})
            chart_exp.set_x_axis({"name": "Expiry Month"})
            chart_exp.set_y_axis({"name": "Quantity"})
            ws_exp.insert_chart("D2", chart_exp)

            # D) TOP-N 临期产品 --------------------------------------------------
            today_date = datetime.now().date()
            near_expiry_df = (
                df[pd.to_datetime(df["Expiry Date"], errors="coerce").dt.date.between(
                    today_date,
                    today_date + timedelta(days=near_expiry_days),
                    inclusive="both")]
                .groupby("Product Name")["Barcode"]
                .count()
                .reset_index(name="Unit Count")
                .sort_values("Unit Count", ascending=False)
                .head(top_n)
            )
            near_expiry_df.rename(columns={"Product Name": "Product"}, inplace=True)
            near_expiry_df.to_excel(writer, sheet_name="TopNearExpiry", index=False)
            ws_near = writer.sheets["TopNearExpiry"]

            chart_near = wb.add_chart({"type": "column"})
            max_row_near = len(near_expiry_df)
            chart_near.add_series({
                "name": f"<= {near_expiry_days} Days",
                "categories": ["TopNearExpiry", 1, 0, max_row_near, 0],
                "values": ["TopNearExpiry", 1, 1, max_row_near, 1],
            })
            chart_near.set_title({"name": f"TOP-{top_n} Near-Expiry Products (<= {near_expiry_days} Days)"})
            chart_near.set_x_axis({"name": "Product"})
            chart_near.set_y_axis({"name": "Quantity"})
            ws_near.insert_chart("D2", chart_near)

            # Workbook saved automatically on exit

        # Clean up temporary files after Excel is written
        print(f"Cleaning up {len(temp_files_to_cleanup)} temporary files...")
        for temp_file in temp_files_to_cleanup:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
                    print(f"Cleaned up: {temp_file}")
            except Exception as cleanup_err:
                print(f"Failed to cleanup {temp_file}: {cleanup_err}")

        # ------------------------------------------------------------------
        # Persist excel file name back to DB for today's products
        # ------------------------------------------------------------------
        cur.execute(
            "UPDATE products SET excel_name = ? "
            "WHERE DATE(arrival_date) = DATE('now','localtime')",
            (file_name,)
        )
        conn.commit()
        print(f"✅ Excel exported with embedded barcode images: {file_name}")

    except Exception as e:
        print(f"❌ Error exporting today's products with images: {e}")
    finally:
        if conn:
            conn.close()


def export_product_with_unit_barcodes_to_excel_with_images(product_id: int):
    """Export a product and all its unit barcodes to an Excel file with embedded barcode images.

    Parameters
    ----------
    product_id : int
        ID of the product to export.

    The function creates ``product_<product_id>_unit_barcodes_with_images.xlsx`` in the
    current working directory. If the product does not exist, a warning is
    printed and nothing is generated.
    """
    import pandas as pd  # Local import to avoid heavy dependency at startup

    conn = None
    try:
        conn = sqlite3.connect("testing_system.db", timeout=10.0)
        cur = conn.cursor()

        # ------------------------------------------------------------------
        # Fetch product-level information
        # ------------------------------------------------------------------
        cur.execute(
            """
            SELECT p.product_id, p.product_name, p.batch, p.sku,
                   p.arrival_date, p.manufacture_date, p.expired_date,
                   COALESCE(rl.rack_location_name, p.location, '') AS location,
                   p.branch_id, COALESCE(p.status, '') AS status,
                   COALESCE(p.product_desc, '') AS description,
                   COALESCE(p.rejection_comment, '') AS rejection_comment
            FROM products p
            LEFT JOIN racklocations rl ON p.rack_location_id = rl.rack_location_id
            WHERE p.product_id = ?
            """,
            (product_id,),
        )
        prod = cur.fetchone()
        if not prod:
            print(f"⚠️  Product ID {product_id} not found – export skipped.")
            return

        (prod_id, prod_name, batch, sku, arrival_date, mfg_date, exp_date,
         location, branch_id, status, description, rejection_comment) = prod

        # ------------------------------------------------------------------
        # Fetch unit barcodes for this product
        # ------------------------------------------------------------------
        cur.execute(
            "SELECT seq_no, barcode, barcode_image FROM product_barcodes WHERE product_id = ? ORDER BY seq_no",
            (product_id,),
        )
        barcode_rows = cur.fetchall()
        if not barcode_rows:
            print(f"⚠️  No unit barcodes found for product {product_id}. Excel not created.")
            return

        # ------------------------------------------------------------------
        # Build rows for DataFrame
        # ------------------------------------------------------------------
        rows = []
        for seq_no, barcode, barcode_image in barcode_rows:
            rows.append({
                "product_id": prod_id,
                "product_name": prod_name,
                "batch": batch,
                "sku": sku,
                "arrival_date": arrival_date,
                "manufacture_date": mfg_date,
                "expired_date": exp_date,
                "location": location,
                "branch_id": branch_id,
                "status": status,
                "seq_no": seq_no,
                "barcode": barcode,
                "barcode_image_path": barcode_image,
                "description": description,
                "rejection_comment": rejection_comment,
            })

        df = pd.DataFrame(rows)
        file_name = f"product_{product_id}_unit_barcodes_with_images.xlsx"

        # ------------------------------------------------------------------
        # Write to Excel with embedded images using xlsxwriter
        # ------------------------------------------------------------------
        with pd.ExcelWriter(file_name, engine="xlsxwriter") as writer:
            wb = writer.book

            # Write dataframe to excel
            df.to_excel(writer, sheet_name="Product_Barcodes", index=False)
            ws = writer.sheets["Product_Barcodes"]

            # Set row height for barcode images
            ws.set_default_row(45)

            # Find the barcode image path column index
            barcode_img_col = df.columns.get_loc("barcode_image_path")

            # Add a new column for actual barcode images
            ws.write(0, barcode_img_col + 1, "Barcode Image")

            # Insert barcode images
            for idx, row in df.iterrows():
                barcode_img_path = row["barcode_image_path"]
                excel_row = idx + 1  # +1 because Excel is 1-indexed and we have headers

                if barcode_img_path and os.path.exists(barcode_img_path):
                    try:
                        # Insert image with proper sizing
                        ws.insert_image(
                            excel_row, barcode_img_col + 1,  # row, col
                            barcode_img_path,
                            {
                                'x_scale': 0.8,  # Scale down to fit in cell
                                'y_scale': 0.8,
                                'x_offset': 5,  # Small offset from cell border
                                'y_offset': 5
                            }
                        )
                    except Exception as img_err:
                        print(f"Failed to insert image {barcode_img_path}: {img_err}")
                        # Write "Image Error" text as fallback
                        ws.write(excel_row, barcode_img_col + 1, "Image Error")
                else:
                    # Write "No Image" if path doesn't exist
                    ws.write(excel_row, barcode_img_col + 1, "No Image")

            # Set column widths for better display
            ws.set_column(0, barcode_img_col, 15)  # Standard columns
            ws.set_column(barcode_img_col + 1, barcode_img_col + 1, 25)  # Barcode image column wider

        print(f"✅ Excel exported with embedded barcode images: {file_name} ({len(rows)} rows)")
    except Exception as e:
        print(f"❌ Error exporting product {product_id} barcodes with images: {e}")
    finally:
        if conn:
            conn.close()


def export_unit_barcodes_to_excel_with_images(product_id: int):
    """Export only the unit-level barcode table for *product_id* to Excel with embedded images.

    Output file: ``product_<product_id>_unit_barcodes_with_images_only.xlsx``
    Columns: ``seq_no, barcode, created_at, barcode_image_path, barcode_image``
    """
    import pandas as pd

    conn = None
    try:
        conn = sqlite3.connect("testing_system.db", timeout=10.0)
        cur = conn.cursor()

        # Ensure table has a created_at column (timestamp)
        try:
            cur.execute("ALTER TABLE product_barcodes ADD COLUMN created_at DATETIME DEFAULT CURRENT_TIMESTAMP")
        except sqlite3.OperationalError:
            pass  # column already exists

        cur.execute(
            "SELECT seq_no, barcode, created_at, barcode_image FROM product_barcodes WHERE product_id = ? ORDER BY seq_no",
            (product_id,),
        )
        rows = cur.fetchall()
        if not rows:
            print(f"⚠️  No unit barcodes found for product {product_id}. Export skipped.")
            return

        df = pd.DataFrame(rows, columns=["seq_no", "barcode", "created_at", "barcode_image_path"])
        file_name = f"product_{product_id}_unit_barcodes_with_images_only.xlsx"

        # ------------------------------------------------------------------
        # Write to Excel with embedded images using xlsxwriter
        # ------------------------------------------------------------------
        with pd.ExcelWriter(file_name, engine="xlsxwriter") as writer:
            wb = writer.book

            # Write dataframe to excel
            df.to_excel(writer, sheet_name="Unit_Barcodes", index=False)
            ws = writer.sheets["Unit_Barcodes"]

            # Set row height for barcode images
            ws.set_default_row(45)

            # Find the barcode image path column index
            barcode_img_col = df.columns.get_loc("barcode_image_path")

            # Add a new column for actual barcode images
            ws.write(0, barcode_img_col + 1, "Barcode Image")

            # Insert barcode images
            for idx, row in df.iterrows():
                barcode_img_path = row["barcode_image_path"]
                excel_row = idx + 1  # +1 because Excel is 1-indexed and we have headers

                if barcode_img_path and os.path.exists(barcode_img_path):
                    try:
                        # Insert image with proper sizing
                        ws.insert_image(
                            excel_row, barcode_img_col + 1,  # row, col
                            barcode_img_path,
                            {
                                'x_scale': 0.8,  # Scale down to fit in cell
                                'y_scale': 0.8,
                                'x_offset': 5,  # Small offset from cell border
                                'y_offset': 5
                            }
                        )
                    except Exception as img_err:
                        print(f"Failed to insert image {barcode_img_path}: {img_err}")
                        # Write "Image Error" text as fallback
                        ws.write(excel_row, barcode_img_col + 1, "Image Error")
                else:
                    # Write "No Image" if path doesn't exist
                    ws.write(excel_row, barcode_img_col + 1, "No Image")

            # Set column widths for better display
            ws.set_column(0, barcode_img_col, 15)  # Standard columns
            ws.set_column(barcode_img_col + 1, barcode_img_col + 1, 25)  # Barcode image column wider

        print(f"✅ Unit barcodes Excel exported with embedded images: {file_name} ({len(rows)} rows)")
    except Exception as e:
        print(f"❌ Error exporting unit barcodes for product {product_id} with images: {e}")
    finally:
        if conn:
            conn.close()


# === Pending Product Detail Dialog ===
class PendingProductDetailDialog(ProductDetailDialog):
    """Specialized dialog for showing product details from Pending Approval page.
    This version never shows the recall button and has pending-specific styling."""

    def __init__(self, product_data, parent=None):
        # Always set allow_recall to False for pending products
        super().__init__(product_data, parent, allow_recall=False)

        # Override window title for pending context
        self.setWindowTitle("Pending Product Details")

        # Add pending-specific styling
        self.setStyleSheet("""
            QDialog {
                background-color: #FFF8E1;
                border: 2px solid #FF9800;
                border-radius: 15px;
            }
        """)

    def init_ui(self):
        """Initialize the user interface with pending-specific modifications"""
        # Call parent init_ui
        super().init_ui()

        # Find and update the title label to reflect pending status
        title_labels = self.findChildren(QLabel)
        for label in title_labels:
            if "Product Information" in label.text():
                label.setText("⏳ Pending Product Information")
                label.setStyleSheet("""
                    QLabel {
                        font-size: 24px;
                        font-weight: bold;
                        color: #E65100;
                        margin-bottom: 10px;
                    }
                """)
                break


class EnhancedEmailSettingsDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("📬 Enhanced Email Settings")
        self.setFixedSize(1400, 1000)
        self.setStyleSheet("""
            QDialog {
                background-color: #f5f5f5;
                font-size: 14px;
            }
            QTabWidget::pane {
                border: 2px solid #4CAF50;
                border-radius: 8px;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #e8f5e8;
                color: #2e7d32;
                padding: 12px 20px;
                margin: 2px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: bold;
            }
            QTabBar::tab:selected {
                background-color: #4CAF50;
                color: white;
            }
            QTabBar::tab:hover {
                background-color: #66bb6a;
                color: white;
            }
            QLineEdit, QTextEdit, QComboBox, QSpinBox {
                border: 2px solid #4CAF50;
                border-radius: 5px;
                padding: 8px;
                background-color: white;
                font-size: 14px;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #4CAF50;
                border-radius: 5px;
                margin-top: 15px;
                padding-top: 15px;
                font-size: 14px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 10px 0 10px;
                background-color: white;
                color: #2e7d32;
            }
            QLabel {
                font-size: 14px;
                color: #333;
            }
            QCheckBox {
                font-size: 14px;
                color: #333;
            }
        """)

        self.init_ui()
        self.load_current_settings()

    def init_ui(self):
        layout = QVBoxLayout()

        # Create tab widget
        self.tab_widget = QTabWidget()

        # Tab 1: SMTP Configuration
        smtp_tab = self.create_smtp_tab()
        self.tab_widget.addTab(smtp_tab, "🔧 SMTP Settings")

        # Tab 2: Assignment Email Template
        assignment_tab = self.create_assignment_template_tab()
        self.tab_widget.addTab(assignment_tab, "👥 Assignment Emails")

        # Tab 3: Maturation Email Template
        maturation_tab = self.create_maturation_template_tab()
        self.tab_widget.addTab(maturation_tab, "⏰ Maturation Emails")

        layout.addWidget(self.tab_widget)

        # Bottom buttons
        button_layout = QHBoxLayout()

        save_button = QPushButton("💾 Save All Settings")
        save_button.clicked.connect(self.save_all_settings)
        save_button.setStyleSheet("""
            QPushButton {
                background-color: #2E7D32;
                padding: 12px 24px;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1B5E20;
            }
        """)

        test_assignment_button = QPushButton("📧 Test Assignment Email")
        test_assignment_button.clicked.connect(self.test_assignment_email)

        test_maturation_button = QPushButton("📧 Test Maturation Email")
        test_maturation_button.clicked.connect(self.test_maturation_email)

        cancel_button = QPushButton("❌ Cancel")
        cancel_button.clicked.connect(self.reject)

        button_layout.addWidget(save_button)
        button_layout.addWidget(test_assignment_button)
        button_layout.addWidget(test_maturation_button)
        button_layout.addStretch()
        button_layout.addWidget(cancel_button)

        layout.addLayout(button_layout)
        self.setLayout(layout)

    def create_smtp_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()

        # SMTP Settings Group
        smtp_group = QGroupBox("📡 SMTP Configuration (Shared for all emails)")
        smtp_layout = QFormLayout()

        self.smtp_server = QLineEdit()
        self.smtp_server.setPlaceholderText("smtp.gmail.com")
        smtp_layout.addRow("SMTP Server:", self.smtp_server)

        self.smtp_port = QSpinBox()
        self.smtp_port.setRange(1, 65535)
        self.smtp_port.setValue(587)
        smtp_layout.addRow("Port:", self.smtp_port)

        self.sender_email = QLineEdit()
        self.sender_email.setPlaceholderText("your-email@company.com")
        smtp_layout.addRow("Sender Email:", self.sender_email)

        self.sender_password = QLineEdit()
        self.sender_password.setEchoMode(QLineEdit.Password)
        self.sender_password.setPlaceholderText("Your App Password")
        smtp_layout.addRow("Password:", self.sender_password)

        self.sender_name = QLineEdit()
        self.sender_name.setPlaceholderText("Laboratory Management System")
        smtp_layout.addRow("Sender Name:", self.sender_name)

        smtp_group.setLayout(smtp_layout)
        layout.addWidget(smtp_group)

        # Notification Schedule Group
        schedule_group = QGroupBox("📅 Notification Schedule")
        schedule_layout = QFormLayout()

        self.notify_maturation = QCheckBox("Send maturation notifications")
        self.notify_expired = QCheckBox("Send expired product alerts")
        self.notify_assignments = QCheckBox("Send assignment notifications")

        schedule_layout.addRow("Enable notifications for:", self.notify_maturation)
        schedule_layout.addRow("", self.notify_expired)
        schedule_layout.addRow("", self.notify_assignments)

        self.notification_frequency = QComboBox()
        self.notification_frequency.addItems(["Daily", "Weekly", "Monthly"])
        schedule_layout.addRow("Frequency:", self.notification_frequency)

        schedule_group.setLayout(schedule_layout)
        layout.addWidget(schedule_group)

        layout.addStretch()
        widget.setLayout(layout)
        return widget

    def create_assignment_template_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()

        # Template info
        info_label = QLabel("📋 This template is used when testers are assigned to products.")
        info_label.setStyleSheet("color: #666; font-style: italic; margin-bottom: 10px;")
        layout.addWidget(info_label)

        # Subject template
        subject_group = QGroupBox("📝 Email Subject Template")
        subject_layout = QVBoxLayout()

        self.assignment_subject = QLineEdit()
        self.assignment_subject.setPlaceholderText("🧪 New Product Assignment - {PRODUCT_NAME} (Batch: {BATCH})")
        subject_layout.addWidget(self.assignment_subject)

        subject_group.setLayout(subject_layout)
        layout.addWidget(subject_group)

        # Body template
        body_group = QGroupBox("📄 Email Body Template")
        body_layout = QVBoxLayout()

        self.assignment_body = QTextEdit()
        self.assignment_body.setPlainText("""Dear {TESTER_NAME},

You have been assigned a new product for testing in the Laboratory Management System.

📋 ASSIGNMENT DETAILS:
* * * * * * * * * * * *
• Product Name: {PRODUCT_NAME}
• Batch Number: {BATCH}
• Product ID: {PRODUCT_ID}
• Product Owner: {OWNER_NAME} ({OWNER_USERNAME})
• Assigned By: {ASSIGNED_BY}
• Assignment Date: {ASSIGNMENT_DATE}

* * * * * * * * * * * *

Please log into the system to begin testing.

This is an automated message. Please do not reply directly to this email.

Best regards,
Laboratory Management System""")

        body_layout.addWidget(self.assignment_body)
        body_group.setLayout(body_layout)
        layout.addWidget(body_group)

        # Available variables
        variables_group = QGroupBox("🏷️ Available Variables")
        variables_layout = QVBoxLayout()
        variables_text = QLabel("""
Available placeholders for assignment emails:
• {TESTER_NAME} - Name of the assigned tester
• {PRODUCT_NAME} - Name of the product
• {BATCH} - Batch number
• {PRODUCT_ID} - Product ID
• {OWNER_NAME} - Product owner's full name
• {OWNER_USERNAME} - Product owner's username
• {ASSIGNED_BY} - Name of person who made the assignment
• {ASSIGNMENT_DATE} - Date and time of assignment
        """)
        variables_text.setStyleSheet("color: #666; font-family: 'Courier New', monospace;")
        variables_layout.addWidget(variables_text)
        variables_group.setLayout(variables_layout)
        layout.addWidget(variables_group)

        widget.setLayout(layout)
        return widget

    def create_maturation_template_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()

        # Template info
        info_label = QLabel("⏰ This template is used for product expiry/maturation notifications.")
        info_label.setStyleSheet("color: #666; font-style: italic; margin-bottom: 10px;")
        layout.addWidget(info_label)

        # Subject template
        subject_group = QGroupBox("📝 Email Subject Template")
        subject_layout = QVBoxLayout()

        self.maturation_subject = QLineEdit()
        self.maturation_subject.setPlaceholderText(
            "🚨 Product Expiry Alert - {PRODUCT_COUNT} Product(s) Require Attention")
        subject_layout.addWidget(self.maturation_subject)

        subject_group.setLayout(subject_layout)
        layout.addWidget(subject_group)

        # Body template
        body_group = QGroupBox("📄 Email Body Template")
        body_layout = QVBoxLayout()

        self.maturation_body = QTextEdit()
        self.maturation_body.setPlainText("""Dear {OWNER_NAME},

This is an automated notification regarding product(s) under your ownership that are approaching their expiry dates:

{PRODUCT_LIST}

* * * * * * * * * * * *

Please take immediate action to:
1. Review the product condition
2. Process or transfer if still viable
3. Report any issues to the system administrator

This notification was sent on: {NOTIFICATION_DATE}

Best regards,
Laboratory Management System""")

        body_layout.addWidget(self.maturation_body)
        body_group.setLayout(body_layout)
        layout.addWidget(body_group)

        # Available variables
        variables_group = QGroupBox("🏷️ Available Variables")
        variables_layout = QVBoxLayout()
        variables_text = QLabel("""
Available placeholders for maturation emails:
• {OWNER_NAME} - Product owner's name
• {PRODUCT_COUNT} - Number of products nearing expiry
• {PRODUCT_LIST} - Formatted list of products with details
• {NOTIFICATION_DATE} - Date and time of notification
• {NOTIFICATION_TYPE} - "Manual Reminder" or "Automated Alert"
        """)
        variables_text.setStyleSheet("color: #666; font-family: 'Courier New', monospace;")
        variables_layout.addWidget(variables_text)
        variables_group.setLayout(variables_layout)
        layout.addWidget(variables_group)

        widget.setLayout(layout)
        return widget

    def load_current_settings(self):
        """Load current email settings from database"""
        try:
            import sqlite3
            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Load SMTP settings
            cursor.execute("SELECT * FROM email_config WHERE id = 1")
            smtp_config = cursor.fetchone()

            if smtp_config:
                self.smtp_server.setText(smtp_config[1] or "")
                self.smtp_port.setValue(smtp_config[2] or 587)
                self.sender_email.setText(smtp_config[3] or "")
                self.sender_password.setText(smtp_config[4] or "")
                self.sender_name.setText(smtp_config[5] or "Laboratory Management System")

            # Create email_templates table if it doesn't exist
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS email_templates (
                    id INTEGER PRIMARY KEY,
                    template_type TEXT UNIQUE,
                    subject_template TEXT,
                    body_template TEXT,
                    created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    modified_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Load email templates
            cursor.execute("SELECT template_type, subject_template, body_template FROM email_templates")
            templates = cursor.fetchall()

            for template_type, subject, body in templates:
                if template_type == 'assignment':
                    self.assignment_subject.setText(subject or "")
                    self.assignment_body.setPlainText(body or "")
                elif template_type == 'maturation':
                    self.maturation_subject.setText(subject or "")
                    self.maturation_body.setPlainText(body or "")

            conn.close()

        except Exception as e:
            print(f"Error loading email settings: {e}")

    def save_all_settings(self):
        """Save all email settings to database"""
        try:
            import sqlite3
            from PyQt5.QtWidgets import QMessageBox

            conn = sqlite3.connect("testing_system.db")
            cursor = conn.cursor()

            # Save SMTP settings
            cursor.execute("""
                INSERT OR REPLACE INTO email_config 
                (id, smtp_server, smtp_port, sender_email, sender_password, sender_name, use_tls)
                VALUES (1, ?, ?, ?, ?, ?, 1)
            """, (
                self.smtp_server.text(),
                self.smtp_port.value(),
                self.sender_email.text(),
                self.sender_password.text(),
                self.sender_name.text()
            ))

            # Save email templates
            templates = [
                ('assignment', self.assignment_subject.text(), self.assignment_body.toPlainText()),
                ('maturation', self.maturation_subject.text(), self.maturation_body.toPlainText())
            ]

            for template_type, subject, body in templates:
                cursor.execute("""
                    INSERT OR REPLACE INTO email_templates 
                    (template_type, subject_template, body_template, modified_date)
                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                """, (template_type, subject, body))

            conn.commit()
            conn.close()

            QMessageBox.information(self, "Settings Saved",
                                    "✅ All email settings have been saved successfully!\n\n"
                                    "📧 SMTP configuration updated\n"
                                    "📝 Assignment email template saved\n"
                                    "⏰ Maturation email template saved")
            self.accept()

        except Exception as e:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.critical(self, "Save Error", f"Failed to save settings: {str(e)}")

    def test_assignment_email(self):
        """Send a real test assignment email"""
        from PyQt5.QtWidgets import QMessageBox, QInputDialog

        # Get recipient email
        recipient_email, ok = QInputDialog.getText(
            self,
            "📧 Test Assignment Email",
            "Enter recipient email address to send test assignment email:",
            text="test@example.com"
        )

        if not ok or not recipient_email.strip():
            return

        try:
            # Get current template
            subject_template = self.assignment_subject.text() or "🧪 New Product Assignment - {PRODUCT_NAME} (Batch: {BATCH})"
            body_template = self.assignment_body.toPlainText() or """Dear {TESTER_NAME},

This is a test assignment email.

📋 ASSIGNMENT DETAILS:
* * * * * * * * * * * *
• Product Name: {PRODUCT_NAME}
• Batch Number: {BATCH}
• Product ID: {PRODUCT_ID}
• Product Owner: {OWNER_NAME} ({OWNER_USERNAME})
• Assigned By: {ASSIGNED_BY}
• Assignment Date: {ASSIGNMENT_DATE}

* * * * * * * * * * * *

Please log into the system to begin testing.

This is an automated message. Please do not reply directly to this email.

Best regards,
Laboratory Management System"""

            # Fill template with test data
            test_data = {
                'TESTER_NAME': 'Test Tester',
                'PRODUCT_NAME': 'Sample Product XYZ',
                'BATCH': 'TEST001',
                'PRODUCT_ID': '12345',
                'OWNER_NAME': 'Product Owner',
                'OWNER_USERNAME': 'owner123',
                'ASSIGNED_BY': 'Admin User',
                'ASSIGNMENT_DATE': format_malaysia_time()
            }

            subject = subject_template.format(**test_data)
            body = body_template.format(**test_data)

            # Send test email
            success = self.send_test_email(recipient_email, subject, body)

            if success:
                QMessageBox.information(self, "✅ Test Email Sent",
                                        f"Test assignment email sent successfully to:\n{recipient_email}\n\n"
                                        f"Subject: {subject}")
            else:
                QMessageBox.warning(self, "❌ Test Email Failed",
                                    "Failed to send test email. Please check your SMTP configuration.")

        except Exception as e:
            QMessageBox.critical(self, "❌ Error", f"Error sending test email: {str(e)}")

    def test_maturation_email(self):
        """Send a real test maturation email"""
        from PyQt5.QtWidgets import QMessageBox, QInputDialog

        # Get recipient email
        recipient_email, ok = QInputDialog.getText(
            self,
            "📧 Test Maturation Email",
            "Enter recipient email address to send test maturation email:",
            text="test@example.com"
        )

        if not ok or not recipient_email.strip():
            return

        try:
            # Get current template
            subject_template = self.maturation_subject.text() or "🚨 Product Expiry {NOTIFICATION_TYPE} - {PRODUCT_COUNT} Product(s) Require Attention"
            body_template = self.maturation_body.toPlainText() or """Dear {OWNER_NAME},

{INTRO_TEXT} regarding product(s) under your ownership that are approaching their expiry dates:

{PRODUCT_LIST}

* * * * * * * * * * * *

Please take immediate action to:
1. Review the product condition
2. Process or transfer if still viable
3. Report any issues to the system administrator

This notification was sent on: {NOTIFICATION_DATE}

Best regards,
Laboratory Management System"""

            # Fill template with test data
            test_product_list = """
• Product: Test Product A (Batch: TEST001)
  Expiry Date: 2024-02-15
  Days Remaining: 7 days - ⚠️ CRITICAL

• Product: Test Product B (Batch: TEST002)
  Expiry Date: 2024-02-20
  Days Remaining: 12 days - 🔴 WARNING
"""

            test_data = {
                'OWNER_NAME': 'Product Owner',
                'NOTIFICATION_TYPE': 'Manual Test',
                'PRODUCT_COUNT': '2',
                'INTRO_TEXT': 'This is a test notification',
                'PRODUCT_LIST': test_product_list,
                'NOTIFICATION_DATE': format_malaysia_time()
            }

            subject = subject_template.format(**test_data)
            body = body_template.format(**test_data)

            # Send test email
            success = self.send_test_email(recipient_email, subject, body)

            if success:
                QMessageBox.information(self, "✅ Test Email Sent",
                                        f"Test maturation email sent successfully to:\n{recipient_email}\n\n"
                                        f"Subject: {subject}")
            else:
                QMessageBox.warning(self, "❌ Test Email Failed",
                                    "Failed to send test email. Please check your SMTP configuration.")

        except Exception as e:
            QMessageBox.critical(self, "❌ Error", f"Error sending test email: {str(e)}")

    def send_test_email(self, recipient_email, subject, body):
        """Send actual test email using current SMTP settings"""
        try:
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart
            from email.utils import formataddr

            # Get SMTP settings from current form
            smtp_server = self.smtp_server.text() or "smtp.gmail.com"
            smtp_port = self.smtp_port.value() or 587
            sender_email = self.sender_email.text()
            sender_password = self.sender_password.text()
            sender_name = self.sender_name.text() or "Laboratory Management System"

            if not sender_email or not sender_password:
                print("❌ Email configuration incomplete - missing sender email or password")
                return False

            # Create message
            msg = MIMEMultipart()
            msg['From'] = formataddr((sender_name, sender_email))
            msg['To'] = recipient_email
            msg['Subject'] = f"[TEST] {subject}"

            # Add note that this is a test email
            test_body = f"""*** THIS IS A TEST EMAIL ***

{body}

*** END OF TEST EMAIL ***

This test email was sent from the Laboratory Management System email settings configuration."""

            msg.attach(MIMEText(test_body, 'plain'))

            # Create SMTP session
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()  # Enable security
            server.login(sender_email, sender_password)

            # Send email
            text = msg.as_string()
            server.sendmail(sender_email, [recipient_email], text)
            server.quit()

            print(f"✅ Test email sent successfully to {recipient_email}")
            return True

        except Exception as e:
            print(f"❌ Error sending test email: {e}")
            return False


__all__ = [
    "verify_user_login", "get_all_messages", "insert_message", "get_all_users_for_chat", "get_private_messages",
    "insert_private_message", "format_malaysia_time", "create_chat_tables",
    "SuperAdminDialog", "SummaryCard", "ProductCardWithCheckbox", "ProductCard", "ExpiredProductsDialog",
    "EmailSettingsDialog", "EnhancedEmailSettingsDialog", "UserManagementDialog", "SystemSettingsDialog",
    "ExportReportsDialog",
    "ProductDetailDialog", "AddProductDialog", "ManageProductsDialog", "EditProductDialog", "CreateAccountDialog",
    "export_today_products_barcodes_to_excel_with_images", "export_product_with_unit_barcodes_to_excel_with_images",
    "export_unit_barcodes_to_excel_with_images", "PendingProductDetailDialog"
]
