import sys
import sqlite3
import os
import datetime
from datetime import datetime, timedelta, date
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import pandas as pd
from fpdf import FPDF
import platform
from openpyxl.utils import get_column_letter
import yagmail
from PIL import Image
from PyQt5.QtWidgets import QApplication

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
db_path = os.path.join(BASE_DIR, "medical_system.db")

print(f"Database full path: {db_path}")


def create_database_tables():
    """Create all database tables with proper schema"""
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            cursor.executescript('''
                CREATE TABLE IF NOT EXISTS Users (
                    user_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT NOT NULL UNIQUE,
                    password TEXT NOT NULL,
                    role TEXT NOT NULL,
                    email TEXT NOT NULL UNIQUE,
                    phone_no TEXT,
                    branch TEXT,
                    batch TEXT
                );

                CREATE TABLE IF NOT EXISTS Products (
                    product_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    product_name TEXT NOT NULL,
                    manufacture_date DATE NOT NULL,
                    expired_date DATE NOT NULL,
                    arrival_date DATE NOT NULL,
                    location TEXT NOT NULL,
                    batch TEXT NOT NULL,
                    barcode TEXT UNIQUE,
                    excel_name TEXT,
                    sku TEXT,
                    branch TEXT,
                    product_image BLOB,
                    FOREIGN KEY (user_id) REFERENCES Users(user_id)
                );

                CREATE TABLE IF NOT EXISTS Testing (
                    test_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    product_id INTEGER NOT NULL,
                    test_start DATETIME NOT NULL,
                    test_end DATETIME,
                    test_status TEXT NOT NULL,
                    test_result TEXT,
                    test_image TEXT,
                    FOREIGN KEY (user_id) REFERENCES Users(user_id),
                    FOREIGN KEY (product_id) REFERENCES Products(product_id)
                );

                CREATE TABLE IF NOT EXISTS SentEmails (
                    email_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sender_email TEXT NOT NULL,
                    receiver_email TEXT NOT NULL,
                    cc_emails TEXT,
                    email_subject TEXT,
                    email_message TEXT,
                    sent_at DATETIME DEFAULT CURRENT_TIMESTAMP
                );

                CREATE TABLE IF NOT EXISTS Notifications (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    message TEXT,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                    read INTEGER DEFAULT 0
                );

                CREATE TABLE IF NOT EXISTS product_tester_assignments (
                    assignment_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id INTEGER NOT NULL,
                    tester_id INTEGER NOT NULL,
                    FOREIGN KEY (product_id) REFERENCES Products(product_id),
                    FOREIGN KEY (tester_id) REFERENCES Users(user_id),
                    UNIQUE(product_id, tester_id)
                );
            ''')
            print("Database tables created successfully.")
    except sqlite3.Error as e:
        print(f"Database error occurred: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")


class DateEdit(QDateEdit):
    """Custom DateEdit widget to match original functionality"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setDate(QDate.currentDate())
        self.setCalendarPopup(True)
        self.setDisplayFormat("yyyy-MM-dd")


class TesterHomePage(QMainWindow):
    def __init__(self):
        super().__init__()
        self.user_id = self.get_current_user_id()
        self.attached_file_path = None
        self.init_ui()

    def get_current_user_id(self):
        # For testing purposes, return 1 (the tester user)
        return 1

    def init_ui(self):
        self.setWindowTitle("Tester Dashboard")
        self.setGeometry(100, 100, 1400, 900)

        # Create central widget and main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)

        # Create sidebar
        self.create_sidebar(main_layout)

        # Create main area
        self.create_main_area(main_layout)

        # Initialize with Product Assignments
        self.switch_frame("Product Assignments")

    def create_sidebar(self, main_layout):
        sidebar = QWidget()
        sidebar.setFixedWidth(300)
        sidebar.setStyleSheet("background-color: #FCA6FF;")

        sidebar_layout = QVBoxLayout(sidebar)

        # Title
        title = QLabel("Tester Panel")
        title.setStyleSheet("color: #EEEEEE; font-size: 30px; font-weight: bold;")
        title.setAlignment(Qt.AlignCenter)
        sidebar_layout.addWidget(title)
        sidebar_layout.addSpacing(20)

        # Menu options
        options = [
            "Product Assignments",
            "Test Result Update",
            "Test Schedule",
            "Progress Tracker",
            "Test History",
            "Test Analysis",
            "Notifications",
            "Send Email",
        ]

        self.buttons = {}
        for opt in options:
            btn = QPushButton(opt)
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #34495e;
                    color: white;
                    border: none;
                    padding: 15px;
                    font-size: 25px;
                    text-align: left;
                }
                QPushButton:hover {
                    background-color: #2c3e50;
                }
                QPushButton:pressed {
                    background-color: #1a252f;
                }
            """)
            btn.clicked.connect(lambda checked, o=opt: self.switch_frame(o))
            sidebar_layout.addWidget(btn)
            self.buttons[opt] = btn

        sidebar_layout.addStretch()
        main_layout.addWidget(sidebar)

    def create_main_area(self, main_layout):
        self.main_frame = QWidget()
        self.main_frame.setStyleSheet("background-color: white;")
        main_layout.addWidget(self.main_frame, 1)

    def clear_main_frame(self):
        layout = self.main_frame.layout()
        if layout is not None:
            while layout.count():
                child = layout.takeAt(0)
                if child.widget():
                    child.widget().deleteLater()
        else:
            layout = QVBoxLayout()
            self.main_frame.setLayout(layout)

    def switch_frame(self, section):
        self.clear_main_frame()
        layout = self.main_frame.layout()

        if section == "Product Assignments":
            self.show_product_assignments(layout)
        elif section == "Test Result Update":
            self.show_test_result_update(layout)
        elif section == "Test Schedule":
            self.show_test_schedule(layout)
        elif section == "Progress Tracker":
            self.show_progress_tracker(layout)
        elif section == "Test History":
            self.show_test_history(layout)
        elif section == "Test Analysis":
            self.show_test_analysis(layout)
        elif section == "Notifications":
            self.show_notifications(layout)
        elif section == "Send Email":
            self.show_send_email(layout)

    def show_product_assignments(self, layout):
        title_label = QLabel("Assigned Products")
        title_label.setStyleSheet("font-size: 30px; font-weight: bold;")
        layout.addWidget(title_label)

        # Create table
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Product ID", "Product Name", "Status", "Location", "Batch"])

        # Set column widths
        header = table.horizontalHeader()
        header.resizeSection(0, 300)
        header.resizeSection(1, 500)
        header.resizeSection(2, 500)
        header.resizeSection(3, 500)
        header.resizeSection(4, 340)

        try:
            conn = sqlite3.connect('medical_system.db')
            cursor = conn.cursor()
            cursor.execute("""
                    SELECT 
                        P.product_id,
                        P.product_name, 
                        T.test_status,
                        P.location,
                        P.batch
                    FROM Products P
                    INNER JOIN product_tester_assignments A ON P.product_id = A.product_id
                    LEFT JOIN Testing T ON T.product_id = P.product_id AND T.user_id = A.tester_id
                    WHERE A.tester_id = ?
                    ORDER BY P.product_name
                """, (self.user_id,))

            rows = cursor.fetchall()

            if rows:
                table.setRowCount(len(rows))
                for i, row in enumerate(rows):
                    for j, value in enumerate(row):
                        item = QTableWidgetItem(str(value) if value else "")
                        table.setItem(i, j, item)
            else:
                no_data_label = QLabel("No products assigned")
                no_data_label.setAlignment(Qt.AlignCenter)
                no_data_label.setStyleSheet("font-size: 20px; padding: 20px;")
                layout.addWidget(no_data_label)
                return

        except sqlite3.Error as e:
            error_label = QLabel(f"Error fetching product assignments: {e}")
            error_label.setStyleSheet("color: red; font-size: 12px; padding: 20px;")
            layout.addWidget(error_label)
            return
        finally:
            if conn:
                conn.close()

        layout.addWidget(table)


    def show_test_result_update(self, layout):
        title_label = QLabel("Update Test Result Status")
        title_label.setStyleSheet("font-size: 30px; font-weight: bold;")
        layout.addWidget(title_label)

        form_layout = QFormLayout()

        # Product selection
        self.product_combo = QComboBox()
        self.product_combo.setMinimumWidth(400)

        try:
            with sqlite3.connect('medical_system.db') as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT P.product_id, P.product_name 
                    FROM Products P
                    INNER JOIN product_tester_assignments A ON P.product_id = A.product_id
                    WHERE A.tester_id = ?
                """, (self.get_current_user_id(),))
                products = cursor.fetchall()

                if products:
                    for p in products:
                        self.product_combo.addItem(f"{p[0]} - {p[1]}")
                else:
                    self.product_combo.addItem("No products assigned")
        except sqlite3.Error as e:
            print(f"Error loading products: {e}")
            QMessageBox.critical(self, "Database Error", f"Failed to load products: {e}")

        form_layout.addRow("Select Product:", self.product_combo)

        # Status selection
        self.status_combo = QComboBox()
        self.status_combo.addItems(["Not Started", "In Progress", "Completed", "Failed", "Retest Required"])
        form_layout.addRow("Test Status:", self.status_combo)

        # Test result
        self.result_text = QTextEdit()
        self.result_text.setMaximumHeight(100)
        form_layout.addRow("Test Result:", self.result_text)

        # Image upload
        image_layout = QVBoxLayout()

        self.image_label = QLabel()
        self.image_label.setFixedSize(300, 200)
        self.image_label.setStyleSheet("border: 1px solid gray;")
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setText("No image selected")

        self.image_path_label = QLabel("")

        attach_btn = QPushButton("Attach File")
        attach_btn.clicked.connect(self.attach_file)

        image_layout.addWidget(QLabel("Upload Image:"))
        image_layout.addWidget(self.image_label)
        image_layout.addWidget(self.image_path_label)
        image_layout.addWidget(attach_btn)

        # Update button
        update_btn = QPushButton("Update Status")
        update_btn.clicked.connect(self.update_test_status)
        update_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)

        # Add to main layout
        form_widget = QWidget()
        form_widget.setLayout(form_layout)
        layout.addWidget(form_widget)

        image_widget = QWidget()
        image_widget.setLayout(image_layout)
        layout.addWidget(image_widget)

        layout.addWidget(update_btn)
        layout.addStretch()

    def show_test_schedule(self, layout):
        title_label = QLabel("Test Schedule")
        title_label.setStyleSheet("font-size: 30px; font-weight: bold;")
        layout.addWidget(title_label)

        # Date controls
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("From:"))

        self.start_date_entry = DateEdit()
        date_layout.addWidget(self.start_date_entry)

        date_layout.addWidget(QLabel("To:"))

        self.end_date_entry = DateEdit()
        date_layout.addWidget(self.end_date_entry)

        refresh_btn = QPushButton("Show Schedule")
        refresh_btn.clicked.connect(self.load_test_schedule)
        date_layout.addWidget(refresh_btn)

        date_layout.addStretch()

        date_widget = QWidget()
        date_widget.setLayout(date_layout)
        layout.addWidget(date_widget)

        # Schedule table
        self.schedule_table = QTableWidget()
        self.schedule_table.setColumnCount(4)
        self.schedule_table.setHorizontalHeaderLabels(["Product", "Status", "Test Start", "Test End"])
        layout.addWidget(self.schedule_table)

        # Load initial data
        self.load_test_schedule()

    def show_progress_tracker(self, layout):
        title_label = QLabel("Progress Tracker")
        title_label.setStyleSheet("font-size: 30px; font-weight: bold; margin-bottom: 10px;")
        layout.addWidget(title_label)

        # Create matplotlib figure
        fig = Figure(figsize=(8, 4))
        ax = fig.add_subplot(111)

        try:
            conn = sqlite3.connect("medical_system.db")
            cursor = conn.cursor()
            cursor.execute("""
                SELECT test_status, COUNT(*) 
                FROM Testing 
                WHERE user_id = ? 
                GROUP BY test_status
            """, (self.get_current_user_id(),))
            status_data = cursor.fetchall()

            if status_data:
                statuses, counts = zip(*status_data)
                color_map = {
                    "Completed": "green",
                    "In Progress": "blue",
                    "Failed": "red",
                    "Retest Required": "orange",
                    "Not Started": "gray"
                }
                bar_colors = [color_map.get(status, 'black') for status in statuses]
                ax.bar(statuses, counts, color=bar_colors)

                total = sum(counts)
                completed = sum(c for s, c in zip(statuses, counts) if s == "Completed")
                progress_percent = int((completed / total) * 100) if total > 0 else 0

                progress_label = QLabel(f"Total Progress: {progress_percent}%")
                progress_label.setStyleSheet("font-size: 20px; padding: 5px;")
                layout.addWidget(progress_label)
            else:
                ax.bar(['No Data'], [0], color=['gray'])
                no_data_label = QLabel("No test data available.")
                no_data_label.setStyleSheet("font-size: 12px; padding: 5px;")
                layout.addWidget(no_data_label)

        except sqlite3.Error as e:
            ax.bar(['Error'], [0], color=['red'])
            error_label = QLabel(f"Database error: {e}")
            error_label.setStyleSheet("color: red; font-size: 12px; padding: 5px;")
            layout.addWidget(error_label)
        finally:
            if conn:
                conn.close()

        ax.set_title("Test Result Summary")
        ax.set_ylabel("Number of Tests")

        canvas = FigureCanvas(fig)
        layout.addWidget(canvas)

    def show_test_history(self, layout):
        # Title
        title_label = QLabel("Test History")
        title_label.setStyleSheet("font-size: 30px; font-weight: bold; margin-bottom: 10px;")
        layout.addWidget(title_label)

        # Export buttons layout
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)

        export_excel_btn = QPushButton("Export to Excel")
        export_excel_btn.clicked.connect(lambda: self.export_history("excel"))
        button_layout.addWidget(export_excel_btn)

        export_pdf_btn = QPushButton("Export to PDF")
        export_pdf_btn.clicked.connect(lambda: self.export_history("pdf"))
        button_layout.addWidget(export_pdf_btn)

        button_layout.addStretch()

        layout.addLayout(button_layout)

        # Test history table
        self.history_table = QTableWidget()
        self.history_table.setColumnCount(5)
        self.history_table.setHorizontalHeaderLabels(["Product", "Status", "Start Date", "End Date", "Result"])
        self.history_table.setAlternatingRowColors(True)
        self.history_table.setStyleSheet("alternate-background-color: #f2f2f2; background-color: white;")
        self.history_table.horizontalHeader().setStretchLastSection(True)
        self.history_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.history_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.history_table.setEditTriggers(QTableWidget.NoEditTriggers)

        try:
            with sqlite3.connect('medical_system.db') as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT P.product_name, T.test_status, T.test_start, T.test_end, T.test_result
                    FROM Testing T
                    INNER JOIN Products P ON T.product_id = P.product_id
                    WHERE T.user_id = ?
                    ORDER BY T.test_start DESC
                """, (self.get_current_user_id(),))

                history_rows = cursor.fetchall()

                if history_rows:
                    self.history_table.setRowCount(len(history_rows))
                    for i, row in enumerate(history_rows):
                        for j, value in enumerate(row):
                            item = QTableWidgetItem(str(value) if value else "N/A")
                            item.setTextAlignment(Qt.AlignCenter)
                            self.history_table.setItem(i, j, item)
                else:
                    QMessageBox.information(self, "No Data", "You have no test history yet.")

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"Failed to load test history: {e}")

        layout.addWidget(self.history_table)

    def show_test_analysis(self, layout):
        title_label = QLabel("Test Analysis")
        title_label.setStyleSheet("font-size: 30px; font-weight: bold;")
        layout.addWidget(title_label)

        # Summary section
        summary_group = QGroupBox("Summary Statistics")
        summary_layout = QVBoxLayout(summary_group)

        # Charts section
        chart_group = QGroupBox("Charts")
        chart_layout = QVBoxLayout(chart_group)

        # Initialize data
        status_counts = {
            "Not Started": 0,
            "In Progress": 0,
            "Completed": 0,
            "Failed": 0,
            "Retest Required": 0
        }

        total_duration = 0
        duration_count = 0
        product_count = {}

        try:
            with sqlite3.connect('medical_system.db') as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT T.test_status, T.test_start, T.test_end, P.product_name
                    FROM Testing T
                    INNER JOIN Products P ON T.product_id = P.product_id
                    WHERE T.user_id = ?
                """, (self.get_current_user_id(),))

                rows = cursor.fetchall()
                for status, start, end, product in rows:
                    if status in status_counts:
                        status_counts[status] += 1

                    if start and end:
                        try:
                            start_dt = datetime.strptime(start, '%Y-%m-%d')
                            end_dt = datetime.strptime(end, '%Y-%m-%d')
                            duration = (end_dt - start_dt).days
                            if duration >= 0:
                                total_duration += duration
                                duration_count += 1
                        except:
                            pass

                    if product:
                        product_count[product] = product_count.get(product, 0) + 1

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"Failed to fetch analysis data: {e}")
            return

        total_tests = sum(status_counts.values())
        avg_duration = round(total_duration / duration_count, 2) if duration_count else 0

        # Display summary
        summary_text = f"""🧪 Total Tests: {total_tests}
✅ Completed: {status_counts['Completed']}
⏳ In Progress: {status_counts['In Progress']}
💤 Not Started: {status_counts['Not Started']}
❌ Failed: {status_counts['Failed']}
🔁 Retest Required: {status_counts['Retest Required']}
⏱ Avg Duration (for completed): {avg_duration} days"""

        summary_label = QLabel(summary_text)
        summary_label.setStyleSheet("font-family: monospace; padding: 10px;")
        summary_layout.addWidget(summary_label)

        # Create charts
        fig = Figure(figsize=(10, 4))

        # Pie chart
        ax1 = fig.add_subplot(121)
        labels = list(status_counts.keys())
        values = list(status_counts.values())
        colors = ['#9E9E9E', '#2196F3', '#4CAF50', '#F44336', '#FF9800']
        ax1.pie(values, labels=labels, autopct='%1.1f%%', colors=colors)
        ax1.set_title('Test Status Breakdown')

        # Bar chart
        ax2 = fig.add_subplot(122)
        if product_count:
            sorted_products = sorted(product_count.items(), key=lambda x: x[1], reverse=True)[:5]
            prod_labels, prod_values = zip(*sorted_products)
            ax2.bar(prod_labels, prod_values, color='skyblue')
            ax2.set_title('Top 5 Tested Products')
            ax2.tick_params(axis='x', rotation=30)
        else:
            ax2.text(0.5, 0.5, "No Product Data", ha='center', transform=ax2.transAxes)

        fig.tight_layout()
        canvas = FigureCanvas(fig)
        chart_layout.addWidget(canvas)

        # Add to main layout
        layout.addWidget(summary_group)
        layout.addWidget(chart_group)

    def show_notifications(self, layout):
        title_label = QLabel("Notifications")
        title_label.setStyleSheet("font-size: 30px; font-weight: bold;")
        layout.addWidget(title_label)

        current_user_id = self.get_current_user_id()

        # Insert expiry reminders
        self.insert_expiry_reminders(current_user_id)

        # Create table for notifications
        table = QTableWidget()
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(["Message", "Timestamp", "Read"])

        # Set column widths
        header = table.horizontalHeader()
        header.resizeSection(0, 400)
        header.resizeSection(1, 150)
        header.resizeSection(2, 60)

        try:
            conn = sqlite3.connect("medical_system.db")
            cursor = conn.cursor()
            cursor.execute("""
                SELECT message, timestamp, read FROM Notifications
                WHERE user_id = ?
                ORDER BY timestamp DESC
            """, (current_user_id,))
            notifications = cursor.fetchall()

            if notifications:
                table.setRowCount(len(notifications))
                for i, (message, timestamp, read) in enumerate(notifications):
                    table.setItem(i, 0, QTableWidgetItem(message))
                    table.setItem(i, 1, QTableWidgetItem(timestamp))
                    table.setItem(i, 2, QTableWidgetItem("Yes" if read else "No"))

            # Mark all unread notifications as read
            cursor.execute("""
                UPDATE Notifications SET read = 1
                WHERE user_id = ? AND read = 0
            """, (current_user_id,))
            conn.commit()

        except sqlite3.Error as e:
            error_label = QLabel(f"Error loading notifications: {e}")
            error_label.setStyleSheet("color: red; padding: 10px;")
            layout.addWidget(error_label)
            return
        finally:
            if conn:
                conn.close()

        layout.addWidget(table)

    def show_send_email(self, layout):
        title_label = QLabel("Send Email")
        title_label.setStyleSheet("font-size: 30px; font-weight: bold;")
        layout.addWidget(title_label)

        form_layout = QFormLayout()

        # Email input fields
        self.gmail_input = QLineEdit()
        self.gmail_input.setPlaceholderText("your.email@gmail.com")
        form_layout.addRow("Your Gmail Account:", self.gmail_input)

        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.setPlaceholderText("App Password")
        form_layout.addRow("App Password:", self.password_input)

        self.to_email_input = QLineEdit()
        self.to_email_input.setPlaceholderText("recipient@example.com")
        form_layout.addRow("To Email:", self.to_email_input)

        self.cc_input = QLineEdit()
        self.cc_input.setPlaceholderText("cc1@example.com, cc2@example.com")
        form_layout.addRow("CC (comma-separated):", self.cc_input)

        self.subject_input = QLineEdit()
        self.subject_input.setPlaceholderText("Email Subject")
        form_layout.addRow("Subject:", self.subject_input)

        self.message_input = QTextEdit()
        self.message_input.setMaximumHeight(200)
        self.message_input.setPlaceholderText("Type your message here...")
        form_layout.addRow("Message:", self.message_input)

        # Send button
        send_btn = QPushButton("Send Email")
        send_btn.clicked.connect(self.send_email_handler)
        send_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                padding: 10px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)

        # Add to layout
        form_widget = QWidget()
        form_widget.setLayout(form_layout)
        layout.addWidget(form_widget)
        layout.addWidget(send_btn)
        layout.addStretch()


    def attach_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Image",
            "",
            "Image files (*.png *.jpg *.jpeg *.bmp *.gif);;All files (*.*)"
        )

        if file_path:
            self.attached_file_path = file_path
            self.image_path_label.setText(f"Attached: {file_path}")

            # Load and display image
            try:
                pixmap = QPixmap(file_path)
                scaled_pixmap = pixmap.scaled(300, 200, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.image_label.setPixmap(scaled_pixmap)
            except Exception as e:
                self.image_label.setText("Error loading image")
                print(f"Error loading image: {e}")
        else:
            self.attached_file_path = None
            self.image_label.clear()
            self.image_label.setText("No image selected")
            self.image_path_label.setText("")

    def update_test_status(self):
        selected_product = self.product_combo.currentText()
        if not selected_product or selected_product == "No products assigned":
            QMessageBox.critical(self, "Error", "Please select a product")
            return

        test_status = self.status_combo.currentText()
        if not test_status:
            QMessageBox.critical(self, "Error", "Please select a test status")
            return

        try:
            product_id = selected_product.split(" - ")[0].strip()
        except IndexError:
            QMessageBox.critical(self, "Error", "Invalid product selection")
            return

        test_result = self.result_text.toPlainText().strip()
        user_id = self.get_current_user_id()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        image_path = self.attached_file_path if hasattr(self, 'attached_file_path') else None

        try:
            with sqlite3.connect('medical_system.db', timeout=10.0) as conn:
                cursor = conn.cursor()

                # Check if test record exists
                cursor.execute("SELECT test_id FROM Testing WHERE user_id = ? AND product_id = ?",
                               (user_id, product_id))
                row = cursor.fetchone()

                if row:
                    print("Updating existing test record...")
                    cursor.execute("""
                        UPDATE Testing
                        SET test_status = ?, test_result = ?, test_end = ?, test_image = ?
                        WHERE test_id = ?
                    """, (test_status, test_result, now, image_path, row[0]))
                else:
                    print("Inserting new test record...")
                    cursor.execute("""
                        INSERT INTO Testing (user_id, product_id, test_start, test_status, test_result, test_image)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (user_id, product_id, now, test_status, test_result, image_path))

                conn.commit()

                # Verify the update
                cursor.execute("SELECT * FROM Testing WHERE user_id = ? AND product_id = ?", (user_id, product_id))
                print("Current record:", cursor.fetchone())

                QMessageBox.information(self, "Success", "Test status updated successfully!")

                # Clear form
                self.product_combo.setCurrentText("")
                self.status_combo.setCurrentText("")
                self.result_text.clear()
                self.attached_file_path = None

        except sqlite3.Error as e:
            print("Database error:", e)
            QMessageBox.critical(self, "Database Error", f"Failed to update test status: {e}")
        except Exception as e:
            print("Unexpected error:", e)
            QMessageBox.critical(self, "Error", f"An unexpected error occurred: {e}")

    def apply_date_filter(self):
        start_date = self.start_date_entry.get_date()
        end_date = self.end_date_entry.get_date()
        self.load_test_schedule(self.user_id, start_date, end_date)

    def load_test_schedule(self):
        # Get selected dates
        start_date = self.start_date_entry.date().toPyDate()
        end_date = self.end_date_entry.date().toPyDate()

        if start_date > end_date:
            QMessageBox.critical(self, "Date Error", "Start date must not be after end date.")
            return

        self.schedule_table.setRowCount(0)

        # Database connection
        conn = sqlite3.connect('medical_system.db')
        cursor = conn.cursor()

        cursor.execute("""
            SELECT 
                P.product_name, 
                T.test_status,
                T.test_start,
                T.test_end
            FROM Products P
            INNER JOIN product_tester_assignments A ON P.product_id = A.product_id
            LEFT JOIN Testing T ON T.product_id = P.product_id AND T.user_id = A.tester_id
            WHERE A.tester_id = ?
              AND (
                  date(T.test_start) BETWEEN ? AND ?
                  OR date(T.test_end) BETWEEN ? AND ?
              )
            ORDER BY P.product_name
        """, (self.user_id, start_date, end_date, start_date, end_date))

        rows = cursor.fetchall()
        conn.close()

        self.schedule_table.setRowCount(len(rows))

        for row_index, (product_name, status, test_start, test_end) in enumerate(rows):
            overdue = False
            if status and status in ["Not Started", "In Progress"] and test_end:
                try:
                    end_dt = datetime.strptime(test_end, '%Y-%m-%d %H:%M:%S')
                    if end_dt.date() < datetime.today().date():
                        overdue = True
                except Exception:
                    pass

            self.schedule_table.setItem(row_index, 0, QTableWidgetItem(str(product_name or "")))
            self.schedule_table.setItem(row_index, 1, QTableWidgetItem(str(status or "")))
            self.schedule_table.setItem(row_index, 2, QTableWidgetItem(str(test_start or "")))
            self.schedule_table.setItem(row_index, 3, QTableWidgetItem(str(test_end or "")))

            if overdue:
                for col in range(4):
                    item = self.schedule_table.item(row_index, col)
                    if item:
                        item.setBackground(QColor(240, 128, 128))  # lightcoral
                        item.setForeground(QColor(255, 255, 255))  # white

    def export_history(self, table, export_type):
        """Export test history to Excel or PDF"""
        try:
            data = []
            # Extract data from QTableWidget
            for row in range(table.rowCount()):
                row_data = []
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)

            if not data:
                QMessageBox.warning(self, "No Data", "No test history to export")
                return

            # Extract column headers
            columns = []
            for col in range(table.columnCount()):
                header_item = table.horizontalHeaderItem(col)
                columns.append(header_item.text() if header_item else f"Column {col + 1}")

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"test_history_{timestamp}"

            if export_type == "excel":
                filename += ".xlsx"
                self.export_to_excel(columns, data, filename)
            elif export_type == "pdf":
                filename += ".pdf"
                self.export_to_pdf(columns, data, filename)

            QMessageBox.information(self, "Export Successful", f"Exported to {filename}")

            # Open file after export
            if platform.system() == "Windows":
                os.startfile(filename)
            elif platform.system() == "Darwin":  # macOS
                os.system(f"open {filename}")
            else:  # Linux
                os.system(f"xdg-open {filename}")

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export: {str(e)}")

    def export_to_excel(self, columns, data, filename):
        """Export data to Excel file"""
        df = pd.DataFrame(data, columns=columns)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')

            worksheet = writer.sheets['Sheet1']
            for i, col in enumerate(df.columns):
                col_width = max(df[col].astype(str).map(len).max(), len(col))
                worksheet.column_dimensions[get_column_letter(i + 1)].width = col_width + 2

    def export_to_pdf(self, columns, data, filename):
        pdf = FPDF()
        pdf.add_page()

        pdf.set_font("Arial" if platform.system() == "Windows" else "Helvetica", size=10)

        pdf.cell(200, 10, txt="Test History Report", ln=1, align='C')
        pdf.cell(200, 10, txt=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=1, align='C')

        col_widths = [40, 30, 30, 30, 60]

        # Header
        for i, col in enumerate(columns):
            pdf.cell(col_widths[i], 10, col, border=1)
        pdf.ln()

        # Rows
        for row in data:
            for i, item in enumerate(row):
                pdf.cell(col_widths[i], 10, str(item), border=1)
            pdf.ln()

        pdf.output(filename)

    def export_history(self, format_type):
        # Get column headers
        columns = [self.history_table.horizontalHeaderItem(i).text() for i in range(self.history_table.columnCount())]

        # Get table data
        data = []
        for row in range(self.history_table.rowCount()):
            row_data = []
            for col in range(self.history_table.columnCount()):
                item = self.history_table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)

        options = QFileDialog.Options()
        if format_type == "excel":
            filename, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)",
                                                      options=options)
            if filename:
                self.export_to_excel(columns, data, filename)
                QMessageBox.information(self, "Export", f"Excel file saved to:\n{filename}")
                self.open_file(filename)  # 👈 open after saving

        elif format_type == "pdf":
            filename, _ = QFileDialog.getSaveFileName(self, "Save PDF File", "", "PDF Files (*.pdf)", options=options)
            if filename:
                self.export_to_pdf(columns, data, filename)
                QMessageBox.information(self, "Export", f"PDF file saved to:\n{filename}")
                self.open_file(filename)  # 👈 open after saving

    def open_file(self, filepath):
        import subprocess
        import platform
        if platform.system() == 'Windows':
            os.startfile(filepath)
        elif platform.system() == 'Darwin':  # macOS
            subprocess.call(['open', filepath])
        else:  # Linux
            subprocess.call(['xdg-open', filepath])

    def insert_expiry_reminders(self, user_id):
        today = date.today()
        upcoming = today + timedelta(days=2)

        try:
            conn = sqlite3.connect("medical_system.db")
            cursor = conn.cursor()

            cursor.execute("""
                SELECT ap.product_id, ap.expired_date, p.product_name
                FROM Products ap
                JOIN Products p ON ap.product_id = p.product_id
                WHERE ap.user_id = ? AND DATE(ap.expired_date) BETWEEN DATE(?) AND DATE(?)
            """, (user_id, today, upcoming))

            rows = cursor.fetchall()

            for product_id, expiry_date, product_name in rows:
                # Prevent duplicate reminders
                cursor.execute("""
                    SELECT 1 FROM Notifications
                    WHERE user_id = ? AND message LIKE ? AND DATE(timestamp) = DATE(?)
                """, (user_id, f"%{product_name}%expires%", today))
                exists = cursor.fetchone()

                if not exists:
                    message = f"Reminder: Product '{product_name}' (ID: {product_id}) expires on {expiry_date}. Please test it soon."
                    cursor.execute("""
                        INSERT INTO Notifications (user_id, message, timestamp, read)
                        VALUES (?, ?, ?, 0)
                    """, (user_id, message, datetime.now()))

            conn.commit()

        except sqlite3.Error as e:
            print("Reminder insertion failed:", e)
            # Optional: Show error message to user if this is called from UI context
            if hasattr(self, 'parent') and self.parent:
                QMessageBox.critical(self.parent, "Database Error", f"Failed to insert expiry reminders: {e}")

        finally:
            if conn:
                conn.close()

    def test_insert_expiry_reminder(self):
        """Test method for expiry reminder insertion"""
        # Setup test data
        user_id = 1
        product_id = 200
        product_name = "Test Kit B"
        expiry_date = (date.today() + timedelta(days=1)).isoformat()

        try:
            conn = sqlite3.connect("medical_system.db")
            cursor = conn.cursor()
            cursor.execute("INSERT OR IGNORE INTO Products (product_id, product_name, description) VALUES (?, ?, ?)",
                           (product_id, product_name, "Test kit"))
            cursor.execute(
                "INSERT OR REPLACE INTO AssignedProducts (product_id, tester_id, expiry_date) VALUES (?, ?, ?)",
                (product_id, user_id, expiry_date))
            conn.commit()
            conn.close()

            # Run the method
            self.insert_expiry_reminders(user_id)

            # Check the notification
            conn = sqlite3.connect("medical_system.db")
            cursor = conn.cursor()
            cursor.execute("SELECT message FROM Notifications WHERE user_id = ? AND message LIKE ?",
                           (user_id, f"%{product_name}%"))
            result = cursor.fetchone()
            conn.close()

            if result:
                print("✅ Reminder successfully inserted:", result[0])
                # Optional: Show success message in UI
                if hasattr(self, 'parent') and self.parent:
                    QMessageBox.information(self.parent, "Test Successful",
                                            f"Reminder successfully inserted: {result[0]}")
                return True
            else:
                print("❌ Reminder not inserted.")
                # Optional: Show failure message in UI
                if hasattr(self, 'parent') and self.parent:
                    QMessageBox.warning(self.parent, "Test Failed", "Reminder not inserted.")
                return False

        except Exception as e:
            print(f"❌ Test failed with error: {e}")
            # Optional: Show error message in UI
            if hasattr(self, 'parent') and self.parent:
                QMessageBox.critical(self.parent, "Test Error", f"Test failed with error: {e}")
            return False

    def send_email_handler(self):
        from_email = self.gmail_input.text().strip()
        password = self.password_input.text().strip()
        to_email = self.to_email_input.text().strip()
        cc_raw = self.cc_input.text().strip()
        subject = self.subject_input.text().strip()
        body = self.message_input.toPlainText().strip()

        cc_emails = [e.strip() for e in cc_raw.split(",") if e.strip()]
        all_recipients = [to_email] + cc_emails

        try:
            # Register app password ONCE (comment this out after first run)
            yagmail.register(from_email, password)

            # Login using app password
            yag = yagmail.SMTP(from_email)

            # Send email
            yag.send(
                to=all_recipients,
                subject=subject,
                contents=body
            )

            QMessageBox.information(self, "Success", "Email sent successfully!")

            # Save to DB
            conn = sqlite3.connect("medical_system.db")
            cursor = conn.cursor()

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS SentEmails (
                    email_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sender_email TEXT NOT NULL,
                    receiver_email TEXT NOT NULL,
                    cc_emails TEXT,
                    email_subject TEXT,
                    email_message TEXT,
                    sent_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """)

            cursor.execute("""
                INSERT INTO SentEmails (sender_email, receiver_email, cc_emails, email_subject, email_message)
                VALUES (?, ?, ?, ?, ?)
            """, (from_email, to_email, ', '.join(cc_emails), subject, body))

            conn.commit()
            conn.close()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to send email:\n{e}")


if __name__ == '__main__':
    import sys
    from PyQt5.QtWidgets import QApplication

    app = QApplication(sys.argv)
    window = TesterHomePage()
    window.show()
    sys.exit(app.exec_())
